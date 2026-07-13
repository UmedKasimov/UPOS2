from __future__ import annotations

import calendar
import json
import logging
import io
import math
import os
import re
import secrets
import socket
import uuid
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlencode, urlparse
from urllib.error import HTTPError, URLError
from urllib.request import Request as UrlRequest, urlopen
from zoneinfo import ZoneInfo

from contextlib import asynccontextmanager

from fastapi import BackgroundTasks, FastAPI, Form, Query, Request, UploadFile, File
from fastapi.exception_handlers import http_exception_handler
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, or_, select, text
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm.attributes import flag_modified
from starlette.exceptions import HTTPException as StarletteHTTPException
from starlette.middleware.sessions import SessionMiddleware

from upos.config import get_settings
from upos.csrf import csrf_matches_session, ensure_csrf_token, rotate_csrf_token
from upos.db import get_engine, init_db, list_public_tables, session_scope, startup_timings_ms
from upos.db_models import (
    Counterparty,
    CrmRecord,
    EmployeeAccountAccess,
    FinanceAccount,
    FinanceCategory,
    Product,
    PurchaseDocument,
    SaleDocument,
    TelegramChat,
    TelegramSubscriber,
    Transaction,
    User,
    UserAuthSession,
    Warehouse,
    WarehouseOperation,
)
from upos.billing_maintenance import (
    clear_workspace_database,
    clear_workspace_transactions,
    export_workspace_database,
    restore_workspace_database,
)
from upos.user_cache import get_cached as get_user_cached
from upos.i18n import (
    apply_locale_cookie,
    context_i18n,
    localize_treasury_templates,
    localized_timezone_groups,
    normalize_locale,
    resolve_locale,
    translate,
)
from upos.timezones import (
    curated_zone_ids,
    current_month_local_bounds_utc,
    normalize_workspace_timezone,
    period_local_bounds_utc,
)
from upos.storage import (
    load_legacy_settings,
    load_workspace_settings,
    save_legacy_settings,
    save_workspace_settings,
    valid_workspace_owner_id,
)
from upos.treasury_store import (
    TreasuryPostingError,
    aggregate_director_treasury,
    load_treasury,
    patch_display_currency,
    save_treasury,
    validate_and_clean_treasury,
)
from upos.treasury_templates import list_templates_public
from upos.auth_sessions_store import (
    block_auth_session,
    create_auth_session,
    ensure_auth_session_for_user,
    list_user_devices,
    revoke_auth_session,
    touch_auth_session,
    validate_auth_session,
)
from upos.transactions_store import (
    director_transaction_filter_hints,
    get_director_consolidated_pnl,
    is_report_locked_category,
    list_director_consolidated_transactions,
    list_transfers,
    list_transactions,
    create_transaction,
    resolve_pending_transfer,
    set_transaction_status,
    update_transaction,
    delete_transaction,
    list_categories,
    seed_default_categories,
    create_category,
    add_category_subcategory,
    update_category,
    delete_category,
    delete_category_subcategory,
    get_account_movements,
    get_pnl_data,
)
from PIL import Image
from upos.clopos_client import CloposError, test_clopos_connection
from upos.greenwhite_client import GreenWhiteError
from upos.greenwhite_store import (
    last_greenwhite_status,
    sync_greenwhite,
    test_greenwhite_connection,
)
from upos.integrations import (
    CLOPOS_DEFAULT_API_BASE_URL,
    INTEGRATION_PROVIDERS,
    integration_badges,
    integration_configured,
)
from upos.organizations_store import (
    create_organization,
    default_organization,
    ensure_organizations_for_existing_owners,
    get_organization,
    list_organizations,
    sync_common_settings,
    update_organization,
)
from upos.shipments_store import (
    COURIER_PAYMENT_CATEGORY,
    confirm_delivery_shipment,
    confirm_delivery_shipment_document,
    create_delivery_shipments,
    create_hr_employee,
    create_position,
    delete_delivery_shipment,
    delete_delivery_shipment_document,
    delete_position,
    delete_hr_employee_permanently,
    dismiss_hr_employee,
    list_courier_debts,
    list_delivery_shipments,
    list_hr_employees,
    list_positions,
    recompute_delivery_debts,
    restore_hr_employee,
    set_hr_attendance,
    set_hr_attendance_day,
    shipment_totals,
    update_delivery_shipment_document,
    update_hr_employee,
    update_position,
)
from upos.users_store import (
    ROLE_BUTTON_PERMISSION_LABELS,
    ROLE_PERMISSION_KEYS,
    ROLE_PERMISSION_LABELS,
    STAFF_ROLE_LABELS,
    add_billing_account,
    add_employee,
    bootstrap_from_env,
    delete_employee,
    ensure_all_user_ids,
    ensure_account_ids,
    ensure_default_roles,
    get_by_username,
    get_employee_for_owner,
    list_employee_organizations_safe,
    list_employees_safe,
    list_users_safe,
    list_roles_safe,
    migrate_legacy_superuser_flag,
    reset_employee_password,
    reset_billing_account_password,
    save_user_avatar,
    session_payload,
    set_billing_account_password,
    set_employee_frozen,
    update_employee,
    update_billing_account_name,
    update_role_permissions,
    update_self_account,
    update_user,
    user_count,
    verify_login,
)

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
templates.env.auto_reload = True
templates.env.cache = {}
CLIENT_WORKSPACES_DIR = BASE_DIR.parent / "client_workspaces"


def billing_root_credentials() -> tuple[str, str]:
    settings = get_settings()
    user = (settings.billing_root_user or settings.admin_basic_user or "root").strip()
    password = (settings.billing_root_password or settings.admin_basic_password or "root12345").strip()
    return user, password


def billing_logged_in(request: Request) -> bool:
    return bool((request.session.get("billing_root") or {}).get("ok"))


def client_workspace_path(account_key: str) -> Path:
    safe_id = "".join(ch for ch in str(account_key or "") if ch.isalnum() or ch == "-")
    return CLIENT_WORKSPACES_DIR / safe_id


def ensure_client_workspace(account_key: str) -> str:
    path = client_workspace_path(account_key)
    path.mkdir(parents=True, exist_ok=True)
    return str(path)


def format_bytes(size: int) -> str:
    n = float(max(0, int(size or 0)))
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if n < 1024 or unit == "TB":
            if unit == "B":
                return f"{int(n)} {unit}"
            return f"{n:.1f} {unit}"
        n /= 1024
    return f"{int(size or 0)} B"


def folder_size_bytes(path: Path) -> int:
    total = 0
    if not path.exists():
        return 0
    for item in path.rglob("*"):
        try:
            if item.is_file():
                total += item.stat().st_size
        except OSError:
            continue
    return total


WORKSPACE_SIZE_TABLES = (
    "organizations",
    "roles",
    "positions",
    "hr_employees",
    "hr_attendance",
    "delivery_shipments",
    "delivery_shipment_items",
    "workspace_settings",
    "treasuries",
    "transactions",
    "finance_accounts",
    "account_balances",
    "finance_categories",
    "counterparties",
    "branches",
    "transaction_entries",
    "external_records",
    "integration_sync_runs",
    "products",
    "warehouses",
    "sale_documents",
    "purchase_documents",
    "payment_documents",
    "expense_documents",
    "telegram_bot_configs",
    "telegram_chats",
    "telegram_subscribers",
    "telegram_delivery_log",
)


def db_account_size_bytes(owner_id: str) -> int:
    owner = str(owner_id or "").strip()
    if not owner:
        return 0
    total = 0
    with get_engine().connect() as conn:
        try:
            total += int(
                conn.execute(
                    text(
                        'SELECT COALESCE(SUM(pg_column_size(t)), 0) '
                        'FROM public."users" AS t '
                        "WHERE t.id = :owner OR t.employer_user_id = :owner",
                    ),
                    {"owner": owner},
                ).scalar()
                or 0,
            )
        except Exception:
            pass
        for table in WORKSPACE_SIZE_TABLES:
            try:
                total += int(
                    conn.execute(
                        text(
                            f'SELECT COALESCE(SUM(pg_column_size(t)), 0) '
                            f'FROM public."{table}" AS t '
                            "WHERE t.workspace_owner_id = :owner",
                        ),
                        {"owner": owner},
                    ).scalar()
                    or 0,
                )
            except Exception:
                continue
    return total


TELEPHONY_SETTING_LIST_KEYS = (
    "telephony_calls",
    "telephony_numbers",
    "telephony_providers",
    "telephony_devices",
    "telephony_recordings",
)


def telephony_account_dashboard(workspace_owner_id: str) -> dict[str, Any]:
    data = load_workspace_settings(workspace_owner_id)
    calls = data.get("telephony_calls") if isinstance(data.get("telephony_calls"), list) else []
    numbers = data.get("telephony_numbers") if isinstance(data.get("telephony_numbers"), list) else []
    providers = data.get("telephony_providers") if isinstance(data.get("telephony_providers"), list) else []
    devices = data.get("telephony_devices") if isinstance(data.get("telephony_devices"), list) else []
    recordings = data.get("telephony_recordings") if isinstance(data.get("telephony_recordings"), list) else []

    def recent_key(item: dict[str, Any]) -> str:
        return str(
            item.get("last_seen_at")
            or item.get("updated_at")
            or item.get("started_at")
            or item.get("created_at")
            or ""
        )

    def is_online(item: dict[str, Any]) -> bool:
        return str(item.get("status") or "").strip().lower() in {"online", "connected", "active"}

    answered = sum(1 for row in calls if str(row.get("status") or "") == "answered")
    missed = sum(1 for row in calls if str(row.get("status") or "") == "missed")
    incoming = sum(1 for row in calls if str(row.get("direction") or "") == "incoming")
    outgoing = sum(1 for row in calls if str(row.get("direction") or "") == "outgoing")
    connected_devices = sum(1 for row in devices if is_online(row))
    online_providers = sum(1 for row in providers if is_online(row))

    return {
        "total_calls": len(calls),
        "answered_calls": answered,
        "missed_calls": missed,
        "incoming_calls": incoming,
        "outgoing_calls": outgoing,
        "numbers_count": len(numbers),
        "providers_count": len(providers),
        "online_providers": online_providers,
        "devices_count": len(devices),
        "connected_devices": connected_devices,
        "recordings_count": len(recordings),
        "latest_calls": sorted([dict(row) for row in calls], key=recent_key, reverse=True)[:10],
        "devices": sorted([dict(row) for row in devices], key=recent_key, reverse=True)[:20],
        "recordings": sorted([dict(row) for row in recordings], key=recent_key, reverse=True)[:10],
    }


def billing_clients_context() -> list[dict[str, Any]]:
    clients = []
    for row in list_users_safe():
        if str(row.get("role") or "") == "user":
            account_key = str(row.get("account_id") or row.get("id") or "")
            owner_id = str(row.get("id") or "")
            clients.append({
                **row,
                "workspace_path": ensure_client_workspace(account_key),
                "telephony": telephony_account_dashboard(owner_id) if owner_id else {},
            })
    return clients


def billing_account_detail(account_key: str, *, temp_password: str = "") -> dict[str, Any] | None:
    key = str(account_key or "").strip()
    if not key:
        return None
    with session_scope() as session:
        owner = session.scalar(
            select(User).where(
                User.role == "user",
                User.employer_user_id.is_(None),
                or_(User.account_id == key, User.id == key),
            ),
        )
        if owner is None:
            return None
        users = session.scalars(
            select(User)
            .where(or_(User.id == owner.id, User.employer_user_id == owner.id))
            .order_by(User.employer_user_id.is_not(None), func.lower(User.username)),
        ).all()
        ids = [str(row.id) for row in users]
        last_seen_map: dict[str, datetime | None] = {}
        if ids:
            for user_id, seen_at in session.execute(
                select(UserAuthSession.user_id, func.max(UserAuthSession.last_seen_at))
                .where(UserAuthSession.user_id.in_(ids))
                .group_by(UserAuthSession.user_id),
            ).all():
                last_seen_map[str(user_id)] = seen_at
        user_rows = [
            {
                "id": row.id,
                "account_id": row.account_id or "",
                "username": row.username,
                "name": row.name or row.username,
                "email": row.email or "",
                "role": row.role or "user",
                "is_owner": row.id == owner.id,
                "last_seen_at": last_seen_map.get(str(row.id)).isoformat() if last_seen_map.get(str(row.id)) else "",
            }
            for row in users
        ]
        last_seen_values = [last_seen_map[str(row.id)] for row in users if last_seen_map.get(str(row.id))]
        owner_payload = {
            "id": owner.id,
            "account_id": owner.account_id or owner.id,
            "username": owner.username,
            "name": owner.name or owner.username,
            "email": owner.email or "",
        }

    account_id = str(owner_payload["account_id"])
    workspace_path = Path(ensure_client_workspace(account_id))
    folder_bytes = folder_size_bytes(workspace_path)
    db_bytes = db_account_size_bytes(str(owner_payload["id"]))
    total_bytes = folder_bytes + db_bytes
    return {
        "client": owner_payload,
        "users": user_rows,
        "last_seen_at": max(last_seen_values).isoformat() if last_seen_values else "",
        "workspace_path": str(workspace_path),
        "folder_bytes": folder_bytes,
        "db_bytes": db_bytes,
        "total_bytes": total_bytes,
        "folder_size": format_bytes(folder_bytes),
        "db_size": format_bytes(db_bytes),
        "total_size": format_bytes(total_bytes),
        "temp_password": temp_password,
        "telephony": telephony_account_dashboard(str(owner_payload["id"])),
    }


def safe_internal_path(raw: str | None, default: str = "/") -> str:
    """Только известные внутренние пути — иначе произвольный `next` после входа даёт 404 JSON."""
    if raw is None or not isinstance(raw, str):
        return default
    p = raw.strip()
    if not p.startswith("/") or p.startswith("//"):
        return default
    path_only = p.split("?", 1)[0].split("#", 1)[0]
    if path_only == "/":
        return p if ("?" in p or "#" in p) else "/"
    if path_only == "/settings" or path_only.startswith("/settings/"):
        return p
    if path_only == "/employees" or path_only.startswith("/employees/"):
        return p
    if path_only == "/organizations" or path_only.startswith("/organizations/"):
        return p
    if path_only == "/schet" or path_only.startswith("/schet/"):
        return p
    if path_only == "/kassa" or path_only.startswith("/kassa/"):
        return p
    if path_only == "/reports" or path_only.startswith("/reports/"):
        return p
    if path_only == "/shipments" or path_only.startswith("/shipments/"):
        return p
    if path_only == "/hr" or path_only.startswith("/hr/"):
        return p
    if path_only == "/admin" or path_only.startswith("/admin/"):
        return p
    return default


def next_query(request: Request) -> str:
    n = request.query_params.get("next") or request.query_params.get("callbackUrl")
    return safe_internal_path(n, "/schet")


def _admin_post_login_path(target: str) -> str:
    """Учётка администратора не использует пользовательское рабочее место (/ и /settings)."""
    path_only = target.split("?", 1)[0].split("#", 1)[0]
    if path_only == "/admin/users" or path_only.startswith("/admin/users/"):
        return "/admin"
    if (
        path_only == "/"
        or path_only == "/schet"
        or path_only == "/kassa"
        or path_only == "/reports"
        or path_only == "/shipments"
        or path_only == "/hr"
        or path_only == "/organizations"
        or path_only == "/settings"
        or path_only.startswith("/organizations/")
        or path_only.startswith("/settings/")
    ):
        return "/admin"
    if path_only.startswith("/admin"):
        return target
    return "/admin"


def post_login_redirect(user: dict, next_raw: str | None) -> str:
    if user.get("role") == "admin":
        default = "/admin"
    elif user.get("is_employee"):
        perms = user.get("employee_permissions") if isinstance(user.get("employee_permissions"), dict) else {}
        default = (
            "/schet"
            if perms.get("schet")
            else "/kassa"
            if perms.get("kassa")
            else "/reports"
            if perms.get("reports")
            else "/settings?tab=employees"
            if perms.get("employees")
            else "/settings"
            if perms.get("settings")
            else "/auth"
        )
    else:
        default = "/schet"
    target = safe_internal_path(next_raw, default)
    if user.get("role") == "admin":
        return _admin_post_login_path(target)
    return target


def create_app() -> FastAPI:
    settings = get_settings()

    @asynccontextmanager
    async def lifespan(app: FastAPI):
        from upos.telegram_routes import reregister_webhooks_later
        from upos.telegram_scheduler import start_scheduler, stop_scheduler
        from upos.telegram_store import list_active_configs

        try:
            if list_active_configs():
                start_scheduler()
                reregister_webhooks_later()
            else:
                logger.info("[telegram] no active bots — scheduler and webhook refresh skipped")
        except Exception:
            logger.exception(
                "[telegram] startup skipped due to error (check DB migrations / notification_prefs column)",
            )
        yield
        stop_scheduler()

    app = FastAPI(title="UPOS FINANCE", version="0.3", lifespan=lifespan)

    @app.exception_handler(StarletteHTTPException)
    async def starlette_http_handler(request: Request, exc: StarletteHTTPException):
        if exc.status_code == 404:
            accept = request.headers.get("accept", "")
            if "text/html" in accept:
                loc = resolve_locale(request, request.session.get("user"))
                ctx_nf = {
                    "request": request,
                    "path": request.url.path,
                    "body_theme": "light",
                    "error_octopus_id": "oct404",
                    **context_i18n(loc),
                }
                return templates.TemplateResponse(
                    request,
                    "not_found.html",
                    ctx_nf,
                    status_code=404,
                )
        return await http_exception_handler(request, exc)

    if settings.auth_secret in ("", "dev-only-change-me"):
        logger.warning(
            "[upos] Задайте AUTH_SECRET в .env.local (секрет для подписи cookie-сессии).",
        )

    init_db()

    bootstrap_from_env(
        settings.admin_basic_user or "",
        settings.admin_basic_password or "",
        settings.admin_display_name or "Администратор",
    )
    migrate_legacy_superuser_flag()
    ensure_all_user_ids()
    ensure_account_ids()
    ensure_organizations_for_existing_owners()

    app.mount(
        "/static",
        StaticFiles(directory=str(BASE_DIR / "static")),
        name="static",
    )

    ORG_GENERAL_VALUE = "__general__"

    def _is_director(user: dict | None) -> bool:
        u = user or {}
        return bool(str(u.get("role") or "") == "user" and not u.get("is_employee"))

    def _session_user_id(user: dict | None) -> str:
        u = user or {}
        return str(u.get("user_id") or u.get("id") or "").strip()

    def _visible_employee_id(user: dict | None) -> str | None:
        u = user or {}
        if not u.get("is_employee"):
            return None
        if _employee_role_key(u) == "general_director":
            return None
        uid = _session_user_id(u)
        return uid or None

    def _employee_role_key(user: dict | None) -> str:
        return str((user or {}).get("employee_role_key") or "").strip()

    def _employee_permissions(user: dict | None) -> dict[str, bool]:
        raw = (user or {}).get("employee_permissions")
        src = raw if isinstance(raw, dict) else {}
        out = {key: bool(src.get(key)) for key in ROLE_PERMISSION_KEYS}
        if "shipments" not in src:
            out["shipments"] = bool(src.get("kassa") or src.get("reports"))
        if "hr" not in src:
            out["hr"] = bool(src.get("employees"))
        return out

    def _employee_button_access(user: dict | None) -> dict[str, dict[str, bool]]:
        raw = (user or {}).get("employee_permissions")
        perms = raw if isinstance(raw, dict) else {}
        raw_button_access = perms.get("button_access")
        button_src = raw_button_access if isinstance(raw_button_access, dict) else {}
        access: dict[str, dict[str, bool]] = {}
        for section, actions in ROLE_BUTTON_PERMISSION_LABELS.items():
            raw_section = button_src.get(section)
            section_src = raw_section if isinstance(raw_section, dict) else {}
            access[section] = {
                action: bool(section_src[action]) if action in section_src else True
                for action in actions
            }
        return access

    def _employee_category_access(user: dict | None) -> dict[str, Any] | None:
        u = user or {}
        if not u.get("is_employee"):
            return None
        raw_perms = u.get("employee_permissions")
        perms = raw_perms if isinstance(raw_perms, dict) else {}
        raw_access = perms.get("category_access")
        access = raw_access if isinstance(raw_access, dict) else {}
        if not access.get("enabled"):
            return None
        category_ids = [
            str(x).strip()
            for x in access.get("category_ids", [])
            if str(x or "").strip()
        ]
        subcategories_raw = access.get("subcategories")
        subcategories_src = subcategories_raw if isinstance(subcategories_raw, dict) else {}
        subcategories = {
            str(cat_id).strip(): [
                str(name).strip()
                for name in values
                if str(name or "").strip()
            ]
            for cat_id, values in subcategories_src.items()
            if str(cat_id or "").strip() and isinstance(values, list)
        }
        return {
            "category_ids": category_ids,
            "subcategories": subcategories,
        }

    def _filter_categories_for_user(user: dict | None, categories: list[dict[str, Any]]) -> list[dict[str, Any]]:
        access = _employee_category_access(user)
        if access is None:
            return categories
        allowed = set(access.get("category_ids") or [])
        return [cat for cat in categories if str(cat.get("id") or "") in allowed]

    def _category_filters_for_user(user: dict | None, workspace_owner_id: str) -> tuple[list[str] | None, list[str] | None]:
        access = _employee_category_access(user)
        if access is None:
            return None, None
        allowed = set(access.get("category_ids") or [])
        categories = list_categories(workspace_owner_id)
        rows = [cat for cat in categories if str(cat.get("id") or "") in allowed]
        return (
            [str(cat.get("id") or "") for cat in rows if str(cat.get("id") or "")],
            [str(cat.get("name") or "") for cat in rows if str(cat.get("name") or "")],
        )

    def _category_payload_allowed(user: dict | None, workspace_owner_id: str, payload: dict[str, Any]) -> bool:
        access = _employee_category_access(user)
        if access is None:
            return True
        allowed_ids = set(access.get("category_ids") or [])
        cat_id = str(payload.get("category_id") or "").strip()
        cat_name = str(payload.get("category") or "").strip()
        if cat_id:
            return cat_id in allowed_ids
        if not cat_name:
            return True
        categories = list_categories(workspace_owner_id)
        return any(str(cat.get("id") or "") in allowed_ids and str(cat.get("name") or "") == cat_name for cat in categories)

    def _has_permission(user: dict | None, key: str) -> bool:
        u = user or {}
        if not u:
            return False
        if str(u.get("role") or "") == "admin":
            return False
        if not u.get("is_employee"):
            return str(u.get("role") or "") == "user"
        return bool(_employee_permissions(u).get(key))

    def _has_button_permission(user: dict | None, section: str, action: str) -> bool:
        u = user or {}
        if not u:
            return False
        if str(u.get("role") or "") == "admin":
            return False
        if not u.get("is_employee"):
            return str(u.get("role") or "") == "user"
        if not _has_permission(u, section):
            return False
        return bool(_employee_button_access(u).get(section, {}).get(action, True))

    def _has_any_permission(user: dict | None, keys: tuple[str, ...]) -> bool:
        return any(_has_permission(user, key) for key in keys)

    def _is_employee_general_director(user: dict | None) -> bool:
        return bool((user or {}).get("is_employee") and _employee_role_key(user) == "general_director")

    def _is_employee_adminish(user: dict | None) -> bool:
        return bool(_employee_role_key(user) in {"general_director", "administrator", "hr_manager"})

    def _first_allowed_user_path(user: dict | None) -> str:
        if _has_permission(user, "schet"):
            return "/schet"
        if _has_permission(user, "kassa"):
            return "/kassa"
        if _has_permission(user, "reports"):
            return "/reports"
        if _has_permission(user, "adjustments"):
            return "/adjustments"
        if _has_permission(user, "shipments"):
            return "/shipments"
        if _has_permission(user, "hr"):
            return "/hr"
        if _has_permission(user, "employees"):
            return "/settings?tab=employees"
        if _has_permission(user, "dictionary"):
            return "/settings?tab=dictionary"
        if _has_permission(user, "settings"):
            return "/settings"
        return "/auth"

    def _path_permission_requirement(path: str, method: str) -> tuple[str, ...]:
        m = method.upper()
        if path == "/schet" or path.startswith("/schet/"):
            return ("schet",)
        if path == "/kassa" or path.startswith("/kassa/"):
            return ("kassa",)
        if path == "/reports" or path.startswith("/reports/"):
            return ("reports",)
        if path == "/adjustments" or path.startswith("/adjustments/"):
            return ("adjustments",)
        if path == "/shipments" or path.startswith("/shipments/"):
            return ("shipments",)
        if path == "/hr" or path.startswith("/hr/"):
            return ("hr",)
        if path == "/employees" or path.startswith("/employees/"):
            return ("employees",)
        if path == "/settings" or path.startswith("/settings/"):
            return ("settings", "employees", "dictionary")
        if path.startswith("/api/employees"):
            return ("employees",)
        if path.startswith("/api/hr"):
            return ("hr",)
        if path.startswith("/api/adjustments"):
            return ("adjustments",)
        if path.startswith("/api/categories"):
            return ("settings", "dictionary")
        if path == "/api/settings/preferences":
            return ("settings", "dictionary")
        if path.startswith("/api/settings"):
            return ("settings",)
        if path.startswith("/api/kassa/sms-report"):
            return ("kassa",)
        if path.startswith("/api/reports"):
            return ("reports",)
        if path.startswith("/api/wallets"):
            return ("kassa", "schet", "reports") if m == "GET" else ("schet",)
        if path == "/api/treasury":
            return ("kassa", "schet", "reports") if m == "GET" else ("schet",)
        if path.startswith("/api/transactions"):
            return ("kassa", "reports") if m == "GET" else ("kassa",)
        if path.startswith("/api/transfers"):
            return ("kassa", "schet")
        return ()

    def _button_permission_requirement(path: str, method: str) -> tuple[str, str] | None:
        if method.upper() not in {"POST", "PUT", "PATCH", "DELETE"}:
            return None
        clean = path.rstrip("/") or "/"
        if clean == "/api/kassa/sms-report":
            return ("kassa", "sms_report")
        if clean == "/api/transactions":
            return ("kassa", "create")
        if clean.startswith("/api/transactions/"):
            if clean.endswith("/status"):
                return ("kassa", "confirm")
            if method.upper() == "DELETE":
                return ("kassa", "delete")
            return ("kassa", "edit")
        if clean.startswith("/api/transfers"):
            return ("kassa", "transfer")
        if clean == "/api/wallets":
            return ("schet", "create")
        if clean.startswith("/api/wallets/"):
            return ("schet", "edit")
        if clean == "/api/treasury":
            return ("schet", "balance")
        if clean.startswith("/api/adjustments"):
            return ("adjustments", "save")
        if clean in {"/shipments/create", "/organizations/shipments/create"}:
            return ("shipments", "create")
        if "/shipments/day/" in clean and clean.endswith("/update"):
            return ("shipments", "save")
        if "/shipments/day/" in clean and clean.endswith("/confirm"):
            return ("shipments", "confirm")
        if "/shipments/day/" in clean and clean.endswith("/delete"):
            return ("shipments", "delete")
        if "/shipments/shipment/" in clean and clean.endswith("/confirm"):
            return ("shipments", "confirm")
        if "/shipments/shipment/" in clean and clean.endswith("/delete"):
            return ("shipments", "delete")
        if "/hr/employees/create" in clean:
            return ("hr", "create")
        if "/hr/employees/update" in clean:
            return ("hr", "edit")
        if "/hr/employees/" in clean and clean.endswith("/dismiss"):
            return ("hr", "dismiss")
        if "/hr/employees/" in clean and clean.endswith("/restore"):
            return ("hr", "restore")
        if "/hr/employees/" in clean and clean.endswith("/delete"):
            return ("hr", "delete")
        if clean.endswith("/hr/attendance/report"):
            return ("hr", "attendance_report")
        if clean.endswith("/hr/attendance"):
            return ("hr", "attendance")
        if clean.endswith("/hr/salary-act/save"):
            return ("hr", "salary_act")
        if clean.endswith("/hr/salary-adjustment/save"):
            return ("hr", "salary_adjustment")
        if clean == "/employees/create":
            return ("employees", "create")
        if clean == "/employees/update":
            return ("employees", "edit")
        if clean == "/employees/delete":
            return ("employees", "delete")
        if clean.startswith("/api/employees/") and clean.endswith("/reset-password"):
            return ("employees", "reset_password")
        if clean.startswith("/api/employees/") and "/devices" in clean:
            return ("employees", "devices")
        if clean == "/api/settings/roles":
            return ("settings", "roles")
        if clean == "/api/settings/integrations":
            return ("settings", "integrations")
        if clean == "/api/settings/preferences":
            return ("settings", "preferences")
        if clean == "/api/settings/account":
            return ("settings", "profile")
        if clean == "/api/categories/create":
            return ("dictionary", "category_create")
        if clean == "/api/categories/update":
            return ("dictionary", "category_edit")
        if clean == "/api/categories/delete":
            return ("dictionary", "category_delete")
        if clean.startswith("/api/categories/subcategories/"):
            return ("dictionary", "subcategory")
        return None

    def _can_manage_wallets(user: dict | None) -> bool:
        u = user or {}
        return bool(
            u
            and str(u.get("role") or "") != "admin"
            and (not u.get("is_employee") or (_is_employee_adminish(u) and _has_permission(u, "schet")))
            and u.get("org_scope") != "general"
        )

    def _can_manage_employees(user: dict | None) -> bool:
        u = user or {}
        return bool(
            u
            and str(u.get("role") or "") != "admin"
            and (not u.get("is_employee") or (_is_employee_adminish(u) and _has_permission(u, "employees")))
        )

    def _can_modify_transactions(user: dict | None) -> bool:
        u = user or {}
        if not u.get("is_employee"):
            return True
        return bool(_has_permission(u, "kassa") and _employee_role_key(u) != "accountant")

    def _employee_can_access_transaction(user: dict | None, workspace_owner_id: str, tx_id: str) -> bool:
        employee_id = _visible_employee_id(user)
        if not employee_id:
            return True
        category_ids, category_names = _category_filters_for_user(user, workspace_owner_id)
        return any(
            str(tx.get("id") or "") == str(tx_id)
            for tx in list_transactions(
                workspace_owner_id,
                limit=5000,
                visible_employee_id=employee_id,
                allowed_category_ids=category_ids,
                allowed_category_names=category_names,
            )
        )

    def _devices_owner_only(request: Request) -> tuple[dict | None, JSONResponse | None]:
        user = request.session.get("user") or {}
        if not _is_director(user):
            return None, JSONResponse({"error": "forbidden"}, status_code=403)
        uid = _session_user_id(user)
        if not uid:
            return None, JSONResponse({"error": "unauthorized"}, status_code=401)
        return user, None

    def _director_employee_or_error(
        request: Request,
        employee_id: str,
    ) -> tuple[dict[str, Any] | None, JSONResponse | None]:
        user, err = _devices_owner_only(request)
        if err:
            return None, err
        assert user is not None
        owner_id = str(user.get("account_owner_id") or user.get("user_id") or "").strip()
        active_org_id = str(
            user.get("active_org_id") or user.get("workspace_owner_id") or owner_id,
        )
        emp = get_employee_for_owner(owner_id, employee_id, active_org_id)
        if not emp:
            return None, JSONResponse({"error": "not_found"}, status_code=404)
        return emp, None

    def _can_manage_telegram(user: dict | None) -> bool:
        return _is_director(user)

    def _attach_organization_context(
        payload: dict[str, object],
        previous: dict | None = None,
    ) -> dict[str, object]:
        prev = previous or {}
        if str(payload.get("role") or "") == "admin":
            return payload

        owner_id = str(
            payload.get("account_owner_id")
            or payload.get("workspace_owner_id")
            or payload.get("user_id")
            or "",
        ).strip()
        if not valid_workspace_owner_id(owner_id):
            return payload

        payload["account_owner_id"] = owner_id
        if payload.get("is_employee"):
            requested_org_id = str(prev.get("active_org_id") or payload.get("organization_id") or "").strip()
            employee_id = str(payload.get("user_id") or "").strip()
            employee_orgs = list_employee_organizations_safe(owner_id, employee_id) if employee_id else []
            org = next((item for item in employee_orgs if str(item.get("id")) == requested_org_id), None)
            org = org or (employee_orgs[0] if employee_orgs else None) or default_organization(owner_id)
            if org:
                roles = ensure_default_roles(str(org["id"]))
                rid = str(payload.get("employee_role_id") or "").strip()
                role = next((row for row in roles if str(row.get("id") or "") == rid), None)
                if not role:
                    role = next(
                        (
                            row
                            for row in roles
                            if str(row.get("key") or "") == str(payload.get("employee_role_key") or "")
                        ),
                        None,
                    )
                if role:
                    payload["employee_role_id"] = str(role.get("id") or "")
                    payload["employee_role_key"] = str(role.get("key") or "")
                    payload["employee_role_name"] = str(role.get("name") or "")
                    payload["employee_permissions"] = role.get("permissions") or {}
            if _is_employee_general_director(payload):
                orgs = list_organizations(owner_id)
                fallback = next((item for item in orgs if item.get("is_default")), None) or (orgs[0] if orgs else None)
                previous_active_id = str(prev.get("active_org_id") or "").strip()
                active = next((item for item in orgs if str(item.get("id")) == previous_active_id), None) or org or fallback
                if active:
                    payload["workspace_owner_id"] = str(active["id"])
                    payload["active_org_id"] = str(active["id"])
                    payload["active_organization_name"] = str(active.get("name") or "")
                payload["organizations"] = orgs
                payload["org_scope"] = "general" if prev.get("org_scope") == "general" else "organization"
                payload["can_switch_organizations"] = True
                payload["organization_switch_general_value"] = ORG_GENERAL_VALUE
                return payload
            if org:
                payload["workspace_owner_id"] = str(org["id"])
                payload["active_org_id"] = str(org["id"])
                payload["active_organization_name"] = str(org.get("name") or "")
            payload["organizations"] = employee_orgs
            payload["org_scope"] = "organization"
            payload["can_switch_organizations"] = len(employee_orgs) > 1
            payload["organization_switch_general_value"] = ORG_GENERAL_VALUE
            return payload

        if str(payload.get("role") or "") != "user":
            return payload

        orgs = list_organizations(owner_id)
        fallback = next((org for org in orgs if org.get("is_default")), None) or (orgs[0] if orgs else None)
        previous_active_id = str(prev.get("active_org_id") or "").strip()
        active = next((org for org in orgs if str(org.get("id")) == previous_active_id), None) or fallback
        if active:
            payload["workspace_owner_id"] = str(active["id"])
            payload["active_org_id"] = str(active["id"])
            payload["active_organization_name"] = str(active.get("name") or "")
        payload["organizations"] = orgs
        payload["org_scope"] = "general" if prev.get("org_scope") == "general" else "organization"
        payload["can_switch_organizations"] = True
        payload["organization_switch_general_value"] = ORG_GENERAL_VALUE
        return payload

    def _general_scope_blocked(path: str) -> bool:
        if path in {
            "/",
            "/organizations",
            "/organizations/settings",
            "/settings",
            "/api/csrf-token",
            "/api/settings/preferences",
            "/api/settings/account",
            "/api/settings/roles",
            "/api/user/avatar",
        }:
            return False
        if path.startswith("/organizations/"):
            return False
        if path == "/employees" or path.startswith("/employees/"):
            return False
        if path.startswith("/api/employees"):
            return False
        if path.startswith("/api/me/devices"):
            return False
        if path.startswith("/api/telegram"):
            return False
        if path.startswith("/api/hr/"):
            return False
        if path == "/api/director/consolidated-transactions":
            return False
        if path == "/api/director/consolidated-treasury":
            return False
        if path == "/api/director/organization-wallets":
            return False
        if path == "/api/director/consolidated-pnl":
            return False
        return True

    @app.middleware("http")
    async def require_login(request: Request, call_next):
        path = request.url.path
        if (
            path.startswith("/static")
            or path in {"/favicon.ico", "/health", "/health/db"}
            or path.startswith("/auth")
            or path.startswith("/billing")
            or path.startswith("/api/telegram/webhook/")
            or path == "/api/telephony/calls/ingest"
            or path == "/api/telephony/contacts/sync"
            or path == "/api/telephony/recordings/upload"
            or path == "/api/telephony/config"
            or path == "/api/telephony/softphone/login"
        ):
            return await call_next(request)
        user = request.session.get("user")
        if not user:
            nxt = quote(path + (f"?{request.url.query}" if request.url.query else ""))
            return RedirectResponse(url=f"/auth?next={nxt}", status_code=302)

        db_rec = None
        if user.get("username"):
            db_rec = get_user_cached(str(user["username"]), get_by_username)
            if db_rec and db_rec.get("employer_user_id") and db_rec.get("is_frozen"):
                request.session.clear()
                if path.startswith("/api/"):
                    return JSONResponse({"error": "employee_frozen"}, status_code=403)
                return RedirectResponse(url="/auth?error=frozen", status_code=302)
            if db_rec:
                request.session["user"] = _attach_organization_context(session_payload(db_rec), user)
                user = request.session["user"]

        if path.startswith("/admin"):
            if not db_rec or str(db_rec.get("role") or "") != "admin":
                su = request.session.get("user") or {}
                wid = su.get("workspace_owner_id") or su.get("user_id")
                theme = (
                    (load_workspace_settings(str(wid)).get("theme") or "light")
                    if wid
                    else (load_legacy_settings().get("theme") or "light")
                )
                loc = resolve_locale(request, user)
                return templates.TemplateResponse(
                    request,
                    "forbidden.html",
                    {
                        "request": request,
                        "user": user,
                        "body_theme": theme,
                        "variant": "user",
                        "active": "",
                        "csrf_token": ensure_csrf_token(request),
                        **context_i18n(loc),
                    },
                    status_code=403,
                )
        elif db_rec and str(db_rec.get("role") or "") == "admin":
            if (
                path in {"/", "/employees", "/schet", "/kassa"}
                or path == "/reports"
                or path == "/organizations"
                or path.startswith("/settings/")
                or path.startswith("/employees/")
                or path.startswith("/organizations/")
                or path == "/settings"
            ):
                return RedirectResponse(url="/admin", status_code=302)

        user = request.session.get("user") or {}
        if (
            _is_director(user)
            and user.get("org_scope") == "general"
            and _general_scope_blocked(path)
        ):
            if path.startswith("/api/"):
                return JSONResponse({"error": "organization_required"}, status_code=409)
            return RedirectResponse(url="/organizations", status_code=302)

        if user.get("is_employee"):
            required_permissions = _path_permission_requirement(path, request.method)
            if required_permissions and not _has_any_permission(user, required_permissions):
                if path.startswith("/api/"):
                    return JSONResponse({"error": "forbidden"}, status_code=403)
                return RedirectResponse(url=_first_allowed_user_path(user), status_code=302)
            button_permission = _button_permission_requirement(path, request.method)
            if button_permission and not _has_button_permission(user, button_permission[0], button_permission[1]):
                if path.startswith("/api/"):
                    return JSONResponse({"error": "button_forbidden"}, status_code=403)
                return RedirectResponse(url=_first_allowed_user_path(user), status_code=302)

        uid = _session_user_id(user)
        if uid and not path.startswith("/static"):
            auth_sid = str(request.session.get("auth_session_id") or "").strip()
            if not auth_sid:
                auth_sid = ensure_auth_session_for_user(uid, request, None)
                request.session["auth_session_id"] = auth_sid
            else:
                ok, _reason = validate_auth_session(auth_sid, uid)
                if not ok:
                    request.session.clear()
                    if path.startswith("/api/"):
                        return JSONResponse({"error": "session_invalid"}, status_code=401)
                    return RedirectResponse(url="/auth?error=session", status_code=302)
                if path not in {"/auth/logout"} and not path.startswith("/api/me/devices"):
                    touch_auth_session(auth_sid, uid)

        return await call_next(request)

    def tpl(request: Request, name: str, **extra: object):
        sess_user = request.session.get("user")
        ctx = {
            "request": request,
            "user": sess_user,
            "body_theme": "light",
        }
        if sess_user:
            if extra.get("variant") == "admin" or sess_user.get("role") == "admin":
                ctx["body_theme"] = load_legacy_settings().get("theme") or "light"
            else:
                wid = (
                    sess_user.get("account_owner_id")
                    if sess_user.get("org_scope") == "general"
                    else None
                ) or sess_user.get("workspace_owner_id") or sess_user.get("user_id")
                ctx["body_theme"] = "light"
                if wid:
                    try:
                        ctx["body_theme"] = load_workspace_settings(str(wid)).get("theme") or "light"
                    except Exception:
                        logger.exception(
                            "[upos] load_workspace_settings failed in template context (theme); wid=%s",
                            wid,
                        )
        ctx.update(extra)
        ctx.setdefault("role_button_permission_labels", ROLE_BUTTON_PERMISSION_LABELS)
        ctx["can_use_button"] = lambda section, action: _has_button_permission(sess_user, section, action)
        ctx["csrf_token"] = ensure_csrf_token(request)
        ctx.update(context_i18n(resolve_locale(request, sess_user)))
        return templates.TemplateResponse(request, name, ctx)

    @app.get("/health")
    def health():
        timings = startup_timings_ms()
        payload: dict[str, object] = {"ok": True}
        if timings:
            payload["startup_ms"] = timings
        return payload

    @app.get("/api/csrf-token")
    def api_csrf_token(request: Request):
        """Актуальный CSRF из cookie-сессии (meta на странице может устареть после входа в другой вкладке)."""
        if not request.session.get("user"):
            return JSONResponse({"error": "unauthorized"}, status_code=401)
        return {"csrf_token": ensure_csrf_token(request)}

    @app.get("/health/db")
    def health_db():
        """Диагностика: приложение действительно подключено к той же БД (таблицы в `public`)."""
        expected = frozenset(
            {
                "users",
                "roles",
                "organizations",
                "employee_organizations",
                "employee_account_access",
                "hr_employees",
                "hr_attendance",
                "delivery_shipments",
                "delivery_shipment_items",
                "workspace_settings",
                "treasuries",
                "global_settings",
                "transactions",
                "finance_accounts",
                "account_balances",
                "finance_categories",
                "counterparties",
                "branches",
                "transaction_entries",
                "external_records",
                "integration_sync_runs",
                "products",
                "warehouses",
                "sale_documents",
                "purchase_documents",
                "payment_documents",
                "expense_documents",
                "telegram_bot_configs",
                "telegram_chats",
                "telegram_subscribers",
                "telegram_delivery_log",
            },
        )
        try:
            eng = get_engine()
            with eng.connect() as conn:
                conn.execute(text("SELECT 1"))
                dbname_raw = conn.execute(text("SELECT current_database()")).scalar_one()
            dbname = str(dbname_raw)
            tables = list_public_tables()
            have = expected.issubset(set(tables))
            timings = startup_timings_ms()
            out: dict[str, object] = {
                "ok": have,
                "database": dbname,
                "schema": "public",
                "tables": tables,
            }
            if timings:
                out["startup_ms"] = timings
            return out
        except Exception as exc:
            logger.exception("[upos] /health/db failed")
            return JSONResponse(
                {"ok": False, "error": type(exc).__name__, "detail": str(exc)},
                status_code=503,
            )

    @app.get("/auth", response_class=HTMLResponse)
    def auth_page(request: Request):
        if request.session.get("user"):
            u = request.session["user"]
            nxt = request.query_params.get("next") or request.query_params.get("callbackUrl") or ""
            return RedirectResponse(url=post_login_redirect(u, nxt), status_code=302)
        err = request.query_params.get("error")
        nu = next_query(request)
        loc = resolve_locale(request, None)
        return templates.TemplateResponse(
            request,
            "auth.html",
            {
                "request": request,
                "error": err,
                "next_url": nu,
                "csrf_token": ensure_csrf_token(request),
                **context_i18n(loc),
            },
        )

    @app.post("/auth/login")
    def login(
        request: Request,
        csrf_token: str = Form(default=""),
        account_id: str = Form(...),
        login: str = Form(...),
        password: str = Form(...),
        next_url: str = Form(default="/schet"),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/auth?error=csrf", status_code=302)
        if user_count() == 0:
            return RedirectResponse(
                url="/auth?error=no_users",
                status_code=302,
            )
        payload = verify_login(login, password, account_id)
        if not payload:
            return RedirectResponse(url="/auth?error=credentials", status_code=302)
        request.session["user"] = payload
        uid = _session_user_id(payload)
        if uid:
            request.session["auth_session_id"] = create_auth_session(uid, request)
        rotate_csrf_token(request)
        return RedirectResponse(url=post_login_redirect(payload, next_url), status_code=302)

    @app.get("/billing", response_class=HTMLResponse)
    def billing_page(request: Request):
        if not billing_logged_in(request):
            return RedirectResponse(url="/billing/login", status_code=302)
        loc = resolve_locale(request, None)
        detail_key = request.query_params.get("detail") or ""
        return templates.TemplateResponse(
            request,
            "billing.html",
            {
                "request": request,
                "error": request.query_params.get("error") or "",
                "ok": request.query_params.get("ok") or "",
                "created": None,
                "clients": billing_clients_context(),
                "detail": billing_account_detail(detail_key),
                "csrf_token": ensure_csrf_token(request),
                **context_i18n(loc),
            },
        )

    @app.get("/billing/login", response_class=HTMLResponse)
    def billing_login_page(request: Request):
        if billing_logged_in(request):
            return RedirectResponse(url="/billing", status_code=302)
        loc = resolve_locale(request, None)
        return templates.TemplateResponse(
            request,
            "billing_login.html",
            {
                "request": request,
                "error": request.query_params.get("error") or "",
                "csrf_token": ensure_csrf_token(request),
                **context_i18n(loc),
            },
        )

    @app.post("/billing/login")
    def billing_login(
        request: Request,
        csrf_token: str = Form(default=""),
        login: str = Form(...),
        password: str = Form(...),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/billing/login?error=csrf", status_code=302)
        root_user, root_password = billing_root_credentials()
        if login.strip() != root_user or password.strip() != root_password:
            return RedirectResponse(url="/billing/login?error=credentials", status_code=302)
        request.session["billing_root"] = {"ok": True, "login": root_user}
        rotate_csrf_token(request)
        return RedirectResponse(url="/billing", status_code=302)

    @app.post("/billing/logout")
    def billing_logout(request: Request, csrf_token: str = Form(default="")):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/billing?error=csrf", status_code=302)
        request.session.pop("billing_root", None)
        rotate_csrf_token(request)
        return RedirectResponse(url="/billing/login", status_code=302)

    @app.post("/billing/create", response_class=HTMLResponse)
    def billing_create(
        request: Request,
        csrf_token: str = Form(default=""),
        username: str = Form(...),
        password: str = Form(...),
        name: str = Form(default=""),
        email: str = Form(default=""),
    ):
        if not billing_logged_in(request):
            return RedirectResponse(url="/billing/login", status_code=302)
        loc = resolve_locale(request, None)
        error = ""
        created = None
        if not csrf_matches_session(request, csrf_token):
            error = "Сессия формы устарела. Обновите страницу и попробуйте снова."
        else:
            ok, msg, rec = add_billing_account(username, password, name, email)
            if ok:
                created = rec
                ensure_client_workspace(str(rec.get("account_id") or rec.get("id") or ""))
                rotate_csrf_token(request)
            else:
                error = msg
        detail = billing_account_detail(str(created.get("account_id") or created.get("id") or "")) if created else None
        return templates.TemplateResponse(
            request,
            "billing.html",
            {
                "request": request,
                "error": error,
                "ok": "",
                "created": created,
                "clients": billing_clients_context(),
                "detail": detail,
                "csrf_token": ensure_csrf_token(request),
                **context_i18n(loc),
            },
        )

    @app.post("/billing/update-name")
    def billing_update_name(
        request: Request,
        csrf_token: str = Form(default=""),
        account_id: str = Form(...),
        name: str = Form(default=""),
    ):
        if not billing_logged_in(request):
            return RedirectResponse(url="/billing/login", status_code=302)
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/billing?error=csrf", status_code=302)
        ok, msg = update_billing_account_name(account_id, name)
        if not ok:
            return RedirectResponse(url=f"/billing?error={quote(msg)}", status_code=302)
        rotate_csrf_token(request)
        return RedirectResponse(url="/billing?ok=name", status_code=302)

    @app.post("/billing/reset-password", response_class=HTMLResponse)
    async def billing_reset_password(request: Request):
        if not billing_logged_in(request):
            return RedirectResponse(url="/billing/login", status_code=302)
        loc = resolve_locale(request, None)
        error = ""
        temp_password = ""
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        account_id = str(form.get("account_id") or "")
        new_password = str(form.get("new_password") or form.get("manual_password") or form.get("password") or "")
        if not csrf_matches_session(request, csrf_token):
            error = "Сессия формы устарела. Обновите страницу и попробуйте снова."
        else:
            if new_password.strip():
                ok, msg, temp_password = set_billing_account_password(account_id, new_password)
            else:
                ok, msg, temp_password = reset_billing_account_password(account_id)
            if ok:
                rotate_csrf_token(request)
            else:
                error = msg
        detail = billing_account_detail(account_id, temp_password=temp_password)
        return templates.TemplateResponse(
            request,
            "billing.html",
            {
                "request": request,
                "error": error,
                "ok": "password" if temp_password else "",
                "created": None,
                "clients": billing_clients_context(),
                "detail": detail,
                "csrf_token": ensure_csrf_token(request),
                **context_i18n(loc),
            },
        )

    @app.post("/billing/export-database")
    async def billing_export_database(request: Request):
        if not billing_logged_in(request):
            return RedirectResponse(url="/billing/login", status_code=302)
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        account_id = str(form.get("account_id") or "").strip()
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url=f"/billing?detail={quote(account_id)}&error=csrf", status_code=302)
        detail = billing_account_detail(account_id)
        if not detail:
            return RedirectResponse(url="/billing?error=account_not_found", status_code=302)
        try:
            backup = export_workspace_database(str(detail["client"]["id"]))
        except Exception:
            logger.exception("[billing] failed to export account database")
            return RedirectResponse(url=f"/billing?detail={quote(account_id)}&error=export_failed", status_code=302)
        stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
        safe_account = "".join(ch for ch in str(detail["client"]["account_id"] or account_id) if ch.isalnum() or ch in ("-", "_"))
        filename = f"upos-finance-{safe_account or 'account'}-{stamp}.json"
        content = json.dumps(backup, ensure_ascii=False, indent=2).encode("utf-8")
        return Response(
            content=content,
            media_type="application/json; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @app.post("/billing/import-database", response_class=HTMLResponse)
    async def billing_import_database(
        request: Request,
        backup_file: UploadFile = File(...),
    ):
        if not billing_logged_in(request):
            return RedirectResponse(url="/billing/login", status_code=302)
        loc = resolve_locale(request, None)
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        account_id = str(form.get("account_id") or "").strip()
        root_login = str(form.get("root_login") or "").strip()
        root_password = str(form.get("root_password") or "")
        expected_login, expected_password = billing_root_credentials()
        error = ""
        restore_result: dict[str, Any] | None = None
        detail = billing_account_detail(account_id)
        if not csrf_matches_session(request, csrf_token):
            error = "Сессия формы устарела. Обновите страницу и попробуйте снова."
        elif not detail:
            error = "Аккаунт не найден."
        elif root_login != expected_login or root_password != expected_password:
            error = "Root-логин или пароль биллинга неверный."
        elif not backup_file or not str(backup_file.filename or "").strip():
            error = "Выберите файл бэкапа."
        else:
            try:
                raw = await backup_file.read(100 * 1024 * 1024 + 1)
                if len(raw) > 100 * 1024 * 1024:
                    raise ValueError("backup_too_large")
                backup = json.loads(raw.decode("utf-8-sig"))
                restore_result = restore_workspace_database(str(detail["client"]["id"]), backup)
                rotate_csrf_token(request)
            except json.JSONDecodeError:
                error = "Файл бэкапа не является корректным JSON."
            except ValueError as exc:
                code = str(exc)
                error = {
                    "backup_format": "Файл не похож на бэкап UPOS Finance.",
                    "backup_version": "Версия бэкапа не поддерживается.",
                    "backup_account_mismatch": "Бэкап относится к другому аккаунту.",
                    "backup_tables": "В файле бэкапа повреждён список таблиц.",
                    "backup_too_large": "Файл бэкапа слишком большой.",
                    "account_not_found": "Аккаунт не найден.",
                }.get(code, "Не удалось восстановить базу из файла.")
            except Exception:
                logger.exception("[billing] failed to import account database")
                error = "Не удалось восстановить базу из файла."
        detail = billing_account_detail(account_id)
        if detail is not None and restore_result is not None:
            detail = {**detail, "restore_result": restore_result}
        return templates.TemplateResponse(
            request,
            "billing.html",
            {
                "request": request,
                "error": error,
                "ok": "database-imported" if restore_result is not None else "",
                "created": None,
                "clients": billing_clients_context(),
                "detail": detail,
                "csrf_token": ensure_csrf_token(request),
                **context_i18n(loc),
            },
        )

    @app.post("/billing/clear-transactions", response_class=HTMLResponse)
    async def billing_clear_transactions(request: Request):
        if not billing_logged_in(request):
            return RedirectResponse(url="/billing/login", status_code=302)
        loc = resolve_locale(request, None)
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        account_id = str(form.get("account_id") or "").strip()
        error = ""
        clear_result: dict[str, Any] | None = None
        detail = billing_account_detail(account_id)
        if not csrf_matches_session(request, csrf_token):
            error = "Сессия формы устарела. Обновите страницу и попробуйте снова."
        elif not detail:
            error = "Аккаунт не найден."
        else:
            try:
                clear_result = clear_workspace_transactions(str(detail["client"]["id"]))
                rotate_csrf_token(request)
            except Exception:
                logger.exception("[billing] failed to clear account transactions")
                error = "Не удалось очистить транзакции. Попробуйте ещё раз."
        detail = billing_account_detail(account_id)
        if detail is not None and clear_result is not None:
            detail = {**detail, "clear_result": clear_result}
        return templates.TemplateResponse(
            request,
            "billing.html",
            {
                "request": request,
                "error": error,
                "ok": "transactions-cleared" if clear_result is not None else "",
                "created": None,
                "clients": billing_clients_context(),
                "detail": detail,
                "csrf_token": ensure_csrf_token(request),
                **context_i18n(loc),
            },
        )

    @app.post("/billing/clear-database", response_class=HTMLResponse)
    async def billing_clear_database(request: Request):
        if not billing_logged_in(request):
            return RedirectResponse(url="/billing/login", status_code=302)
        loc = resolve_locale(request, None)
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        account_id = str(form.get("account_id") or "").strip()
        root_login = str(form.get("root_login") or "").strip()
        root_password = str(form.get("root_password") or "")
        expected_login, expected_password = billing_root_credentials()
        error = ""
        clear_result: dict[str, Any] | None = None
        detail = billing_account_detail(account_id)
        if not csrf_matches_session(request, csrf_token):
            error = "Сессия формы устарела. Обновите страницу и попробуйте снова."
        elif not detail:
            error = "Аккаунт не найден."
        elif root_login != expected_login or root_password != expected_password:
            error = "Root-логин или пароль биллинга неверный."
        else:
            try:
                clear_result = clear_workspace_database(str(detail["client"]["id"]))
                rotate_csrf_token(request)
            except Exception:
                logger.exception("[billing] failed to clear account database")
                error = "Не удалось полностью очистить базу. Попробуйте ещё раз."
        detail = billing_account_detail(account_id)
        if detail is not None and clear_result is not None:
            detail = {**detail, "clear_result": clear_result}
        return templates.TemplateResponse(
            request,
            "billing.html",
            {
                "request": request,
                "error": error,
                "ok": "database-cleared" if clear_result is not None else "",
                "created": None,
                "clients": billing_clients_context(),
                "detail": detail,
                "csrf_token": ensure_csrf_token(request),
                **context_i18n(loc),
            },
        )

    @app.post("/auth/logout")
    def logout(request: Request, csrf_token: str = Form(default="")):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/schet?err=csrf", status_code=302)
        user = request.session.get("user") or {}
        uid = _session_user_id(user)
        auth_sid = str(request.session.get("auth_session_id") or "").strip()
        if uid and auth_sid:
            revoke_auth_session(uid, auth_sid)
        request.session.clear()
        return RedirectResponse(url="/auth", status_code=302)

    @app.get("/api/me/devices")
    def api_me_devices_list(request: Request):
        user, err = _devices_owner_only(request)
        if err:
            return err
        assert user is not None
        uid = _session_user_id(user)
        sid = str(request.session.get("auth_session_id") or "").strip()
        sid = ensure_auth_session_for_user(uid, request, sid or None)
        request.session["auth_session_id"] = sid
        return {"devices": list_user_devices(uid, current_session_id=sid)}

    @app.post("/api/me/devices/heartbeat")
    async def api_me_devices_heartbeat(request: Request):
        user = request.session.get("user") or {}
        uid = _session_user_id(user)
        if not uid:
            return JSONResponse({"error": "unauthorized"}, status_code=401)
        sid = str(request.session.get("auth_session_id") or "").strip()
        if not sid:
            sid = ensure_auth_session_for_user(uid, request, None)
            request.session["auth_session_id"] = sid
        client_meta: dict[str, Any] = {}
        try:
            body = await request.json()
            if isinstance(body, dict) and isinstance(body.get("client_meta"), dict):
                client_meta = body["client_meta"]
        except Exception:
            pass
        touch_auth_session(sid, uid, client_meta=client_meta or None)
        return {"ok": True}

    @app.post("/api/me/devices/{session_id}/revoke")
    def api_me_devices_revoke(request: Request, session_id: str):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        user, err = _devices_owner_only(request)
        if err:
            return err
        assert user is not None
        if not revoke_auth_session(_session_user_id(user), session_id):
            return JSONResponse({"error": "not_found"}, status_code=404)
        return {"ok": True}

    @app.post("/api/me/devices/{session_id}/block")
    def api_me_devices_block(request: Request, session_id: str):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        user, err = _devices_owner_only(request)
        if err:
            return err
        assert user is not None
        if not block_auth_session(_session_user_id(user), session_id):
            return JSONResponse({"error": "not_found"}, status_code=404)
        cur = str(request.session.get("auth_session_id") or "").strip()
        if cur == session_id:
            request.session.clear()
            return {"ok": True, "logged_out": True}
        return {"ok": True}

    @app.get("/api/employees/{employee_id}/devices")
    def api_employee_devices_list(request: Request, employee_id: str):
        emp, err = _director_employee_or_error(request, employee_id)
        if err:
            return err
        assert emp is not None
        return {"devices": list_user_devices(str(emp["id"]), current_session_id=None)}

    @app.post("/api/employees/{employee_id}/devices/{session_id}/revoke")
    def api_employee_devices_revoke(request: Request, employee_id: str, session_id: str):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        emp, err = _director_employee_or_error(request, employee_id)
        if err:
            return err
        assert emp is not None
        if not revoke_auth_session(str(emp["id"]), session_id):
            return JSONResponse({"error": "not_found"}, status_code=404)
        return {"ok": True}

    @app.post("/api/employees/{employee_id}/devices/{session_id}/block")
    def api_employee_devices_block(request: Request, employee_id: str, session_id: str):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        emp, err = _director_employee_or_error(request, employee_id)
        if err:
            return err
        assert emp is not None
        if not block_auth_session(str(emp["id"]), session_id):
            return JSONResponse({"error": "not_found"}, status_code=404)
        return {"ok": True}

    @app.post("/api/employees/{employee_id}/reset-password")
    def api_employee_reset_password(request: Request, employee_id: str):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        user = request.session.get("user") or {}
        if not _can_manage_employees(user):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        owner_id = str(user.get("account_owner_id") or user.get("user_id") or "").strip()
        active_org_id = str(user.get("active_org_id") or user.get("workspace_owner_id") or owner_id)
        ok, msg, temp_password = reset_employee_password(owner_id, employee_id, active_org_id)
        if not ok:
            return JSONResponse({"error": msg or "not_found"}, status_code=404)
        return {"ok": True, "temporary_password": temp_password}

    @app.get("/", response_class=HTMLResponse)
    def home(request: Request):
        u = request.session.get("user") or {}
        dest = "/organizations" if u.get("org_scope") == "general" else "/schet"
        if request.url.query:
            dest = f"{dest}?{request.url.query}"
        return RedirectResponse(url=dest, status_code=302)

    def _treasury_workspace_owner(request: Request) -> tuple[str | None, JSONResponse | None]:
        u = request.session.get("user") or {}
        if u.get("role") == "admin":
            return None, JSONResponse({"error": "forbidden"}, status_code=403)
        requested_org_id = str(request.query_params.get("organization_id") or "").strip()
        if requested_org_id and valid_workspace_owner_id(requested_org_id):
            return requested_org_id, None
        if u.get("org_scope") == "general":
            if requested_org_id:
                session_orgs = u.get("organizations") if isinstance(u.get("organizations"), list) else []
                if any(str(org.get("id")) == requested_org_id for org in session_orgs if isinstance(org, dict)):
                    return requested_org_id, None
                owner_candidates = [
                    str(u.get("account_owner_id") or "").strip(),
                    str(u.get("employer_user_id") or "").strip(),
                    str(u.get("user_id") or "").strip(),
                    str(u.get("workspace_owner_id") or "").strip(),
                ]
                for owner_id in dict.fromkeys(owner_candidates):
                    if valid_workspace_owner_id(owner_id):
                        active_orgs = list_organizations(owner_id)
                        if any(str(org.get("id")) == requested_org_id for org in active_orgs):
                            return requested_org_id, None
                if valid_workspace_owner_id(requested_org_id):
                    return requested_org_id, None
            return None, JSONResponse({"error": "organization_required"}, status_code=409)
        oid = str(u.get("workspace_owner_id") or u.get("user_id") or "").strip()
        if not valid_workspace_owner_id(oid):
            return None, JSONResponse({"error": "workspace"}, status_code=400)
        return oid, None

    def _positions_scope(request: Request, requested_org_id: str = "") -> tuple[str | None, JSONResponse | None]:
        u = request.session.get("user") or {}
        if u.get("role") == "admin":
            return None, JSONResponse({"error": "forbidden"}, status_code=403)
        if u.get("org_scope") == "general":
            if not _can_manage_employees(u):
                return None, JSONResponse({"error": "forbidden"}, status_code=403)
            owner_id = str(u.get("account_owner_id") or u.get("user_id") or "").strip()
            oid = str(requested_org_id or u.get("active_org_id") or u.get("workspace_owner_id") or "").strip()
            if owner_id and valid_workspace_owner_id(owner_id):
                active_orgs = list_organizations(owner_id)
                if oid and not any(str(org.get("id")) == oid for org in active_orgs):
                    oid = ""
                if not oid:
                    active_org = next((org for org in active_orgs if org.get("is_default")), active_orgs[0] if active_orgs else None)
                    oid = str(active_org.get("id") if active_org else "").strip()
        else:
            oid = str(u.get("workspace_owner_id") or u.get("user_id") or "").strip()
        if not valid_workspace_owner_id(oid):
            return None, JSONResponse({"error": "organization_required"}, status_code=409)
        return oid, None

    def _csrf_header_ok(request: Request) -> bool:
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        return csrf_matches_session(request, tok)

    @app.post("/api/reports/courier-debt-limits")
    async def api_reports_courier_debt_limits(request: Request):
        if not _csrf_header_ok(request):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        if not _has_permission(request.session.get("user") or {}, "reports"):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        try:
            body = await request.json()
        except Exception:
            body = {}
        raw_limits = body.get("limits") if isinstance(body, dict) and isinstance(body.get("limits"), dict) else {}
        limits: dict[str, float] = {}
        for raw_key, raw_value in raw_limits.items():
            key = str(raw_key or "").strip()
            if not key:
                continue
            try:
                amount = float(str(raw_value or "0").replace(" ", "").replace("\u202f", "").replace(",", "."))
            except (TypeError, ValueError):
                continue
            if amount > 0:
                limits[key[:260]] = round(amount, 2)
        data = load_workspace_settings(oid)
        data["delivery_debt_limits"] = limits
        save_workspace_settings(oid, data)
        return {"ok": True, "limits": limits}

    @app.get("/api/positions")
    def api_positions_list(request: Request):
        oid, err = _positions_scope(request, str(request.query_params.get("organization_id") or ""))
        if err:
            return err
        assert oid is not None
        return {"positions": list_positions(oid)}

    @app.post("/api/positions")
    async def api_positions_create(request: Request):
        if not _csrf_header_ok(request):
            return JSONResponse({"error": "csrf"}, status_code=403)
        try:
            body = await request.json()
        except Exception:
            body = {}
        oid, err = _positions_scope(request, str((body or {}).get("organization_id") or ""))
        if err:
            return err
        assert oid is not None
        try:
            row = create_position(oid, str((body or {}).get("name") or ""))
        except ValueError as exc:
            return JSONResponse({"error": str(exc) or "position"}, status_code=400)
        return {"position": row}

    @app.put("/api/positions/{position_id}")
    async def api_positions_update(request: Request, position_id: str):
        if not _csrf_header_ok(request):
            return JSONResponse({"error": "csrf"}, status_code=403)
        try:
            body = await request.json()
        except Exception:
            body = {}
        oid, err = _positions_scope(request, str((body or {}).get("organization_id") or ""))
        if err:
            return err
        assert oid is not None
        try:
            row = update_position(oid, position_id, str((body or {}).get("name") or ""))
        except ValueError as exc:
            return JSONResponse({"error": str(exc) or "position"}, status_code=400)
        if row is None:
            return JSONResponse({"error": "not_found"}, status_code=404)
        return {"position": row}

    @app.delete("/api/positions/{position_id}")
    def api_positions_delete(request: Request, position_id: str):
        if not _csrf_header_ok(request):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _positions_scope(request, str(request.query_params.get("organization_id") or ""))
        if err:
            return err
        assert oid is not None
        try:
            ok = delete_position(oid, position_id)
        except ValueError as exc:
            return JSONResponse({"error": str(exc) or "position"}, status_code=400)
        if not ok:
            return JSONResponse({"error": "not_found"}, status_code=404)
        return {"ok": True}

    def _period_from_request(
        request: Request,
        tz_name: str,
        *,
        default: str,
    ) -> dict[str, object]:
        qp = request.query_params
        preset = (
            qp.get("period")
            or qp.get("movement_preset")
            or qp.get("preset")
            or default
        )
        start, end, now_local, mode, start_label, end_label = period_local_bounds_utc(
            tz_name,
            preset,
            qp.get("date_from"),
            qp.get("date_to"),
        )
        return {
            "start": start,
            "end": end,
            "now_local": now_local,
            "preset": mode,
            "date_from": start_label or "",
            "date_to": end_label or "",
            "is_all": mode == "all",
        }

    def _period_label(loc: str, period: dict[str, object]) -> str:
        preset = str(period.get("preset") or "today")
        if preset == "all":
            return translate(loc, "period.all_time")
        if preset == "today":
            return translate(loc, "period.today")
        if preset == "month":
            now_local = period.get("now_local")
            month_keys = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
            if hasattr(now_local, "month") and hasattr(now_local, "year"):
                return f"{translate(loc, f'month.{month_keys[now_local.month - 1]}')} {now_local.year}"
            return translate(loc, "period.this_month")
        start = str(period.get("date_from") or "")
        end = str(period.get("date_to") or "")
        if start and end and start != end:
            return f"{start} - {end}"
        return start or translate(loc, "period.custom")

    def _treasury_payload(request: Request, oid: str) -> dict[str, object]:
        sess_user = request.session.get("user") or {}
        ws_data = load_workspace_settings(oid)
        loc = resolve_locale(request, sess_user)
        tz_name = normalize_workspace_timezone(str(ws_data.get("timezone") or ""))
        period = _period_from_request(request, tz_name, default="today")
        employee_id = _visible_employee_id(sess_user)
        owner_id = str(sess_user.get("account_owner_id") or sess_user.get("user_id") or "").strip()
        employees = list_employees_safe(owner_id, oid) if _can_manage_wallets(sess_user) else []
        visible_treasury = load_treasury(oid, visible_employee_id=employee_id)
        all_treasury = load_treasury(oid)
        visible_pocket_ids = {str(p.get("id") or "") for p in visible_treasury.get("pockets") or [] if isinstance(p, dict)}
        transfer_targets = []
        for pocket in all_treasury.get("pockets") or []:
            if not isinstance(pocket, dict):
                continue
            target = dict(pocket)
            if employee_id and str(target.get("id") or "") not in visible_pocket_ids:
                target["entries"] = []
            transfer_targets.append(target)
        return {
            "treasury": visible_treasury,
            "templates": localize_treasury_templates(loc, list_templates_public()),
            "settings": {
                "enabled_currencies": ws_data.get("enabled_currencies")
                or ["UZS", "USD", "RUB", "EUR"],
                "timezone": tz_name,
            },
            "employees": employees,
            "transfer_targets": transfer_targets,
            "movement": get_account_movements(
                oid,
                period["start"],
                period["end"],
                visible_employee_id=employee_id,
            ),
            "movement_period": {
                "preset": period["preset"],
                "date_from": period["date_from"],
                "date_to": period["date_to"],
                "label": _period_label(loc, period),
                "is_all": period["is_all"],
            },
        }

    @app.get("/api/treasury")
    def api_treasury_get(request: Request):
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        return _treasury_payload(request, oid)

    @app.put("/api/treasury")
    async def api_treasury_put(request: Request):
        tok = (
            request.headers.get("X-CSRF-Token")
            or request.headers.get("x-csrf-token")
            or ""
        )
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        if not _can_manage_wallets(request.session.get("user") or {}):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "json"}, status_code=400)
        clean, msg = validate_and_clean_treasury(body)
        if clean is None:
            return JSONResponse({"error": msg}, status_code=400)
        try:
            save_treasury(oid, clean)
        except ValueError as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        return {"ok": True, "treasury": load_treasury(oid)}

    @app.patch("/api/treasury")
    async def api_treasury_patch(request: Request):
        tok = (
            request.headers.get("X-CSRF-Token")
            or request.headers.get("x-csrf-token")
            or ""
        )
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        if not _can_modify_transactions(request.session.get("user") or {}):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "json"}, status_code=400)
        if not isinstance(body, dict):
            return JSONResponse({"error": "json"}, status_code=400)
        dc = body.get("display_currency")
        if not isinstance(dc, str) or not dc.strip():
            return JSONResponse({"error": "display_currency"}, status_code=400)
        try:
            patch_display_currency(oid, dc)
        except ValueError as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        return {
            "ok": True,
            "treasury": load_treasury(oid),
            "templates": localize_treasury_templates(
                resolve_locale(request, request.session.get("user") or {}),
                list_templates_public(),
            ),
        }

    @app.get("/api/wallets")
    def api_wallets_get(request: Request, mine: bool = Query(default=False)):
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        sess_user = request.session.get("user") or {}
        employee_id = _visible_employee_id(sess_user) if (mine or sess_user.get("is_employee")) else None
        return {"wallets": load_treasury(oid, visible_employee_id=employee_id).get("pockets") or []}

    @app.post("/api/wallets")
    async def api_wallets_post(request: Request):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        if not _can_manage_wallets(request.session.get("user") or {}):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "json"}, status_code=400)
        if not isinstance(body, dict):
            return JSONResponse({"error": "json"}, status_code=400)
        wallet_id = str(body.get("id") or uuid.uuid4())
        current = load_treasury(oid)
        pockets = list(current.get("pockets") or [])
        pocket = {
            "id": wallet_id,
            "template_id": str(body.get("template_id") or body.get("kind") or "custom"),
            "label": str(body.get("label") or body.get("name") or "Счёт").strip() or "Счёт",
            "note": str(body.get("note") or ""),
            "icon": str(body.get("icon") or ""),
            "owner_employee_id": str(body.get("owner_employee_id") or "").strip(),
            "entries": body.get("entries") if isinstance(body.get("entries"), list) else [],
        }
        pockets.append(pocket)
        clean, msg = validate_and_clean_treasury(
            {
                "version": 2,
                "display_currency": current.get("display_currency") or "USD",
                "pockets": pockets,
            },
        )
        if clean is None:
            return JSONResponse({"error": msg}, status_code=400)
        try:
            save_treasury(oid, clean)
        except ValueError as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        saved = load_treasury(oid)
        wallet = next((p for p in saved.get("pockets") or [] if str(p.get("id")) == wallet_id), pocket)
        return {"ok": True, "wallet": wallet}

    @app.patch("/api/wallets/{wallet_id}")
    async def api_wallets_patch(request: Request, wallet_id: str):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        if not _can_manage_wallets(request.session.get("user") or {}):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "json"}, status_code=400)
        if not isinstance(body, dict):
            return JSONResponse({"error": "json"}, status_code=400)
        current = load_treasury(oid)
        pockets = list(current.get("pockets") or [])
        found = None
        for pocket in pockets:
            if str(pocket.get("id") or "") == wallet_id:
                found = pocket
                break
        if found is None:
            return JSONResponse({"error": "not_found"}, status_code=404)
        if "owner_employee_id" in body:
            found["owner_employee_id"] = str(body.get("owner_employee_id") or "").strip()
        if "label" in body or "name" in body:
            found["label"] = str(body.get("label") or body.get("name") or found.get("label") or "Счёт").strip() or "Счёт"
        if "note" in body:
            found["note"] = str(body.get("note") or "")
        if "template_id" in body or "kind" in body:
            found["template_id"] = str(body.get("template_id") or body.get("kind") or found.get("template_id") or "custom")
        if "icon" in body:
            found["icon"] = str(body.get("icon") or "")
        if isinstance(body.get("entries"), list):
            found["entries"] = body["entries"]
        clean, msg = validate_and_clean_treasury(
            {
                "version": 2,
                "display_currency": current.get("display_currency") or "USD",
                "pockets": pockets,
            },
        )
        if clean is None:
            return JSONResponse({"error": msg}, status_code=400)
        try:
            save_treasury(oid, clean)
        except ValueError as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        saved = load_treasury(oid)
        wallet = next((p for p in saved.get("pockets") or [] if str(p.get("id")) == wallet_id), found)
        return {"ok": True, "wallet": wallet}

    @app.get("/api/transfers")
    def api_transfers_get(
        request: Request,
        status: str | None = Query(default=None),
        direction: str | None = Query(default=None),
    ):
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        return {
            "transfers": list_transfers(
                oid,
                status=status,
                direction=direction,
                actor=request.session.get("user") or {},
            ),
        }

    def _queue_transaction_telegram_notification(
        request: Request,
        background_tasks: BackgroundTasks,
        workspace_owner_id: str,
        tx: dict | None,
    ) -> None:
        if not tx or not tx.get("is_confirmed"):
            return
        notify_tx = dict(tx)
        notify_tx["actor_name"] = str((request.session.get("user") or {}).get("name") or "")
        notify_tx["organization_name"] = str(
            (request.session.get("user") or {}).get("active_organization_name") or "",
        )
        from upos.telegram_notifier import notify_transaction_created

        background_tasks.add_task(notify_transaction_created, workspace_owner_id, notify_tx)

    def _limit_amount_decimal(raw: Any) -> Decimal:
        clean = str(raw or "").replace("\u202f", "").replace(" ", "").replace(",", ".").strip()
        if not clean:
            return Decimal("0")
        return Decimal(clean)

    def _apply_telegram_limit_approval_if_needed(
        workspace_owner_id: str,
        body: dict[str, Any],
        actor: dict[str, Any],
    ) -> dict[str, Any] | None:
        tx_type = str(body.get("type") or "income").strip().lower()
        if tx_type not in {"income", "expense"}:
            return None
        try:
            amount = _limit_amount_decimal(body.get("amount"))
        except Exception:
            return None
        if amount <= 0:
            return None
        try:
            from upos.telegram_store import delivery_targets, get_notification_prefs

            prefs = get_notification_prefs(workspace_owner_id)
        except Exception:
            return None
        reports = prefs.get("reports") if isinstance(prefs.get("reports"), dict) else {}
        if not bool(reports.get("limits", True)):
            return None
        try:
            if not delivery_targets(workspace_owner_id, "limits"):
                return None
        except Exception:
            return None
        limits = prefs.get("limits") if isinstance(prefs.get("limits"), dict) else {}
        if not bool(limits.get("enabled")):
            return None
        rule = limits.get(tx_type) if isinstance(limits.get(tx_type), dict) else {}
        if not bool(rule.get("enabled", True)):
            return None
        try:
            limit_amount = _limit_amount_decimal(rule.get("amount"))
        except Exception:
            return None
        if limit_amount <= 0 or amount <= limit_amount:
            return None
        data = body.get("data") if isinstance(body.get("data"), dict) else {}
        data = dict(data)
        approval = {
            "token": secrets.token_hex(6),
            "status": "pending",
            "limit_type": tx_type,
            "limit_amount": str(limit_amount),
            "requested_amount": str(amount),
            "requested_by": str(actor.get("name") or actor.get("username") or "").strip(),
        }
        data["telegram_limit_approval"] = approval
        body["data"] = data
        body["is_confirmed"] = False
        body["status"] = "pending"
        body["requires_confirmation"] = True
        return approval

    def _queue_transaction_telegram_limit_approval(
        request: Request,
        background_tasks: BackgroundTasks,
        workspace_owner_id: str,
        tx: dict | None,
    ) -> None:
        if not tx:
            return
        notify_tx = dict(tx)
        notify_tx["actor_name"] = str((request.session.get("user") or {}).get("name") or "")
        notify_tx["organization_name"] = str(
            (request.session.get("user") or {}).get("active_organization_name") or "",
        )
        from upos.telegram_notifier import notify_transaction_limit_approval_required

        background_tasks.add_task(notify_transaction_limit_approval_required, workspace_owner_id, notify_tx)

    @app.post("/api/transfers")
    async def api_transfers_post(request: Request, background_tasks: BackgroundTasks):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        if not _can_modify_transactions(request.session.get("user") or {}):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "json"}, status_code=400)
        if not isinstance(body, dict):
            return JSONResponse({"error": "json"}, status_code=400)
        body = dict(body)
        body["type"] = "transfer"
        if body.get("from_wallet_id") and not body.get("from_pocket_id"):
            body["from_pocket_id"] = body.get("from_wallet_id")
        if body.get("to_wallet_id") and not body.get("to_pocket_id"):
            body["to_pocket_id"] = body.get("to_wallet_id")
        try:
            tx = create_transaction(oid, body, actor=request.session.get("user"))
        except PermissionError as exc:
            return JSONResponse({"error": str(exc) or "forbidden"}, status_code=403)
        except (TreasuryPostingError, ValueError) as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        _queue_transaction_telegram_notification(request, background_tasks, oid, tx)
        return {"ok": True, "transfer": tx}

    @app.post("/api/transfers/{transfer_id}/confirm")
    async def api_transfer_confirm(request: Request, transfer_id: str, background_tasks: BackgroundTasks):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        try:
            body = await request.json()
        except Exception:
            body = {}
        if not isinstance(body, dict):
            body = {}
        target_account_id = str(
            body.get("to_account_id")
            or body.get("to_pocket_id")
            or body.get("target_account_id")
            or "",
        ).strip()
        try:
            tx = resolve_pending_transfer(
                oid,
                transfer_id,
                action="confirm",
                actor=request.session.get("user") or {},
                to_account_id=target_account_id,
            )
        except PermissionError as exc:
            return JSONResponse({"error": str(exc) or "forbidden"}, status_code=403)
        except ValueError as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        if not tx:
            return JSONResponse({"error": "not_found"}, status_code=404)
        _queue_transaction_telegram_notification(request, background_tasks, oid, tx)
        return {"ok": True, "transfer": tx}

    @app.post("/api/transfers/{transfer_id}/reject")
    async def api_transfer_reject(request: Request, transfer_id: str):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        try:
            tx = resolve_pending_transfer(
                oid,
                transfer_id,
                action="reject",
                actor=request.session.get("user") or {},
            )
        except PermissionError as exc:
            return JSONResponse({"error": str(exc) or "forbidden"}, status_code=403)
        except ValueError as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        if not tx:
            return JSONResponse({"error": "not_found"}, status_code=404)
        return {"ok": True, "transfer": tx}

    @app.get("/api/fx/rates")
    def api_fx_rates():
        from upos.fx import get_usd_rates

        return get_usd_rates()



    @app.get("/api/transactions")
    def api_transactions_get(request: Request):
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        try:
            _ensure_sales_cash_transactions(oid)
        except Exception:
            logger.exception("[sales] failed to ensure sales cash transactions for %s", oid)
        try:
            limit = int(request.query_params.get("limit") or 500)
        except (TypeError, ValueError):
            limit = 500
        try:
            offset = int(request.query_params.get("offset") or 0)
        except (TypeError, ValueError):
            offset = 0
        category_ids, category_names = _category_filters_for_user(request.session.get("user") or {}, oid)
        return {
            "transactions": list_transactions(
                oid,
                limit=limit,
                offset=offset,
                visible_employee_id=_visible_employee_id(request.session.get("user") or {}),
                allowed_category_ids=category_ids,
                allowed_category_names=category_names,
            )
        }

    @app.get("/api/director/consolidated-transactions")
    def api_director_consolidated_transactions(request: Request):
        """Консолидированная касса: только владелец бизнеса (директор), не сотрудники и не admin."""
        u = request.session.get("user") or {}
        if u.get("role") == "admin":
            return JSONResponse({"error": "forbidden"}, status_code=403)
        if not _is_director(u):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        owner_id = str(u.get("account_owner_id") or u.get("user_id") or "").strip()
        if not valid_workspace_owner_id(owner_id):
            return JSONResponse({"error": "workspace"}, status_code=400)
        qp = request.query_params
        organization_id = (qp.get("organization_id") or "").strip() or None
        orgs = list_organizations(owner_id)
        default_org = next((o for o in orgs if o.get("is_default")), orgs[0] if orgs else None)
        tz_name = normalize_workspace_timezone(str(qp.get("timezone") or "").strip())
        if not (qp.get("timezone") or "").strip() and default_org:
            tz_name = normalize_workspace_timezone(
                str(load_workspace_settings(str(default_org["id"])).get("timezone") or ""),
            )
        period = _period_from_request(request, tz_name, default="month")
        try:
            lim = int(qp.get("limit") or "2500")
        except ValueError:
            lim = 2500
        try:
            payload = list_director_consolidated_transactions(
                owner_id,
                organization_id=organization_id,
                period_start_utc=period["start"],  # type: ignore[arg-type]
                period_end_utc=period["end"],  # type: ignore[arg-type]
                tx_type=(qp.get("type") or "").strip() or None,
                currency=(qp.get("currency") or "").strip() or None,
                category=(qp.get("category") or "").strip() or None,
                source=(qp.get("source") or "").strip() or None,
                limit=lim,
            )
        except ValueError as exc:
            if str(exc) == "organization_not_allowed":
                return JSONResponse({"error": "organization_not_allowed"}, status_code=400)
            raise
        hints = director_transaction_filter_hints(owner_id)
        loc = resolve_locale(request, u)
        return {
            **payload,
            "filter_hints": hints,
            "period_meta": {
                "preset": period["preset"],
                "date_from": period["date_from"],
                "date_to": period["date_to"],
                "label": _period_label(loc, period),
                "is_all": period["is_all"],
                "timezone": tz_name,
            },
        }

    @app.get("/api/director/consolidated-treasury")
    def api_director_consolidated_treasury(request: Request):
        """Консолидированные остатки по счетам организаций владельца (только просмотр)."""
        u = request.session.get("user") or {}
        if u.get("role") == "admin":
            return JSONResponse({"error": "forbidden"}, status_code=403)
        if not _is_director(u):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        owner_id = str(u.get("account_owner_id") or u.get("user_id") or "").strip()
        if not valid_workspace_owner_id(owner_id):
            return JSONResponse({"error": "workspace"}, status_code=400)
        qp = request.query_params
        organization_id = (qp.get("organization_id") or "").strip() or None
        dc = (qp.get("display_currency") or "USD").strip().upper()
        try:
            payload = aggregate_director_treasury(
                owner_id,
                organization_id=organization_id,
                display_currency=dc,
            )
            employees_by_org = {
                str(org.get("id") or ""): [
                    {
                        "id": str(emp.get("id") or ""),
                        "name": str(emp.get("name") or emp.get("username") or ""),
                        "position": str(emp.get("position") or ""),
                    }
                    for emp in list_employees_safe(owner_id, str(org.get("id") or ""))
                    if str(emp.get("id") or "")
                ]
                for org in list_organizations(owner_id)
            }
            for org_row in payload.get("organizations") or []:
                if isinstance(org_row, dict):
                    org_row["employees"] = employees_by_org.get(str(org_row.get("organization_id") or ""), [])
            return payload
        except ValueError as exc:
            if str(exc) == "organization_not_allowed":
                return JSONResponse({"error": "organization_not_allowed"}, status_code=400)
            raise

    @app.post("/api/director/organization-wallets")
    async def api_director_organization_wallets_post(request: Request):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        u = request.session.get("user") or {}
        if u.get("role") == "admin" or not _is_director(u):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        owner_id = str(u.get("account_owner_id") or u.get("user_id") or "").strip()
        if not valid_workspace_owner_id(owner_id):
            return JSONResponse({"error": "workspace"}, status_code=400)
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "json"}, status_code=400)
        if not isinstance(body, dict):
            return JSONResponse({"error": "json"}, status_code=400)

        org_id = str(body.get("organization_id") or "").strip()
        org = get_organization(owner_id, org_id)
        if not org:
            return JSONResponse({"error": "organization_not_allowed"}, status_code=400)

        employee_id = str(body.get("owner_employee_id") or "").strip()
        if employee_id and get_employee_for_owner(owner_id, employee_id, org_id) is None:
            return JSONResponse({"error": "employee_not_allowed"}, status_code=400)

        wallet_id = str(body.get("id") or uuid.uuid4())
        current = load_treasury(org_id)
        pockets = list(current.get("pockets") or [])
        pocket = {
            "id": wallet_id,
            "template_id": str(body.get("template_id") or body.get("kind") or "custom"),
            "label": str(body.get("label") or body.get("name") or "Счёт").strip() or "Счёт",
            "note": str(body.get("note") or ""),
            "icon": str(body.get("icon") or ""),
            "owner_employee_id": employee_id,
            "entries": body.get("entries") if isinstance(body.get("entries"), list) else [],
        }
        pockets.append(pocket)
        clean, msg = validate_and_clean_treasury(
            {
                "version": 2,
                "display_currency": current.get("display_currency") or "USD",
                "pockets": pockets,
            },
        )
        if clean is None:
            return JSONResponse({"error": msg}, status_code=400)
        try:
            save_treasury(org_id, clean)
            if employee_id:
                with session_scope() as session:
                    exists = session.scalar(
                        select(EmployeeAccountAccess.account_id).where(
                            EmployeeAccountAccess.employee_id == employee_id,
                            EmployeeAccountAccess.account_id == wallet_id,
                        ),
                    )
                    if not exists:
                        session.add(EmployeeAccountAccess(employee_id=employee_id, account_id=wallet_id))
        except ValueError as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        saved = load_treasury(org_id)
        wallet = next((p for p in saved.get("pockets") or [] if str(p.get("id")) == wallet_id), pocket)
        return {"ok": True, "wallet": wallet}

    @app.get("/api/director/consolidated-pnl")
    def api_director_consolidated_pnl(request: Request):
        """ОиУ по всем организациям владельца за период."""
        u = request.session.get("user") or {}
        if u.get("role") == "admin":
            return JSONResponse({"error": "forbidden"}, status_code=403)
        if not _is_director(u):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        owner_id = str(u.get("account_owner_id") or u.get("user_id") or "").strip()
        if not valid_workspace_owner_id(owner_id):
            return JSONResponse({"error": "workspace"}, status_code=400)
        qp = request.query_params
        organization_id = (qp.get("organization_id") or "").strip() or None
        orgs = list_organizations(owner_id)
        default_org = next((o for o in orgs if o.get("is_default")), orgs[0] if orgs else None)
        tz_name = normalize_workspace_timezone(str(qp.get("timezone") or "").strip())
        if not (qp.get("timezone") or "").strip() and default_org:
            tz_name = normalize_workspace_timezone(
                str(load_workspace_settings(str(default_org["id"])).get("timezone") or ""),
            )
        period = _period_from_request(request, tz_name, default="month")
        try:
            payload = get_director_consolidated_pnl(
                owner_id,
                period["start"],  # type: ignore[arg-type]
                period["end"],  # type: ignore[arg-type]
                organization_id=organization_id,
            )
        except ValueError as exc:
            if str(exc) == "organization_not_allowed":
                return JSONResponse({"error": "organization_not_allowed"}, status_code=400)
            raise
        loc = resolve_locale(request, u)
        return {
            **payload,
            "period_meta": {
                "preset": period["preset"],
                "date_from": period["date_from"],
                "date_to": period["date_to"],
                "label": _period_label(loc, period),
                "is_all": period["is_all"],
                "timezone": tz_name,
            },
        }

    @app.post("/api/kassa/sms-report")
    async def api_kassa_sms_report(request: Request):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        if not _has_permission(request.session.get("user") or {}, "kassa"):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "json"}, status_code=400)
        if not isinstance(body, dict):
            return JSONResponse({"error": "json"}, status_code=400)

        kind = str(body.get("kind") or "daily_couriers").strip()
        allowed_kinds = {"daily_couriers", "daily_expenses", "daily_transfers"}
        if kind not in allowed_kinds:
            return JSONResponse({"error": "unknown_report"}, status_code=400)

        report_date = str(body.get("date") or "").strip()[:10]
        user = request.session.get("user") or {}
        actor_name = str(user.get("name") or user.get("username") or "")
        from upos.telegram_notifier import (
            send_kassa_sms_daily_courier_report,
            send_kassa_sms_daily_expense_report,
            send_kassa_sms_daily_transfer_report,
        )

        senders = {
            "daily_couriers": send_kassa_sms_daily_courier_report,
            "daily_expenses": send_kassa_sms_daily_expense_report,
            "daily_transfers": send_kassa_sms_daily_transfer_report,
        }
        result = senders[kind](oid, report_date, actor_name=actor_name)
        status = 200 if result.get("ok") else 400
        return JSONResponse(result, status_code=status)

    @app.post("/api/transactions")
    async def api_transactions_post(request: Request, background_tasks: BackgroundTasks):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        if not _can_modify_transactions(request.session.get("user") or {}):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "json"}, status_code=400)
        if not isinstance(body, dict):
            return JSONResponse({"error": "json"}, status_code=400)
        if not _category_payload_allowed(request.session.get("user") or {}, oid, body if isinstance(body, dict) else {}):
            return JSONResponse({"error": "category_forbidden"}, status_code=403)
        body = dict(body)
        limit_approval = _apply_telegram_limit_approval_if_needed(oid, body, request.session.get("user") or {})
        
        try:
            tx = create_transaction(oid, body, actor=request.session.get("user"))
        except PermissionError as exc:
            return JSONResponse({"error": str(exc) or "forbidden"}, status_code=403)
        except (TreasuryPostingError, ValueError) as exc:
            logger.warning(f"[api] create_transaction validation failed: {exc}")
            return JSONResponse({"error": str(exc)}, status_code=400)
        except SQLAlchemyError as exc:
            logger.warning("[api] create_transaction database layer: %s", exc, exc_info=True)
            return JSONResponse(
                {
                    "error": (
                        "Не удалось записать операцию в базу. Сократите длинный текст "
                        "(категория, филиал, клиент) или обновите страницу."
                    ),
                },
                status_code=400,
            )
        except Exception as exc:
            logger.error(f"[api] create_transaction CRITICAL ERROR: {exc}", exc_info=True)
            return JSONResponse({"error": "internal_error"}, status_code=500)
        if limit_approval:
            _queue_transaction_telegram_limit_approval(request, background_tasks, oid, tx)
        else:
            _queue_transaction_telegram_notification(request, background_tasks, oid, tx)
        return {"ok": True, "transaction": tx}

    @app.put("/api/transactions/{tx_id}")
    async def api_transactions_put(request: Request, tx_id: str):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        if not _can_modify_transactions(request.session.get("user") or {}):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        if not _employee_can_access_transaction(request.session.get("user") or {}, oid, tx_id):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "json"}, status_code=400)
        if not _category_payload_allowed(request.session.get("user") or {}, oid, body if isinstance(body, dict) else {}):
            return JSONResponse({"error": "category_forbidden"}, status_code=403)
            
        try:
            tx = update_transaction(oid, tx_id, body, actor=request.session.get("user") or {})
        except PermissionError as exc:
            return JSONResponse({"error": str(exc) or "forbidden"}, status_code=403)
        except (TreasuryPostingError, ValueError) as exc:
            logger.warning(f"[api] update_transaction validation failed: {exc}")
            return JSONResponse({"error": str(exc)}, status_code=400)
        except SQLAlchemyError as exc:
            logger.warning("[api] update_transaction database layer: %s", exc, exc_info=True)
            return JSONResponse(
                {
                    "error": (
                        "Не удалось записать операцию в базу. Сократите длинный текст "
                        "(категория, филиал, клиент) или обновите страницу."
                    ),
                },
                status_code=400,
            )
        except Exception as exc:
            logger.error(f"[api] update_transaction CRITICAL ERROR: {exc}", exc_info=True)
            return JSONResponse({"error": "internal_error"}, status_code=500)
        if not tx:
            return JSONResponse({"error": "not found"}, status_code=404)
        return {"ok": True, "transaction": tx}

    @app.patch("/api/transactions/{tx_id}/status")
    async def api_transactions_status_patch(request: Request, tx_id: str):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        if not _can_modify_transactions(request.session.get("user") or {}):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        if not _employee_can_access_transaction(request.session.get("user") or {}, oid, tx_id):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "json"}, status_code=400)
        if not isinstance(body, dict):
            return JSONResponse({"error": "json"}, status_code=400)
        try:
            tx = set_transaction_status(
                oid,
                tx_id,
                str(body.get("status") or ""),
                actor=request.session.get("user") or {},
            )
        except PermissionError as exc:
            return JSONResponse({"error": str(exc) or "forbidden"}, status_code=403)
        except (TreasuryPostingError, ValueError) as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        if not tx:
            return JSONResponse({"error": "not found"}, status_code=404)
        return {"ok": True, "transaction": tx}

    @app.delete("/api/transactions/{tx_id}")
    def api_transactions_delete(request: Request, tx_id: str, background_tasks: BackgroundTasks):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        if not _can_modify_transactions(request.session.get("user") or {}):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        if not _employee_can_access_transaction(request.session.get("user") or {}, oid, tx_id):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        
        try:
            deleted_tx = delete_transaction(oid, tx_id)
        except (TreasuryPostingError, ValueError) as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        if not deleted_tx:
            return JSONResponse({"error": "not found"}, status_code=404)
        sess_user = request.session.get("user") or {}
        notify_tx = dict(deleted_tx)
        notify_tx["deleted_by_name"] = str(sess_user.get("name") or sess_user.get("username") or "")
        notify_tx["deleted_at"] = datetime.now(timezone.utc).isoformat()
        notify_tx["organization_name"] = str(sess_user.get("active_organization_name") or "")
        from upos.telegram_notifier import notify_transaction_deleted

        background_tasks.add_task(notify_transaction_deleted, oid, notify_tx)
        return {"ok": True}

    @app.get("/schet", response_class=HTMLResponse)
    def home_schet(request: Request):
        u = request.session.get("user") or {}
        wid = str(u.get("workspace_owner_id") or u.get("user_id") or "").strip()
        loc = resolve_locale(request, u)
        bootstrap = None
        if u.get("role") != "admin" and valid_workspace_owner_id(wid):
            bootstrap = _treasury_payload(request, wid)
        schet_bootstrap_json = "null"
        if bootstrap:
            raw = json.dumps(bootstrap, ensure_ascii=False)
            schet_bootstrap_json = raw.replace("</", "<\\/")
        return tpl(
            request,
            "home_schet.html",
            variant="user",
            active="home_account",
            schet_bootstrap_json=schet_bootstrap_json,
        )

    @app.get("/kassa", response_class=HTMLResponse)
    def home_kassa(request: Request):
        u = request.session.get("user") or {}
        wid = str(u.get("workspace_owner_id") or u.get("user_id") or "").strip()
        workspace_timezone = normalize_workspace_timezone("")
        if valid_workspace_owner_id(wid):
            workspace_timezone = normalize_workspace_timezone(
                str(load_workspace_settings(wid).get("timezone") or "")
            )
        return tpl(
            request,
            "home_kassa.html",
            variant="user",
            active="home_kassa",
            workspace_timezone=workspace_timezone,
        )

    def _business_module_context(key: str) -> dict[str, Any]:
        modules: dict[str, dict[str, Any]] = {
            "products": {
                "title": "Товары",
                "kicker": "Торговый контур",
                "subtitle": "Номенклатура, цены, штрихкоды и остатки",
                "heading": "Каталог товаров",
                "action": "+ Товар",
                "filters": ["Поиск", "Категория", "Группа", "Бренд", "Статус"],
                "columns": ["Название", "На складе", "Цена продажи", "Штрихкод", "Статус"],
                "list_title": "Список товаров",
                "status": "Каркас",
                "empty": "Здесь будет список товаров с логикой создания как в ibox.",
                "logic": [
                    "Создание простого товара и товара-коллекции.",
                    "Категория, папка, единицы хранения, артикул и изображение.",
                    "Несколько прайс-листов и продажных цен.",
                    "Начальные остатки по складам, партиям и датам.",
                    "Штрихкоды, упаковки, минимальные остатки и характеристики.",
                ],
            },
            "warehouse": {
                "title": "Склад",
                "kicker": "Остатки и движение",
                "subtitle": "Склады, перемещения, корректировки и списания",
                "heading": "Складской учет",
                "action": "+ Операция",
                "launcher": [
                    {"id": "warehouses", "title": "Склады", "subtitle": "Создание и настройки", "icon": "warehouse"},
                    {"id": "transfers", "title": "Перемещения", "subtitle": "Между складами", "icon": "transfer"},
                    {"id": "purchases", "title": "Закупки", "subtitle": "Приход от поставщиков", "icon": "purchase"},
                    {"id": "adjustments", "title": "Корректировки", "subtitle": "Приход и списание", "icon": "adjustment"},
                ],
                "filters": ["Поиск", "Склад", "Товар", "Дата", "Тип операции"],
                "columns": ["Документ", "Дата", "Склад", "Товар", "Количество", "Статус"],
                "list_title": "Складские документы",
                "status": "Каркас",
                "empty": "Здесь будут остатки, перемещения и корректировки как в ibox.",
                "logic": [
                    "Остатки товаров по каждому складу.",
                    "Приход, расход, перемещение и корректировка склада.",
                    "Списание остатков при продаже или отгрузке.",
                    "История движений по товару и складу.",
                    "Контроль минимального остатка.",
                ],
            },
            "clients": {
                "title": "Клиенты",
                "kicker": "CRM и продажи",
                "subtitle": "Клиенты, должники, маршруты и балансы",
                "heading": "База клиентов",
                "action": "+ Клиент",
                "launcher": [
                    {"id": "clients", "title": "Клиенты", "subtitle": "База контрагентов", "icon": "clients"},
                    {"id": "balances", "title": "Должники", "subtitle": "Клиенты с долгом", "icon": "balance"},
                    {"id": "routes", "title": "Маршруты", "subtitle": "Территории и график", "icon": "route"},
                ],
                "filters": ["Поиск", "Территория", "Категория", "Маршрут", "Статус"],
                "columns": ["ИД", "Название", "Баланс", "Телефон", "Категория", "Маршрут"],
                "list_title": "Список клиентов",
                "status": "Каркас",
                "empty": "Здесь будет клиентская база с балансами и связями с продажами.",
                "logic": [
                    "Карточка клиента с контактами, телефоном и Telegram.",
                    "Категории, территории, маршруты и статус активности.",
                    "Баланс клиента и история оплат/отгрузок.",
                    "Связь клиента с заказами, сделками и задачами CRM.",
                    "Один контрагент может быть клиентом и поставщиком.",
                ],
            },
            "suppliers": {
                "title": "Поставщики",
                "kicker": "Закупки и кредиторка",
                "subtitle": "Поставщики, категории, балансы и закупки",
                "heading": "База поставщиков",
                "action": "+ Поставщик",
                "launcher": [
                    {"id": "suppliers", "title": "Поставщики", "subtitle": "База контрагентов", "icon": "suppliers"},
                    {"id": "purchases", "title": "Закупки", "subtitle": "Документы прихода", "icon": "purchase"},
                    {"id": "payables", "title": "Кредиторка", "subtitle": "Балансы и оплаты", "icon": "balance"},
                ],
                "filters": ["Поиск", "Статус", "Категория"],
                "columns": ["Название", "Баланс", "Создал", "Дата создания", "Категория"],
                "list_title": "Список поставщиков",
                "status": "Каркас",
                "empty": "Здесь будет база поставщиков с логикой закупок как в ibox.",
                "logic": [
                    "Карточка поставщика с категорией и статусом.",
                    "Баланс и кредиторская задолженность.",
                    "Связь с закупками, оплатами и возвратами поставщику.",
                    "История документов по каждому поставщику.",
                    "Общий контрагент может совмещать роли клиента и поставщика.",
                ],
            },
            "crm": {
                "title": "CRM",
                "kicker": "Задачи и сделки",
                "subtitle": "Контроль клиентов, ответственных и этапов сделки",
                "heading": "CRM-центр",
                "action": "+ Сделка",
                "launcher": [
                    {"id": "tasks", "title": "Задачи", "subtitle": "Сроки и исполнители", "icon": "tasks"},
                    {"id": "deals", "title": "Сделки", "subtitle": "Клиенты и этапы", "icon": "deals"},
                    {"id": "history", "title": "История", "subtitle": "Контакты и звонки", "icon": "history"},
                ],
                "filters": ["Клиент", "Ответственный", "Статус", "Срок", "Тип задачи"],
                "columns": ["Название", "Клиент", "Ответственный", "Срок", "Статус"],
                "list_title": "Задачи и сделки",
                "status": "Каркас",
                "empty": "Здесь будут задачи и сделки, связанные с клиентами и продажами.",
                "logic": [
                    "Задачи со сроком, исполнителем, типом и статусом.",
                    "Сделки с клиентом, ответственным и этапом.",
                    "Связь CRM с клиентской карточкой и заказом.",
                    "Фильтры по статусу, сроку и ответственному.",
                    "История взаимодействий по клиенту.",
                ],
            },
            "telephony": {
                "title": "Телефония",
                "kicker": "Звонки и номера",
                "subtitle": "История звонков, рабочие номера и интеграции связи",
                "heading": "Центр телефонии",
                "action": "+ Звонок",
                "launcher": [
                    {"id": "dashboard", "title": "Дашборд", "subtitle": "Телефоны и записи", "icon": "deals"},
                    {"id": "calls", "title": "Звонки", "subtitle": "Журнал и статусы", "icon": "phone"},
                    {"id": "numbers", "title": "Номера", "subtitle": "Линии и сотрудники", "icon": "numbers"},
                    {"id": "integrations", "title": "Интеграции", "subtitle": "АТС и SIP", "icon": "adjustment"},
                    {"id": "operators", "title": "Операторы", "subtitle": "Команда связи", "icon": "clients", "view_id": "numbers"},
                ],
                "filters": ["Поиск", "Провайдер", "Ответственный", "Статус"],
                "columns": ["Клиент", "Номер", "Направление", "Ответственный", "Статус"],
                "list_title": "Журнал звонков",
                "status": "Каркас",
                "empty": "Здесь будет журнал входящих и исходящих звонков с привязкой к CRM и клиентам.",
                "logic": [
                    "Входящие и исходящие звонки по клиентам и ответственным.",
                    "Привязка номеров к сотрудникам, отделам и каналам продаж.",
                    "История звонков в карточке клиента и CRM.",
                    "Интеграции с SIP/АТС и статусами пропущенных звонков.",
                ],
            },
            "messengers": {
                "title": "Месенджеры",
                "kicker": "Диалоги и рассылки",
                "subtitle": "Чаты, кампании, шаблоны и каналы связи",
                "heading": "Центр сообщений",
                "action": "+ Сообщение",
                "launcher": [
                    {"id": "inbox", "title": "Диалоги", "subtitle": "Чаты и обращения", "icon": "chat"},
                    {"id": "campaigns", "title": "Рассылки", "subtitle": "Кампании и сегменты", "icon": "broadcast"},
                    {"id": "templates", "title": "Шаблоны", "subtitle": "Тексты и сценарии", "icon": "template"},
                ],
                "filters": ["Поиск", "Канал", "Ответственный", "Статус"],
                "columns": ["Канал", "Контакт", "Последнее сообщение", "Ответственный", "Статус"],
                "list_title": "Коммуникации",
                "status": "Каркас",
                "empty": "Здесь будут диалоги, рассылки и шаблоны сообщений в стиле ibox.",
                "logic": [
                    "Единый список диалогов по клиентам и каналам связи.",
                    "Шаблоны сообщений для типовых ответов и уведомлений.",
                    "Рассылки по сегментам клиентов и статусам сделок.",
                    "Связь сообщений с CRM, клиентами и продажами.",
                    "История контактов по каждому каналу.",
                ],
            },
        }
        return modules[key]

    def _product_workspace_owner(request: Request) -> tuple[str | None, RedirectResponse | None]:
        u = request.session.get("user") or {}
        if u.get("role") == "admin":
            return None, RedirectResponse(url="/admin", status_code=302)
        wid = str(u.get("workspace_owner_id") or u.get("user_id") or "").strip()
        if not valid_workspace_owner_id(wid):
            return None, RedirectResponse(url="/auth", status_code=302)
        return wid, None

    def _json_object(raw: Any) -> dict[str, Any]:
        return raw if isinstance(raw, dict) else {}

    def _entity_slug(value: str) -> str:
        base = "".join(ch.lower() if ch.isalnum() else "-" for ch in str(value or "").strip())
        base = "-".join(part for part in base.split("-") if part)
        return base[:150] or str(uuid.uuid4())

    def _counterparty_role_flags(kind: str) -> tuple[bool, bool]:
        normalized = str(kind or "").strip().lower()
        return normalized in {"client", "both"}, normalized in {"supplier", "both"}

    def _counterparty_kind_from_flags(is_client: bool, is_supplier: bool) -> str:
        if is_client and is_supplier:
            return "both"
        if is_supplier:
            return "supplier"
        return "client"

    def _manual_counterparty_external_id(name: str, tax_id: str = "") -> str:
        suffix = _entity_slug(tax_id or name)
        return f"manual:{suffix}"[:180]

    def _manual_warehouse_external_id(name: str) -> str:
        return f"manual:{_entity_slug(name)}"[:180]

    def _counterparty_extra(row: Counterparty) -> dict[str, Any]:
        return _json_object(getattr(row, "data", {}))

    def _resolve_counterparty(
        session: Any,
        workspace_owner_id: str,
        *,
        counterparty_id: str = "",
        name: str = "",
        role: str = "client",
    ) -> Counterparty | None:
        row = session.get(Counterparty, counterparty_id) if counterparty_id else None
        if row and row.workspace_owner_id == workspace_owner_id:
            return row
        clean_name = str(name or "").strip()
        if not clean_name:
            return None
        manual_ext = _manual_counterparty_external_id(clean_name)
        row = session.execute(
            select(Counterparty).where(
                Counterparty.workspace_owner_id == workspace_owner_id,
                or_(
                    func.lower(Counterparty.name) == clean_name.lower(),
                    Counterparty.external_id == manual_ext,
                ),
            )
        ).scalars().first()
        if row is None:
            return None
        wants_client = role in {"client", "both"}
        wants_supplier = role in {"supplier", "both"}
        has_client, has_supplier = _counterparty_role_flags(row.kind)
        row.kind = _counterparty_kind_from_flags(has_client or wants_client, has_supplier or wants_supplier)
        return row

    def _ensure_counterparty(
        session: Any,
        workspace_owner_id: str,
        *,
        counterparty_id: str = "",
        name: str,
        role: str,
        phone: str = "",
        tax_id: str = "",
        data: dict[str, Any] | None = None,
    ) -> Counterparty:
        clean_name = str(name or "").strip()
        if not clean_name:
            raise ValueError("Counterparty name is required")
        row = _resolve_counterparty(
            session,
            workspace_owner_id,
            counterparty_id=counterparty_id,
            name=clean_name,
            role=role,
        )
        wants_client = role in {"client", "both"}
        wants_supplier = role in {"supplier", "both"}
        extra = dict(data or {})
        if row is None:
            row = Counterparty(
                id=str(uuid.uuid4()),
                workspace_owner_id=workspace_owner_id,
                kind=_counterparty_kind_from_flags(wants_client, wants_supplier),
                name=clean_name,
                phone=phone[:64],
                tax_id=tax_id[:64],
                external_source="manual",
                external_id=_manual_counterparty_external_id(clean_name, tax_id),
                data=extra,
            )
            session.add(row)
            session.flush()
            return row
        has_client, has_supplier = _counterparty_role_flags(row.kind)
        row.kind = _counterparty_kind_from_flags(has_client or wants_client, has_supplier or wants_supplier)
        row.name = clean_name
        row.phone = phone[:64] or row.phone
        row.tax_id = tax_id[:64] or row.tax_id
        current_extra = _counterparty_extra(row)
        current_extra.update({k: v for k, v in extra.items() if v not in (None, "")})
        current_extra["is_client"] = row.kind in {"client", "both"}
        current_extra["is_supplier"] = row.kind in {"supplier", "both"}
        row.data = current_extra
        row.external_source = row.external_source or "manual"
        row.external_id = row.external_id or _manual_counterparty_external_id(clean_name, tax_id)
        return row

    def _drop_counterparty_role(row: Counterparty, role: str) -> bool:
        has_client, has_supplier = _counterparty_role_flags(row.kind)
        if role == "client":
            has_client = False
        elif role == "supplier":
            has_supplier = False
        remaining = has_client or has_supplier
        if not remaining:
            return False
        row.kind = _counterparty_kind_from_flags(has_client, has_supplier)
        extra = _counterparty_extra(row)
        extra["is_client"] = has_client
        extra["is_supplier"] = has_supplier
        row.data = extra
        return True

    def _resolve_product_row(session: Any, workspace_owner_id: str, raw_value: str) -> Product | None:
        value = str(raw_value or "").strip()
        if not value:
            return None
        row = session.get(Product, value)
        if row and row.workspace_owner_id == workspace_owner_id:
            return row
        lowered = value.lower()
        return session.execute(
            select(Product).where(
                Product.workspace_owner_id == workspace_owner_id,
                or_(
                    func.lower(Product.name) == lowered,
                    func.lower(Product.sku) == lowered,
                    func.lower(Product.barcode) == lowered,
                ),
            )
        ).scalars().first()

    def _resolve_warehouse(session: Any, workspace_owner_id: str, raw_value: str) -> Warehouse | None:
        value = str(raw_value or "").strip()
        if not value:
            return None
        row = session.get(Warehouse, value)
        if row and row.workspace_owner_id == workspace_owner_id:
            return row
        lowered = value.lower()
        return session.execute(
            select(Warehouse).where(
                Warehouse.workspace_owner_id == workspace_owner_id,
                or_(
                    func.lower(Warehouse.name) == lowered,
                    Warehouse.external_id == _manual_warehouse_external_id(value),
                ),
            )
        ).scalars().first()

    def _ensure_warehouse(
        session: Any,
        workspace_owner_id: str,
        *,
        warehouse_id: str = "",
        name: str,
        data: dict[str, Any] | None = None,
    ) -> Warehouse:
        clean_name = str(name or "").strip()
        if not clean_name:
            raise ValueError("Warehouse name is required")
        row = _resolve_warehouse(session, workspace_owner_id, warehouse_id or clean_name)
        if row is None:
            row = Warehouse(
                id=str(uuid.uuid4()),
                workspace_owner_id=workspace_owner_id,
                name=clean_name,
                external_source="manual",
                external_id=_manual_warehouse_external_id(clean_name),
                data=dict(data or {}),
            )
            session.add(row)
            session.flush()
            return row
        row.name = clean_name
        merged = _json_object(row.data)
        merged.update({k: v for k, v in dict(data or {}).items() if v not in (None, "")})
        row.data = merged
        row.external_source = row.external_source or "manual"
        row.external_id = row.external_id or _manual_warehouse_external_id(clean_name)
        return row

    def _product_uses_stock(row: Product | None) -> bool:
        if row is None:
            return False
        kind = str(_json_object(row.data).get("kind") or "product")
        return kind != "service"

    def _product_stocks(row: Product) -> list[dict[str, Any]]:
        raw = _json_object(row.data).get("stocks")
        return [dict(item) for item in raw if isinstance(item, dict)] if isinstance(raw, list) else []

    def _product_available_in_warehouse(row: Product, warehouse_name: str) -> Decimal:
        clean_name = str(warehouse_name or "").strip().lower()
        total = Decimal("0")
        for stock in _product_stocks(row):
            stock_name = str(stock.get("warehouse") or "").strip().lower()
            if clean_name and stock_name != clean_name:
                continue
            total += _sales_decimal(stock.get("quantity"))
        return total

    def _write_product_stocks(row: Product, stocks: list[dict[str, Any]]) -> None:
        data = dict(_json_object(row.data))
        data["stocks"] = stocks
        row.data = data

    def _apply_product_stock_change(
        row: Product,
        warehouse_name: str,
        delta: Decimal,
        *,
        price: str = "",
        op_date: str = "",
        allow_negative_stock: bool = False,
    ) -> None:
        if not delta:
            return
        clean_name = str(warehouse_name or "").strip() or "Основной склад"
        stocks = _product_stocks(row)
        matches = [
            stock for stock in stocks
            if str(stock.get("warehouse") or "").strip().lower() == clean_name.lower()
        ]
        if delta < 0:
            required = abs(delta)
            available = sum((_sales_decimal(stock.get("quantity")) for stock in matches), Decimal("0"))
            if available < required and not allow_negative_stock:
                raise ValueError(f"Недостаточно остатка по товару {row.name}")
            remaining = required
            for stock in matches:
                current = _sales_decimal(stock.get("quantity"))
                if current <= 0:
                    continue
                take = current if current <= remaining else remaining
                stock["quantity"] = str((current - take).normalize() if current - take else "0")
                remaining -= take
                if remaining <= 0:
                    break
            if remaining > 0 and allow_negative_stock:
                target = matches[0] if matches else None
                if target is None:
                    target = {
                        "warehouse": clean_name,
                        "quantity": "0",
                        "price": "",
                        "date": "",
                    }
                    stocks.append(target)
                current = _sales_decimal(target.get("quantity"))
                target["warehouse"] = clean_name
                target["quantity"] = str((current - remaining).normalize() if current - remaining else "0")
                if price:
                    target["price"] = str(price)
                if op_date:
                    target["date"] = str(op_date)
        else:
            target = matches[0] if matches else None
            if target is None:
                target = {
                    "warehouse": clean_name,
                    "quantity": "0",
                    "price": "",
                    "date": "",
                }
                stocks.append(target)
            current = _sales_decimal(target.get("quantity"))
            target["warehouse"] = clean_name
            target["quantity"] = str((current + delta).normalize() if current + delta else "0")
            if price:
                target["price"] = str(price)
            if op_date:
                target["date"] = str(op_date)
        _write_product_stocks(row, stocks)

    def _sync_product_lines(
        session: Any,
        workspace_owner_id: str,
        *,
        warehouse_name: str,
        lines: list[dict[str, Any]],
        delta_sign: int,
        op_date: str = "",
        require_stock_product: bool = False,
        allow_negative_stock: bool = False,
    ) -> list[dict[str, Any]]:
        resolved_lines: list[dict[str, Any]] = []
        managed_rows: list[tuple[Product, dict[str, Any], Decimal, str]] = []
        required_by_product_warehouse: dict[tuple[str, str], Decimal] = {}
        for raw_line in lines:
            line = dict(raw_line)
            line_warehouse = str(line.get("warehouse") or warehouse_name or "Основной склад").strip() or "Основной склад"
            line["warehouse"] = line_warehouse
            quantity = _sales_decimal(line.get("quantity"))
            if quantity <= 0:
                resolved_lines.append(line)
                continue
            product_row = None
            product_id = str(line.get("product_id") or "").strip()
            if product_id:
                row = session.get(Product, product_id)
                if row and row.workspace_owner_id == workspace_owner_id:
                    product_row = row
            if product_row is None:
                product_row = _resolve_product_row(session, workspace_owner_id, str(line.get("product") or ""))
            if product_row is None:
                product_name = str(line.get("product") or "").strip()
                raise ValueError(f"Позиция не найдена: {product_name or 'без названия'}")
            if require_stock_product and not _product_uses_stock(product_row):
                raise ValueError(f"Позиция {product_row.name} является услугой и не может менять склад")
            line["product_id"] = product_row.id
            line["product"] = product_row.name
            line["unit"] = str(_json_object(product_row.data).get("unit") or "Штука")
            if _product_uses_stock(product_row):
                managed_rows.append((product_row, line, quantity, line_warehouse))
                if delta_sign < 0:
                    key = (product_row.id, line_warehouse.lower())
                    required_by_product_warehouse[key] = required_by_product_warehouse.get(key, Decimal("0")) + quantity
            resolved_lines.append(line)
        if delta_sign < 0 and not allow_negative_stock:
            checked: set[tuple[str, str]] = set()
            for product_row, _, _, line_warehouse in managed_rows:
                key = (product_row.id, line_warehouse.lower())
                if key in checked:
                    continue
                checked.add(key)
                required = required_by_product_warehouse.get(key, Decimal("0"))
                if required <= 0:
                    continue
                available = _product_available_in_warehouse(product_row, line_warehouse)
                if available < required:
                    raise ValueError(f"Недостаточно остатка для товара {product_row.name} на складе {line_warehouse}")
        for product_row, line, quantity, line_warehouse in managed_rows:
            _apply_product_stock_change(
                product_row,
                line_warehouse,
                Decimal(delta_sign) * quantity,
                price=str(line.get("price") or ""),
                op_date=op_date,
                allow_negative_stock=allow_negative_stock,
            )
        return resolved_lines

    def _sales_return_quantity_map(lines: list[dict[str, Any]], default_warehouse: str = "") -> dict[tuple[str, str], Decimal]:
        quantities: dict[tuple[str, str], Decimal] = {}
        for raw_line in lines:
            if not isinstance(raw_line, dict):
                continue
            product_name = str(raw_line.get("product") or raw_line.get("product_name") or "").strip().casefold()
            if not product_name:
                continue
            warehouse = str(raw_line.get("warehouse") or default_warehouse or "Основной склад").strip().casefold()
            quantity = _sales_decimal(raw_line.get("quantity"))
            if quantity <= 0:
                continue
            key = (product_name, warehouse)
            quantities[key] = quantities.get(key, Decimal("0")) + quantity
        return quantities

    def _validate_sales_return(session: Any, workspace_owner_id: str, data: dict[str, Any]) -> None:
        source_sale_id = str(data.get("source_sale_id") or "").strip()
        if not source_sale_id:
            raise ValueError("Создайте возврат из исходной продажи в журнале продаж")
        source_row = session.get(SaleDocument, source_sale_id)
        if not source_row or source_row.workspace_owner_id != workspace_owner_id:
            raise ValueError("Исходная продажа для возврата не найдена")
        source_data = _json_object(source_row.data)
        if str(source_data.get("doc_type") or "sale") != "sale":
            raise ValueError("Возврат можно оформить только из продажи")
        source_client = str(source_data.get("client") or "").strip().casefold()
        return_client = str(data.get("client") or "").strip().casefold()
        if source_client and source_client != return_client:
            raise ValueError("Клиент возврата не совпадает с исходной продажей")
        source_quantities = _sales_return_quantity_map(
            list(source_data.get("lines") or []),
            str(source_data.get("warehouse") or "Основной склад"),
        )
        returned_quantities: dict[tuple[str, str], Decimal] = {}
        for return_row in session.execute(
            select(SaleDocument).where(SaleDocument.workspace_owner_id == workspace_owner_id)
        ).scalars():
            return_data = _json_object(return_row.data)
            if str(return_data.get("doc_type") or "") != "return":
                continue
            if str(return_data.get("source_sale_id") or "") != source_sale_id:
                continue
            for key, quantity in _sales_return_quantity_map(
                list(return_data.get("lines") or []),
                str(return_data.get("warehouse") or "Основной склад"),
            ).items():
                returned_quantities[key] = returned_quantities.get(key, Decimal("0")) + quantity
        requested_quantities = _sales_return_quantity_map(
            list(data.get("lines") or []),
            str(data.get("warehouse") or "Основной склад"),
        )
        for key, requested in requested_quantities.items():
            sold = source_quantities.get(key, Decimal("0"))
            already_returned = returned_quantities.get(key, Decimal("0"))
            available = max(Decimal("0"), sold - already_returned)
            if requested > available:
                product_name = next(
                    (
                        str(line.get("product") or line.get("product_name") or "Товар")
                        for line in data.get("lines") or []
                        if str(line.get("product") or line.get("product_name") or "").strip().casefold() == key[0]
                    ),
                    "Товар",
                )
                raise ValueError(
                    f"Нельзя вернуть {requested.normalize()} {product_name}: доступно к возврату {available.normalize()}"
                )
        data["source_sale_number"] = source_row.number

    def _apply_product_price_change(row: Product, price_type: dict[str, Any], price: Any, currency: str) -> None:
        price_value = _sales_decimal(price)
        if price_value <= 0:
            return
        price_type_id = str(price_type.get("id") or "").strip()
        price_type_name = str(price_type.get("name") or "").strip() or "Продажная цена"
        if not price_type_id:
            return
        data = dict(_json_object(row.data))
        prices = [dict(item) for item in data.get("prices", []) if isinstance(item, dict)]
        entry = next((item for item in prices if str(item.get("price_type_id") or "") == price_type_id), None)
        if entry is None:
            entry = next((item for item in prices if str(item.get("name") or "").strip().lower() == price_type_name.lower()), None)
        if entry is None:
            entry = {}
            prices.append(entry)
        entry.update(
            {
                "price_type_id": price_type_id,
                "name": price_type_name,
                "price": str(price_value.normalize() if price_value else ""),
                "currency": str(currency or price_type.get("convert_to_currency") or "UZS").strip().upper() or "UZS",
            }
        )
        data["prices"] = prices
        row.data = data
        flag_modified(row, "data")

    def _sales_signed_balance(item: dict[str, Any]) -> Decimal:
        doc_type = str(item.get("doc_type") or "sale")
        amount = _sales_decimal(item.get("amount"))
        paid = _sales_decimal(item.get("paid_amount"))
        sign = Decimal("-1") if doc_type == "return" else Decimal("1")
        return sign * (amount - paid)

    def _warehouse_operation_type_label(operation_type: str) -> str:
        return {
            "in": "Приход",
            "out": "Списание",
            "adjustment": "Корректировка",
            "transfer": "Перемещение",
        }.get(str(operation_type or ""), "Операция")

    def _crm_type_label(item_type: str) -> str:
        return {
            "task": "Задача",
            "deal": "Сделка",
            "history": "История",
        }.get(str(item_type or ""), "CRM")

    def _crm_status_label(status: str) -> str:
        return {
            "new": "Новый",
            "in_progress": "В работе",
            "won": "Успешно",
            "lost": "Потеряно",
            "done": "Завершено",
            "planned": "Запланировано",
            "archived": "Архив",
        }.get(str(status or ""), "Новый")

    CRM_DEFAULT_STAGES = (
        {"id": "leads", "title": "Лиды", "hint": "Сайт, Instagram, Telegram"},
        {"id": "qualification", "title": "Квалификация", "hint": "Проверка потребности"},
        {"id": "negotiation", "title": "Переговоры", "hint": "Контакт и предложение"},
        {"id": "invoice", "title": "Счёт", "hint": "Заказ, счёт, ожидание оплаты"},
        {"id": "won", "title": "Успешно", "hint": "Сделка закрыта"},
        {"id": "lost", "title": "Потеряно", "hint": "Отказ или пауза"},
    )
    CRM_LEAD_SOURCES = ("Сайт", "Instagram", "Telegram", "WhatsApp", "Facebook", "Звонок", "Ручной ввод")
    CRM_SERVICE_TYPES = (
        "Консультация",
        "Диагностика",
        "Внедрение",
        "Обслуживание",
        "Поддержка",
        "Доставка",
        "Другое",
    )
    CRM_PRIORITY_LABELS = {
        "low": "Низкий",
        "normal": "Обычный",
        "high": "Высокий",
        "urgent": "Срочный",
    }

    def _crm_priority_label(priority: str) -> str:
        clean = str(priority or "normal").strip()
        return CRM_PRIORITY_LABELS.get(clean, CRM_PRIORITY_LABELS["normal"])

    def _crm_probability_value(value: Any) -> str:
        raw = str(value or "").strip().replace("%", "")
        if not raw:
            return ""
        try:
            number = int(Decimal(raw))
        except Exception:
            return ""
        number = max(0, min(100, number))
        return str(number)

    def _crm_due_state(due_date: str, status: str, today_iso: str) -> dict[str, str]:
        if str(status or "") in {"done", "won", "lost", "archived"}:
            return {"id": "closed", "label": "Закрыто"}
        raw = str(due_date or "").strip()
        if not raw:
            return {"id": "", "label": ""}
        try:
            due = datetime.strptime(raw[:10], "%Y-%m-%d").date()
            today = datetime.strptime(today_iso[:10], "%Y-%m-%d").date()
        except ValueError:
            return {"id": "", "label": ""}
        delta = (due - today).days
        if delta < 0:
            return {"id": "overdue", "label": f"Просрочено {abs(delta)} дн."}
        if delta == 0:
            return {"id": "today", "label": "Сегодня"}
        if delta <= 3:
            return {"id": "soon", "label": f"Через {delta} дн."}
        return {"id": "planned", "label": raw[:10]}

    def _crm_stage_slug(value: str, fallback: str = "stage") -> str:
        base = "".join(ch.lower() if ch.isalnum() else "-" for ch in str(value or "").strip())
        base = "-".join(part for part in base.split("-") if part)
        return (base or fallback)[:80]

    def _crm_clean_stages(raw: Any) -> list[dict[str, str]]:
        source = raw if isinstance(raw, list) else []
        stages: list[dict[str, str]] = []
        seen: set[str] = set()
        for index, item in enumerate(source, start=1):
            if isinstance(item, dict):
                title = str(item.get("title") or item.get("name") or "").strip()
                stage_id = str(item.get("id") or "").strip()
                hint = str(item.get("hint") or "").strip()
            else:
                title = str(item or "").strip()
                stage_id = ""
                hint = ""
            if not title:
                continue
            stage_id = _crm_stage_slug(stage_id or title, f"stage-{index}")
            if stage_id in seen:
                stage_id = f"{stage_id}-{index}"
            seen.add(stage_id)
            stages.append({"id": stage_id, "title": title[:80], "hint": hint[:120]})
        return stages or [dict(item) for item in CRM_DEFAULT_STAGES]

    def _crm_workspace_stages(workspace_owner_id: str) -> list[dict[str, str]]:
        data = load_workspace_settings(workspace_owner_id)
        return _crm_clean_stages(data.get("crm_pipeline_stages"))

    def _crm_stage_lookup(stages: list[dict[str, str]]) -> dict[str, dict[str, str]]:
        lookup: dict[str, dict[str, str]] = {}
        for stage in stages:
            lookup[str(stage["id"]).lower()] = stage
            lookup[str(stage["title"]).lower()] = stage
        return lookup

    def _crm_stage_for_value(value: Any, stages: list[dict[str, str]]) -> dict[str, str]:
        lookup = _crm_stage_lookup(stages)
        raw = str(value or "").strip().lower()
        if raw in lookup:
            return lookup[raw]
        return stages[0] if stages else dict(CRM_DEFAULT_STAGES[0])

    def _crm_apply_stage(data: dict[str, Any], stages: list[dict[str, str]], raw_stage: Any = "") -> dict[str, str]:
        stage = _crm_stage_for_value(raw_stage or data.get("stage_id") or data.get("stage"), stages)
        data["stage_id"] = stage["id"]
        data["stage"] = stage["title"]
        return stage

    CLIENT_CRM_STATUS_LABELS = {
        "new_lead": "Новый лид",
        "qualification": "Квалификация",
        "negotiation": "Переговоры",
        "invoice": "Счёт",
        "our_client": "Наш клиент",
        "lost": "Потеряно",
        "in_work": "В работе",
        "paused": "Пауза",
    }

    CLIENT_CRM_STATUS_STAGE_IDS = {
        "new_lead": "leads",
        "qualification": "qualification",
        "negotiation": "negotiation",
        "invoice": "invoice",
        "our_client": "won",
        "lost": "lost",
        "in_work": "negotiation",
        "paused": "lost",
    }

    def _client_crm_status_from_stage(stage: dict[str, str] | None) -> str:
        raw_id = str((stage or {}).get("id") or "").strip().lower()
        raw_title = str((stage or {}).get("title") or "").strip().lower()
        combined = f"{raw_id} {raw_title}"
        if raw_id in {"leads", "lead", "new_lead"} or "лид" in combined:
            return "new_lead"
        if raw_id == "qualification" or "квалиф" in combined:
            return "qualification"
        if raw_id == "negotiation" or "перег" in combined:
            return "negotiation"
        if raw_id == "invoice" or "сч" in combined:
            return "invoice"
        if raw_id == "won" or "усп" in combined or "закрыт" in combined:
            return "our_client"
        if raw_id == "lost" or "потер" in combined or "отказ" in combined:
            return "lost"
        return "new_lead"

    def _crm_stage_for_client_status(status: str, stages: list[dict[str, str]]) -> dict[str, str]:
        status = status if status in CLIENT_CRM_STATUS_LABELS else "new_lead"
        lookup = _crm_stage_lookup(stages)
        for candidate in (CLIENT_CRM_STATUS_STAGE_IDS.get(status, ""), CLIENT_CRM_STATUS_LABELS.get(status, "")):
            key = str(candidate or "").strip().lower()
            if key and key in lookup:
                return lookup[key]
        if status in {"qualification", "negotiation", "invoice", "our_client", "lost"}:
            for stage in stages:
                if _client_crm_status_from_stage(stage) == status:
                    return stage
        return stages[0] if stages else dict(CRM_DEFAULT_STAGES[0])

    def _crm_record_client_keys(row: CrmRecord) -> set[str]:
        data = _json_object(row.data)
        keys = {
            str(row.counterparty_id or "").strip(),
            str(data.get("counterparty_id") or "").strip(),
            str(data.get("client") or "").strip().lower(),
        }
        keys.discard("")
        return keys

    def _crm_status_maps_from_records(rows: list[CrmRecord], stages: list[dict[str, str]]) -> tuple[dict[str, str], dict[str, str]]:
        by_id: dict[str, str] = {}
        by_name: dict[str, str] = {}
        for row in rows:
            data = _json_object(row.data)
            stage = _crm_stage_for_value(data.get("stage_id") or data.get("stage"), stages)
            status = _client_crm_status_from_stage(stage)
            if row.counterparty_id and row.counterparty_id not in by_id:
                by_id[row.counterparty_id] = status
            name = str(data.get("client") or "").strip().lower()
            if name and name not in by_name:
                by_name[name] = status
        return by_id, by_name

    def _sync_counterparty_crm_status_from_stage(
        session: Any,
        workspace_owner_id: str,
        row: CrmRecord,
        stages: list[dict[str, str]],
    ) -> None:
        data = _json_object(row.data)
        stage = _crm_stage_for_value(data.get("stage_id") or data.get("stage"), stages)
        crm_status = _client_crm_status_from_stage(stage)
        counterparty = session.get(Counterparty, row.counterparty_id) if row.counterparty_id else None
        if counterparty is None and data.get("client"):
            counterparty = _resolve_counterparty(session, workspace_owner_id, name=str(data.get("client") or ""), role="client")
        if counterparty is None or counterparty.workspace_owner_id != workspace_owner_id:
            return
        extra = _counterparty_extra(counterparty)
        extra["crm_status"] = crm_status
        counterparty.data = extra
        flag_modified(counterparty, "data")
        row.counterparty_id = counterparty.id
        data["client"] = counterparty.name
        row.data = data
        flag_modified(row, "data")

    def _sync_crm_records_for_counterparty_status(
        session: Any,
        workspace_owner_id: str,
        counterparty: Counterparty,
        crm_status: str,
        stages: list[dict[str, str]],
    ) -> None:
        stage = _crm_stage_for_client_status(crm_status, stages)
        extra = _counterparty_extra(counterparty)
        names = {
            str(counterparty.name or "").strip().lower(),
            str(extra.get("official_name") or "").strip().lower(),
            str(extra.get("legal_name") or "").strip().lower(),
        }
        names.discard("")
        rows = session.execute(
            select(CrmRecord)
            .where(CrmRecord.workspace_owner_id == workspace_owner_id)
            .order_by(CrmRecord.updated_at.desc())
        ).scalars()
        for row in rows:
            data = _json_object(row.data)
            row_name = str(data.get("client") or "").strip().lower()
            if row.counterparty_id != counterparty.id and row_name not in names:
                continue
            _crm_apply_stage(data, stages, stage["id"])
            data["client"] = counterparty.name
            row.data = data
            flag_modified(row, "data")
            row.counterparty_id = counterparty.id
            if stage["id"] == "won":
                row.status = "won"
            elif stage["id"] == "lost":
                row.status = "lost"
            elif row.status in {"new", "won", "lost"}:
                row.status = "in_progress"

    def _crm_stage_lines(stages: list[dict[str, str]]) -> str:
        return "\n".join(stage["title"] for stage in stages)

    def _crm_activity_settings(data: dict[str, Any]) -> dict[str, int]:
        raw = data.get("crm_activity") if isinstance(data.get("crm_activity"), dict) else {}
        try:
            yellow_hours = max(1, min(int(raw.get("yellow_hours") or 24), 720))
        except (TypeError, ValueError):
            yellow_hours = 24
        try:
            red_hours = max(yellow_hours + 1, min(int(raw.get("red_hours") or 48), 1440))
        except (TypeError, ValueError):
            red_hours = 48
        return {"yellow_hours": yellow_hours, "red_hours": red_hours}

    def _crm_activity_state(row: CrmRecord, settings: dict[str, int]) -> dict[str, str]:
        now = datetime.now(timezone.utc)
        touched = row.updated_at or row.created_at or now
        if touched.tzinfo is None:
            touched = touched.replace(tzinfo=timezone.utc)
        age_hours = max((now - touched).total_seconds() / 3600, 0)
        age_days = math.floor(age_hours / 24)
        age_label = "NEW" if age_days < 1 else f"{age_days}д прошло"
        yellow_at = max(settings.get("yellow_hours", 24), 1)
        red_at = max(settings.get("red_hours", 48), yellow_at + 1)
        if age_hours >= red_at:
            return {"state": "red", "fill": "100", "label": "Нет контакта больше лимита", "age_label": age_label}
        if age_hours >= yellow_at:
            fill = round(((age_hours - yellow_at) / (red_at - yellow_at)) * 100)
            return {"state": "yellow", "fill": str(max(12, min(fill, 100))), "label": "Нужно связаться", "age_label": age_label}
        fill = round((age_hours / yellow_at) * 100)
        return {"state": "green", "fill": str(max(8, min(fill, 100))), "label": "Свежая карточка", "age_label": age_label}

    def _crm_actor_name(request: Request) -> str:
        user = request.session.get("user") or {}
        return str(user.get("name") or user.get("username") or user.get("login") or "UPOS").strip() or "UPOS"

    def _crm_append_activity(data: dict[str, Any], action: str, actor: str = "", detail: str = "") -> None:
        raw = data.get("activity_log") if isinstance(data.get("activity_log"), list) else []
        events = [item for item in raw if isinstance(item, dict)][-49:]
        events.append(
            {
                "at": datetime.now(timezone.utc).isoformat(),
                "action": str(action or "Изменение").strip()[:120],
                "actor": str(actor or "UPOS").strip()[:120],
                "detail": str(detail or "").strip()[:300],
            }
        )
        data["activity_log"] = events

    def _crm_open_record_error(data: dict[str, Any], status: str, stages: list[dict[str, str]]) -> str:
        item_type = str(data.get("item_type") or "task")
        if item_type == "history":
            return ""
        stage = _crm_stage_for_value(data.get("stage_id") or data.get("stage"), stages)
        is_lost = status == "lost" or _client_crm_status_from_stage(stage) == "lost"
        if is_lost and not str(data.get("lost_reason") or "").strip():
            return "Укажите причину потери сделки"
        if status in {"done", "won", "lost", "archived"} or is_lost:
            return ""
        if not str(data.get("next_step") or "").strip():
            return "Укажите следующий шаг"
        if not str(data.get("due_date") or "").strip():
            return "Укажите дату следующего шага"
        return ""

    def _product_data(row: Product) -> dict[str, Any]:
        data = row.data if isinstance(row.data, dict) else {}
        prices = data.get("prices") if isinstance(data.get("prices"), list) else []
        stocks = data.get("stocks") if isinstance(data.get("stocks"), list) else []
        purchase_history = data.get("purchase_history") if isinstance(data.get("purchase_history"), list) else []
        raw_variations = data.get("variations") if isinstance(data.get("variations"), list) else []
        variations: list[dict[str, Any]] = []
        for item in raw_variations:
            if not isinstance(item, dict):
                continue
            values = item.get("values")
            if isinstance(values, list):
                values_text = ", ".join(str(part).strip() for part in values if str(part).strip())
            else:
                values_text = str(values or "").strip()
                values = [part.strip() for part in re.split(r"[\n,;]+", values_text) if part.strip()]
            variations.append(
                {
                    "attribute": str(item.get("attribute") or "").strip(),
                    "values": values,
                    "values_text": values_text,
                }
            )
        qty = Decimal("0")
        for item in stocks:
            if not isinstance(item, dict):
                continue
            try:
                qty += Decimal(str(item.get("quantity") or "0"))
            except Exception:
                pass
        if not purchase_history and stocks:
            purchase_history = [
                {
                    "warehouse": str(item.get("warehouse") or ""),
                    "quantity": str(item.get("quantity") or ""),
                    "price": str(item.get("price") or ""),
                    "date": str(item.get("date") or ""),
                }
                for item in stocks
                if isinstance(item, dict) and any(item.get(key) for key in ("warehouse", "quantity", "price", "date"))
            ]
        sale_price = ""
        sale_currency = "UZS"
        if prices and isinstance(prices[0], dict):
            sale_price = str(prices[0].get("price") or "")
            sale_currency = str(prices[0].get("currency") or "UZS").upper()
        return {
            "id": row.id,
            "name": row.name,
            "sku": row.sku,
            "barcode": row.barcode,
            "kind": str(data.get("kind") or "product"),
            "category": str(data.get("category") or ""),
            "folder": str(data.get("folder") or ""),
            "group": str(data.get("group") or ""),
            "brand": str(data.get("brand") or ""),
            "unit": str(data.get("unit") or "Штука"),
            "second_unit": str(data.get("second_unit") or ""),
            "status": str(data.get("status") or "active"),
            "barcode_type": str(data.get("barcode_type") or "EAN13"),
            "batch_tracking": bool(data.get("batch_tracking")),
            "packages": str(data.get("packages") or ""),
            "min_stock": str(data.get("min_stock") or ""),
            "characteristics": str(data.get("characteristics") or ""),
            "classification": str(data.get("classification") or ""),
            "owner": str(data.get("owner") or ""),
            "photo_url": str(data.get("photo_url") or ""),
            "prices": prices,
            "stocks": stocks,
            "purchase_history": purchase_history,
            "purchase_history_json": json.dumps(purchase_history, ensure_ascii=False, indent=2) if purchase_history else "",
            "variations": variations,
            "quantity": str(qty.normalize() if qty else 0),
            "sale_price": sale_price,
            "sale_currency": sale_currency,
            "updated_at": row.updated_at.isoformat() if row.updated_at else "",
        }

    def _product_json_rows(raw: Any, *, note_key: str = "note") -> list[dict[str, Any]]:
        if isinstance(raw, list):
            return [dict(item) for item in raw if isinstance(item, dict)]
        text = str(raw or "").strip()
        if not text:
            return []
        try:
            parsed = json.loads(text)
        except Exception:
            return [{note_key: text}]
        if isinstance(parsed, dict):
            return [parsed]
        if isinstance(parsed, list):
            return [dict(item) for item in parsed if isinstance(item, dict)]
        return []

    def _product_form_payload(form: Any) -> dict[str, Any]:
        def val(name: str, default: str = "") -> str:
            return str(form.get(name) or default).strip()

        prices: list[dict[str, str]] = []
        price_names = list(form.getlist("price_name"))
        price_values = list(form.getlist("price_value"))
        price_currencies = list(form.getlist("price_currency"))
        for idx in range(max(len(price_names), len(price_values), len(price_currencies), 1)):
            name = str(price_names[idx] if idx < len(price_names) else "").strip() or "Продажная цена"
            price = str(price_values[idx] if idx < len(price_values) else "").strip()
            currency = str(price_currencies[idx] if idx < len(price_currencies) else "UZS").strip().upper() or "UZS"
            if name or price:
                prices.append({"name": name, "price": price, "currency": currency})

        stocks: list[dict[str, str]] = []
        warehouses = list(form.getlist("stock_warehouse"))
        quantities = list(form.getlist("stock_quantity"))
        stock_prices = list(form.getlist("stock_price"))
        stock_dates = list(form.getlist("stock_date"))
        for idx in range(max(len(warehouses), len(quantities), len(stock_prices), len(stock_dates), 1)):
            warehouse = str(warehouses[idx] if idx < len(warehouses) else "").strip()
            quantity = str(quantities[idx] if idx < len(quantities) else "").strip()
            price = str(stock_prices[idx] if idx < len(stock_prices) else "").strip()
            date = str(stock_dates[idx] if idx < len(stock_dates) else "").strip()
            if warehouse or quantity or price or date:
                stocks.append({"warehouse": warehouse or "Основной склад", "quantity": quantity, "price": price, "date": date})

        purchase_history: list[dict[str, str]] = []
        purchase_dates = list(form.getlist("purchase_date"))
        purchase_warehouses = list(form.getlist("purchase_warehouse"))
        purchase_quantities = list(form.getlist("purchase_quantity"))
        purchase_prices = list(form.getlist("purchase_price"))
        purchase_suppliers = list(form.getlist("purchase_supplier"))
        purchase_total = max(
            len(purchase_dates),
            len(purchase_warehouses),
            len(purchase_quantities),
            len(purchase_prices),
            len(purchase_suppliers),
            0,
        )
        for idx in range(purchase_total):
            date = str(purchase_dates[idx] if idx < len(purchase_dates) else "").strip()
            warehouse = str(purchase_warehouses[idx] if idx < len(purchase_warehouses) else "").strip()
            quantity = str(purchase_quantities[idx] if idx < len(purchase_quantities) else "").strip()
            price = str(purchase_prices[idx] if idx < len(purchase_prices) else "").strip()
            supplier = str(purchase_suppliers[idx] if idx < len(purchase_suppliers) else "").strip()
            if date or warehouse or quantity or price or supplier:
                purchase_history.append(
                    {
                        "date": date,
                        "warehouse": warehouse or "Основной склад",
                        "quantity": quantity,
                        "price": price,
                        "supplier": supplier,
                    }
                )
        if not purchase_history:
            purchase_history = _product_json_rows(form.get("purchase_history_json"))

        variations: list[dict[str, Any]] = []
        variation_attributes = list(form.getlist("variation_attribute"))
        variation_values = list(form.getlist("variation_values"))
        for idx in range(max(len(variation_attributes), len(variation_values), 0)):
            attribute = str(variation_attributes[idx] if idx < len(variation_attributes) else "").strip()
            raw_values = str(variation_values[idx] if idx < len(variation_values) else "").strip()
            values = [part.strip() for part in re.split(r"[\n,;]+", raw_values) if part.strip()]
            if attribute or values:
                variations.append({"attribute": attribute, "values": values})

        data = {
            "kind": val("kind", "product"),
            "category": val("category"),
            "folder": val("folder"),
            "group": val("group"),
            "brand": val("brand"),
            "unit": val("unit", "Штука"),
            "second_unit": val("second_unit"),
            "status": val("status", "active"),
            "barcode_type": val("barcode_type", "EAN13"),
            "batch_tracking": val("batch_tracking") == "1",
            "packages": val("packages"),
            "min_stock": val("min_stock"),
            "characteristics": val("characteristics"),
            "classification": val("classification"),
            "owner": val("owner"),
            "photo_url": val("photo_url"),
            "prices": prices,
            "stocks": stocks,
            "purchase_history": purchase_history,
            "variations": variations,
        }
        return data

    def _product_filters_payload(
        q: str = "",
        category: str = "",
        group: str = "",
        brand: str = "",
        folder: str = "",
        status: str = "active",
        kind: str = "product",
        category_values: list[str] | None = None,
        group_values: list[str] | None = None,
        brand_values: list[str] | None = None,
        folder_values: list[str] | None = None,
        status_values: list[str] | None = None,
        kind_values: list[str] | None = None,
    ) -> dict[str, Any]:
        def clean_values(values: list[str] | None, fallback: str = "") -> list[str]:
            source = values if values is not None else [fallback]
            out: list[str] = []
            seen: set[str] = set()
            for value in source:
                text = str(value or "").strip()
                if not text or text in seen:
                    continue
                out.append(text)
                seen.add(text)
            return out

        selected_kind = clean_values(kind_values, kind.strip() or "product")
        if not selected_kind:
            selected_kind = ["product"]
        selected_status = clean_values(status_values, status.strip() or "active")
        if not selected_status:
            selected_status = ["active"]
        selected_category = clean_values(category_values, category)
        selected_group = clean_values(group_values, group)
        selected_brand = clean_values(brand_values, brand)
        selected_folder = clean_values(folder_values, folder)
        return {
            "q": q.strip(),
            "category": selected_category[0] if selected_category else "",
            "group": selected_group[0] if selected_group else "",
            "brand": selected_brand[0] if selected_brand else "",
            "folder": selected_folder[0] if selected_folder else "",
            "status": selected_status[0] if selected_status else "active",
            "kind": selected_kind[0] if selected_kind else "product",
            "category_values": selected_category,
            "group_values": selected_group,
            "brand_values": selected_brand,
            "folder_values": selected_folder,
            "status_values": selected_status,
            "kind_values": selected_kind,
        }

    def _product_page_size(raw: Any = 100) -> int:
        try:
            value = int(str(raw or "100").strip())
        except Exception:
            value = 100
        return value if value in {100, 500, 1000, 10000} else 100

    def _positive_int(raw: Any, default: int = 1) -> int:
        try:
            value = int(str(raw or str(default)).strip())
        except Exception:
            value = default
        return max(1, value)

    PRODUCT_PRICE_TYPE_DEFAULTS = [
        {
            "id": "1",
            "name": "ПРОДАЖНАЯ ЦЕНА",
            "sort_order": 1,
            "is_for_sales": True,
            "is_for_purchases": False,
            "is_active": True,
            "pricing_method": "manual",
            "created_by": "Admin123",
            "updated_at": "2026-06-21",
            "base_price_type_id": "",
            "markup_type": "markup",
            "markup_value": "",
            "convert_to_currency": "UZS",
            "rounding": "1.0",
        },
        {
            "id": "2",
            "name": "PRIXOD CENA",
            "sort_order": 2,
            "is_for_sales": False,
            "is_for_purchases": True,
            "is_active": False,
            "pricing_method": "dependent",
            "created_by": "Admin123",
            "updated_at": "2026-06-21",
            "base_price_type_id": "last_purchase_price",
            "markup_type": "markup",
            "markup_value": "30",
            "convert_to_currency": "USD",
            "rounding": "0.001",
        },
        {
            "id": "3",
            "name": "СКИДОЧНЫЕ ЦЕНА",
            "sort_order": 3,
            "is_for_sales": True,
            "is_for_purchases": False,
            "is_active": True,
            "pricing_method": "manual",
            "created_by": "Admin123",
            "updated_at": "2026-06-21",
            "base_price_type_id": "",
            "markup_type": "discount",
            "markup_value": "",
            "convert_to_currency": "UZS",
            "rounding": "1.0",
        },
        {
            "id": "4",
            "name": "СТАНДАРТНАЯ ЦЕНА",
            "sort_order": 4,
            "is_for_sales": True,
            "is_for_purchases": False,
            "is_active": False,
            "pricing_method": "manual",
            "created_by": "Admin123",
            "updated_at": "2026-06-21",
            "base_price_type_id": "",
            "markup_type": "markup",
            "markup_value": "",
            "convert_to_currency": "UZS",
            "rounding": "1.0",
        },
        {
            "id": "retail",
            "name": "Розничная цена",
            "sort_order": 5,
            "is_for_sales": True,
            "is_for_purchases": False,
            "is_active": True,
            "pricing_method": "manual",
            "created_by": "Admin123",
            "updated_at": "2026-07-07",
            "base_price_type_id": "",
            "markup_type": "markup",
            "markup_value": "",
            "convert_to_currency": "UZS",
            "rounding": "100.0",
        },
        {
            "id": "wholesale",
            "name": "Оптовая цена",
            "sort_order": 6,
            "is_for_sales": True,
            "is_for_purchases": False,
            "is_active": True,
            "pricing_method": "manual",
            "created_by": "Admin123",
            "updated_at": "2026-07-07",
            "base_price_type_id": "",
            "markup_type": "markup",
            "markup_value": "",
            "convert_to_currency": "UZS",
            "rounding": "100.0",
        },
        {
            "id": "vip",
            "name": "VIP цена",
            "sort_order": 7,
            "is_for_sales": True,
            "is_for_purchases": False,
            "is_active": True,
            "pricing_method": "manual",
            "created_by": "Admin123",
            "updated_at": "2026-07-07",
            "base_price_type_id": "",
            "markup_type": "markup",
            "markup_value": "",
            "convert_to_currency": "UZS",
            "rounding": "100.0",
        },
        {
            "id": "marketplace",
            "name": "Маркетплейс",
            "sort_order": 8,
            "is_for_sales": True,
            "is_for_purchases": False,
            "is_active": True,
            "pricing_method": "manual",
            "created_by": "Admin123",
            "updated_at": "2026-07-07",
            "base_price_type_id": "",
            "markup_type": "markup",
            "markup_value": "",
            "convert_to_currency": "UZS",
            "rounding": "100.0",
        },
        {
            "id": "promo",
            "name": "Акционная цена",
            "sort_order": 9,
            "is_for_sales": True,
            "is_for_purchases": False,
            "is_active": True,
            "pricing_method": "manual",
            "created_by": "Admin123",
            "updated_at": "2026-07-07",
            "base_price_type_id": "",
            "markup_type": "discount",
            "markup_value": "",
            "convert_to_currency": "UZS",
            "rounding": "100.0",
        },
    ]
    PRODUCT_PRICE_ROUNDINGS = ["0.0001", "0.001", "0.01", "0.1", "0.5", "1.0", "5.0", "10.0", "50.0", "100.0", "500.0", "1000.0", "5000.0", "10000.0"]
    PRODUCT_USD_RATE = Decimal("12000")

    def _workspace_usd_uzs_rate(workspace_owner_id: str) -> Decimal:
        try:
            data = load_workspace_settings(workspace_owner_id)
        except Exception:
            data = {}
        rate = _sales_decimal(data.get("usd_uzs_rate") if isinstance(data, dict) else "")
        return rate if rate > 0 else PRODUCT_USD_RATE

    def _normalize_price_type(raw: dict[str, Any], fallback: dict[str, Any] | None = None) -> dict[str, Any]:
        base = dict(fallback or {})
        base.update(raw if isinstance(raw, dict) else {})
        price_type_id = str(base.get("id") or uuid.uuid4()).strip()
        name = str(base.get("name") or "Новый прайс-лист").strip()
        method = str(base.get("pricing_method") or "manual").strip()
        if method not in {"manual", "dependent"}:
            method = "manual"
        markup_type = str(base.get("markup_type") or "markup").strip()
        if markup_type not in {"markup", "discount"}:
            markup_type = "markup"
        currency = str(base.get("convert_to_currency") or "UZS").strip().upper()
        if currency not in {"UZS", "USD"}:
            currency = "UZS"
        rounding = str(base.get("rounding") or "1.0").strip()
        if rounding not in PRODUCT_PRICE_ROUNDINGS:
            rounding = "1.0"
        try:
            sort_order = int(str(base.get("sort_order") or "0"))
        except Exception:
            sort_order = 0
        return {
            "id": price_type_id,
            "name": name,
            "sort_order": sort_order,
            "is_for_sales": bool(base.get("is_for_sales")),
            "is_for_purchases": bool(base.get("is_for_purchases")),
            "is_active": bool(base.get("is_active")),
            "pricing_method": method,
            "created_by": str(base.get("created_by") or "Admin123"),
            "updated_at": str(base.get("updated_at") or datetime.now(timezone.utc).date().isoformat()),
            "base_price_type_id": str(base.get("base_price_type_id") or "").strip(),
            "markup_type": markup_type,
            "markup_value": str(base.get("markup_value") or "").strip(),
            "convert_to_currency": currency,
            "rounding": rounding,
        }

    def _workspace_price_types(workspace_owner_id: str) -> list[dict[str, Any]]:
        settings = load_workspace_settings(workspace_owner_id)
        stored = settings.get("product_price_types") if isinstance(settings.get("product_price_types"), list) else []
        by_id: dict[str, dict[str, Any]] = {
            item["id"]: _normalize_price_type(item)
            for item in PRODUCT_PRICE_TYPE_DEFAULTS
        }
        for item in stored:
            if not isinstance(item, dict):
                continue
            normalized = _normalize_price_type(item, by_id.get(str(item.get("id") or "")))
            by_id[normalized["id"]] = normalized
        result = sorted(by_id.values(), key=lambda item: (int(item.get("sort_order") or 0), str(item.get("name") or "")))
        if stored != result:
            settings["product_price_types"] = result
            save_workspace_settings(workspace_owner_id, settings)
        return result

    def _save_workspace_price_types(workspace_owner_id: str, price_types: list[dict[str, Any]]) -> None:
        settings = load_workspace_settings(workspace_owner_id)
        settings["product_price_types"] = sorted(
            [_normalize_price_type(item) for item in price_types],
            key=lambda item: (int(item.get("sort_order") or 0), str(item.get("name") or "")),
        )
        save_workspace_settings(workspace_owner_id, settings)

    def _price_type_by_id(price_types: list[dict[str, Any]], price_type_id: str) -> dict[str, Any]:
        return next((item for item in price_types if str(item.get("id")) == str(price_type_id)), price_types[0])

    def _product_price_entry(product: dict[str, Any], price_type: dict[str, Any]) -> dict[str, Any] | None:
        target_id = str(price_type.get("id") or "")
        target_name = str(price_type.get("name") or "").strip().lower()
        for price in product.get("prices") or []:
            if not isinstance(price, dict):
                continue
            if str(price.get("price_type_id") or "") == target_id:
                return price
        for price in product.get("prices") or []:
            if isinstance(price, dict) and str(price.get("name") or "").strip().lower() == target_name:
                return price
        if target_id == "1":
            for price in product.get("prices") or []:
                if isinstance(price, dict):
                    return price
        return None

    def _product_last_purchase(product: dict[str, Any]) -> tuple[Decimal, str]:
        rows = product.get("purchase_history") if isinstance(product.get("purchase_history"), list) else []
        for item in reversed(rows):
            if not isinstance(item, dict):
                continue
            value = _sales_decimal(item.get("price"))
            if value:
                return value, str(item.get("currency") or "UZS").upper()
        stocks = product.get("stocks") if isinstance(product.get("stocks"), list) else []
        for item in reversed(stocks):
            if not isinstance(item, dict):
                continue
            value = _sales_decimal(item.get("price"))
            if value:
                return value, str(item.get("currency") or "UZS").upper()
        return Decimal("0"), "UZS"

    def _convert_product_currency(value: Decimal, from_currency: str, to_currency: str, usd_rate: Decimal | None = None) -> Decimal:
        source = str(from_currency or "UZS").upper()
        target = str(to_currency or "UZS").upper()
        rate = usd_rate if usd_rate and usd_rate > 0 else PRODUCT_USD_RATE
        if source == target:
            return value
        if source == "USD" and target == "UZS":
            return value * rate
        if source == "UZS" and target == "USD":
            return value / rate
        return value

    def _round_product_price(value: Decimal, step_raw: Any) -> Decimal:
        step = _sales_decimal(step_raw)
        if step <= 0:
            return value
        return (value / step).to_integral_value(rounding=ROUND_HALF_UP) * step

    def _calculated_product_price(
        product: dict[str, Any],
        price_type: dict[str, Any],
        price_types: list[dict[str, Any]],
        usd_rate: Decimal | None = None,
    ) -> tuple[str, str]:
        if price_type.get("pricing_method") != "dependent":
            entry = _product_price_entry(product, price_type)
            if not entry:
                return "", str(price_type.get("convert_to_currency") or "UZS")
            return str(entry.get("price") or ""), str(entry.get("currency") or "UZS").upper()
        base_id = str(price_type.get("base_price_type_id") or "")
        if base_id == "last_purchase_price":
            base, base_currency = _product_last_purchase(product)
        else:
            base_type = next((item for item in price_types if str(item.get("id")) == base_id), None)
            if not base_type:
                base, base_currency = Decimal("0"), "UZS"
            else:
                raw_value, base_currency = _calculated_product_price(product, base_type, price_types, usd_rate)
                base = _sales_decimal(raw_value)
        if not base:
            return "", str(price_type.get("convert_to_currency") or "UZS")
        percent = _sales_decimal(price_type.get("markup_value"))
        factor = Decimal("1") + (percent / Decimal("100"))
        if price_type.get("markup_type") == "discount":
            factor = Decimal("1") - (percent / Decimal("100"))
        value = base * factor
        currency = str(price_type.get("convert_to_currency") or base_currency or "UZS").upper()
        value = _convert_product_currency(value, base_currency, currency, usd_rate)
        value = _round_product_price(value, price_type.get("rounding"))
        return _decimal_plain_text(value), currency

    def _price_type_product_rows(products: list[dict[str, Any]], price_type: dict[str, Any], price_types: list[dict[str, Any]]) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for product in products:
            price, currency = _calculated_product_price(product, price_type, price_types)
            rows.append({**product, "price_type_price": price, "price_type_currency": currency, "has_price": bool(price)})
        return rows

    def _product_search_variants(raw: Any) -> list[str]:
        text = str(raw or "").lower()
        spaced = re.sub(r"[^0-9a-zа-яё]+", " ", text, flags=re.IGNORECASE).strip()
        compact = re.sub(r"[^0-9a-zа-яё]+", "", text, flags=re.IGNORECASE)
        parts = re.split(r"\s+", spaced) if spaced else []
        return [item for item in (text, spaced, compact, *parts) if item]

    def _product_query_terms(raw: Any) -> list[str]:
        text = str(raw or "").strip().lower()
        if not text:
            return []
        terms = re.split(r"\s+", text)
        out: list[str] = []
        seen: set[str] = set()
        for term in terms:
            for variant in _product_search_variants(term):
                if variant and variant not in seen:
                    out.append(variant)
                    seen.add(variant)
        return out

    def _product_matches_filters(item: dict[str, Any], filters: dict[str, Any]) -> bool:
        kind_values = [str(value) for value in filters.get("kind_values", []) if str(value)]
        status_values = [str(value) for value in filters.get("status_values", []) if str(value)]
        category_values = [str(value) for value in filters.get("category_values", []) if str(value)]
        group_values = [str(value) for value in filters.get("group_values", []) if str(value)]
        brand_values = [str(value) for value in filters.get("brand_values", []) if str(value)]
        folder_values = [str(value) for value in filters.get("folder_values", []) if str(value)]
        if kind_values and "all" not in kind_values and item["kind"] not in kind_values:
            return False
        if status_values and "all" not in status_values and item["status"] not in status_values:
            return False
        if category_values and item["category"] not in category_values:
            return False
        if group_values and item["group"] not in group_values:
            return False
        if brand_values and item["brand"] not in brand_values:
            return False
        if folder_values and item["folder"] not in folder_values:
            return False
        hay_parts = [
            item["name"],
            item["sku"],
            item["barcode"],
            item["category"],
            item["brand"],
            item["folder"],
        ]
        hay = " ".join(
            variant
            for part in hay_parts
            for variant in _product_search_variants(part)
        )
        q_terms = _product_query_terms(filters["q"])
        if q_terms and not all(term in hay for term in q_terms):
            return False
        return True

    def _collect_products_view_data(
        session: Any,
        workspace_owner_id: str,
        filters: dict[str, str],
        *,
        edit: str = "",
    ) -> tuple[list[dict[str, Any]], dict[str, Any] | None, dict[str, list[str]]]:
        products: list[dict[str, Any]] = []
        all_items: list[dict[str, Any]] = []
        edit_product = None
        rows = list(
            session.execute(
                select(Product)
                .where(Product.workspace_owner_id == workspace_owner_id)
                .order_by(Product.updated_at.desc())
            ).scalars()
        )
        for row in rows:
            item = _product_data(row)
            all_items.append(item)
            if not _product_matches_filters(item, filters):
                continue
            products.append(item)
            if edit and row.id == edit:
                edit_product = item
        if edit and edit_product is None:
            found = session.get(Product, edit)
            if found and found.workspace_owner_id == workspace_owner_id:
                edit_product = _product_data(found)
        settings_payload = load_workspace_settings(workspace_owner_id)
        price_lists = _workspace_price_types(workspace_owner_id)
        for price_type in price_lists:
            rows_with_price = _price_type_product_rows(all_items, price_type, price_lists)
            price_type["product_count"] = sum(1 for row in rows_with_price if row["has_price"])
            price_type["type"] = "Зависимая" if price_type.get("pricing_method") == "dependent" else "Ручная"
            price_type["currency"] = str(price_type.get("convert_to_currency") or "UZS")
            price_type["default_value"] = ""
        stored_categories = [
            str(item or "").strip()
            for item in settings_payload.get("product_categories", [])
            if str(item or "").strip()
        ]
        category_counts: dict[str, int] = {}
        for product in all_items:
            category_name = str(product.get("category") or "").strip()
            if category_name:
                category_counts[category_name] = category_counts.get(category_name, 0) + 1
        category_names = sorted({*stored_categories, *category_counts.keys()})
        options = {
            "categories": category_names,
            "category_rows": [
                {"name": name, "product_count": category_counts.get(name, 0)}
                for name in category_names
            ],
            "groups": sorted({p["group"] for p in all_items if p["group"]}),
            "brands": sorted({p["brand"] for p in all_items if p["brand"]}),
            "folders": sorted({p["folder"] for p in all_items if p["folder"]}),
            "price_types": sorted(
                {
                    str(price.get("name") or "")
                    for p in products
                    for price in p["prices"]
                    if isinstance(price, dict) and price.get("name")
                }
            ),
            "price_lists": sorted(price_lists, key=lambda item: (int(item.get("sort_order") or 0), str(item.get("name") or ""))),
        }
        return products, edit_product, options

    def _save_product_categories(workspace_owner_id: str, names: list[str]) -> None:
        clean = sorted({str(name or "").strip() for name in names if str(name or "").strip()})
        data = load_workspace_settings(workspace_owner_id)
        data["product_categories"] = clean
        save_workspace_settings(workspace_owner_id, data)

    def _price_type_redirect(price_type_id: str = "", *, msg: str = "saved") -> RedirectResponse:
        suffix = f"?price_type={quote(str(price_type_id))}&msg={quote(msg)}" if price_type_id else f"?msg={quote(msg)}"
        anchor = "price-type-detail" if price_type_id else "price-types"
        return RedirectResponse(url=f"/products{suffix}#{anchor}", status_code=302)

    def _product_excel_text(raw: Any) -> str:
        return str(raw or "").strip()

    def _decimal_plain_text(value: Decimal) -> str:
        if not value:
            return ""
        text = format(value, "f")
        if "." in text:
            text = text.rstrip("0").rstrip(".")
        return text

    def _product_excel_decimal_text(raw: Any) -> str:
        value = _sales_decimal(raw)
        return _decimal_plain_text(value)

    def _product_excel_bool(raw: Any) -> bool:
        return str(raw or "").strip().lower() in {"1", "true", "yes", "y", "да", "активно"}

    def _product_excel_kind(raw: Any, fallback: str = "product") -> str:
        normalized = str(raw or "").strip().lower()
        mapping = {
            "product": "product",
            "товар": "product",
            "simple": "product",
            "service": "service",
            "услуга": "service",
            "collection": "collection",
            "комплект": "collection",
            "коллекция": "collection",
        }
        return mapping.get(normalized, fallback)

    def _product_excel_status(raw: Any, fallback: str = "active") -> str:
        normalized = str(raw or "").strip().lower()
        mapping = {
            "active": "active",
            "активный": "active",
            "inactive": "inactive",
            "неактивный": "inactive",
        }
        return mapping.get(normalized, fallback)

    def _product_import_redirect_url(kind: str, *, msg: str = "", count: int = 0, error: str = "") -> str:
        query: dict[str, str] = {}
        clean_kind = str(kind or "").strip()
        if clean_kind and clean_kind != "product":
            query["kind"] = clean_kind
        if msg:
            query["msg"] = msg
        if count > 0:
            query["count"] = str(count)
        if error:
            query["error"] = error
        suffix = f"?{urlencode(query)}" if query else ""
        target_hash = "#service" if clean_kind == "service" else "#catalog"
        return f"/products{suffix}{target_hash}"

    def _product_list_redirect_url(
        *,
        q: str = "",
        category: str = "",
        group: str = "",
        brand: str = "",
        folder: str = "",
        status: str = "active",
        kind: str = "product",
        msg: str = "",
        count: int = 0,
        error: str = "",
        anchor: str = "catalog",
        category_values: list[str] | None = None,
        group_values: list[str] | None = None,
        brand_values: list[str] | None = None,
        folder_values: list[str] | None = None,
        status_values: list[str] | None = None,
        kind_values: list[str] | None = None,
        **_unused: Any,
    ) -> str:
        query: dict[str, Any] = {}
        def values_or_single(values: list[str] | None, fallback: str = "") -> list[str]:
            source = values if values is not None else [fallback]
            out: list[str] = []
            seen: set[str] = set()
            for item in source:
                text = str(item or "").strip()
                if not text or text in seen:
                    continue
                out.append(text)
                seen.add(text)
            return out

        if q.strip():
            query["q"] = q.strip()
        selected_category = values_or_single(category_values, category)
        selected_group = values_or_single(group_values, group)
        selected_brand = values_or_single(brand_values, brand)
        selected_folder = values_or_single(folder_values, folder)
        selected_status = values_or_single(status_values, status.strip() or "active")
        selected_kind = values_or_single(kind_values, kind.strip() or "product")
        if selected_category:
            query["category"] = selected_category
        if selected_group:
            query["group"] = selected_group
        if selected_brand:
            query["brand"] = selected_brand
        if selected_folder:
            query["folder"] = selected_folder
        if selected_status:
            query["status"] = selected_status
        if selected_kind:
            query["kind"] = selected_kind
        if msg:
            query["msg"] = msg
        if count > 0:
            query["count"] = str(count)
        if error:
            query["error"] = error
        suffix = f"?{urlencode(query, doseq=True)}" if query else ""
        target_hash = f"#{anchor}" if anchor else ""
        return f"/products{suffix}{target_hash}"

    @app.get("/products", response_class=HTMLResponse, name="products_get")
    def products_get(
        request: Request,
        q: str = "",
        category: str = "",
        group: str = "",
        brand: str = "",
        folder: str = "",
        status: str = "active",
        kind: str = "product",
        page_size: str = "100",
        page: str = "1",
        edit: str = "",
        price_type: str = "",
        price_page: str = "1",
        price_q: str = "",
        price_category: str = "",
        price_group: str = "",
        price_brand: str = "",
        price_status: str = "all",
    ):
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        query = request.query_params
        filters = _product_filters_payload(
            q,
            category,
            group,
            brand,
            folder,
            status,
            kind,
            category_values=query.getlist("category"),
            group_values=query.getlist("group"),
            brand_values=query.getlist("brand"),
            folder_values=query.getlist("folder"),
            status_values=query.getlist("status"),
            kind_values=query.getlist("kind"),
        )
        product_page_size = _product_page_size(page_size)
        with session_scope() as session:
            products, edit_product, options = _collect_products_view_data(session, wid, filters, edit=edit)
            price_products = [
                _product_data(row)
                for row in session.execute(
                    select(Product)
                    .where(Product.workspace_owner_id == wid)
                    .order_by(Product.name.asc())
                ).scalars()
            ]
        products_total = len(products)
        product_total_pages = max(1, math.ceil(products_total / product_page_size))
        product_page = min(_positive_int(page, 1), product_total_pages)
        product_page_start = (product_page - 1) * product_page_size
        product_page_end = product_page_start + product_page_size
        products = products[product_page_start:product_page_end]

        def product_page_url(target_page: int) -> str:
            pairs = [
                (str(key), str(value))
                for key, value in request.query_params.multi_items()
                if str(key) != "page"
            ]
            pairs.append(("page", str(target_page)))
            return f"{request.url.path}?{urlencode(pairs, doseq=True)}#catalog"

        product_pagination_pages = [
            page_no
            for page_no in range(max(1, product_page - 2), min(product_total_pages, product_page + 2) + 1)
        ]
        price_types = list(options.get("price_lists") or _workspace_price_types(wid))
        selected_price_type = _price_type_by_id(price_types, price_type) if price_types else {}
        price_rows = _price_type_product_rows(price_products, selected_price_type, price_types) if selected_price_type else []
        price_q_clean = str(price_q or "").strip().lower()
        if price_q_clean:
            price_rows = [
                row for row in price_rows
                if price_q_clean in " ".join([row.get("name", ""), row.get("sku", ""), row.get("barcode", "")]).lower()
            ]
        if price_category:
            price_rows = [row for row in price_rows if row.get("category") == price_category]
        if price_group:
            price_rows = [row for row in price_rows if row.get("group") == price_group]
        if price_brand:
            price_rows = [row for row in price_rows if row.get("brand") == price_brand]
        if price_status == "with":
            price_rows = [row for row in price_rows if row.get("has_price")]
        elif price_status == "without":
            price_rows = [row for row in price_rows if not row.get("has_price")]
        price_page_size = 15
        price_rows_total = len(price_rows)
        price_total_pages = max(1, math.ceil(price_rows_total / price_page_size))
        price_current_page = min(_positive_int(price_page, 1), price_total_pages)
        price_start = (price_current_page - 1) * price_page_size
        price_end = price_start + price_page_size

        def price_page_url(target_page: int) -> str:
            pairs = [
                (str(key), str(value))
                for key, value in request.query_params.multi_items()
                if str(key) != "price_page"
            ]
            pairs.extend(
                [
                    ("price_type", str(selected_price_type.get("id") or "")),
                    ("price_page", str(target_page)),
                ]
            )
            return f"{request.url.path}?{urlencode(pairs, doseq=True)}#price-type-detail"

        price_type_state = {
            "price_types": price_types,
            "selected": selected_price_type,
            "rows": price_rows[price_start:price_end],
            "filters": {
                "q": price_q,
                "category": price_category,
                "group": price_group,
                "brand": price_brand,
                "status": price_status,
            },
            "total": price_rows_total,
            "page": price_current_page,
            "total_pages": price_total_pages,
            "from": (price_start + 1) if price_rows_total else 0,
            "to": min(price_end, price_rows_total),
            "prev_url": price_page_url(max(1, price_current_page - 1)),
            "next_url": price_page_url(min(price_total_pages, price_current_page + 1)),
            "page_urls": {page_no: price_page_url(page_no) for page_no in range(max(1, price_current_page - 2), min(price_total_pages, price_current_page + 2) + 1)},
            "roundings": PRODUCT_PRICE_ROUNDINGS,
        }
        return tpl(
            request,
            "home_products.html",
            variant="user",
            active="products",
            products=products,
            products_total=products_total,
            product_page_size=product_page_size,
            product_page=product_page,
            product_total_pages=product_total_pages,
            product_page_from=(product_page_start + 1) if products_total else 0,
            product_page_to=min(product_page_end, products_total),
            product_pagination_pages=product_pagination_pages,
            product_prev_page_url=product_page_url(max(1, product_page - 1)),
            product_next_page_url=product_page_url(min(product_total_pages, product_page + 1)),
            product_page_urls={page_no: product_page_url(page_no) for page_no in product_pagination_pages},
            product_filters=filters,
            product_options=options,
            price_type_state=price_type_state,
            edit_product=edit_product,
            flash_ok=request.query_params.get("msg"),
            imported_count=request.query_params.get("count") or "0",
            bulk_updated_count=request.query_params.get("count") or "0",
            flash_err=request.query_params.get("error"),
        )

    @app.get("/products/export", name="products_export")
    def products_export(
        request: Request,
        q: str = "",
        category: str = "",
        group: str = "",
        brand: str = "",
        folder: str = "",
        status: str = "active",
        kind: str = "product",
    ):
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        query = request.query_params
        filters = _product_filters_payload(
            q,
            category,
            group,
            brand,
            folder,
            status,
            kind,
            category_values=query.getlist("category"),
            group_values=query.getlist("group"),
            brand_values=query.getlist("brand"),
            folder_values=query.getlist("folder"),
            status_values=query.getlist("status"),
            kind_values=query.getlist("kind"),
        )
        with session_scope() as session:
            products, _, _ = _collect_products_view_data(session, wid, filters)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Products"
        headers = [
            "ID",
            "Тип",
            "Название",
            "Артикул",
            "Штрихкод",
            "Категория",
            "Папка",
            "Группа",
            "Бренд",
            "Единица",
            "Вторая единица",
            "Статус",
            "Цена продажи",
            "Валюта цены",
            "Склад",
            "Количество",
            "Себестоимость",
            "Дата остатка",
            "Тип штрихкода",
            "Партийный учет",
            "Упаковки",
            "Мин. остаток",
            "Характеристики",
            "Классификация",
            "Владелец",
            "Цены JSON",
            "Остатки JSON",
        ]
        headers.extend(["Photo URL", "Purchase History JSON"])
        sheet.append(headers)
        for product in products:
            prices = product["prices"] if isinstance(product.get("prices"), list) else []
            stocks = product["stocks"] if isinstance(product.get("stocks"), list) else []
            primary_price = prices[0] if prices and isinstance(prices[0], dict) else {}
            primary_stock = stocks[0] if stocks and isinstance(stocks[0], dict) else {}
            export_row = [
                    product["id"],
                    product["kind"],
                    product["name"],
                    product["sku"],
                    product["barcode"],
                    product["category"],
                    product["folder"],
                    product["group"],
                    product["brand"],
                    product["unit"],
                    product["second_unit"],
                    product["status"],
                    str(primary_price.get("price") or ""),
                    str(primary_price.get("currency") or product["sale_currency"] or "UZS"),
                    str(primary_stock.get("warehouse") or ""),
                    product["quantity"],
                    str(primary_stock.get("price") or ""),
                    str(primary_stock.get("date") or ""),
                    product["barcode_type"],
                    "1" if product["batch_tracking"] else "0",
                    product["packages"],
                    product["min_stock"],
                    product["characteristics"],
                    product["classification"],
                    product["owner"],
                    json.dumps(prices, ensure_ascii=False),
                    json.dumps(stocks, ensure_ascii=False),
                ]
            export_row.extend(
                [
                    product["photo_url"],
                    json.dumps(product["purchase_history"], ensure_ascii=False),
                ]
            )
            sheet.append(export_row)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        stream = io.BytesIO()
        workbook.save(stream)
        filename = f"products-{filters['kind'] or 'all'}-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"
        return Response(
            content=stream.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
        )

    @app.post("/products/import", name="products_import")
    async def products_import(
        request: Request,
        csrf_token: str = Form(default=""),
        kind: str = Form(default="product"),
        excel_file: UploadFile = File(...),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url=_product_import_redirect_url(kind, error="Форма устарела. Обновите страницу и повторите."), status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        if not excel_file.filename or not excel_file.filename.lower().endswith(".xlsx"):
            return RedirectResponse(url=_product_import_redirect_url(kind, error="Нужен файл Excel в формате .xlsx"), status_code=302)
        try:
            payload = await excel_file.read()
            workbook = load_workbook(io.BytesIO(payload), data_only=True)
        except Exception:
            return RedirectResponse(url=_product_import_redirect_url(kind, error="Не удалось прочитать Excel-файл"), status_code=302)

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return RedirectResponse(url=_product_import_redirect_url(kind, error="Excel-файл пустой"), status_code=302)

        header_map = {
            _product_excel_text(value).lower(): index
            for index, value in enumerate(rows[0])
            if _product_excel_text(value)
        }

        def cell(row_values: tuple[Any, ...], *names: str) -> Any:
            for name in names:
                idx = header_map.get(name.lower())
                if idx is not None and idx < len(row_values):
                    return row_values[idx]
            return ""

        imported_count = 0
        with session_scope() as session:
            for row_values in rows[1:]:
                name = _product_excel_text(cell(row_values, "Название", "Name"))
                if not name:
                    continue
                imported_count += 1
                imported_kind = _product_excel_kind(cell(row_values, "Тип", "Kind"), fallback=str(kind or "product"))
                sku = _product_excel_text(cell(row_values, "Артикул", "SKU"))
                barcode = _product_excel_text(cell(row_values, "Штрихкод", "Barcode"))
                product_id = _product_excel_text(cell(row_values, "ID"))
                row = session.get(Product, product_id) if product_id else None
                if row and row.workspace_owner_id != wid:
                    row = None
                if row is None and sku:
                    row = session.execute(
                        select(Product).where(
                            Product.workspace_owner_id == wid,
                            Product.sku == sku,
                        )
                    ).scalars().first()
                if row is None and barcode:
                    row = session.execute(
                        select(Product).where(
                            Product.workspace_owner_id == wid,
                            Product.barcode == barcode,
                        )
                    ).scalars().first()
                if row is None:
                    row = session.execute(
                        select(Product).where(
                            Product.workspace_owner_id == wid,
                            func.lower(Product.name) == name.lower(),
                        )
                    ).scalars().first()

                prices_json = _product_excel_text(cell(row_values, "Цены JSON"))
                prices: list[dict[str, str]] = []
                if prices_json:
                    try:
                        parsed_prices = json.loads(prices_json)
                        if isinstance(parsed_prices, list):
                            prices = [dict(item) for item in parsed_prices if isinstance(item, dict)]
                    except Exception:
                        prices = []
                primary_price = _product_excel_text(cell(row_values, "Цена продажи"))
                primary_currency = _product_excel_text(cell(row_values, "Валюта цены")) or "UZS"
                if primary_price:
                    if prices:
                        prices[0]["name"] = str(prices[0].get("name") or "Продажная цена")
                        prices[0]["price"] = primary_price
                        prices[0]["currency"] = primary_currency.upper()
                    else:
                        prices = [{"name": "Продажная цена", "price": primary_price, "currency": primary_currency.upper()}]

                stocks_json = _product_excel_text(cell(row_values, "Остатки JSON"))
                stocks: list[dict[str, str]] = []
                if stocks_json:
                    try:
                        parsed_stocks = json.loads(stocks_json)
                        if isinstance(parsed_stocks, list):
                            stocks = [dict(item) for item in parsed_stocks if isinstance(item, dict)]
                    except Exception:
                        stocks = []
                raw_stock_warehouse = _product_excel_text(cell(row_values, "Склад"))
                stock_warehouse = raw_stock_warehouse or "Основной склад"
                stock_quantity = _product_excel_decimal_text(cell(row_values, "Количество"))
                stock_price = _product_excel_decimal_text(cell(row_values, "Себестоимость"))
                stock_date = _product_excel_text(cell(row_values, "Дата остатка"))
                if raw_stock_warehouse or stock_quantity or stock_price or stock_date:
                    if stocks:
                        stocks[0]["warehouse"] = stock_warehouse
                        stocks[0]["quantity"] = stock_quantity
                        stocks[0]["price"] = stock_price
                        stocks[0]["date"] = stock_date
                    else:
                        stocks = [
                            {
                                "warehouse": stock_warehouse,
                                "quantity": stock_quantity,
                                "price": stock_price,
                                "date": stock_date,
                            }
                        ]
                purchase_history = _product_json_rows(cell(row_values, "Purchase History JSON"))

                data = {
                    "kind": imported_kind,
                    "category": _product_excel_text(cell(row_values, "Категория")),
                    "folder": _product_excel_text(cell(row_values, "Папка")),
                    "group": _product_excel_text(cell(row_values, "Группа")),
                    "brand": _product_excel_text(cell(row_values, "Бренд")),
                    "unit": _product_excel_text(cell(row_values, "Единица")) or "Штука",
                    "second_unit": _product_excel_text(cell(row_values, "Вторая единица")),
                    "status": _product_excel_status(cell(row_values, "Статус")),
                    "barcode_type": _product_excel_text(cell(row_values, "Тип штрихкода")) or "EAN13",
                    "batch_tracking": _product_excel_bool(cell(row_values, "Партийный учет")),
                    "packages": _product_excel_text(cell(row_values, "Упаковки")),
                    "min_stock": _product_excel_decimal_text(cell(row_values, "Мин. остаток")),
                    "characteristics": _product_excel_text(cell(row_values, "Характеристики")),
                    "classification": _product_excel_text(cell(row_values, "Классификация")),
                    "owner": _product_excel_text(cell(row_values, "Владелец")),
                    "photo_url": _product_excel_text(cell(row_values, "Photo URL")),
                    "prices": prices,
                    "stocks": stocks,
                    "purchase_history": purchase_history,
                }

                if row is None:
                    next_id = product_id or str(uuid.uuid4())
                    row = Product(
                        id=next_id,
                        workspace_owner_id=wid,
                        name=name,
                        sku=sku,
                        barcode=barcode,
                        external_source="local",
                        external_id=next_id,
                        data=data,
                    )
                    session.add(row)
                else:
                    row.name = name
                    row.sku = sku
                    row.barcode = barcode
                    row.external_source = row.external_source or "local"
                    row.external_id = row.external_id or row.id
                    row.data = data
        if imported_count <= 0:
            return RedirectResponse(url=_product_import_redirect_url(kind, error="В Excel не найдено ни одной строки товара"), status_code=302)
        return RedirectResponse(url=_product_import_redirect_url(kind, msg="imported", count=imported_count), status_code=302)

    async def _save_product_photo_upload(workspace_owner_id: str, upload: Any) -> str:
        filename = str(getattr(upload, "filename", "") or "").strip()
        if not filename:
            return ""
        suffix = Path(filename).suffix.lower()
        if suffix not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
            return ""
        content_type = str(getattr(upload, "content_type", "") or "").lower()
        if content_type and not content_type.startswith("image/"):
            return ""
        payload = await upload.read()
        if not payload:
            return ""
        target_dir = BASE_DIR / "static" / "uploads" / "products" / re.sub(r"[^a-zA-Z0-9_-]", "_", workspace_owner_id)
        target_dir.mkdir(parents=True, exist_ok=True)
        file_name = f"{uuid.uuid4().hex}{suffix}"
        target = target_dir / file_name
        target.write_bytes(payload)
        return f"/static/uploads/products/{target_dir.name}/{file_name}"

    @app.post("/products/save", name="products_save")
    async def products_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/products?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        product_id = str(form.get("product_id") or "").strip()
        name = str(form.get("name") or "").strip()
        if not name:
            return RedirectResponse(url="/products?error=" + quote("Название товара обязательно"), status_code=302)
        sku = str(form.get("sku") or "").strip()
        barcode = str(form.get("barcode") or "").strip()
        data = _product_form_payload(form)
        uploaded_photo_url = await _save_product_photo_upload(wid, form.get("photo_file"))
        if uploaded_photo_url:
            data["photo_url"] = uploaded_photo_url
        with session_scope() as session:
            row = session.get(Product, product_id) if product_id else None
            if row and row.workspace_owner_id != wid:
                return RedirectResponse(url="/products?error=" + quote("Товар не найден"), status_code=302)
            duplicate = session.execute(
                select(Product)
                .where(
                    Product.workspace_owner_id == wid,
                    func.lower(Product.name) == name.lower(),
                )
                .limit(1)
            ).scalars().first()
            if duplicate and duplicate.id != product_id:
                return RedirectResponse(url="/products?error=" + quote("Товар с таким названием уже есть") + "#product-form", status_code=302)
            if sku:
                duplicate = session.execute(
                    select(Product)
                    .where(Product.workspace_owner_id == wid, func.lower(Product.sku) == sku.lower())
                    .limit(1)
                ).scalars().first()
                if duplicate and duplicate.id != product_id:
                    return RedirectResponse(url="/products?error=" + quote("Товар с таким артикулом уже есть") + "#product-form", status_code=302)
            if barcode:
                duplicate = session.execute(
                    select(Product)
                    .where(Product.workspace_owner_id == wid, Product.barcode == barcode)
                    .limit(1)
                ).scalars().first()
                if duplicate and duplicate.id != product_id:
                    return RedirectResponse(url="/products?error=" + quote("Товар с таким штрихкодом уже есть") + "#product-form", status_code=302)
            if row is None:
                product_id = str(uuid.uuid4())
                row = Product(
                    id=product_id,
                    workspace_owner_id=wid,
                    name=name,
                    sku=sku,
                    barcode=barcode,
                    external_source="local",
                    external_id=product_id,
                    data=data,
                )
                session.add(row)
            else:
                row.name = name
                row.sku = sku
                row.barcode = barcode
                row.external_source = row.external_source or "local"
                row.external_id = row.external_id or row.id
                row.data = data
        return RedirectResponse(url="/products?msg=saved", status_code=302)

    @app.post("/products/categories/save", name="products_category_save")
    async def products_category_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/products?err=csrf#categories", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        old_name = str(form.get("old_name") or "").strip()
        name = str(form.get("name") or "").strip()
        if not name:
            return RedirectResponse(url="/products?error=" + quote("Название категории обязательно") + "#categories", status_code=302)
        with session_scope() as session:
            rows = list(session.execute(select(Product).where(Product.workspace_owner_id == wid)).scalars())
            used_names = []
            for row in rows:
                data = dict(row.data if isinstance(row.data, dict) else {})
                category_name = str(data.get("category") or "").strip()
                if category_name:
                    used_names.append(category_name)
                if old_name and category_name == old_name:
                    data["category"] = name
                    row.data = data
            stored = [
                str(item or "").strip()
                for item in load_workspace_settings(wid).get("product_categories", [])
                if str(item or "").strip()
            ]
            stored = [name if item == old_name else item for item in stored]
            _save_product_categories(wid, [*stored, *used_names, name])
        return RedirectResponse(url="/products?msg=saved#categories", status_code=302)

    @app.post("/products/categories/delete", name="products_category_delete")
    async def products_category_delete(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/products?err=csrf#categories", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        name = str(form.get("name") or "").strip()
        if not name:
            return RedirectResponse(url="/products#categories", status_code=302)
        with session_scope() as session:
            rows = list(session.execute(select(Product).where(Product.workspace_owner_id == wid)).scalars())
            remaining_names = []
            for row in rows:
                data = dict(row.data if isinstance(row.data, dict) else {})
                category_name = str(data.get("category") or "").strip()
                if category_name == name:
                    data["category"] = ""
                    row.data = data
                elif category_name:
                    remaining_names.append(category_name)
            stored = [
                str(item or "").strip()
                for item in load_workspace_settings(wid).get("product_categories", [])
                if str(item or "").strip() and str(item or "").strip() != name
            ]
            _save_product_categories(wid, [*stored, *remaining_names])
        return RedirectResponse(url="/products?msg=deleted#categories", status_code=302)

    @app.post("/products/price-types/save", name="products_price_type_save")
    async def products_price_type_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/products?err=csrf#price-types", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        price_types = _workspace_price_types(wid)
        price_type_id = str(form.get("price_type_id") or "").strip() or str(uuid.uuid4())
        existing = next((item for item in price_types if str(item.get("id")) == price_type_id), None)
        payload = _normalize_price_type(
            {
                "id": price_type_id,
                "name": str(form.get("name") or "").strip() or "Новый прайс-лист",
                "sort_order": str(form.get("sort_order") or "0"),
                "is_for_sales": str(form.get("is_for_sales") or "") == "1",
                "is_for_purchases": str(form.get("is_for_purchases") or "") == "1",
                "is_active": str(form.get("is_active") or "") == "1",
                "pricing_method": str(form.get("pricing_method") or "manual"),
                "base_price_type_id": str(form.get("base_price_type_id") or ""),
                "markup_type": str(form.get("markup_type") or "markup"),
                "markup_value": str(form.get("markup_value") or ""),
                "convert_to_currency": str(form.get("convert_to_currency") or "UZS"),
                "rounding": str(form.get("rounding") or "1.0"),
                "created_by": (existing or {}).get("created_by") or "Admin123",
                "updated_at": datetime.now(timezone.utc).date().isoformat(),
            },
            existing,
        )
        next_items = [item for item in price_types if str(item.get("id")) != price_type_id]
        next_items.append(payload)
        _save_workspace_price_types(wid, next_items)
        return _price_type_redirect(price_type_id)

    @app.post("/products/price-types/toggle", name="products_price_type_toggle")
    async def products_price_type_toggle(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/products?err=csrf#price-types", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        price_type_id = str(form.get("price_type_id") or "").strip()
        price_types = _workspace_price_types(wid)
        for item in price_types:
            if str(item.get("id")) == price_type_id:
                item["is_active"] = not bool(item.get("is_active"))
                item["updated_at"] = datetime.now(timezone.utc).date().isoformat()
        _save_workspace_price_types(wid, price_types)
        return _price_type_redirect(price_type_id)

    @app.post("/products/price-types/delete", name="products_price_type_delete")
    async def products_price_type_delete(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/products?err=csrf#price-types", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        price_type_id = str(form.get("price_type_id") or "").strip()
        price_types = _workspace_price_types(wid)
        if price_type_id in {"1", "2", "3", "4"}:
            return _price_type_redirect(price_type_id, msg="base")
        _save_workspace_price_types(wid, [item for item in price_types if str(item.get("id")) != price_type_id])
        return _price_type_redirect("", msg="deleted")

    @app.post("/products/price-types/prices", name="products_price_type_prices_save")
    async def products_price_type_prices_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/products?err=csrf#price-types", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        price_types = _workspace_price_types(wid)
        price_type_id = str(form.get("price_type_id") or "").strip()
        price_type = _price_type_by_id(price_types, price_type_id)
        if price_type.get("pricing_method") != "manual":
            return _price_type_redirect(price_type_id)
        product_ids = list(form.getlist("product_id"))
        values = list(form.getlist("price_value"))
        currencies = list(form.getlist("price_currency"))
        with session_scope() as session:
            for idx, product_id in enumerate(product_ids):
                row = session.get(Product, str(product_id or ""))
                if not row or row.workspace_owner_id != wid:
                    continue
                data = dict(row.data if isinstance(row.data, dict) else {})
                prices = [dict(item) for item in data.get("prices", []) if isinstance(item, dict)]
                value = str(values[idx] if idx < len(values) else "").strip()
                currency = str(currencies[idx] if idx < len(currencies) else "UZS").strip().upper()
                if currency not in {"UZS", "USD"}:
                    currency = "UZS"
                entry = next((item for item in prices if str(item.get("price_type_id") or "") == price_type_id), None)
                if entry is None:
                    entry = next((item for item in prices if str(item.get("name") or "").strip().lower() == str(price_type.get("name") or "").strip().lower()), None)
                if entry is None:
                    entry = {}
                    prices.append(entry)
                entry.update({"price_type_id": price_type_id, "name": price_type["name"], "price": value, "currency": currency})
                data["prices"] = prices
                row.data = data
                flag_modified(row, "data")
        return _price_type_redirect(price_type_id)

    @app.post("/products/bulk-update", name="products_bulk_update")
    async def products_bulk_update(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/products?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None

        filters = _product_filters_payload(
            str(form.get("q") or ""),
            str(form.get("category") or ""),
            str(form.get("group") or ""),
            str(form.get("brand") or ""),
            str(form.get("folder") or ""),
            str(form.get("status") or "active"),
            str(form.get("kind") or "product"),
            category_values=[str(item or "") for item in form.getlist("category")],
            group_values=[str(item or "") for item in form.getlist("group")],
            brand_values=[str(item or "") for item in form.getlist("brand")],
            folder_values=[str(item or "") for item in form.getlist("folder")],
            status_values=[str(item or "") for item in form.getlist("status")],
            kind_values=[str(item or "") for item in form.getlist("kind")],
        )
        product_ids = [str(item or "").strip() for item in form.getlist("product_ids") if str(item or "").strip()]
        updates = {
            "category": str(form.get("bulk_category") or "").strip(),
            "group": str(form.get("bulk_group") or "").strip(),
            "brand": str(form.get("bulk_brand") or "").strip(),
            "folder": str(form.get("bulk_folder") or "").strip(),
        }
        updates = {key: value for key, value in updates.items() if value}

        if not product_ids:
            return RedirectResponse(
                url=_product_list_redirect_url(**filters, error="Отметьте хотя бы один товар для массового редактирования"),
                status_code=302,
            )
        if not updates:
            return RedirectResponse(
                url=_product_list_redirect_url(**filters, error="Заполните хотя бы одно поле: категория, группа, бренд или папка"),
                status_code=302,
            )

        updated_count = 0
        with session_scope() as session:
            rows = list(
                session.execute(
                    select(Product).where(
                        Product.workspace_owner_id == wid,
                        Product.id.in_(product_ids),
                    )
                ).scalars()
            )
            for row in rows:
                data = dict(row.data) if isinstance(row.data, dict) else {}
                for key, value in updates.items():
                    data[key] = value
                row.data = data
                row.external_source = row.external_source or "local"
                row.external_id = row.external_id or row.id
                updated_count += 1

        return RedirectResponse(
            url=_product_list_redirect_url(**filters, msg="bulk_saved", count=updated_count),
            status_code=302,
        )

    @app.post("/products/{product_id}/delete", name="products_delete")
    async def products_delete(request: Request, product_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/products?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        with session_scope() as session:
            row = session.get(Product, product_id)
            if row and row.workspace_owner_id == wid:
                product = _product_data(row)
                if _sales_decimal(product.get("quantity")) > 0:
                    return RedirectResponse(
                        url="/products?error=" + quote("Нельзя удалить товар, который есть в наличии") + "#catalog",
                        status_code=302,
                    )
                session.delete(row)
        return RedirectResponse(url="/products?msg=deleted", status_code=302)

    def _sales_decimal(raw: Any) -> Decimal:
        value = re.sub(r"\s+", "", str(raw or "").strip()).replace(",", ".")
        if not value:
            return Decimal("0")
        try:
            return Decimal(value)
        except Exception:
            return Decimal("0")

    def _sales_decimal_strict(raw: Any, field_label: str) -> Decimal:
        raw_text = str(raw or "").strip()
        if not raw_text:
            return Decimal("0")
        value = re.sub(r"\s+", "", raw_text).replace(",", ".")
        if not re.fullmatch(r"-?\d+(?:\.\d+)?", value):
            raise ValueError(f"{field_label}: введите число")
        try:
            return Decimal(value)
        except Exception as exc:
            raise ValueError(f"{field_label}: введите число") from exc

    def _sales_money_label(raw: Any) -> str:
        amount = _sales_decimal(raw)
        if amount == amount.to_integral():
            return f"{int(amount):,}".replace(",", " ")
        return f"{amount:,.2f}".replace(",", " ").rstrip("0").rstrip(".")

    def _sales_status_label(status: str) -> str:
        return {
            "new": "Новый",
            "reserved": "Резерв",
            "paid": "Оплачен",
            "partial": "Частично",
            "debt": "Долг",
            "return": "Возврат",
        }.get(status or "", "Новый")

    def _sales_doc_type_label(doc_type: str) -> str:
        return {
            "sale": "Продажа",
            "order": "Заказ",
            "return": "Возврат",
        }.get(doc_type or "", "Продажа")

    def _sales_number_prefix(doc_type: str) -> str:
        return {"sale": "a", "order": "b", "return": "c"}.get(doc_type or "", "a")

    def _sales_document_number(prefix: str, index: int) -> str:
        return f"{prefix}{index:02d}"

    def _next_sales_document_number(session: Session, workspace_owner_id: str, doc_type: str) -> str:
        prefix = _sales_number_prefix(doc_type)
        existing_numbers = session.execute(
            select(SaleDocument.number).where(SaleDocument.workspace_owner_id == workspace_owner_id)
        ).scalars().all()
        pattern = re.compile(rf"^{re.escape(prefix)}(\d+)$", re.IGNORECASE)
        used_indexes = [
            int(match.group(1))
            for number in existing_numbers
            for match in [pattern.match(str(number or ""))]
            if match
        ]
        index = (max(used_indexes) + 1) if used_indexes else 1
        while True:
            number = _sales_document_number(prefix, index)
            exists = session.execute(
                select(SaleDocument.id)
                .where(
                    SaleDocument.workspace_owner_id == workspace_owner_id,
                    SaleDocument.number == number,
                )
                .limit(1)
            ).scalar_one_or_none()
            if not exists:
                return number
            index += 1

    def _sales_document_data(row: SaleDocument) -> dict[str, Any]:
        data = row.data if isinstance(row.data, dict) else {}
        paid_amount = _sales_decimal(data.get("paid_amount"))
        amount_value = _sales_decimal(row.amount)
        debt_amount = amount_value - paid_amount
        doc_type = str(data.get("doc_type") or "sale")
        status = str(data.get("status") or ("return" if doc_type == "return" else "new"))
        date_from = str(data.get("date") or "")
        date_to = str(data.get("date_to") or "")
        date_label = f"{date_from} - {date_to}" if date_from and date_to and date_to != date_from else date_from
        item = {
            "id": row.id,
            "number": row.number,
            "date": date_from,
            "date_to": date_to,
            "date_label": date_label,
            "doc_type": doc_type,
            "doc_type_label": _sales_doc_type_label(doc_type),
            "client": str(data.get("client") or ""),
            "warehouse": str(data.get("warehouse") or ""),
            "amount": _sales_money_label(row.amount),
            "amount_value": str(amount_value),
            "currency": row.currency,
            "paid_amount": _sales_money_label(paid_amount),
            "paid_value": str(paid_amount),
            "debt_amount": _sales_money_label(debt_amount if debt_amount > 0 else 0),
            "debt_value": str(debt_amount if debt_amount > 0 else 0),
            "payment_type": str(data.get("payment_type") or ""),
            "status": status,
            "status_label": _sales_status_label(status),
            "manager": str(data.get("manager") or ""),
            "note": str(data.get("note") or ""),
            "lines": data.get("lines") if isinstance(data.get("lines"), list) else [],
            "updated_at": row.updated_at,
        }
        item["detail_json"] = {
            key: value
            for key, value in item.items()
            if key != "updated_at"
        }
        return item

    def _sales_product_option(
        row: Product,
        active_sales_price_types: list[dict[str, Any]],
        price_types: list[dict[str, Any]],
        usd_rate: Decimal,
    ) -> dict[str, Any]:
        row_data = _json_object(row.data)
        product_view = {
            "id": row.id,
            "name": row.name,
            "kind": str(row_data.get("kind") or "product"),
            "category": str(row_data.get("category") or ""),
            "sku": row.sku,
            "barcode": row.barcode,
            "unit": str(row_data.get("unit") or "Штука"),
            "stocks": _product_stocks(row),
            "prices": row_data.get("prices") if isinstance(row_data.get("prices"), list) else [],
            "price_by_type": {},
        }
        for price_type in active_sales_price_types:
            price, price_currency = _calculated_product_price(product_view, price_type, price_types, usd_rate)
            product_view["price_by_type"][str(price_type.get("id") or "")] = {
                "price": price,
                "currency": price_currency,
            }
        return product_view

    def _sales_document_payload(form: Any) -> tuple[dict[str, Any], Decimal, str]:
        doc_type = str(form.get("doc_type") or "sale").strip()
        if doc_type not in {"sale", "order", "return"}:
            doc_type = "sale"
        currency = str(form.get("currency") or "UZS").strip().upper()[:3] or "UZS"
        products = list(form.getlist("line_product"))
        warehouses = list(form.getlist("line_warehouse"))
        quantities = list(form.getlist("line_quantity"))
        prices = list(form.getlist("line_price"))
        lines: list[dict[str, str]] = []
        total = Decimal("0")
        first_warehouse = str(form.get("warehouse") or "").strip()
        for idx in range(max(len(products), len(warehouses), len(quantities), len(prices), 1)):
            product = str(products[idx] if idx < len(products) else "").strip()
            warehouse = str(warehouses[idx] if idx < len(warehouses) else "").strip() or "Основной склад"
            quantity = _sales_decimal_strict(quantities[idx] if idx < len(quantities) else "", "Количество")
            price = _sales_decimal_strict(prices[idx] if idx < len(prices) else "", "Цена")
            if not product and not quantity and not price:
                continue
            if not product:
                raise ValueError("В строке позиции не выбран товар или услуга")
            if quantity < 0:
                raise ValueError("Количество не может быть отрицательным")
            if price < 0:
                raise ValueError("Цена не может быть отрицательной")
            if quantity <= 0:
                quantity = Decimal("1")
            line_total = quantity * price
            if not first_warehouse:
                first_warehouse = warehouse
            total += line_total
            lines.append(
                {
                    "product": product,
                    "warehouse": warehouse,
                    "quantity": str(quantity.normalize() if quantity else ""),
                    "price": str(price.normalize() if price else ""),
                    "total": str(line_total.normalize() if line_total else "0"),
                }
            )
        if not lines:
            raise ValueError("Добавьте хотя бы один товар или услугу")
        if total <= 0:
            raise ValueError("Сумма продажи должна быть больше 0")
        amount = total
        paid_amount = _sales_decimal_strict(form.get("paid_amount"), "Оплачено")
        if paid_amount < 0:
            raise ValueError("Оплачено не может быть отрицательным")
        if doc_type == "return":
            status = "return"
        else:
            status = "paid" if paid_amount and paid_amount >= amount else "partial" if paid_amount else "new"
        payment_lines: list[dict[str, str]] = []
        raw_payment_lines = str(form.get("payment_lines") or "").strip()
        if raw_payment_lines:
            try:
                parsed_payment_lines = json.loads(raw_payment_lines)
            except Exception:
                parsed_payment_lines = []
            if isinstance(parsed_payment_lines, list):
                for payment in parsed_payment_lines:
                    if not isinstance(payment, dict):
                        continue
                    payment_amount = _sales_decimal(payment.get("amount"))
                    if payment_amount <= 0:
                        continue
                    payment_currency = str(payment.get("currency") or currency or "UZS").strip().upper() or "UZS"
                    payment_lines.append(
                        {
                            "account_id": str(payment.get("account_id") or "").strip(),
                            "account": str(payment.get("account") or "").strip(),
                            "type": str(payment.get("type") or "").strip(),
                            "currency": payment_currency[:8],
                            "amount": str(payment_amount.normalize() if payment_amount else "0"),
                        }
                    )
        payment_type = str(form.get("payment_type") or "").strip()
        if not payment_type and payment_lines:
            payment_type = ", ".join(
                dict.fromkeys(str(item.get("type") or "").strip() for item in payment_lines if str(item.get("type") or "").strip())
            )
        data = {
            "doc_type": doc_type,
            "date": str(form.get("date") or "").strip(),
            "date_to": str(form.get("date_to") or form.get("date") or "").strip(),
            "client": str(form.get("client") or "").strip(),
            "warehouse": first_warehouse or "Основной склад",
            "status": status,
            "price_type_id": str(form.get("price_type_id") or "").strip(),
            "price_type": str(form.get("price_type_name") or "").strip(),
            "paid_amount": str(paid_amount.normalize() if paid_amount else "0"),
            "payment_type": payment_type,
            "payment_lines": payment_lines,
            "manager": str(form.get("manager") or "").strip(),
            "note": str(form.get("note") or "").strip(),
            "crm_record_id": str(form.get("crm_record_id") or "").strip(),
            "source_sale_id": str(form.get("source_sale_id") or "").strip(),
            "lines": lines,
        }
        return data, amount, currency

    def _normalize_sales_payment_lines(
        workspace_owner_id: str,
        payment_lines: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        if not payment_lines:
            return []
        with session_scope() as session:
            accounts = list(
                session.execute(
                    select(FinanceAccount)
                    .where(
                        FinanceAccount.workspace_owner_id == workspace_owner_id,
                        FinanceAccount.is_active == True,  # noqa: E712
                    )
                    .order_by(FinanceAccount.created_at.asc())
                ).scalars()
            )
        if not accounts:
            raise ValueError("Создайте активный финансовый счёт перед добавлением оплаты")

        accounts_by_id = {str(account.id): account for account in accounts}
        accounts_by_name = {
            str(account.name or "").strip().casefold(): account
            for account in accounts
            if str(account.name or "").strip()
        }
        normalized: list[dict[str, Any]] = []
        for payment in payment_lines:
            if not isinstance(payment, dict) or _sales_decimal(payment.get("amount")) <= 0:
                continue
            clean_id = str(payment.get("account_id") or "").strip()
            clean_label = str(payment.get("account") or payment.get("type") or "").strip()
            account = accounts_by_id.get(clean_id) if clean_id else None
            if account is None and clean_label:
                account = accounts_by_name.get(clean_label.casefold())
            if clean_id and account is None:
                raise ValueError("Выбранный счёт оплаты не найден или отключён")
            if account is None:
                # Legacy demo payments used generic labels such as cash/bank.
                account = accounts[0]

            item = dict(payment)
            item["account_id"] = str(account.id)
            item["account"] = str(account.name or clean_label)
            item["type"] = str(payment.get("type") or account.name or "Оплата").strip() or "Оплата"
            normalized.append(item)
        return normalized

    def _document_payment_entries_from_form(
        form: Any,
        document_currency: str,
        debt_amount: Decimal,
        workspace_owner_id: str,
    ) -> tuple[list[dict[str, str]], Decimal, str]:
        currency = str(document_currency or "UZS").strip().upper() or "UZS"
        rate = _workspace_usd_uzs_rate(workspace_owner_id)
        raw_entries: list[dict[str, Any]] = []
        raw_payment_lines = str(form.get("payment_lines") or "").strip()
        if raw_payment_lines:
            try:
                parsed_payment_lines = json.loads(raw_payment_lines)
            except Exception:
                parsed_payment_lines = []
            if isinstance(parsed_payment_lines, list):
                raw_entries = [item for item in parsed_payment_lines if isinstance(item, dict)]
        if not raw_entries:
            raw_entries = [
                {
                    "amount": form.get("payment_amount"),
                    "currency": form.get("payment_currency") or currency,
                    "type": form.get("payment_type") or form.get("payment_account") or "Оплата",
                    "account": form.get("payment_account") or "Наличные",
                    "account_id": form.get("payment_account_id") or "",
                }
            ]
        payment_lines: list[dict[str, str]] = []
        paid_in_document_currency = Decimal("0")
        for item in raw_entries:
            payment_amount = _sales_decimal(item.get("amount"))
            if payment_amount <= 0:
                continue
            payment_currency = str(item.get("currency") or currency).strip().upper() or currency
            converted_amount = _convert_product_currency(payment_amount, payment_currency, currency, rate)
            if converted_amount <= 0:
                continue
            remaining = debt_amount - paid_in_document_currency
            if remaining <= 0:
                break
            if converted_amount > remaining:
                converted_amount = remaining
                payment_amount = _convert_product_currency(converted_amount, currency, payment_currency, rate)
            account_label = str(item.get("account") or item.get("type") or "Наличные").strip() or "Наличные"
            payment_type = str(item.get("type") or account_label or "Оплата").strip() or "Оплата"
            payment_lines.append(
                {
                    "amount": _decimal_plain_text(payment_amount),
                    "currency": payment_currency[:8],
                    "type": payment_type,
                    "account": account_label,
                    "account_id": str(item.get("account_id") or "").strip(),
                }
            )
            paid_in_document_currency += converted_amount
        if not payment_lines and debt_amount > 0:
            payment_lines.append(
                {
                    "amount": _decimal_plain_text(debt_amount),
                    "currency": currency,
                    "type": "Оплата",
                    "account": "Наличные",
                    "account_id": "",
                }
            )
            paid_in_document_currency = debt_amount
        payment_lines = [
            {
                **item,
                "amount": _decimal_plain_text(_sales_decimal(item.get("amount"))),
            }
            for item in _normalize_sales_payment_lines(workspace_owner_id, payment_lines)
        ]
        payment_type = ", ".join(
            dict.fromkeys(str(item.get("type") or "").strip() for item in payment_lines if str(item.get("type") or "").strip())
        )
        return payment_lines, paid_in_document_currency, payment_type or "Оплата"

    def _delete_sales_cash_transactions(workspace_owner_id: str, sale_id: str) -> None:
        clean_sale_id = str(sale_id or "").strip()
        if not clean_sale_id:
            return
        with session_scope() as session:
            tx_ids = list(
                session.execute(
                    select(Transaction.id).where(
                        Transaction.workspace_owner_id == workspace_owner_id,
                        Transaction.data.op("->>")("source") == "sales",
                        Transaction.data.op("->>")("sales_document_id") == clean_sale_id,
                    )
                ).scalars()
            )
        for tx_id in tx_ids:
            delete_transaction(workspace_owner_id, str(tx_id))

    def _sync_sales_cash_transactions(
        workspace_owner_id: str,
        *,
        sale_id: str,
        sale_number: str,
        data: dict[str, Any],
        amount: Decimal,
        currency: str,
        user: dict[str, Any] | None = None,
    ) -> None:
        doc_type = str(data.get("doc_type") or "sale").strip()
        payment_rows: list[dict[str, Any]] = []
        for payment in data.get("payment_lines") if isinstance(data.get("payment_lines"), list) else []:
            if not isinstance(payment, dict):
                continue
            payment_amount = _sales_decimal(payment.get("amount"))
            if payment_amount <= 0:
                continue
            payment_rows.append(
                {
                    "amount": payment_amount,
                    "account_id": str(payment.get("account_id") or "").strip(),
                    "account": str(payment.get("account") or "").strip(),
                    "type": str(payment.get("type") or "").strip(),
                    "currency": str(payment.get("currency") or currency or "UZS").strip().upper() or "UZS",
                }
            )
        paid_amount = _sales_decimal(data.get("paid_amount"))
        if not payment_rows and paid_amount > 0:
            payment_rows.append(
                {
                    "amount": paid_amount,
                    "account_id": "",
                    "account": "",
                    "type": str(data.get("payment_type") or "Оплата").strip(),
                    "currency": currency,
                }
            )
        if not payment_rows:
            return
        payment_rows = _normalize_sales_payment_lines(workspace_owner_id, payment_rows)
        _delete_sales_cash_transactions(workspace_owner_id, sale_id)
        tx_type = "expense" if doc_type == "return" else "income"
        actor_name = str((user or {}).get("name") or (user or {}).get("username") or "").strip()
        employee_id = str((user or {}).get("user_id") or (user or {}).get("id") or "").strip()
        for idx, payment in enumerate(payment_rows, start=1):
            account_id = str(payment.get("account_id") or "").strip()
            tx_data: dict[str, Any] = {
                "amount": payment["amount"],
                "currency": str(payment.get("currency") or currency or "UZS").strip().upper() or "UZS",
                "type": tx_type,
                "status": "confirmed",
                "is_confirmed": True,
                "category": "Возврат продажи" if doc_type == "return" else "Продажа",
                "client": str(data.get("client") or ""),
                "employee_id": employee_id,
                "created_at": str(data.get("date") or ""),
                "note": f"{'Возврат' if doc_type == 'return' else 'Продажа'} {sale_number}",
                "data": {
                    "source": "sales",
                    "sales_document_id": sale_id,
                    "sales_number": sale_number,
                    "sales_doc_type": doc_type,
                    "sales_amount": _decimal_plain_text(amount),
                    "payment_index": idx,
                    "payment_type": str(payment.get("type") or ""),
                    "payment_account_id": account_id,
                    "payment_account": str(payment.get("account") or ""),
                    "manager": actor_name,
                },
            }
            if tx_type == "expense":
                tx_data["from_pocket_id"] = account_id
                tx_data["from_account_id"] = account_id
            else:
                tx_data["to_pocket_id"] = account_id
                tx_data["to_account_id"] = account_id
            create_transaction(workspace_owner_id, tx_data)

    def _sales_cash_payment_count(data: dict[str, Any]) -> int:
        count = 0
        for payment in data.get("payment_lines") if isinstance(data.get("payment_lines"), list) else []:
            if isinstance(payment, dict) and _sales_decimal(payment.get("amount")) > 0:
                count += 1
        if count:
            return count
        return 1 if _sales_decimal(data.get("paid_amount")) > 0 else 0

    def _ensure_sales_cash_transactions(workspace_owner_id: str) -> None:
        with session_scope() as session:
            sale_rows = list(
                session.execute(
                    select(SaleDocument)
                    .where(SaleDocument.workspace_owner_id == workspace_owner_id)
                    .order_by(SaleDocument.updated_at.desc())
                    .limit(500)
                ).scalars()
            )
            sales_doc_id = Transaction.data.op("->>")("sales_document_id")
            tx_counts = {
                str(doc_id or ""): int(count or 0)
                for doc_id, count in session.execute(
                    select(sales_doc_id, func.count(Transaction.id))
                    .where(
                        Transaction.workspace_owner_id == workspace_owner_id,
                        Transaction.data.op("->>")("source") == "sales",
                    )
                    .group_by(sales_doc_id)
                ).all()
                if str(doc_id or "").strip()
            }
        for row in sale_rows:
            data = _json_object(row.data)
            expected = _sales_cash_payment_count(data)
            if expected <= 0:
                continue
            if tx_counts.get(str(row.id), 0) == expected:
                continue
            _sync_sales_cash_transactions(
                workspace_owner_id,
                sale_id=str(row.id),
                sale_number=str(row.number or ""),
                data=data,
                amount=_sales_decimal(row.amount),
                currency=str(row.currency or data.get("currency") or "UZS").strip().upper() or "UZS",
                user=None,
            )

    @app.get("/sales", response_class=HTMLResponse, name="sales_get")
    def sales_get(
        request: Request,
        q: str = "",
        doc_type: str = "all",
        status: str = "all",
        client: str = "",
        crm_record_id: str = "",
    ):
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        q_clean = q.strip().lower()
        filters = {
            "q": q.strip(),
            "doc_type": doc_type.strip() or "all",
            "status": status.strip() or "all",
            "client": client.strip(),
        }
        sales: list[dict[str, Any]] = []
        product_names: list[str] = []
        product_options: list[dict[str, Any]] = []
        clients: list[str] = []
        client_options: list[dict[str, Any]] = []
        warehouses: list[str] = []
        warehouse_options: list[dict[str, Any]] = []
        price_type_options: list[dict[str, Any]] = []
        payment_accounts: list[dict[str, str]] = []
        next_numbers: dict[str, str] = {}
        sales_prefill = {"crm_record_id": "", "client": ""}
        usd_rate = _workspace_usd_uzs_rate(wid)
        with session_scope() as session:
            requested_crm_id = str(crm_record_id or "").strip()
            if requested_crm_id:
                crm_row = session.get(CrmRecord, requested_crm_id)
                if crm_row and crm_row.workspace_owner_id == wid and crm_row.item_type in {"deal", "task"}:
                    crm_data = _json_object(crm_row.data)
                    crm_client = str(crm_data.get("client") or "").strip()
                    if not crm_client and crm_row.counterparty_id:
                        crm_counterparty = session.get(Counterparty, crm_row.counterparty_id)
                        if crm_counterparty and crm_counterparty.workspace_owner_id == wid:
                            crm_client = str(crm_counterparty.name or "").strip()
                    sales_prefill = {"crm_record_id": crm_row.id, "client": crm_client}
            rows = list(
                session.execute(
                    select(SaleDocument)
                    .where(SaleDocument.workspace_owner_id == wid)
                    .order_by(SaleDocument.updated_at.desc())
                ).scalars()
            )
            for row in rows:
                item = _sales_document_data(row)
                if filters["doc_type"] != "all" and item["doc_type"] != filters["doc_type"]:
                    continue
                if filters["status"] != "all" and item["status"] != filters["status"]:
                    continue
                if filters["client"] and item["client"] != filters["client"]:
                    continue
                hay = " ".join([item["number"], item["client"], item["warehouse"], item["status_label"], item["doc_type_label"]]).lower()
                if q_clean and q_clean not in hay:
                    continue
                sales.append(item)
            sales_journal_totals_by_currency: dict[str, dict[str, Decimal]] = {}
            for item in sales:
                currency = str(item.get("currency") or "UZS").upper()
                bucket = sales_journal_totals_by_currency.setdefault(
                    currency,
                    {"amount": Decimal("0"), "paid": Decimal("0"), "debt": Decimal("0")},
                )
                bucket["amount"] += _sales_decimal(item.get("amount_value"))
                bucket["paid"] += _sales_decimal(item.get("paid_value"))
                bucket["debt"] += _sales_decimal(item.get("debt_value"))
            sales_journal_totals = [
                {
                    "currency": currency,
                    "amount": _sales_money_label(values["amount"]),
                    "paid": _sales_money_label(values["paid"]),
                    "debt": _sales_money_label(values["debt"]),
                }
                for currency, values in sorted(sales_journal_totals_by_currency.items())
            ]
            price_types = _workspace_price_types(wid)
            active_sales_price_types = [
                item
                for item in price_types
                if item.get("is_for_sales") and item.get("is_active")
            ] or [item for item in price_types if item.get("is_for_sales")] or price_types
            price_type_options = [
                {
                    "id": str(item.get("id") or ""),
                    "name": str(item.get("name") or ""),
                    "currency": str(item.get("convert_to_currency") or "UZS").upper(),
                }
                for item in active_sales_price_types
            ]
            product_rows = list(
                session.execute(
                    select(Product)
                    .where(Product.workspace_owner_id == wid)
                    .order_by(Product.name.asc())
                ).scalars()
            )
            product_names = [str(row.name) for row in product_rows]
            for row in product_rows:
                product_options.append(_sales_product_option(row, active_sales_price_types, price_types, usd_rate))
            client_rows = list(
                session.execute(
                    select(Counterparty)
                    .where(
                        Counterparty.workspace_owner_id == wid,
                        Counterparty.kind.in_(["client", "both"]),
                    )
                    .order_by(Counterparty.name.asc())
                ).scalars()
            )
            client_balance_by_id: dict[str, dict[str, Decimal]] = {}
            client_balance_by_name: dict[str, dict[str, Decimal]] = {}
            for sale_row in session.execute(
                select(SaleDocument).where(SaleDocument.workspace_owner_id == wid)
            ).scalars():
                sale_data = _json_object(sale_row.data)
                sale_client = str(sale_data.get("client") or "").strip()
                if not sale_client:
                    continue
                sale_currency = str(sale_row.currency or "UZS").upper()
                sale_doc_type = str(sale_data.get("doc_type") or "sale")
                sale_sign = Decimal("-1") if sale_doc_type == "return" else Decimal("1")
                sale_balance = sale_sign * (_sales_decimal(sale_row.amount) - _sales_decimal(sale_data.get("paid_amount")))
                sale_counterparty_id = str(sale_data.get("counterparty_id") or "").strip()
                if sale_counterparty_id:
                    bucket = client_balance_by_id.setdefault(sale_counterparty_id, {})
                    bucket[sale_currency] = bucket.get(sale_currency, Decimal("0")) + sale_balance
                bucket = client_balance_by_name.setdefault(sale_client.lower(), {})
                bucket[sale_currency] = bucket.get(sale_currency, Decimal("0")) + sale_balance
            clients = [str(row.name) for row in client_rows]
            for row in client_rows:
                extra = _counterparty_extra(row)
                balance_map = client_balance_by_id.get(row.id) or client_balance_by_name.get(row.name.lower()) or {}
                balance_total = sum(balance_map.values(), Decimal("0"))
                balance_kind = "debt" if balance_total > 0 else "advance" if balance_total < 0 else "zero"
                balance_lines = [
                    {
                        "currency": currency,
                        "amount": _sales_money_label(abs(amount)),
                        "kind": "debt" if amount > 0 else "advance" if amount < 0 else "zero",
                    }
                    for currency, amount in sorted(balance_map.items())
                    if amount
                ]
                client_options.append(
                    {
                        "id": row.id,
                        "name": row.name,
                        "phone": row.phone,
                        "tax_id": row.tax_id,
                        "price_type": str(extra.get("price_type") or ""),
                        "balance_kind": balance_kind,
                        "balance_note": "Баланс: нам должны" if balance_kind == "debt" else "Баланс: мы должны" if balance_kind == "advance" else "Баланс: 0",
                        "balance": _sales_money_label(abs(balance_total)),
                        "balance_lines": balance_lines,
                    }
                )
            warehouse_rows = list(
                session.execute(
                    select(Warehouse)
                    .where(Warehouse.workspace_owner_id == wid)
                    .order_by(Warehouse.name.asc())
                ).scalars()
            )
            warehouses = [str(row.name) for row in warehouse_rows]
            warehouse_options = [_warehouse_view_data(row) for row in warehouse_rows]
            next_numbers = {
                "sale": _next_sales_document_number(session, wid, "sale"),
                "order": _next_sales_document_number(session, wid, "order"),
                "return": _next_sales_document_number(session, wid, "return"),
            }
        clients = sorted({item for item in [*clients, *[sale["client"] for sale in sales if sale["client"]]] if item})
        warehouses = sorted({item for item in [*warehouses, *[sale["warehouse"] for sale in sales if sale["warehouse"]]] if item}) or ["Основной склад"]
        if not warehouse_options:
            warehouse_options = [{"id": "", "name": item, "manager": "", "status": "active", "note": ""} for item in warehouses]
        treasury = load_treasury(wid)
        for pocket in treasury.get("pockets") or []:
            if not isinstance(pocket, dict):
                continue
            label = str(pocket.get("label") or "").strip()
            if not label:
                continue
            payment_accounts.append(
                {
                    "id": str(pocket.get("id") or "").strip(),
                    "label": label,
                    "template_id": str(pocket.get("template_id") or "").strip(),
                }
            )
        if not payment_accounts:
            payment_accounts = [
                {"id": "cash", "label": "Наличные", "template_id": "cash"},
                {"id": "card", "label": "Карта", "template_id": "card"},
            ]
        return tpl(
            request,
            "home_sales.html",
            variant="user",
            active="sales",
            sales=sales,
            sales_journal_totals=sales_journal_totals,
            sales_filters=filters,
            sales_options={
                "clients": clients,
                "client_rows": client_options,
                "warehouses": warehouses,
                "warehouse_rows": warehouse_options,
                "products": product_names,
                "product_rows": product_options,
                "price_types": price_type_options,
                "payment_accounts": payment_accounts,
                "currencies": sorted({*(item["currency"] for item in price_type_options), "UZS", "USD"}),
                "next_numbers": next_numbers,
                "fx": {"USD_UZS": _decimal_plain_text(usd_rate)},
            },
            today=datetime.now(timezone.utc).date().isoformat(),
            flash_ok=request.query_params.get("msg"),
            flash_err=request.query_params.get("error") or ("Форма устарела. Обновите страницу и повторите." if request.query_params.get("err") == "csrf" else ""),
            sales_saved_id=request.query_params.get("saved_id") or "",
            sales_prefill=sales_prefill,
        )

    @app.get("/api/sales/clients/{client_id}/card", name="sales_client_card_api")
    def sales_client_card_api(request: Request, client_id: str):
        wid, redir = _product_workspace_owner(request)
        if redir or not wid:
            return JSONResponse({"error": "workspace"}, status_code=403)
        with session_scope() as session:
            balance_by_id, balance_by_name, last_date_by_id, last_date_by_name = _sales_rollup_maps(session, wid)
            balance_currency_by_id, balance_currency_by_name = _sales_rollup_currency_maps(session, wid)
            crm_rows = list(
                session.execute(
                    select(CrmRecord)
                    .where(CrmRecord.workspace_owner_id == wid)
                    .order_by(CrmRecord.updated_at.desc())
                ).scalars()
            )
            crm_status_by_id, crm_status_by_name = _crm_status_maps_from_records(
                crm_rows,
                _crm_workspace_stages(wid),
            )
            card = _client_card_context(
                session,
                wid,
                client_id,
                balance_by_id=balance_by_id,
                balance_by_name=balance_by_name,
                last_date_by_id=last_date_by_id,
                last_date_by_name=last_date_by_name,
                balance_currency_by_id=balance_currency_by_id,
                balance_currency_by_name=balance_currency_by_name,
                crm_status_by_id=crm_status_by_id,
                crm_status_by_name=crm_status_by_name,
            )
        if not card:
            return JSONResponse({"error": "client_not_found"}, status_code=404)
        return card

    @app.post("/sales/products/quick-save", name="sales_product_quick_save")
    async def sales_product_quick_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return JSONResponse({"error": "csrf"}, status_code=403)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return JSONResponse({"error": "workspace"}, status_code=403)
        assert wid is not None
        kind = _product_excel_kind(form.get("kind"), "product")
        if kind not in {"product", "service"}:
            kind = "product"
        name = str(form.get("name") or "").strip()
        if not name:
            return JSONResponse({"error": "Название обязательно"}, status_code=400)
        sku = str(form.get("sku") or "").strip()
        barcode = str(form.get("barcode") or "").strip()
        data = _product_form_payload(form)
        data["kind"] = kind
        if kind == "service":
            data["stocks"] = []
            data["unit"] = data.get("unit") or "Услуга"
        with session_scope() as session:
            duplicate = session.execute(
                select(Product).where(
                    Product.workspace_owner_id == wid,
                    func.lower(Product.name) == name.lower(),
                )
            ).scalar_one_or_none()
            if duplicate:
                return JSONResponse({"error": "Такая позиция уже есть"}, status_code=409)
            if sku:
                duplicate = session.execute(
                    select(Product)
                    .where(Product.workspace_owner_id == wid, func.lower(Product.sku) == sku.lower())
                    .limit(1)
                ).scalars().first()
                if duplicate:
                    return JSONResponse({"error": "Позиция с таким артикулом уже есть"}, status_code=409)
            if barcode:
                duplicate = session.execute(
                    select(Product)
                    .where(Product.workspace_owner_id == wid, Product.barcode == barcode)
                    .limit(1)
                ).scalars().first()
                if duplicate:
                    return JSONResponse({"error": "Позиция с таким штрихкодом уже есть"}, status_code=409)
            product_id = str(uuid.uuid4())
            row = Product(
                id=product_id,
                workspace_owner_id=wid,
                name=name,
                sku=sku,
                barcode=barcode,
                external_source="local",
                external_id=product_id,
                data=data,
            )
            session.add(row)
            session.flush()
            price_types = _workspace_price_types(wid)
            active_sales_price_types = [
                item
                for item in price_types
                if item.get("is_for_sales") and item.get("is_active")
            ] or [item for item in price_types if item.get("is_for_sales")] or price_types
            product_view = _sales_product_option(row, active_sales_price_types, price_types, _workspace_usd_uzs_rate(wid))
        return JSONResponse({"ok": True, "product": product_view})

    @app.post("/sales/save", name="sales_save")
    async def sales_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/sales?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        try:
            data, amount, currency = _sales_document_payload(form)
            data["payment_lines"] = _normalize_sales_payment_lines(
                wid,
                list(data.get("payment_lines") or []),
            )
        except ValueError as exc:
            return RedirectResponse(url="/sales?error=" + quote(str(exc)) + "#sales-form", status_code=302)
        session_user = request.session.get("user") or {}
        data["manager"] = str(session_user.get("name") or session_user.get("username") or "").strip()
        if not data["client"]:
            return RedirectResponse(url="/sales?error=" + quote("Клиент обязателен") + "#sales-form", status_code=302)
        saved_sale_id = ""
        saved_sale_number = ""
        with session_scope() as session:
            linked_crm_row = None
            linked_crm_id = str(data.get("crm_record_id") or "").strip()
            if linked_crm_id:
                candidate = session.get(CrmRecord, linked_crm_id)
                if candidate and candidate.workspace_owner_id == wid and candidate.item_type in {"deal", "task"}:
                    candidate_data = _json_object(candidate.data)
                    candidate_client = str(candidate_data.get("client") or "").strip()
                    if candidate_client and candidate_client.casefold() != str(data.get("client") or "").strip().casefold():
                        return RedirectResponse(
                            url=f"/sales?doc_type={quote(str(data.get('doc_type') or 'order'))}&crm_record_id={quote(candidate.id)}&error={quote('Клиент заказа не совпадает с карточкой CRM')}#sales-form",
                            status_code=302,
                        )
                    linked_crm_row = candidate
                else:
                    data["crm_record_id"] = ""
            try:
                client_row = _ensure_counterparty(
                    session,
                    wid,
                    name=data["client"],
                    role="client",
                )
                warehouse_row = _ensure_warehouse(
                    session,
                    wid,
                    name=data["warehouse"],
                )
                for line in data.get("lines") or []:
                    line_warehouse = str(line.get("warehouse") or "").strip()
                    if line_warehouse and line_warehouse != warehouse_row.name:
                        _ensure_warehouse(session, wid, name=line_warehouse)
                if data["doc_type"] == "return":
                    _validate_sales_return(session, wid, data)
                delta_sign = -1 if data["doc_type"] == "sale" else 1 if data["doc_type"] == "return" else 0
                data["lines"] = _sync_product_lines(
                    session,
                    wid,
                    warehouse_name=warehouse_row.name,
                    lines=list(data.get("lines") or []),
                    delta_sign=delta_sign,
                    op_date=str(data.get("date") or ""),
                    allow_negative_stock=False,
                )
                data["warehouse_id"] = warehouse_row.id
                data["counterparty_id"] = client_row.id
            except ValueError as exc:
                return RedirectResponse(url="/sales?error=" + quote(str(exc)) + "#sales-form", status_code=302)
            number = _next_sales_document_number(session, wid, data["doc_type"])
            doc_id = str(uuid.uuid4())
            row = SaleDocument(
                id=doc_id,
                workspace_owner_id=wid,
                number=number,
                amount=amount,
                currency=currency,
                counterparty_id=client_row.id,
                external_source="local",
                external_id=doc_id,
                data=data,
            )
            session.add(row)
            saved_sale_id = doc_id
            saved_sale_number = number
            if linked_crm_row is not None:
                crm_data = _json_object(linked_crm_row.data).copy()
                crm_data["related_sale_id"] = doc_id
                crm_data["related_sale_number"] = number
                crm_data["related_sale_type"] = str(data.get("doc_type") or "order")
                _crm_append_activity(
                    crm_data,
                    "Заказ добавлен" if data.get("doc_type") == "order" else "Продажа добавлена",
                    _crm_actor_name(request),
                    f"{number} · {_sales_money_label(amount)} {currency}",
                )
                linked_crm_row.data = crm_data
                flag_modified(linked_crm_row, "data")
        try:
            _sync_sales_cash_transactions(
                wid,
                sale_id=saved_sale_id,
                sale_number=saved_sale_number,
                data=data,
                amount=amount,
                currency=currency,
                user=session_user,
            )
        except Exception:
            logger.exception("[sales] failed to sync sale %s with kassa", saved_sale_id)
            return RedirectResponse(
                url="/sales?error=" + quote("Продажа сохранена, но операция кассы не записалась") + "#sales-form",
                status_code=302,
            )
        saved_message = "order_saved" if data.get("doc_type") == "order" else "return_saved" if data.get("doc_type") == "return" else "saved"
        return RedirectResponse(url=f"/sales?msg={saved_message}&saved_id={quote(saved_sale_id)}#sales-journal", status_code=302)

    @app.post("/sales/{sale_id}/delete", name="sales_delete")
    async def sales_delete(request: Request, sale_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/sales?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        deleted_sale_id = ""
        with session_scope() as session:
            row = session.get(SaleDocument, sale_id)
            if row and row.workspace_owner_id == wid:
                data = _json_object(row.data)
                doc_type = str(data.get("doc_type") or "sale")
                delta_sign = 1 if doc_type == "sale" else -1 if doc_type == "return" else 0
                try:
                    _sync_product_lines(
                        session,
                        wid,
                        warehouse_name=str(data.get("warehouse") or "Основной склад"),
                        lines=list(data.get("lines") or []),
                        delta_sign=delta_sign,
                        op_date=str(data.get("date") or ""),
                    )
                except ValueError as exc:
                    return RedirectResponse(url="/sales?error=" + quote(str(exc)) + "#sales-journal", status_code=302)
                deleted_sale_id = row.id
                session.delete(row)
        if deleted_sale_id:
            try:
                _delete_sales_cash_transactions(wid, deleted_sale_id)
            except Exception:
                logger.exception("[sales] failed to delete kassa transactions for sale %s", deleted_sale_id)
        return RedirectResponse(url="/sales?msg=deleted#sales-journal", status_code=302)

    @app.post("/sales/{sale_id}/status", name="sales_status_update")
    async def sales_status_update(request: Request, sale_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/sales?err=csrf#sales-journal", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        status = str(form.get("status") or "").strip()
        if status not in {"new", "reserved", "paid", "partial", "debt", "return"}:
            return RedirectResponse(url="/sales?error=" + quote("Неверный статус") + "#sales-journal", status_code=302)
        with session_scope() as session:
            row = session.get(SaleDocument, sale_id)
            if row and row.workspace_owner_id == wid:
                data = _json_object(row.data)
                if status == "return" and str(data.get("doc_type") or "sale") != "return":
                    return RedirectResponse(
                        url="/sales?error=" + quote("Оформите возврат отдельным документом из продажи") + "#sales-journal",
                        status_code=302,
                    )
                data["status"] = status
                data["status_manual"] = True
                row.data = data
                flag_modified(row, "data")
                session.add(row)
        return RedirectResponse(url="/sales?msg=status#sales-journal", status_code=302)

    @app.post("/sales/{sale_id}/pay", name="sales_pay")
    async def sales_pay(request: Request, sale_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/sales?err=csrf#sales-journal", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        session_user = request.session.get("user") or {}
        saved_sale_id = ""
        saved_sale_number = ""
        data: dict[str, Any] = {}
        amount = Decimal("0")
        currency = "UZS"
        with session_scope() as session:
            row = session.get(SaleDocument, sale_id)
            if not row or row.workspace_owner_id != wid:
                return RedirectResponse(url="/sales?error=" + quote("Продажа не найдена") + "#sales-journal", status_code=302)
            data = _json_object(row.data)
            doc_type = str(data.get("doc_type") or "sale").strip()
            if doc_type == "return":
                return RedirectResponse(url="/sales?error=" + quote("Возврат нельзя оплатить как продажу") + "#sales-journal", status_code=302)
            amount = _sales_decimal(row.amount)
            paid_amount = _sales_decimal(data.get("paid_amount"))
            debt_amount = amount - paid_amount
            if debt_amount <= 0:
                return RedirectResponse(url=f"/sales?msg=paid&saved_id={quote(row.id)}#sales-journal", status_code=302)
            currency = str(row.currency or data.get("currency") or "UZS").strip().upper() or "UZS"
            raw_payment_lines = data.get("payment_lines") if isinstance(data.get("payment_lines"), list) else []
            payment_lines = [dict(item) for item in raw_payment_lines if isinstance(item, dict)]
            lines_total = sum((_sales_decimal(item.get("amount")) for item in payment_lines), Decimal("0"))
            if paid_amount > lines_total:
                payment_lines.append(
                    {
                        "amount": _decimal_plain_text(paid_amount - lines_total),
                        "currency": currency,
                        "type": str(data.get("payment_type") or "Оплата"),
                        "account": "Наличные",
                        "account_id": "",
                    }
                )
            new_payment_lines, payment_amount, payment_type = _document_payment_entries_from_form(form, currency, debt_amount, wid)
            payment_lines.extend(new_payment_lines)
            next_paid_amount = paid_amount + payment_amount
            if next_paid_amount > amount:
                next_paid_amount = amount
            data["paid_amount"] = _decimal_plain_text(next_paid_amount)
            data["payment_type"] = payment_type or str(data.get("payment_type") or "Оплата")
            data["payment_lines"] = payment_lines
            data["status"] = "paid" if next_paid_amount >= amount else "partial"
            row.data = data
            flag_modified(row, "data")
            session.add(row)
            saved_sale_id = row.id
            saved_sale_number = row.number
        try:
            _sync_sales_cash_transactions(
                wid,
                sale_id=saved_sale_id,
                sale_number=saved_sale_number,
                data=data,
                amount=amount,
                currency=currency,
                user=session_user,
            )
        except Exception:
            logger.exception("[sales] failed to sync payment for sale %s", saved_sale_id)
            return RedirectResponse(
                url="/sales?error=" + quote("Оплата внесена, но операция кассы не записалась") + "#sales-journal",
                status_code=302,
            )
        return RedirectResponse(url=f"/sales?msg=paid&saved_id={quote(saved_sale_id)}#sales-journal", status_code=302)

    def _purchase_status_label(status: str) -> str:
        return {
            "new": "Новый",
            "partial": "Частично",
            "paid": "Оплачен",
            "debt": "Долг",
        }.get(str(status or ""), "Новый")

    def _purchase_document_data(row: PurchaseDocument) -> dict[str, Any]:
        data = _json_object(row.data)
        paid_amount = _sales_decimal(data.get("paid_amount"))
        debt_amount = _sales_decimal(row.amount) - paid_amount
        raw_lines = data.get("lines") if isinstance(data.get("lines"), list) else []
        safe_lines: list[dict[str, str]] = []
        for line in raw_lines:
            if not isinstance(line, dict):
                continue
            safe_lines.append(
                {
                    "product": str(line.get("product") or ""),
                    "quantity": str(line.get("quantity") or ""),
                    "price": str(line.get("price") or ""),
                    "sale_price": str(line.get("sale_price") or ""),
                    "total": str(line.get("total") or ""),
                }
            )
        item = {
            "id": row.id,
            "number": row.number,
            "date": str(data.get("date") or ""),
            "supplier": str(data.get("supplier") or ""),
            "warehouse": str(data.get("warehouse") or ""),
            "amount": _sales_money_label(row.amount),
            "currency": row.currency,
            "paid_amount": _sales_money_label(paid_amount),
            "debt_amount": _sales_money_label(debt_amount if debt_amount > 0 else 0),
            "status": str(data.get("status") or "new"),
            "status_label": _purchase_status_label(str(data.get("status") or "new")),
            "payment_type": str(data.get("payment_type") or ""),
            "price_type_id": str(data.get("price_type_id") or ""),
            "price_type_name": str(data.get("price_type_name") or ""),
            "price_type_currency": str(data.get("price_type_currency") or row.currency or "UZS"),
            "note": str(data.get("note") or ""),
            "lines": safe_lines,
            "updated_at": row.updated_at.isoformat() if row.updated_at else "",
        }
        item["detail_json"] = {
            "id": str(item["id"] or ""),
            "number": str(item["number"] or ""),
            "date": str(item["date"] or ""),
            "supplier": str(item["supplier"] or ""),
            "warehouse": str(item["warehouse"] or ""),
            "amount": str(item["amount"] or "0"),
            "currency": str(item["currency"] or "UZS"),
            "paid_amount": str(item["paid_amount"] or "0"),
            "debt_amount": str(item["debt_amount"] or "0"),
            "status": str(item["status"] or "new"),
            "status_label": str(item["status_label"] or "Новый"),
            "payment_type": str(item["payment_type"] or ""),
            "price_type_id": str(item["price_type_id"] or ""),
            "price_type_name": str(item["price_type_name"] or ""),
            "price_type_currency": str(item["price_type_currency"] or item["currency"] or "UZS"),
            "note": str(item["note"] or ""),
            "lines": safe_lines,
            "updated_at": str(item["updated_at"] or ""),
        }
        return item

    def _purchase_document_payload(form: Any) -> tuple[dict[str, Any], Decimal, str]:
        currency = str(form.get("currency") or "UZS").strip().upper()[:3] or "UZS"
        products = list(form.getlist("line_product"))
        quantities = list(form.getlist("line_quantity"))
        prices = list(form.getlist("line_price"))
        sale_prices = list(form.getlist("line_sale_price"))
        lines: list[dict[str, str]] = []
        total = Decimal("0")
        for idx in range(max(len(products), len(quantities), len(prices), len(sale_prices), 1)):
            product = str(products[idx] if idx < len(products) else "").strip()
            quantity = _sales_decimal_strict(quantities[idx] if idx < len(quantities) else "", "Количество")
            price = _sales_decimal_strict(prices[idx] if idx < len(prices) else "", "Цена")
            sale_price = _sales_decimal_strict(sale_prices[idx] if idx < len(sale_prices) else "", "Продажная цена")
            if not product and not quantity and not price and not sale_price:
                continue
            if not product:
                raise ValueError("В строке прихода не выбран товар")
            if quantity < 0:
                raise ValueError("Количество не может быть отрицательным")
            if price < 0:
                raise ValueError("Цена не может быть отрицательной")
            if sale_price < 0:
                raise ValueError("Продажная цена не может быть отрицательной")
            if quantity <= 0:
                quantity = Decimal("1")
            line_total = quantity * price
            total += line_total
            lines.append(
                {
                    "product": product,
                    "quantity": str(quantity.normalize() if quantity else ""),
                    "price": str(price.normalize() if price else ""),
                    "sale_price": str(sale_price.normalize() if sale_price else ""),
                    "total": str(line_total.normalize() if line_total else "0"),
                }
            )
        manual_amount = _sales_decimal_strict(form.get("amount"), "Итог")
        amount = manual_amount if manual_amount else total
        paid_amount = _sales_decimal_strict(form.get("paid_amount"), "Оплачено")
        if amount < 0:
            raise ValueError("Итог не может быть отрицательным")
        if paid_amount < 0:
            raise ValueError("Оплачено не может быть отрицательным")
        payment_lines: list[dict[str, str]] = []
        raw_payment_lines = str(form.get("payment_lines") or "").strip()
        if raw_payment_lines:
            try:
                parsed_payment_lines = json.loads(raw_payment_lines)
            except Exception:
                parsed_payment_lines = []
            if isinstance(parsed_payment_lines, list):
                for payment in parsed_payment_lines:
                    if not isinstance(payment, dict):
                        continue
                    payment_amount = _sales_decimal(payment.get("amount"))
                    if payment_amount <= 0:
                        continue
                    payment_currency = str(payment.get("currency") or currency or "UZS").strip().upper() or "UZS"
                    payment_lines.append(
                        {
                            "account_id": str(payment.get("account_id") or "").strip(),
                            "account": str(payment.get("account") or "").strip(),
                            "type": str(payment.get("type") or "").strip(),
                            "currency": payment_currency[:8],
                            "amount": str(payment_amount.normalize() if payment_amount else "0"),
                        }
                    )
        status = str(form.get("status") or "").strip()
        if not status:
            status = "paid" if paid_amount and paid_amount >= amount else "partial" if paid_amount else "new"
        payment_type = str(form.get("payment_type") or "").strip()
        if not payment_type and payment_lines:
            payment_type = ", ".join(
                dict.fromkeys(str(item.get("type") or "").strip() for item in payment_lines if str(item.get("type") or "").strip())
            )
        data = {
            "date": str(form.get("date") or "").strip(),
            "supplier": str(form.get("supplier") or "").strip(),
            "warehouse": str(form.get("warehouse") or "").strip() or "Основной склад",
            "status": status,
            "paid_amount": str(paid_amount.normalize() if paid_amount else "0"),
            "payment_type": payment_type,
            "payment_lines": payment_lines,
            "price_type_id": str(form.get("purchase_price_type_id") or "").strip(),
            "price_type_name": str(form.get("purchase_price_type_name") or "").strip(),
            "save_prices_to_price_type": str(form.get("save_purchase_prices") or "").strip().lower() in {"1", "on", "true", "yes"},
            "note": str(form.get("note") or "").strip(),
            "lines": lines,
        }
        return data, amount, currency

    def _sales_rollup_maps(session: Any, workspace_owner_id: str) -> tuple[dict[str, Decimal], dict[str, Decimal], dict[str, str], dict[str, str]]:
        balance_by_id: dict[str, Decimal] = {}
        balance_by_name: dict[str, Decimal] = {}
        last_date_by_id: dict[str, str] = {}
        last_date_by_name: dict[str, str] = {}
        rows = session.execute(
            select(SaleDocument)
            .where(SaleDocument.workspace_owner_id == workspace_owner_id)
            .order_by(SaleDocument.updated_at.desc())
        ).scalars()
        for row in rows:
            data = _json_object(row.data)
            doc_type = str(data.get("doc_type") or "sale")
            signed = Decimal("-1") if doc_type == "return" else Decimal("1")
            balance = signed * (_sales_decimal(row.amount) - _sales_decimal(data.get("paid_amount")))
            name = str(data.get("client") or "").strip()
            counterparty_id = str(data.get("counterparty_id") or row.counterparty_id or "").strip()
            doc_date = str(data.get("date") or "")
            if counterparty_id:
                balance_by_id[counterparty_id] = balance_by_id.get(counterparty_id, Decimal("0")) + balance
                if doc_date and doc_date >= last_date_by_id.get(counterparty_id, ""):
                    last_date_by_id[counterparty_id] = doc_date
            if name:
                lowered = name.lower()
                balance_by_name[lowered] = balance_by_name.get(lowered, Decimal("0")) + balance
                if doc_date and doc_date >= last_date_by_name.get(lowered, ""):
                    last_date_by_name[lowered] = doc_date
        return balance_by_id, balance_by_name, last_date_by_id, last_date_by_name

    def _sales_rollup_currency_maps(session: Any, workspace_owner_id: str) -> tuple[dict[str, dict[str, Decimal]], dict[str, dict[str, Decimal]]]:
        balance_by_id: dict[str, dict[str, Decimal]] = {}
        balance_by_name: dict[str, dict[str, Decimal]] = {}
        rows = session.execute(
            select(SaleDocument)
            .where(SaleDocument.workspace_owner_id == workspace_owner_id)
            .order_by(SaleDocument.updated_at.desc())
        ).scalars()
        for row in rows:
            data = _json_object(row.data)
            doc_type = str(data.get("doc_type") or "sale")
            signed = Decimal("-1") if doc_type == "return" else Decimal("1")
            balance = signed * (_sales_decimal(row.amount) - _sales_decimal(data.get("paid_amount")))
            currency = str(row.currency or "UZS").strip().upper() or "UZS"
            name = str(data.get("client") or "").strip()
            counterparty_id = str(data.get("counterparty_id") or row.counterparty_id or "").strip()
            if counterparty_id:
                bucket = balance_by_id.setdefault(counterparty_id, {})
                bucket[currency] = bucket.get(currency, Decimal("0")) + balance
            if name:
                bucket = balance_by_name.setdefault(name.lower(), {})
                bucket[currency] = bucket.get(currency, Decimal("0")) + balance
        return balance_by_id, balance_by_name

    def _purchase_rollup_maps(session: Any, workspace_owner_id: str) -> tuple[dict[str, Decimal], dict[str, Decimal], dict[str, str], dict[str, str]]:
        balance_by_id: dict[str, Decimal] = {}
        balance_by_name: dict[str, Decimal] = {}
        last_date_by_id: dict[str, str] = {}
        last_date_by_name: dict[str, str] = {}
        rows = session.execute(
            select(PurchaseDocument)
            .where(PurchaseDocument.workspace_owner_id == workspace_owner_id)
            .order_by(PurchaseDocument.updated_at.desc())
        ).scalars()
        for row in rows:
            data = _json_object(row.data)
            balance = _sales_decimal(row.amount) - _sales_decimal(data.get("paid_amount"))
            name = str(data.get("supplier") or "").strip()
            counterparty_id = str(data.get("counterparty_id") or row.counterparty_id or "").strip()
            doc_date = str(data.get("date") or "")
            if counterparty_id:
                balance_by_id[counterparty_id] = balance_by_id.get(counterparty_id, Decimal("0")) + balance
                if doc_date and doc_date >= last_date_by_id.get(counterparty_id, ""):
                    last_date_by_id[counterparty_id] = doc_date
            if name:
                lowered = name.lower()
                balance_by_name[lowered] = balance_by_name.get(lowered, Decimal("0")) + balance
                if doc_date and doc_date >= last_date_by_name.get(lowered, ""):
                    last_date_by_name[lowered] = doc_date
        return balance_by_id, balance_by_name, last_date_by_id, last_date_by_name

    def _purchase_rollup_currency_maps(session: Any, workspace_owner_id: str) -> tuple[dict[str, dict[str, Decimal]], dict[str, dict[str, Decimal]]]:
        balance_by_id: dict[str, dict[str, Decimal]] = {}
        balance_by_name: dict[str, dict[str, Decimal]] = {}
        rows = session.execute(
            select(PurchaseDocument)
            .where(PurchaseDocument.workspace_owner_id == workspace_owner_id)
            .order_by(PurchaseDocument.updated_at.desc())
        ).scalars()
        for row in rows:
            data = _json_object(row.data)
            balance = _sales_decimal(row.amount) - _sales_decimal(data.get("paid_amount"))
            currency = str(row.currency or "UZS").strip().upper() or "UZS"
            name = str(data.get("supplier") or "").strip()
            counterparty_id = str(data.get("counterparty_id") or row.counterparty_id or "").strip()
            if counterparty_id:
                bucket = balance_by_id.setdefault(counterparty_id, {})
                bucket[currency] = bucket.get(currency, Decimal("0")) + balance
            if name:
                bucket = balance_by_name.setdefault(name.lower(), {})
                bucket[currency] = bucket.get(currency, Decimal("0")) + balance
        return balance_by_id, balance_by_name

    def _counterparty_view_data(
        row: Counterparty,
        *,
        balance_by_id: dict[str, Decimal],
        balance_by_name: dict[str, Decimal],
        last_date_by_id: dict[str, str],
        last_date_by_name: dict[str, str],
        balance_currency_by_id: dict[str, dict[str, Decimal]] | None = None,
        balance_currency_by_name: dict[str, dict[str, Decimal]] | None = None,
        crm_status_by_id: dict[str, str] | None = None,
        crm_status_by_name: dict[str, str] | None = None,
    ) -> dict[str, Any]:
        extra = _counterparty_extra(row)
        balance = balance_by_id.get(row.id, balance_by_name.get(row.name.lower(), Decimal("0")))
        last_date = last_date_by_id.get(row.id) or last_date_by_name.get(row.name.lower(), "")
        has_client, has_supplier = _counterparty_role_flags(row.kind)
        balance_kind = "debt" if balance > 0 else "advance" if balance < 0 else "zero"
        currency_balances = (
            (balance_currency_by_id or {}).get(row.id)
            or (balance_currency_by_name or {}).get(row.name.lower())
            or {"UZS": balance}
        )
        balance_lines = []
        for currency, amount in sorted(currency_balances.items()):
            amount = _sales_decimal(amount)
            if not amount and len(currency_balances) > 1:
                continue
            kind = "debt" if amount > 0 else "advance" if amount < 0 else "zero"
            balance_lines.append(
                {
                    "currency": currency,
                    "amount": _sales_money_label(amount),
                    "amount_abs": _sales_money_label(abs(amount)),
                    "kind": kind,
                    "note": "Нам должны" if kind == "debt" else "Мы должны" if kind == "advance" else "Нет долга",
                }
            )
        if not balance_lines:
            balance_lines.append({"currency": "UZS", "amount": "0", "amount_abs": "0", "kind": "zero", "note": "Нет долга"})
        raw_programs = extra.get("programs")
        if isinstance(raw_programs, list):
            programs = [str(item).strip() for item in raw_programs if str(item).strip()]
        elif isinstance(raw_programs, str) and raw_programs.strip():
            programs = [item.strip() for item in raw_programs.split(",") if item.strip()]
        else:
            programs = []
        legacy_program = str(extra.get("program") or extra.get("software") or extra.get("category") or "").strip()
        if legacy_program and legacy_program not in programs:
            programs.append(legacy_program)
        crm_status = str(
            (crm_status_by_id or {}).get(row.id)
            or (crm_status_by_name or {}).get(row.name.lower())
            or extra.get("crm_status")
            or "new_lead"
        ).strip()
        if crm_status not in CLIENT_CRM_STATUS_LABELS:
            crm_status = "new_lead"
        map_icon = str(extra.get("map_icon") or "").strip()
        if not map_icon or map_icon == "default":
            map_icon = str(extra.get("industry") or "default").strip() or "default"
        return {
            "id": row.id,
            "name": row.name,
            "official_name": str(extra.get("official_name") or row.name),
            "legal_name": str(extra.get("legal_name") or extra.get("official_name") or row.name),
            "legal_address": str(extra.get("legal_address") or ""),
            "photo_url": str(extra.get("photo_url") or extra.get("photo") or ""),
            "phone": row.phone,
            "phone_primary": row.phone,
            "tax_id": row.tax_id,
            "inn": str(extra.get("inn") or row.tax_id or ""),
            "pinfl": str(extra.get("pinfl") or ""),
            "okved": str(extra.get("okved") or ""),
            "client_type": str(extra.get("client_type") or "company"),
            "territory": str(extra.get("territory") or ""),
            "category": str(extra.get("category") or ""),
            "industry": str(extra.get("industry") or ""),
            "map_icon": map_icon,
            "program": ", ".join(programs),
            "programs": programs,
            "route": str(extra.get("route") or ""),
            "status": str(extra.get("status") or "active"),
            "crm_status": crm_status,
            "crm_status_label": CLIENT_CRM_STATUS_LABELS[crm_status],
            "is_blacklisted": bool(extra.get("is_blacklisted")),
            "sort_order": str(extra.get("sort_order") or "0"),
            "code": str(extra.get("code") or ""),
            "telegram": str(extra.get("telegram") or ""),
            "telegram_phone": str(extra.get("telegram_phone") or extra.get("telegram") or ""),
            "note": str(extra.get("note") or ""),
            "comment": str(extra.get("comment") or extra.get("note") or ""),
            "email": str(extra.get("email") or ""),
            "address": str(extra.get("address") or ""),
            "latitude": str(extra.get("latitude") or ""),
            "longitude": str(extra.get("longitude") or ""),
            "price_type": str(extra.get("price_type") or ""),
            "credit_limit": str(extra.get("credit_limit") or ""),
            "consignment_days": str(extra.get("consignment_days") or ""),
            "cashback_percent": str(extra.get("cashback_percent") or ""),
            "delivery_frequency_days": str(extra.get("delivery_frequency_days") or ""),
            "owner_user": str(extra.get("owner_user") or ""),
            "owner_group": str(extra.get("owner_group") or ""),
            "balance_value": balance,
            "balance": _sales_money_label(balance),
            "balance_abs": _sales_money_label(abs(balance)),
            "balance_lines": balance_lines,
            "balance_kind": balance_kind,
            "balance_note": "Нам должны" if balance_kind == "debt" else "Мы должны" if balance_kind == "advance" else "Нет долга",
            "last_date": last_date,
            "created_at": row.created_at.date().isoformat() if getattr(row, "created_at", None) else "",
            "updated_at": row.updated_at.date().isoformat() if getattr(row, "updated_at", None) else "",
            "is_client": has_client,
            "is_supplier": has_supplier,
        }

    def _warehouse_view_data(row: Warehouse) -> dict[str, Any]:
        data = _json_object(row.data)
        return {
            "id": row.id,
            "name": row.name,
            "manager": str(data.get("manager") or ""),
            "status": str(data.get("status") or "active"),
            "note": str(data.get("note") or ""),
        }

    def _warehouse_operation_payload(form: Any) -> tuple[dict[str, Any], Decimal]:
        operation_type = str(form.get("operation_type") or "adjustment").strip()
        if operation_type not in {"in", "out", "adjustment", "transfer"}:
            operation_type = "adjustment"
        quantity = _sales_decimal_strict(form.get("quantity"), "Количество")
        amount = _sales_decimal_strict(form.get("amount"), "Сумма")
        price = _sales_decimal_strict(form.get("price"), "Цена")
        if price < 0:
            raise ValueError("Цена не может быть отрицательной")
        if amount < 0:
            raise ValueError("Сумма не может быть отрицательной")
        data = {
            "date": str(form.get("date") or "").strip(),
            "operation_type": operation_type,
            "warehouse": str(form.get("warehouse") or "").strip() or "Основной склад",
            "from_warehouse": str(form.get("from_warehouse") or "").strip() or "Основной склад",
            "to_warehouse": str(form.get("to_warehouse") or "").strip() or "Основной склад",
            "product": str(form.get("product") or "").strip(),
            "price": str(price.normalize() if price else ""),
            "responsible": str(form.get("responsible") or "").strip(),
            "note": str(form.get("note") or "").strip(),
            "quantity": str(quantity.normalize() if quantity else "0"),
        }
        return data, amount

    def _warehouse_operation_data(row: WarehouseOperation) -> dict[str, Any]:
        data = _json_object(row.data)
        quantity = _sales_decimal(row.quantity)
        return {
            "id": row.id,
            "number": row.number,
            "date": str(data.get("date") or ""),
            "operation_type": str(row.operation_type or data.get("operation_type") or "adjustment"),
            "operation_label": _warehouse_operation_type_label(str(row.operation_type or data.get("operation_type") or "adjustment")),
            "warehouse": str(data.get("warehouse") or ""),
            "from_warehouse": str(data.get("from_warehouse") or ""),
            "to_warehouse": str(data.get("to_warehouse") or ""),
            "product": str(data.get("product") or ""),
            "quantity": str(quantity.normalize() if quantity else "0"),
            "amount": _sales_money_label(row.amount),
            "currency": row.currency,
            "responsible": str(data.get("responsible") or ""),
            "note": str(data.get("note") or ""),
        }

    def _crm_record_payload(form: Any) -> tuple[dict[str, Any], Decimal, str]:
        item_type = str(form.get("item_type") or "task").strip()
        if item_type not in {"task", "deal", "history"}:
            item_type = "task"
        amount = _sales_decimal(form.get("amount"))
        currency = str(form.get("currency") or "UZS").strip().upper()[:3] or "UZS"

        def clean_tags(raw: Any) -> list[str]:
            if isinstance(raw, (list, tuple, set)):
                parts = [str(item or "") for item in raw]
            else:
                parts = re.split(r"[,;\n]+", str(raw or ""))
            tags: list[str] = []
            seen: set[str] = set()
            for part in parts:
                tag = re.sub(r"\s+", " ", part).strip(" #\t\r\n")
                if not tag:
                    continue
                tag = tag[:24]
                key = tag.casefold()
                if key in seen:
                    continue
                seen.add(key)
                tags.append(tag)
                if len(tags) >= 8:
                    break
            return tags

        data = {
            "item_type": item_type,
            "client": str(form.get("client") or "").strip(),
            "responsible": str(form.get("responsible") or "").strip(),
            "date": str(form.get("date") or "").strip(),
            "due_date": str(form.get("due_date") or "").strip(),
            "stage": str(form.get("stage") or "").strip(),
            "stage_id": str(form.get("stage_id") or "").strip(),
            "related_deal_id": str(form.get("related_deal_id") or "").strip(),
            "lead_source": str(form.get("lead_source") or form.get("contact_type") or "").strip(),
            "contact_type": str(form.get("contact_type") or "").strip(),
            "chat_ref": str(form.get("chat_ref") or "").strip(),
            "service_type": str(form.get("service_type") or "").strip(),
            "priority": str(form.get("priority") or "normal").strip(),
            "next_step": str(form.get("next_step") or "").strip(),
            "probability": _crm_probability_value(form.get("probability")),
            "note": str(form.get("note") or "").strip(),
            "lost_reason": str(form.get("lost_reason") or "").strip()[:300],
            "tags": clean_tags(form.get("tags")),
        }
        if data["priority"] not in CRM_PRIORITY_LABELS:
            data["priority"] = "normal"
        return data, amount, currency

    def _crm_record_data(row: CrmRecord) -> dict[str, Any]:
        data = _json_object(row.data)
        priority = str(data.get("priority") or "normal")
        probability = _crm_probability_value(data.get("probability"))
        tags = [str(item or "").strip() for item in data.get("tags", []) if str(item or "").strip()] if isinstance(data.get("tags"), list) else []
        amount_value = _sales_decimal(row.amount)
        amount_input = str(amount_value.normalize()) if amount_value else ""
        activity_log = [item for item in data.get("activity_log", []) if isinstance(item, dict)] if isinstance(data.get("activity_log"), list) else []
        if not activity_log:
            created_at = row.created_at.isoformat() if row.created_at else ""
            activity_log = [{"at": created_at, "action": "Карточка создана", "actor": "UPOS", "detail": ""}]
        return {
            "id": row.id,
            "title": row.title,
            "item_type": row.item_type,
            "item_type_label": _crm_type_label(row.item_type),
            "status": row.status,
            "status_label": _crm_status_label(row.status),
            "client": str(data.get("client") or ""),
            "responsible": str(data.get("responsible") or ""),
            "date": str(data.get("date") or ""),
            "due_date": row.due_date or str(data.get("due_date") or ""),
            "stage": str(data.get("stage") or ""),
            "stage_id": str(data.get("stage_id") or ""),
            "related_deal_id": str(data.get("related_deal_id") or ""),
            "lead_source": str(data.get("lead_source") or ""),
            "contact_type": str(data.get("contact_type") or ""),
            "chat_ref": str(data.get("chat_ref") or ""),
            "service_type": str(data.get("service_type") or ""),
            "priority": priority if priority in CRM_PRIORITY_LABELS else "normal",
            "priority_label": _crm_priority_label(priority),
            "next_step": str(data.get("next_step") or ""),
            "probability": probability,
            "probability_label": f"{probability}%" if probability else "",
            "note": str(data.get("note") or ""),
            "lost_reason": str(data.get("lost_reason") or ""),
            "activity_log": list(reversed(activity_log[-50:])),
            "tags": tags[:8],
            "tags_text": ", ".join(tags[:8]),
            "amount": _sales_money_label(row.amount),
            "amount_value": amount_value,
            "amount_input": amount_input,
            "currency": row.currency,
            "counterparty_id": row.counterparty_id or "",
        }

    def _crm_record_edit_payload(item: dict[str, Any]) -> dict[str, str]:
        keys = (
            "id",
            "item_type",
            "title",
            "client",
            "responsible",
            "lead_source",
            "stage_id",
            "related_deal_id",
            "service_type",
            "priority",
            "contact_type",
            "chat_ref",
            "date",
            "due_date",
            "status",
            "probability",
            "amount_input",
            "currency",
            "next_step",
            "note",
            "lost_reason",
            "tags",
        )
        payload: dict[str, str] = {}
        for key in keys:
            if key == "tags":
                payload[key] = str(item.get("tags_text") or "")
                continue
            value = item.get(key)
            payload[key] = "" if value is None else str(value)
        return payload

    def _client_card_context(
        session: Any,
        workspace_owner_id: str,
        client_id: str,
        *,
        balance_by_id: dict[str, Decimal],
        balance_by_name: dict[str, Decimal],
        last_date_by_id: dict[str, str],
        last_date_by_name: dict[str, str],
        balance_currency_by_id: dict[str, dict[str, Decimal]] | None = None,
        balance_currency_by_name: dict[str, dict[str, Decimal]] | None = None,
        crm_status_by_id: dict[str, str] | None = None,
        crm_status_by_name: dict[str, str] | None = None,
    ) -> dict[str, Any] | None:
        row = session.get(Counterparty, client_id)
        if not row or row.workspace_owner_id != workspace_owner_id:
            return None
        has_client, _ = _counterparty_role_flags(row.kind)
        if not has_client:
            return None

        client = _counterparty_view_data(
            row,
            balance_by_id=balance_by_id,
            balance_by_name=balance_by_name,
            last_date_by_id=last_date_by_id,
            last_date_by_name=last_date_by_name,
            balance_currency_by_id=balance_currency_by_id,
            balance_currency_by_name=balance_currency_by_name,
            crm_status_by_id=crm_status_by_id,
            crm_status_by_name=crm_status_by_name,
        )
        extra = _counterparty_extra(row)
        client_names = {
            str(client.get("name") or "").strip().lower(),
            str(client.get("official_name") or "").strip().lower(),
        }
        client_names.discard("")

        def digits(value: Any) -> str:
            return "".join(ch for ch in str(value or "") if ch.isdigit())

        def matches_client_name(value: Any) -> bool:
            lowered = str(value or "").strip().lower()
            return bool(lowered and lowered in client_names)

        history_events: list[dict[str, str]] = []

        def add_history_event(
            *,
            date: Any,
            title: str,
            detail: str = "",
            category: str = "system",
            amount: str = "",
            status: str = "",
        ) -> None:
            raw_date = str(date or "").strip()
            if not raw_date:
                raw_date = datetime.now(timezone.utc).isoformat()
            history_events.append(
                {
                    "date": raw_date[:19].replace("T", " "),
                    "sort_date": raw_date,
                    "title": title,
                    "detail": detail,
                    "category": category,
                    "amount": amount,
                    "status": status,
                }
            )

        add_history_event(
            date=row.created_at.isoformat() if getattr(row, "created_at", None) else client.get("created_at"),
            title="Клиент создан",
            detail=f"Карточка клиента {client['name']} добавлена в систему.",
            category="client",
            status=client.get("crm_status_label") or "",
        )
        if getattr(row, "updated_at", None) and getattr(row, "created_at", None) and row.updated_at != row.created_at:
            add_history_event(
                date=row.updated_at.isoformat(),
                title="Карточка обновлена",
                detail="Изменены данные клиента или его настройки.",
                category="client",
                status=client.get("status") or "",
            )
        if client.get("address") or client.get("latitude") or client.get("longitude"):
            add_history_event(
                date=row.updated_at.isoformat() if getattr(row, "updated_at", None) else client.get("updated_at"),
                title="Локация клиента сохранена",
                detail=client.get("address") or f"{client.get('latitude')}, {client.get('longitude')}",
                category="location",
            )
        if client.get("comment") or client.get("note"):
            add_history_event(
                date=row.updated_at.isoformat() if getattr(row, "updated_at", None) else client.get("updated_at"),
                title="Добавлена заметка",
                detail=str(client.get("comment") or client.get("note") or "")[:160],
                category="note",
            )

        sale_rows = list(
            session.execute(
                select(SaleDocument)
                .where(SaleDocument.workspace_owner_id == workspace_owner_id)
                .order_by(SaleDocument.updated_at.desc())
            ).scalars()
        )
        sales: list[dict[str, Any]] = []
        orders: list[dict[str, Any]] = []
        returns: list[dict[str, Any]] = []
        matched_sale_rows: list[SaleDocument] = []

        def sale_document_created_label(document: SaleDocument) -> str:
            created_at = getattr(document, "created_at", None)
            return created_at.isoformat() if created_at else ""

        def sale_document_day_label(document: SaleDocument, fallback_date: Any = "") -> str:
            raw_date = str(fallback_date or "").strip()
            if raw_date:
                return raw_date
            created_at = getattr(document, "created_at", None)
            return created_at.date().isoformat() if created_at else ""

        def sale_document_sort_key(document: SaleDocument) -> tuple[str, str]:
            data = _json_object(document.data)
            return (str(data.get("date") or ""), sale_document_created_label(document))
        for sale_row in sale_rows:
            data = _json_object(sale_row.data)
            counterparty_id = str(data.get("counterparty_id") or sale_row.counterparty_id or "").strip()
            if counterparty_id != client_id and not matches_client_name(data.get("client")):
                continue
            item = _sales_document_data(sale_row)
            matched_sale_rows.append(sale_row)
            add_history_event(
                date=item["date"] or sale_document_created_label(sale_row),
                title=f"{item['doc_type_label']} {item['number']}",
                detail="Документ клиента в продажах.",
                category="document",
                amount=f"{item['amount']} {item['currency']}",
                status=item["status_label"],
            )
            paid_amount = _sales_decimal(data.get("paid_amount"))
            if paid_amount > 0:
                add_history_event(
                    date=item["date"] or (sale_row.updated_at.isoformat() if getattr(sale_row, "updated_at", None) else ""),
                    title=f"Оплата по {item['number']}",
                    detail="Принята оплата по документу клиента.",
                    category="payment",
                    amount=f"{_sales_money_label(paid_amount)} {item['currency']}",
                    status="Оплачено",
                )
            if item["doc_type"] == "order":
                orders.append(item)
            elif item["doc_type"] == "return":
                returns.append(item)
            else:
                sales.append(item)

        reconciliation_rows: list[dict[str, str]] = []
        running_balance = Decimal("0")
        reconciliation_total_debit = Decimal("0")
        reconciliation_total_credit = Decimal("0")
        reconciliation_currency = str(client.get("currency") or "UZS").upper()
        for sale_row in sorted(
            matched_sale_rows,
            key=sale_document_sort_key,
        ):
            data = _json_object(sale_row.data)
            item = _sales_document_data(sale_row)
            amount = _sales_decimal(sale_row.amount)
            paid_amount = _sales_decimal(data.get("paid_amount"))
            if item["doc_type"] == "return":
                debit = Decimal("0")
                credit = amount
                running_balance -= amount
            else:
                debit = amount
                credit = Decimal("0")
                running_balance += amount
            reconciliation_total_debit += debit
            reconciliation_total_credit += credit
            reconciliation_currency = item["currency"] or reconciliation_currency
            reconciliation_rows.append(
                {
                    "date": sale_document_day_label(sale_row, item["date"]),
                    "document": f"{item['doc_type_label']} {item['number']}",
                    "debit": _sales_money_label(debit),
                    "credit": _sales_money_label(credit),
                    "balance": _sales_money_label(running_balance),
                    "currency": item["currency"],
                }
            )
            if paid_amount > 0 and item["doc_type"] != "return":
                running_balance -= paid_amount
                reconciliation_total_credit += paid_amount
                reconciliation_rows.append(
                    {
                        "date": sale_document_day_label(sale_row, item["date"]),
                        "document": f"Оплата по {item['number']}",
                        "debit": _sales_money_label(Decimal("0")),
                        "credit": _sales_money_label(paid_amount),
                        "balance": _sales_money_label(running_balance),
                        "currency": item["currency"],
                    }
                )

        crm_rows = list(
            session.execute(
                select(CrmRecord)
                .where(CrmRecord.workspace_owner_id == workspace_owner_id)
                .order_by(CrmRecord.updated_at.desc())
            ).scalars()
        )
        tasks: list[dict[str, Any]] = []
        deals: list[dict[str, Any]] = []
        correspondence: list[dict[str, Any]] = []
        for crm_row in crm_rows:
            data = _json_object(crm_row.data)
            if crm_row.counterparty_id != client_id and not matches_client_name(data.get("client")):
                continue
            item = _crm_record_data(crm_row)
            add_history_event(
                date=item.get("date") or item.get("due_date") or (crm_row.updated_at.isoformat() if getattr(crm_row, "updated_at", None) else ""),
                title=f"{item['item_type_label']}: {item['title']}",
                detail=item.get("note") or item.get("stage") or item.get("lead_source") or "",
                category="crm",
                amount=f"{item['amount']} {item['currency']}" if item["amount_value"] else "",
                status=item["status_label"],
            )
            if item["item_type"] == "task":
                tasks.append(item)
            elif item["item_type"] == "deal":
                deals.append(item)
            else:
                correspondence.append(item)

        phone_digits = digits(client.get("phone"))
        telegram_value = str(client.get("telegram") or extra.get("telegram") or "").strip()
        telegram_key = telegram_value.lstrip("@").lower()
        subscriber_rows = list(
            session.execute(
                select(TelegramSubscriber)
                .where(TelegramSubscriber.workspace_owner_id == workspace_owner_id)
                .order_by(TelegramSubscriber.requested_at.desc())
            ).scalars()
        )
        matched_subscribers: list[TelegramSubscriber] = []
        for subscriber in subscriber_rows:
            username_key = str(subscriber.username or "").lstrip("@").lower()
            display_name = str(subscriber.display_name or "").strip().lower()
            subscriber_phone_digits = digits(subscriber.phone)
            if (
                (phone_digits and subscriber_phone_digits and subscriber_phone_digits.endswith(phone_digits[-9:]))
                or (telegram_key and username_key == telegram_key)
                or (display_name and display_name in client_names)
            ):
                matched_subscribers.append(subscriber)

        chat_ids = {subscriber.chat_id for subscriber in matched_subscribers}
        chat_rows = []
        if chat_ids:
            chat_rows = list(
                session.execute(
                    select(TelegramChat)
                    .where(
                        TelegramChat.workspace_owner_id == workspace_owner_id,
                        TelegramChat.chat_id.in_(chat_ids),
                    )
                ).scalars()
            )
        chat_by_id = {chat.chat_id: chat for chat in chat_rows}
        conversations: list[dict[str, Any]] = []
        for subscriber in matched_subscribers:
            chat = chat_by_id.get(subscriber.chat_id)
            conversations.append(
                {
                    "channel": "Telegram",
                    "title": subscriber.display_name or client["name"],
                    "username": f"@{subscriber.username}" if subscriber.username else "",
                    "phone": subscriber.phone,
                    "status": subscriber.status,
                    "chat_title": chat.title if chat else "",
                    "chat_id": str(subscriber.chat_id),
                }
            )
            add_history_event(
                date=subscriber.requested_at.isoformat() if getattr(subscriber, "requested_at", None) else "",
                title="Telegram контакт привязан",
                detail=subscriber.display_name or subscriber.username or subscriber.phone or "",
                category="message",
                status=subscriber.status,
            )
        if not conversations and telegram_value:
            conversations.append(
                {
                    "channel": "Telegram",
                    "title": client["name"],
                    "username": telegram_value,
                    "phone": client.get("phone") or "",
                    "status": "linked",
                    "chat_title": "",
                    "chat_id": "",
                }
            )

        telephony_calls_raw = load_workspace_settings(workspace_owner_id).get("telephony_calls")
        telephony_calls = telephony_calls_raw if isinstance(telephony_calls_raw, list) else []
        calls: list[dict[str, Any]] = []
        for raw_call in telephony_calls:
            if not isinstance(raw_call, dict):
                continue
            call_client = str(raw_call.get("client") or "").strip()
            call_phone_digits = digits(raw_call.get("phone"))
            if (
                matches_client_name(call_client)
                or (phone_digits and call_phone_digits and call_phone_digits.endswith(phone_digits[-9:]))
            ):
                calls.append(
                    {
                        "started_at": str(raw_call.get("started_at") or raw_call.get("created_at") or ""),
                        "phone": str(raw_call.get("phone") or ""),
                        "direction_label": _telephony_direction_label(str(raw_call.get("direction") or "")),
                        "status_label": _telephony_status_label("call", str(raw_call.get("status") or "")),
                        "responsible": str(raw_call.get("responsible") or ""),
                        "comment": str(raw_call.get("comment") or raw_call.get("note") or ""),
                    }
                )
                add_history_event(
                    date=str(raw_call.get("started_at") or raw_call.get("created_at") or ""),
                    title="Звонок клиенту",
                    detail=str(raw_call.get("comment") or raw_call.get("note") or raw_call.get("phone") or ""),
                    category="call",
                    status=_telephony_status_label("call", str(raw_call.get("status") or "")),
                )
        calls.sort(key=lambda item: item["started_at"], reverse=True)
        history_events.sort(key=lambda item: item["sort_date"], reverse=True)

        total_sales = sum(
            (_sales_decimal(item.amount) for item in matched_sale_rows if _json_object(item.data).get("doc_type") != "return"),
            Decimal("0"),
        )
        total_returns = sum(
            (_sales_decimal(item.amount) for item in matched_sale_rows if _json_object(item.data).get("doc_type") == "return"),
            Decimal("0"),
        )
        total_paid = sum(
            (_sales_decimal(_json_object(item.data).get("paid_amount")) for item in matched_sale_rows),
            Decimal("0"),
        )

        client.update(
            {
                "orders": orders[:12],
                "sales": sales[:12],
                "returns": returns[:12],
                "calls": calls[:5],
                "tasks": tasks[:12],
                "deals": deals[:12],
                "correspondence": correspondence[:12],
                "conversations": conversations[:8],
                "reconciliation": reconciliation_rows[-18:],
                "reconciliation_totals": {
                    "debit": _sales_money_label(reconciliation_total_debit),
                    "credit": _sales_money_label(reconciliation_total_credit),
                    "balance": _sales_money_label(running_balance),
                    "currency": reconciliation_currency,
                },
                "history": history_events[:80],
                "summary": {
                    "sales": _sales_money_label(total_sales),
                    "returns": _sales_money_label(total_returns),
                    "paid": _sales_money_label(total_paid),
                    "orders_count": len(orders),
                    "tasks_count": len(tasks),
                    "messages_count": len(conversations) + len(correspondence),
                    "calls_count": len(calls),
                },
            }
        )
        return client

    def _supplier_card_context(
        session: Any,
        workspace_owner_id: str,
        supplier_id: str,
        *,
        balance_by_id: dict[str, Decimal],
        balance_by_name: dict[str, Decimal],
        last_date_by_id: dict[str, str],
        last_date_by_name: dict[str, str],
    ) -> dict[str, Any] | None:
        row = session.get(Counterparty, supplier_id)
        if not row or row.workspace_owner_id != workspace_owner_id:
            return None
        _, has_supplier = _counterparty_role_flags(row.kind)
        if not has_supplier:
            return None
        supplier = _counterparty_view_data(
            row,
            balance_by_id=balance_by_id,
            balance_by_name=balance_by_name,
            last_date_by_id=last_date_by_id,
            last_date_by_name=last_date_by_name,
        )
        names = {
            str(supplier.get("name") or "").strip().lower(),
            str(supplier.get("official_name") or "").strip().lower(),
        }
        names.discard("")
        purchase_rows = list(
            session.execute(
                select(PurchaseDocument)
                .where(PurchaseDocument.workspace_owner_id == workspace_owner_id)
                .order_by(PurchaseDocument.updated_at.desc())
            ).scalars()
        )
        purchases: list[dict[str, Any]] = []
        payables: list[dict[str, Any]] = []
        reconciliation_rows: list[dict[str, str]] = []
        total_amount = Decimal("0")
        total_paid = Decimal("0")
        running_balance = Decimal("0")
        reconciliation_total_purchase = Decimal("0")
        reconciliation_total_paid = Decimal("0")
        reconciliation_currency = str(supplier.get("currency") or "UZS").strip().upper() or "UZS"

        def purchase_sort_key(document: PurchaseDocument) -> tuple[str, str]:
            document_data = _json_object(document.data)
            created_at = getattr(document, "created_at", None)
            return (
                str(document_data.get("date") or ""),
                created_at.isoformat() if created_at else "",
            )

        for purchase_row in purchase_rows:
            data = _json_object(purchase_row.data)
            counterparty_id = str(data.get("counterparty_id") or purchase_row.counterparty_id or "").strip()
            supplier_name = str(data.get("supplier") or "").strip().lower()
            if counterparty_id != supplier_id and supplier_name not in names:
                continue
            item = _purchase_document_data(purchase_row)
            purchases.append(item)
            total_amount += _sales_decimal(purchase_row.amount)
            total_paid += _sales_decimal(data.get("paid_amount"))
            if _sales_decimal(item.get("debt_amount")) > 0:
                payables.append(item)
        for purchase_row in sorted(
            [
                row_item
                for row_item in purchase_rows
                if (
                    str(_json_object(row_item.data).get("counterparty_id") or row_item.counterparty_id or "").strip() == supplier_id
                    or str(_json_object(row_item.data).get("supplier") or "").strip().lower() in names
                )
            ],
            key=purchase_sort_key,
        ):
            data = _json_object(purchase_row.data)
            item = _purchase_document_data(purchase_row)
            amount = _sales_decimal(purchase_row.amount)
            paid_amount = _sales_decimal(data.get("paid_amount"))
            currency = str(item.get("currency") or purchase_row.currency or reconciliation_currency or "UZS").strip().upper() or "UZS"
            reconciliation_currency = currency
            running_balance += amount
            reconciliation_total_purchase += amount
            reconciliation_rows.append(
                {
                    "date": str(item.get("date") or ""),
                    "document": f"Закупка {item['number']}",
                    "purchase": _sales_money_label(amount),
                    "payment": _sales_money_label(Decimal("0")),
                    "balance": _sales_money_label(running_balance),
                    "currency": currency,
                }
            )
            if paid_amount > 0:
                running_balance -= paid_amount
                reconciliation_total_paid += paid_amount
                reconciliation_rows.append(
                    {
                        "date": str(item.get("date") or ""),
                        "document": f"Оплата по {item['number']}",
                        "purchase": _sales_money_label(Decimal("0")),
                        "payment": _sales_money_label(paid_amount),
                        "balance": _sales_money_label(running_balance),
                        "currency": currency,
                    }
                )
        supplier.update(
            {
                "purchases": purchases[:10],
                "payables": payables[:10],
                "reconciliation": reconciliation_rows[-24:],
                "reconciliation_totals": {
                    "purchase": _sales_money_label(reconciliation_total_purchase),
                    "payment": _sales_money_label(reconciliation_total_paid),
                    "balance": _sales_money_label(running_balance),
                    "currency": reconciliation_currency,
                },
                "summary": {
                    "purchases": _sales_money_label(total_amount),
                    "paid": _sales_money_label(total_paid),
                    "payable": supplier["balance"],
                    "purchase_count": len(purchases),
                },
            }
        )
        return supplier

    def _module_flash_error(request: Request) -> str:
        return request.query_params.get("error") or ("Форма устарела. Обновите страницу и повторите." if request.query_params.get("err") == "csrf" else "")

    def _purchase_save_redirect(form: Any, workspace_owner_id: str, base_url: str) -> RedirectResponse:
        try:
            data, amount, currency = _purchase_document_payload(form)
        except ValueError as exc:
            return RedirectResponse(url=f"{base_url}?error=" + quote(str(exc)) + "#purchases", status_code=302)
        if not data["supplier"]:
            return RedirectResponse(url=f"{base_url}?error=" + quote("Поставщик обязателен") + "#purchases", status_code=302)
        with session_scope() as session:
            try:
                supplier_row = _ensure_counterparty(
                    session,
                    workspace_owner_id,
                    name=data["supplier"],
                    role="supplier",
                )
                warehouse_row = _ensure_warehouse(
                    session,
                    workspace_owner_id,
                    name=data["warehouse"],
                )
                data["lines"] = _sync_product_lines(
                    session,
                    workspace_owner_id,
                    warehouse_name=warehouse_row.name,
                    lines=list(data.get("lines") or []),
                    delta_sign=1,
                    op_date=str(data.get("date") or ""),
                    require_stock_product=True,
                )
                price_types = _workspace_price_types(workspace_owner_id)
                selected_price_type = next(
                    (item for item in price_types if str(item.get("id") or "") == str(data.get("price_type_id") or "")),
                    None,
                )
                if selected_price_type:
                    data["price_type_name"] = str(selected_price_type.get("name") or data.get("price_type_name") or "")
                    price_currency = str(selected_price_type.get("convert_to_currency") or currency or "UZS").strip().upper() or "UZS"
                    data["price_type_currency"] = price_currency
                if selected_price_type and data.get("save_prices_to_price_type"):
                    for line in data["lines"]:
                        sale_price = _sales_decimal(line.get("sale_price"))
                        product_id = str(line.get("product_id") or "").strip()
                        if sale_price <= 0 or not product_id:
                            continue
                        sale_price = _convert_product_currency(
                            sale_price,
                            currency,
                            price_currency,
                            _workspace_usd_uzs_rate(workspace_owner_id),
                        )
                        product_row = session.get(Product, product_id)
                        if product_row and product_row.workspace_owner_id == workspace_owner_id:
                            _apply_product_price_change(product_row, selected_price_type, sale_price, price_currency)
                data["warehouse_id"] = warehouse_row.id
                data["counterparty_id"] = supplier_row.id
            except ValueError as exc:
                return RedirectResponse(url=f"{base_url}?error=" + quote(str(exc)) + "#purchases", status_code=302)
            count = session.execute(
                select(func.count(PurchaseDocument.id)).where(PurchaseDocument.workspace_owner_id == workspace_owner_id)
            ).scalar_one()
            number = str(form.get("number") or "").strip() or f"P-{datetime.now(timezone.utc).strftime('%Y%m%d')}-{int(count) + 1:03d}"
            row = PurchaseDocument(
                id=str(uuid.uuid4()),
                workspace_owner_id=workspace_owner_id,
                number=number,
                amount=amount,
                currency=currency,
                counterparty_id=supplier_row.id,
                external_source="local",
                external_id=str(uuid.uuid4()),
                data=data,
            )
            session.add(row)
        return RedirectResponse(url=f"{base_url}?msg=saved#purchases", status_code=302)

    def _purchase_delete_redirect(purchase_id: str, workspace_owner_id: str, base_url: str) -> RedirectResponse:
        with session_scope() as session:
            row = session.get(PurchaseDocument, purchase_id)
            if row and row.workspace_owner_id == workspace_owner_id:
                data = _json_object(row.data)
                try:
                    _sync_product_lines(
                        session,
                        workspace_owner_id,
                        warehouse_name=str(data.get("warehouse") or "Основной склад"),
                        lines=list(data.get("lines") or []),
                        delta_sign=-1,
                        op_date=str(data.get("date") or ""),
                    )
                except ValueError as exc:
                    return RedirectResponse(url=f"{base_url}?error=" + quote(str(exc)) + "#purchases", status_code=302)
                session.delete(row)
        return RedirectResponse(url=f"{base_url}?msg=deleted#purchases", status_code=302)

    def _purchase_pay_redirect(purchase_id: str, workspace_owner_id: str, base_url: str) -> RedirectResponse:
        with session_scope() as session:
            row = session.get(PurchaseDocument, purchase_id)
            if not row or row.workspace_owner_id != workspace_owner_id:
                return RedirectResponse(url=f"{base_url}?error=" + quote("Закупка не найдена") + "#purchases", status_code=302)
            data = _json_object(row.data)
            amount = _sales_decimal(row.amount)
            paid_amount = _sales_decimal(data.get("paid_amount"))
            debt_amount = amount - paid_amount
            if debt_amount <= 0:
                return RedirectResponse(url=f"{base_url}?msg=paid#purchases", status_code=302)
            currency = str(row.currency or data.get("currency") or "UZS").strip().upper() or "UZS"
            raw_payment_lines = data.get("payment_lines") if isinstance(data.get("payment_lines"), list) else []
            payment_lines = [dict(item) for item in raw_payment_lines if isinstance(item, dict)]
            lines_total = sum((_sales_decimal(item.get("amount")) for item in payment_lines), Decimal("0"))
            if paid_amount > lines_total:
                payment_lines.append(
                    {
                        "amount": _decimal_plain_text(paid_amount - lines_total),
                        "currency": currency,
                        "type": str(data.get("payment_type") or "Оплата"),
                        "account": "Наличные",
                        "account_id": "",
                    }
                )
            new_payment_lines, payment_amount, payment_type = _document_payment_entries_from_form(form, currency, debt_amount, workspace_owner_id)
            payment_lines.extend(new_payment_lines)
            next_paid_amount = paid_amount + payment_amount
            if next_paid_amount > amount:
                next_paid_amount = amount
            data["paid_amount"] = _decimal_plain_text(next_paid_amount)
            data["payment_type"] = payment_type or str(data.get("payment_type") or "Оплата")
            data["payment_lines"] = payment_lines
            data["status"] = "paid" if next_paid_amount >= amount else "partial"
            row.data = data
            flag_modified(row, "data")
            session.add(row)
        return RedirectResponse(url=f"{base_url}?msg=paid#purchases", status_code=302)

    @app.get("/warehouse", response_class=HTMLResponse, name="warehouse_get")
    def warehouse_get(
        request: Request,
        q: str = "",
        warehouse: str = "",
        product: str = "",
        op_type: str = "all",
        critical: str = "",
        op: str = "",
    ):
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        filters = {
            "q": q.strip(),
            "warehouse": warehouse.strip(),
            "product": product.strip(),
            "op_type": op_type.strip() or "all",
            "critical": "1" if str(critical or "").strip() == "1" else "",
        }
        operation_preset = str(op or "").strip()
        if operation_preset not in {"in", "out", "transfer"}:
            operation_preset = ""
        q_clean = filters["q"].lower()
        warehouse_records: list[dict[str, Any]] = []
        warehouse_stocks: list[dict[str, Any]] = []
        warehouse_operations: list[dict[str, Any]] = []
        warehouse_purchases: list[dict[str, Any]] = []
        product_names: list[str] = []
        supplier_names: list[str] = []
        warehouse_stock_total = Decimal("0")
        warehouse_purchase_totals: dict[str, Decimal] = {}
        with session_scope() as session:
            warehouse_rows = list(
                session.execute(
                    select(Warehouse)
                    .where(Warehouse.workspace_owner_id == wid)
                    .order_by(Warehouse.name.asc())
                ).scalars()
            )
            product_rows = list(
                session.execute(
                    select(Product)
                    .where(Product.workspace_owner_id == wid)
                    .order_by(Product.name.asc())
                ).scalars()
            )
            operation_rows = list(
                session.execute(
                    select(WarehouseOperation)
                    .where(WarehouseOperation.workspace_owner_id == wid)
                    .order_by(WarehouseOperation.updated_at.desc())
                ).scalars()
            )
            purchase_rows = list(
                session.execute(
                    select(PurchaseDocument)
                    .where(PurchaseDocument.workspace_owner_id == wid)
                    .order_by(PurchaseDocument.updated_at.desc())
                ).scalars()
            )
            supplier_names = [
                str(row.name)
                for row in session.execute(
                    select(Counterparty)
                    .where(
                        Counterparty.workspace_owner_id == wid,
                        Counterparty.kind.in_(["supplier", "both"]),
                    )
                    .order_by(Counterparty.name.asc())
                ).scalars()
            ]
            warehouse_records = [_warehouse_view_data(row) for row in warehouse_rows]
            product_names = [str(row.name) for row in product_rows]
            for product_row in product_rows:
                item = _product_data(product_row)
                for stock in item["stocks"]:
                    quantity_value = _sales_decimal(stock.get("quantity"))
                    cost_value = _sales_decimal(stock.get("price"))
                    min_stock_value = _sales_decimal(item["min_stock"])
                    stock_total_value = quantity_value * cost_value
                    is_critical = bool(min_stock_value > 0 and quantity_value <= min_stock_value)
                    stock_row = {
                        "product": item["name"],
                        "warehouse": str(stock.get("warehouse") or "Основной склад"),
                        "quantity": str(quantity_value.normalize() if quantity_value else "0"),
                        "quantity_value": quantity_value,
                        "unit": item["unit"],
                        "category": item["category"],
                        "status": item["status"],
                        "min_stock": item["min_stock"],
                        "cost": _sales_money_label(cost_value),
                        "cost_value": cost_value,
                        "stock_total": _sales_money_label(stock_total_value),
                        "stock_total_value": stock_total_value,
                        "is_critical": is_critical,
                    }
                    hay = " ".join([stock_row["product"], stock_row["warehouse"], stock_row["category"]]).lower()
                    if q_clean and q_clean not in hay:
                        continue
                    if filters["warehouse"] and stock_row["warehouse"] != filters["warehouse"]:
                        continue
                    if filters["product"] and stock_row["product"] != filters["product"]:
                        continue
                    if filters["critical"] == "1" and not is_critical:
                        continue
                    warehouse_stocks.append(stock_row)
                    warehouse_stock_total += stock_total_value
            for row in operation_rows:
                item = _warehouse_operation_data(row)
                hay = " ".join([item["number"], item["product"], item["warehouse"], item["from_warehouse"], item["to_warehouse"], item["note"]]).lower()
                if q_clean and q_clean not in hay:
                    continue
                if filters["warehouse"] and filters["warehouse"] not in {item["warehouse"], item["from_warehouse"], item["to_warehouse"]}:
                    continue
                if filters["product"] and item["product"] != filters["product"]:
                    continue
                if filters["op_type"] != "all" and item["operation_type"] != filters["op_type"]:
                    continue
                warehouse_operations.append(item)
            for row in purchase_rows:
                item = _purchase_document_data(row)
                line_products = " ".join(
                    str(line.get("product") or "")
                    for line in item["lines"]
                    if isinstance(line, dict)
                ).lower()
                hay = " ".join([item["number"], item["supplier"], item["warehouse"], item["status_label"], item["note"], line_products]).lower()
                if q_clean and q_clean not in hay:
                    continue
                if filters["warehouse"] and item["warehouse"] != filters["warehouse"]:
                    continue
                if filters["product"] and filters["product"].lower() not in line_products:
                    continue
                warehouse_purchases.append(item)
                currency = str(row.currency or item["currency"] or "UZS").upper()
                warehouse_purchase_totals[currency] = warehouse_purchase_totals.get(currency, Decimal("0")) + _sales_decimal(row.amount)
        price_types = _workspace_price_types(wid)
        purchase_price_types = [
            item
            for item in price_types
            if item.get("is_active") and item.get("is_for_sales") and item.get("pricing_method") == "manual"
        ] or [
            item
            for item in price_types
            if item.get("is_active") and item.get("is_for_sales")
        ] or price_types
        payment_accounts: list[dict[str, str]] = []
        treasury = load_treasury(wid)
        for pocket in treasury.get("pockets") or []:
            if not isinstance(pocket, dict):
                continue
            label = str(pocket.get("label") or "").strip()
            if not label:
                continue
            payment_accounts.append(
                {
                    "id": str(pocket.get("id") or "").strip(),
                    "label": label,
                    "template_id": str(pocket.get("template_id") or "").strip(),
                }
            )
        if not payment_accounts:
            payment_accounts = [
                {"id": "cash", "label": "Наличные", "template_id": "cash"},
                {"id": "card", "label": "Карта", "template_id": "card"},
            ]
        warehouse_options = {
            "warehouses": [item["name"] for item in warehouse_records] or ["Основной склад"],
            "products": product_names,
            "product_rows": [_product_data(row) for row in product_rows],
            "price_types": sorted(
                [
                    {
                        "id": str(item.get("id") or ""),
                        "name": str(item.get("name") or ""),
                        "currency": str(item.get("convert_to_currency") or "UZS").upper(),
                        "sort_order": int(item.get("sort_order") or 0),
                    }
                    for item in purchase_price_types
                    if str(item.get("id") or "").strip()
                ],
                key=lambda item: (item["sort_order"], item["name"]),
            ),
            "suppliers": supplier_names,
            "payment_accounts": payment_accounts,
            "fx": {"USD_UZS": _decimal_plain_text(_workspace_usd_uzs_rate(wid))},
        }
        return tpl(
            request,
            "home_business_module.html",
            variant="user",
            active="warehouse",
            module=_business_module_context("warehouse"),
            warehouse_filters=filters,
            warehouse_options=warehouse_options,
            warehouse_records=warehouse_records,
            warehouse_stocks=warehouse_stocks,
            warehouse_stock_total=_sales_money_label(warehouse_stock_total),
            warehouse_operations=warehouse_operations,
            warehouse_purchases=warehouse_purchases,
            warehouse_purchase_totals=[
                {"currency": currency, "amount": _sales_money_label(amount)}
                for currency, amount in sorted(warehouse_purchase_totals.items())
            ],
            warehouse_operation_preset=operation_preset,
            today=datetime.now(timezone.utc).date().isoformat(),
            flash_ok=request.query_params.get("msg"),
            flash_err=_module_flash_error(request),
        )

    @app.post("/warehouse/save", name="warehouse_save")
    async def warehouse_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/warehouse?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        name = str(form.get("name") or "").strip()
        if not name:
            return RedirectResponse(url="/warehouse?error=" + quote("Название склада обязательно") + "#warehouses", status_code=302)
        with session_scope() as session:
            _ensure_warehouse(
                session,
                wid,
                name=name,
                data={
                    "manager": str(form.get("manager") or "").strip(),
                    "status": str(form.get("status") or "active").strip(),
                    "note": str(form.get("note") or "").strip(),
                },
            )
        return RedirectResponse(url="/warehouse?msg=saved#warehouses", status_code=302)

    @app.post("/warehouse/operations/save", name="warehouse_operation_save")
    async def warehouse_operation_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/warehouse?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        try:
            data, amount = _warehouse_operation_payload(form)
        except ValueError as exc:
            return RedirectResponse(url="/warehouse?error=" + quote(str(exc)) + "#adjustments", status_code=302)
        if not data["product"]:
            return RedirectResponse(url="/warehouse?error=" + quote("Товар обязателен") + "#adjustments", status_code=302)
        quantity = _sales_decimal(data.get("quantity"))
        if quantity <= 0:
            return RedirectResponse(url="/warehouse?error=" + quote("Количество должно быть больше нуля") + "#adjustments", status_code=302)
        with session_scope() as session:
            product_row = _resolve_product_row(session, wid, data["product"])
            if product_row is None:
                return RedirectResponse(url="/warehouse?error=" + quote("Товар не найден") + "#adjustments", status_code=302)
            target_hash = "transfers" if data["operation_type"] == "transfer" else "adjustments"
            try:
                if data["operation_type"] == "transfer":
                    from_row = _ensure_warehouse(session, wid, name=data["from_warehouse"])
                    to_row = _ensure_warehouse(session, wid, name=data["to_warehouse"])
                    if from_row.name == to_row.name:
                        return RedirectResponse(url="/warehouse?error=" + quote("Склады перемещения должны отличаться") + "#transfers", status_code=302)
                    _sync_product_lines(
                        session,
                        wid,
                        warehouse_name=from_row.name,
                        lines=[{"product": product_row.name, "product_id": product_row.id, "quantity": data["quantity"], "price": data["price"]}],
                        delta_sign=-1,
                        op_date=str(data.get("date") or ""),
                        require_stock_product=True,
                    )
                    _sync_product_lines(
                        session,
                        wid,
                        warehouse_name=to_row.name,
                        lines=[{"product": product_row.name, "product_id": product_row.id, "quantity": data["quantity"], "price": data["price"]}],
                        delta_sign=1,
                        op_date=str(data.get("date") or ""),
                        require_stock_product=True,
                    )
                    data["warehouse_id"] = from_row.id
                else:
                    warehouse_row = _ensure_warehouse(session, wid, name=data["warehouse"])
                    delta_sign = 1 if data["operation_type"] in {"in", "adjustment"} else -1
                    _sync_product_lines(
                        session,
                        wid,
                        warehouse_name=warehouse_row.name,
                        lines=[{"product": product_row.name, "product_id": product_row.id, "quantity": data["quantity"], "price": data["price"]}],
                        delta_sign=delta_sign,
                        op_date=str(data.get("date") or ""),
                        require_stock_product=True,
                    )
                    data["warehouse_id"] = warehouse_row.id
            except ValueError as exc:
                return RedirectResponse(url="/warehouse?error=" + quote(str(exc)) + f"#{target_hash}", status_code=302)
            count = session.execute(
                select(func.count(WarehouseOperation.id)).where(WarehouseOperation.workspace_owner_id == wid)
            ).scalar_one()
            number = f"W-{datetime.now(timezone.utc).strftime('%Y%m%d')}-{int(count) + 1:03d}"
            row = WarehouseOperation(
                id=str(uuid.uuid4()),
                workspace_owner_id=wid,
                number=number,
                operation_type=data["operation_type"],
                warehouse_id=str(data.get("warehouse_id") or "") or None,
                product_id=product_row.id,
                quantity=quantity,
                amount=amount,
                currency="UZS",
                data=data,
            )
            session.add(row)
        return RedirectResponse(url=f"/warehouse?msg=saved#{target_hash}", status_code=302)

    @app.post("/warehouse/operations/{operation_id}/delete", name="warehouse_operation_delete")
    async def warehouse_operation_delete(request: Request, operation_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/warehouse?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        with session_scope() as session:
            row = session.get(WarehouseOperation, operation_id)
            if row and row.workspace_owner_id == wid:
                data = _json_object(row.data)
                quantity = _sales_decimal(row.quantity)
                target_hash = "transfers" if row.operation_type == "transfer" else "adjustments"
                try:
                    if row.operation_type == "transfer":
                        _sync_product_lines(
                            session,
                            wid,
                            warehouse_name=str(data.get("to_warehouse") or ""),
                            lines=[{"product": data.get("product"), "product_id": row.product_id, "quantity": str(quantity), "price": data.get("price")}],
                            delta_sign=-1,
                            op_date=str(data.get("date") or ""),
                        )
                        _sync_product_lines(
                            session,
                            wid,
                            warehouse_name=str(data.get("from_warehouse") or ""),
                            lines=[{"product": data.get("product"), "product_id": row.product_id, "quantity": str(quantity), "price": data.get("price")}],
                            delta_sign=1,
                            op_date=str(data.get("date") or ""),
                        )
                    else:
                        _sync_product_lines(
                            session,
                            wid,
                            warehouse_name=str(data.get("warehouse") or ""),
                            lines=[{"product": data.get("product"), "product_id": row.product_id, "quantity": str(quantity), "price": data.get("price")}],
                            delta_sign=-1 if row.operation_type in {"in", "adjustment"} else 1,
                            op_date=str(data.get("date") or ""),
                        )
                except ValueError as exc:
                    return RedirectResponse(url="/warehouse?error=" + quote(str(exc)) + f"#{target_hash}", status_code=302)
                session.delete(row)
                return RedirectResponse(url=f"/warehouse?msg=deleted#{target_hash}", status_code=302)
        return RedirectResponse(url="/warehouse?msg=deleted#adjustments", status_code=302)

    @app.post("/warehouse/purchases/save", name="warehouse_purchase_save")
    async def warehouse_purchase_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/warehouse?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        return _purchase_save_redirect(form, wid, "/warehouse")

    @app.post("/warehouse/purchases/{purchase_id}/delete", name="warehouse_purchase_delete")
    async def warehouse_purchase_delete(request: Request, purchase_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/warehouse?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        return _purchase_delete_redirect(purchase_id, wid, "/warehouse")

    @app.post("/warehouse/purchases/{purchase_id}/pay", name="warehouse_purchase_pay")
    async def warehouse_purchase_pay(request: Request, purchase_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/warehouse?err=csrf#purchases", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        return _purchase_pay_redirect(purchase_id, wid, "/warehouse")

    @app.get("/clients", response_class=HTMLResponse, name="clients_get")
    def clients_get(
        request: Request,
        q: str = "",
        territory: str = "",
        category: str = "",
        program: str = "",
        route: str = "",
        status: str = "all",
        client: str = "",
    ):
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        filters = {
            "q": q.strip(),
            "territory": territory.strip(),
            "category": category.strip(),
            "program": program.strip(),
            "route": route.strip(),
            "status": status.strip() or "all",
        }
        q_clean = filters["q"].lower()
        clients_records: list[dict[str, Any]] = []
        client_routes: list[dict[str, Any]] = []
        client_balances: list[dict[str, Any]] = []
        selected_client_card: dict[str, Any] | None = None
        with session_scope() as session:
            balance_by_id, balance_by_name, last_date_by_id, last_date_by_name = _sales_rollup_maps(session, wid)
            balance_currency_by_id, balance_currency_by_name = _sales_rollup_currency_maps(session, wid)
            settings_payload = load_workspace_settings(wid)
            raw_calls = settings_payload.get("telephony_calls") if isinstance(settings_payload.get("telephony_calls"), list) else []
            crm_rows_for_contact = list(
                session.execute(
                    select(CrmRecord)
                    .where(CrmRecord.workspace_owner_id == wid)
                    .order_by(CrmRecord.updated_at.desc())
                ).scalars()
            )
            crm_stages_for_clients = _crm_workspace_stages(wid)
            crm_status_by_id, crm_status_by_name = _crm_status_maps_from_records(crm_rows_for_contact, crm_stages_for_clients)
            telegram_subscribers_for_contact = list(
                session.execute(
                    select(TelegramSubscriber)
                    .where(TelegramSubscriber.workspace_owner_id == wid)
                    .order_by(TelegramSubscriber.requested_at.desc())
                ).scalars()
            )

            def digits(value: Any) -> str:
                return "".join(ch for ch in str(value or "") if ch.isdigit())

            def bump_date(current: str, candidate: Any) -> str:
                raw = str(candidate or "").strip()
                if not raw:
                    return current
                return raw if raw > current else current

            def client_last_contact(row: Counterparty, item: dict[str, Any]) -> str:
                names = {
                    str(item.get("name") or "").strip().lower(),
                    str(item.get("official_name") or "").strip().lower(),
                }
                names.discard("")
                best = str(item.get("last_date") or "")
                phone_digits = digits(item.get("phone"))
                telegram_key = str(item.get("telegram") or "").lstrip("@").lower()
                for call in raw_calls:
                    if not isinstance(call, dict):
                        continue
                    call_client = str(call.get("client") or "").strip().lower()
                    call_phone = digits(call.get("phone"))
                    if call_client in names or (phone_digits and call_phone and call_phone.endswith(phone_digits[-9:])):
                        best = bump_date(best, call.get("started_at") or call.get("created_at"))
                for crm_row in crm_rows_for_contact:
                    data = _json_object(crm_row.data)
                    crm_client = str(data.get("client") or "").strip().lower()
                    if crm_row.counterparty_id == row.id or crm_client in names:
                        best = bump_date(best, data.get("date") or data.get("due_date") or crm_row.updated_at.isoformat())
                for subscriber in telegram_subscribers_for_contact:
                    username = str(subscriber.username or "").lstrip("@").lower()
                    sub_phone = digits(subscriber.phone)
                    display_name = str(subscriber.display_name or "").strip().lower()
                    if (
                        (telegram_key and username == telegram_key)
                        or (phone_digits and sub_phone and sub_phone.endswith(phone_digits[-9:]))
                        or (display_name and display_name in names)
                    ):
                        best = bump_date(best, subscriber.requested_at.isoformat() if subscriber.requested_at else "")
                return best

            selected_client_id = client.strip()
            if selected_client_id:
                try:
                    selected_client_card = _client_card_context(
                        session,
                        wid,
                        selected_client_id,
                        balance_by_id=balance_by_id,
                        balance_by_name=balance_by_name,
                        last_date_by_id=last_date_by_id,
                        last_date_by_name=last_date_by_name,
                        balance_currency_by_id=balance_currency_by_id,
                        balance_currency_by_name=balance_currency_by_name,
                        crm_status_by_id=crm_status_by_id,
                        crm_status_by_name=crm_status_by_name,
                    )
                except Exception:
                    logger.exception("Failed to build client card %s", selected_client_id)
                    selected_client_card = None
            rows = list(
                session.execute(
                    select(Counterparty)
                    .where(
                        Counterparty.workspace_owner_id == wid,
                        Counterparty.kind.in_(["client", "both"]),
                    )
                    .order_by(Counterparty.updated_at.desc())
                ).scalars()
            )
            route_map: dict[str, dict[str, Any]] = {}
            for row in rows:
                item = _counterparty_view_data(
                    row,
                    balance_by_id=balance_by_id,
                    balance_by_name=balance_by_name,
                    last_date_by_id=last_date_by_id,
                    last_date_by_name=last_date_by_name,
                    balance_currency_by_id=balance_currency_by_id,
                    balance_currency_by_name=balance_currency_by_name,
                    crm_status_by_id=crm_status_by_id,
                    crm_status_by_name=crm_status_by_name,
                )
                hay = " ".join(
                    [
                        item["id"],
                        item["name"],
                        item["official_name"],
                        item["legal_name"],
                        item["phone"],
                        item["telegram_phone"],
                        item["category"],
                        item["territory"],
                        item["route"],
                        item["inn"],
                        item["pinfl"],
                        item["address"],
                        item["code"],
                    ]
                ).lower()
                if q_clean and q_clean not in hay:
                    continue
                if filters["territory"] and item["territory"] != filters["territory"]:
                    continue
                if filters["category"] and item["category"] != filters["category"]:
                    continue
                if filters["program"] and filters["program"] not in item["programs"]:
                    continue
                if filters["route"] and item["route"] != filters["route"]:
                    continue
                if filters["status"] != "all" and item["status"] != filters["status"]:
                    continue
                clients_records.append(item)
                item["last_contact"] = client_last_contact(row, item)
                route_key = item["route"] or "Без маршрута"
                extra = _counterparty_extra(row)
                summary = route_map.setdefault(
                    route_key,
                    {
                        "route": route_key,
                        "territory": item["territory"] or "-",
                        "count": 0,
                        "balance_value": Decimal("0"),
                        "clients": [],
                        "days": set(),
                        "agents": set(),
                    },
                )
                summary["count"] += 1
                summary["balance_value"] += item["balance_value"]
                summary["clients"].append(item["name"])
                if extra.get("route_day"):
                    summary["days"].add(str(extra.get("route_day")))
                if extra.get("route_agent"):
                    summary["agents"].add(str(extra.get("route_agent")))
                if item["balance_value"] > 0:
                    client_balances.append(item)
            client_routes = [
                {
                    "route": key,
                    "territory": value["territory"],
                    "count": value["count"],
                    "balance": _sales_money_label(value["balance_value"]),
                    "clients": ", ".join(value["clients"][:8]),
                    "day": ", ".join(sorted(value["days"])) or "-",
                    "agent": ", ".join(sorted(value["agents"])) or "-",
                }
                for key, value in sorted(route_map.items())
            ]
            client_balances.sort(key=lambda item: item["balance_value"], reverse=True)
        clients_debt_total = sum(
            (item["balance_value"] for item in clients_records if item["balance_value"] > 0),
            Decimal("0"),
        )
        clients_advance_total = sum(
            (abs(item["balance_value"]) for item in clients_records if item["balance_value"] < 0),
            Decimal("0"),
        )
        clients_balance_summary = {
            "debt": _sales_money_label(clients_debt_total),
            "advance": _sales_money_label(clients_advance_total),
            "net": _sales_money_label(clients_debt_total - clients_advance_total),
            "debt_count": sum(1 for item in clients_records if item["balance_value"] > 0),
            "advance_count": sum(1 for item in clients_records if item["balance_value"] < 0),
        }
        client_options = {
            "territories": sorted({item["territory"] for item in clients_records if item["territory"]}),
            "categories": sorted({item["category"] for item in clients_records if item["category"]}),
            "programs": sorted({program for item in clients_records for program in item["programs"] if program}),
            "routes": sorted({item["route"] for item in clients_records if item["route"]}),
            "clients": sorted({item["name"] for item in clients_records if item["name"]}),
        }
        return tpl(
            request,
            "home_business_module.html",
            variant="user",
            active="clients",
            module=_business_module_context("clients"),
            client_filters=filters,
            client_options=client_options,
            clients_records=clients_records,
            clients_balance_summary=clients_balance_summary,
            client_routes=client_routes,
            client_balances=client_balances,
            selected_client_card=selected_client_card,
            flash_ok=request.query_params.get("msg"),
            flash_err=_module_flash_error(request),
        )

    async def _save_client_photo_upload(workspace_owner_id: str, upload: Any) -> str:
        filename = str(getattr(upload, "filename", "") or "").strip()
        if not filename:
            return ""
        suffix = Path(filename).suffix.lower()
        if suffix not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
            return ""
        content_type = str(getattr(upload, "content_type", "") or "").lower()
        if content_type and not content_type.startswith("image/"):
            return ""
        payload = await upload.read()
        if not payload:
            return ""
        target_dir = BASE_DIR / "static" / "uploads" / "clients" / re.sub(r"[^a-zA-Z0-9_-]", "_", workspace_owner_id)
        target_dir.mkdir(parents=True, exist_ok=True)
        file_name = f"{uuid.uuid4().hex}{suffix}"
        target = target_dir / file_name
        target.write_bytes(payload)
        return f"/static/uploads/clients/{target_dir.name}/{file_name}"

    def _clean_client_coordinate(value: Any, *, minimum: float, maximum: float) -> str:
        raw = str(value or "").strip().replace(",", ".")
        if not raw:
            return ""
        try:
            number = float(raw)
        except ValueError:
            return ""
        if not (minimum <= number <= maximum):
            return ""
        return f"{number:.6f}".rstrip("0").rstrip(".")

    def _geocode_client_address(address: str) -> tuple[str, str]:
        query = str(address or "").strip()
        if not query:
            return "", ""
        url = "https://nominatim.openstreetmap.org/search?" + urlencode(
            {
                "format": "jsonv2",
                "limit": "1",
                "countrycodes": "uz,kz,kg,tj,tm",
                "q": query,
            }
        )
        request = UrlRequest(
            url,
            headers={
                "Accept": "application/json",
                "User-Agent": "UPOS/2.0 client-location-geocoder",
            },
        )
        try:
            with urlopen(request, timeout=5) as response:
                payload = json.loads(response.read().decode("utf-8") or "[]")
        except Exception:
            return "", ""
        if not isinstance(payload, list) or not payload:
            return "", ""
        first = payload[0] if isinstance(payload[0], dict) else {}
        lat = _clean_client_coordinate(first.get("lat"), minimum=-90, maximum=90)
        lon = _clean_client_coordinate(first.get("lon"), minimum=-180, maximum=180)
        return lat, lon

    @app.post("/clients/save", name="clients_save")
    async def clients_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/clients?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        name = str(form.get("name") or "").strip()
        if not name:
            return RedirectResponse(url="/clients?error=" + quote("Название клиента обязательно") + "#clients", status_code=302)
        is_supplier = str(form.get("is_supplier") or "").strip() == "1"
        client_id = str(form.get("client_id") or "").strip()
        uploaded_photo_url = await _save_client_photo_upload(wid, form.get("photo_file"))
        with session_scope() as session:
            client_type = str(form.get("client_type") or "company").strip()
            if client_type not in {"company", "individual"}:
                client_type = "company"
            status = str(form.get("status") or "active").strip()
            if status not in {"active", "inactive"}:
                status = "active"
            crm_status = str(form.get("crm_status") or "new_lead").strip()
            if crm_status not in CLIENT_CRM_STATUS_LABELS:
                crm_status = "new_lead"
            tax_id = str(form.get("tax_id") or form.get("inn") or "").strip()
            industry = str(form.get("industry") or "").strip()
            map_icon = str(form.get("map_icon") or "").strip()
            if not map_icon or map_icon == "default":
                map_icon = industry or "default"
            selected_programs: list[str] = []
            for value in form.getlist("programs"):
                program = str(value or "").strip()
                if program and program not in selected_programs:
                    selected_programs.append(program)
            custom_program = str(form.get("program_custom") or form.get("program") or "").strip()
            if custom_program and custom_program not in selected_programs:
                selected_programs.append(custom_program)
            address = str(form.get("address") or "").strip()
            latitude = _clean_client_coordinate(form.get("latitude"), minimum=-90, maximum=90)
            longitude = _clean_client_coordinate(form.get("longitude"), minimum=-180, maximum=180)
            if address and (not latitude or not longitude):
                geocoded_latitude, geocoded_longitude = _geocode_client_address(address)
                latitude = latitude or geocoded_latitude
                longitude = longitude or geocoded_longitude
            saved_row = _ensure_counterparty(
                session,
                wid,
                counterparty_id=client_id,
                name=name,
                role="both" if is_supplier else "client",
                phone=str(form.get("phone") or "").strip(),
                tax_id=tax_id,
                data={
                    "official_name": str(form.get("official_name") or "").strip(),
                    "legal_name": str(form.get("legal_name") or form.get("official_name") or name).strip(),
                    "legal_address": str(form.get("legal_address") or "").strip(),
                    "photo_url": uploaded_photo_url or str(form.get("photo_url") or "").strip(),
                    "client_type": client_type,
                    "inn": tax_id,
                    "pinfl": str(form.get("pinfl") or "").strip(),
                    "okved": str(form.get("okved") or "").strip(),
                    "territory": str(form.get("territory") or "").strip(),
                    "category": str(form.get("category") or "").strip(),
                    "industry": industry,
                    "map_icon": map_icon,
                    "program": ", ".join(selected_programs),
                    "programs": selected_programs,
                    "route": str(form.get("route") or "").strip(),
                    "status": status,
                    "crm_status": crm_status,
                    "is_blacklisted": str(form.get("is_blacklisted") or "").strip() == "1",
                    "sort_order": str(form.get("sort_order") or "0").strip(),
                    "code": str(form.get("code") or "").strip(),
                    "telegram": str(form.get("telegram") or "").strip(),
                    "telegram_phone": str(form.get("telegram_phone") or form.get("telegram") or "").strip(),
                    "note": str(form.get("note") or "").strip(),
                    "comment": str(form.get("comment") or form.get("note") or "").strip(),
                    "email": str(form.get("email") or "").strip(),
                    "address": address,
                    "latitude": latitude,
                    "longitude": longitude,
                    "price_type": str(form.get("price_type") or "").strip(),
                    "credit_limit": str(form.get("credit_limit") or "").strip(),
                    "consignment_days": str(form.get("consignment_days") or "").strip(),
                    "cashback_percent": str(form.get("cashback_percent") or "").strip(),
                    "delivery_frequency_days": str(form.get("delivery_frequency_days") or "").strip(),
                    "owner_user": str(form.get("owner_user") or "").strip(),
                    "owner_group": str(form.get("owner_group") or "").strip(),
                    "is_client": True,
                    "is_supplier": is_supplier,
                },
            )
            _sync_crm_records_for_counterparty_status(
                session,
                wid,
                saved_row,
                crm_status,
                _crm_workspace_stages(wid),
            )
            client_id = saved_row.id
        if client_id and str(form.get("return_to_card") or "").strip() == "1":
            return RedirectResponse(url=f"/clients?client={quote(client_id)}&msg=saved#client-card", status_code=302)
        return RedirectResponse(url="/clients?msg=saved#clients", status_code=302)

    @app.post("/api/clients/{counterparty_id}/location", name="client_location_save_api")
    async def client_location_save_api(request: Request, counterparty_id: str):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"ok": False, "error": "csrf"}, status_code=403)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return JSONResponse({"ok": False, "error": "auth"}, status_code=401)
        assert wid is not None
        try:
            payload = await request.json()
        except Exception:
            return JSONResponse({"ok": False, "error": "json"}, status_code=400)
        if not isinstance(payload, dict):
            return JSONResponse({"ok": False, "error": "json"}, status_code=400)
        latitude = _clean_client_coordinate(payload.get("latitude"), minimum=-90, maximum=90)
        longitude = _clean_client_coordinate(payload.get("longitude"), minimum=-180, maximum=180)
        if not latitude or not longitude:
            return JSONResponse({"ok": False, "error": "coordinates"}, status_code=400)
        address = str(payload.get("address") or "").strip()
        map_icon = str(payload.get("map_icon") or "").strip()
        with session_scope() as session:
            row = session.get(Counterparty, counterparty_id)
            if not row or row.workspace_owner_id != wid:
                return JSONResponse({"ok": False, "error": "not_found"}, status_code=404)
            has_client, _has_supplier = _counterparty_role_flags(row.kind)
            if not has_client:
                return JSONResponse({"ok": False, "error": "not_found"}, status_code=404)
            extra = _counterparty_extra(row)
            extra["latitude"] = latitude
            extra["longitude"] = longitude
            if address:
                extra["address"] = address
            if map_icon:
                extra["map_icon"] = map_icon
            row.data = extra
            flag_modified(row, "data")
        return JSONResponse(
            {
                "ok": True,
                "latitude": latitude,
                "longitude": longitude,
                "address": address,
                "map_icon": map_icon,
            }
        )

    @app.post("/clients/{counterparty_id}/delete", name="clients_delete")
    async def clients_delete(request: Request, counterparty_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/clients?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        with session_scope() as session:
            row = session.get(Counterparty, counterparty_id)
            if row and row.workspace_owner_id == wid:
                if not _drop_counterparty_role(row, "client"):
                    session.delete(row)
        return RedirectResponse(url="/clients?msg=deleted#clients", status_code=302)

    @app.post("/clients/routes/save", name="clients_route_save")
    async def clients_route_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/clients?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        client_name = str(form.get("client") or "").strip()
        route_name = str(form.get("route") or "").strip()
        if not client_name or not route_name:
            return RedirectResponse(url="/clients?error=" + quote("Клиент и маршрут обязательны") + "#routes", status_code=302)
        with session_scope() as session:
            row = _resolve_counterparty(session, wid, name=client_name, role="client")
            if row is None:
                return RedirectResponse(url="/clients?error=" + quote("Клиент не найден") + "#routes", status_code=302)
            extra = _counterparty_extra(row)
            extra["route"] = route_name
            extra["territory"] = str(form.get("territory") or extra.get("territory") or "").strip()
            extra["route_day"] = str(form.get("route_day") or "").strip()
            extra["route_agent"] = str(form.get("route_agent") or "").strip()
            extra["status"] = str(extra.get("status") or "active")
            row.data = extra
        return RedirectResponse(url="/clients?msg=saved#routes", status_code=302)

    @app.get("/suppliers", response_class=HTMLResponse, name="suppliers_get")
    def suppliers_get(
        request: Request,
        q: str = "",
        status: str = "all",
        category: str = "",
        supplier: str = "",
    ):
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        filters = {
            "q": q.strip(),
            "status": status.strip() or "all",
            "category": category.strip(),
        }
        q_clean = filters["q"].lower()
        supplier_records: list[dict[str, Any]] = []
        supplier_purchases: list[dict[str, Any]] = []
        supplier_payables: list[dict[str, Any]] = []
        selected_supplier_card: dict[str, Any] | None = None
        product_names: list[str] = []
        warehouse_names: list[str] = []
        with session_scope() as session:
            balance_by_id, balance_by_name, last_date_by_id, last_date_by_name = _purchase_rollup_maps(session, wid)
            balance_currency_by_id, balance_currency_by_name = _purchase_rollup_currency_maps(session, wid)
            selected_supplier_id = supplier.strip()
            if selected_supplier_id:
                selected_supplier_card = _supplier_card_context(
                    session,
                    wid,
                    selected_supplier_id,
                    balance_by_id=balance_by_id,
                    balance_by_name=balance_by_name,
                    last_date_by_id=last_date_by_id,
                    last_date_by_name=last_date_by_name,
                    balance_currency_by_id=balance_currency_by_id,
                    balance_currency_by_name=balance_currency_by_name,
                )
            counterparties = list(
                session.execute(
                    select(Counterparty)
                    .where(
                        Counterparty.workspace_owner_id == wid,
                        Counterparty.kind.in_(["supplier", "both"]),
                    )
                    .order_by(Counterparty.updated_at.desc())
                ).scalars()
            )
            for row in counterparties:
                item = _counterparty_view_data(
                    row,
                    balance_by_id=balance_by_id,
                    balance_by_name=balance_by_name,
                    last_date_by_id=last_date_by_id,
                    last_date_by_name=last_date_by_name,
                )
                hay = " ".join([item["name"], item["official_name"], item["phone"], item["category"]]).lower()
                if q_clean and q_clean not in hay:
                    continue
                if filters["status"] != "all" and item["status"] != filters["status"]:
                    continue
                if filters["category"] and item["category"] != filters["category"]:
                    continue
                supplier_records.append(item)
            purchase_rows = list(
                session.execute(
                    select(PurchaseDocument)
                    .where(PurchaseDocument.workspace_owner_id == wid)
                    .order_by(PurchaseDocument.updated_at.desc())
                ).scalars()
            )
            for row in purchase_rows:
                item = _purchase_document_data(row)
                hay = " ".join([item["number"], item["supplier"], item["warehouse"], item["status_label"], item["note"]]).lower()
                if q_clean and q_clean not in hay:
                    continue
                supplier_purchases.append(item)
                if _sales_decimal(item["debt_amount"]) > 0:
                    supplier_payables.append(item)
            product_names = [
                str(row.name)
                for row in session.execute(
                    select(Product)
                    .where(Product.workspace_owner_id == wid)
                    .order_by(Product.name.asc())
                ).scalars()
            ]
            warehouse_names = [
                str(row.name)
                for row in session.execute(
                    select(Warehouse)
                    .where(Warehouse.workspace_owner_id == wid)
                    .order_by(Warehouse.name.asc())
                ).scalars()
            ]
        supplier_options = {
            "categories": sorted({item["category"] for item in supplier_records if item["category"]}),
            "suppliers": sorted({item["name"] for item in supplier_records if item["name"]}),
            "products": product_names,
            "warehouses": warehouse_names or ["Основной склад"],
        }
        return tpl(
            request,
            "home_business_module.html",
            variant="user",
            active="suppliers",
            module=_business_module_context("suppliers"),
            supplier_filters=filters,
            supplier_options=supplier_options,
            supplier_records=supplier_records,
            supplier_purchases=supplier_purchases,
            supplier_payables=supplier_payables,
            selected_supplier_card=selected_supplier_card,
            today=datetime.now(timezone.utc).date().isoformat(),
            flash_ok=request.query_params.get("msg"),
            flash_err=_module_flash_error(request),
        )

    @app.post("/warehouse/suppliers/quick-save", name="warehouse_supplier_quick_save")
    async def warehouse_supplier_quick_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return JSONResponse({"error": "Форма устарела. Обновите страницу и повторите."}, status_code=403)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return JSONResponse({"error": "Нужно войти заново"}, status_code=401)
        assert wid is not None
        name = str(form.get("name") or "").strip()
        if not name:
            return JSONResponse({"error": "Название поставщика обязательно"}, status_code=400)
        phone = str(form.get("phone") or "").strip()
        tax_id = str(form.get("tax_id") or "").strip()
        data = {
            "official_name": str(form.get("official_name") or "").strip(),
            "category": str(form.get("category") or "").strip(),
            "status": str(form.get("status") or "active").strip(),
            "note": str(form.get("note") or "").strip(),
            "email": str(form.get("email") or "").strip(),
            "address": str(form.get("address") or "").strip(),
            "is_client": False,
            "is_supplier": True,
        }
        with session_scope() as session:
            row = _ensure_counterparty(
                session,
                wid,
                name=name,
                role="supplier",
                phone=phone,
                tax_id=tax_id,
                data=data,
            )
            session.flush()
            extra = _counterparty_extra(row)
            supplier = {
                "id": row.id,
                "name": row.name,
                "phone": row.phone or "",
                "tax_id": row.tax_id or "",
                "official_name": str(extra.get("official_name") or ""),
                "category": str(extra.get("category") or ""),
                "status": str(extra.get("status") or "active"),
                "email": str(extra.get("email") or ""),
                "address": str(extra.get("address") or ""),
                "note": str(extra.get("note") or ""),
            }
        return JSONResponse({"ok": True, "supplier": supplier})

    @app.post("/suppliers/save", name="suppliers_save")
    async def suppliers_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/suppliers?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        name = str(form.get("name") or "").strip()
        if not name:
            return RedirectResponse(url="/suppliers?error=" + quote("Название поставщика обязательно") + "#suppliers", status_code=302)
        is_client = str(form.get("is_client") or "").strip() == "1"
        with session_scope() as session:
            _ensure_counterparty(
                session,
                wid,
                name=name,
                role="both" if is_client else "supplier",
                phone=str(form.get("phone") or "").strip(),
                tax_id=str(form.get("tax_id") or "").strip(),
                data={
                    "official_name": str(form.get("official_name") or "").strip(),
                    "category": str(form.get("category") or "").strip(),
                    "status": str(form.get("status") or "active").strip(),
                    "note": str(form.get("note") or "").strip(),
                    "email": str(form.get("email") or "").strip(),
                    "address": str(form.get("address") or "").strip(),
                    "is_client": is_client,
                    "is_supplier": True,
                },
            )
        return RedirectResponse(url="/suppliers?msg=saved#suppliers", status_code=302)

    @app.post("/suppliers/{counterparty_id}/delete", name="suppliers_delete")
    async def suppliers_delete(request: Request, counterparty_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/suppliers?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        with session_scope() as session:
            row = session.get(Counterparty, counterparty_id)
            if row and row.workspace_owner_id == wid:
                if not _drop_counterparty_role(row, "supplier"):
                    session.delete(row)
        return RedirectResponse(url="/suppliers?msg=deleted#suppliers", status_code=302)

    @app.post("/suppliers/purchases/save", name="suppliers_purchase_save")
    async def suppliers_purchase_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/suppliers?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        return _purchase_save_redirect(form, wid, "/suppliers")

    @app.post("/suppliers/purchases/{purchase_id}/delete", name="suppliers_purchase_delete")
    async def suppliers_purchase_delete(request: Request, purchase_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/suppliers?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        return _purchase_delete_redirect(purchase_id, wid, "/suppliers")

    @app.post("/suppliers/purchases/{purchase_id}/pay", name="suppliers_purchase_pay")
    async def suppliers_purchase_pay(request: Request, purchase_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/suppliers?err=csrf#purchases", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        return _purchase_pay_redirect(purchase_id, wid, "/suppliers")

    @app.get("/crm", response_class=HTMLResponse, name="crm_get")
    def crm_get(
        request: Request,
        q: str = "",
        client: str = "",
        responsible: str = "",
        status: str = "all",
        priority: str = "all",
        date_from: str = "",
        date_to: str = "",
    ):
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        filters = {
            "q": q.strip(),
            "client": client.strip(),
            "responsible": responsible.strip(),
            "status": status.strip() or "all",
            "priority": priority.strip() or "all",
            "date_from": date_from.strip(),
            "date_to": date_to.strip(),
        }
        if filters["priority"] not in {"all", *CRM_PRIORITY_LABELS.keys()}:
            filters["priority"] = "all"
        q_clean = filters["q"].lower()
        crm_records: list[dict[str, Any]] = []
        crm_history_records: list[dict[str, Any]] = []
        crm_archive_records: list[dict[str, Any]] = []
        crm_next_actions: list[dict[str, Any]] = []
        crm_summary = {
            "active_count": 0,
            "open_deals": 0,
            "pipeline_total": "0",
            "overdue_count": 0,
            "today_count": 0,
            "won_count": 0,
            "conversion_rate": "0%",
            "missing_next_step_count": 0,
        }
        summary_pipeline_total = Decimal("0")
        summary_won = 0
        summary_lost = 0
        crm_options = {"clients": [], "responsibles": []}
        crm_workspace_settings = load_workspace_settings(wid)
        crm_stages = _crm_clean_stages(crm_workspace_settings.get("crm_pipeline_stages"))
        crm_activity_settings = _crm_activity_settings(crm_workspace_settings)
        crm_stage_map = {stage["id"]: {**stage, "records": [], "total_value": Decimal("0"), "total": "0"} for stage in crm_stages}
        messenger_threads_by_client: dict[str, list[dict[str, Any]]] = {}
        for thread in _messenger_threads_from_sources(wid):
            client_key = str(thread.get("client") or "").strip().casefold()
            if client_key:
                messenger_threads_by_client.setdefault(client_key, []).append(thread)
        today_iso = datetime.now(timezone.utc).date().isoformat()
        with session_scope() as session:
            counterparty_rows = list(
                session.execute(
                    select(Counterparty)
                    .where(
                        Counterparty.workspace_owner_id == wid,
                        Counterparty.kind.in_(["client", "both"]),
                    )
                    .order_by(Counterparty.name.asc())
                ).scalars()
            )
            counterparty_by_id = {row.id: row for row in counterparty_rows}
            counterparty_by_name = {row.name.strip().lower(): row for row in counterparty_rows if row.name}
            sale_by_key: dict[str, dict[str, Any]] = {}
            sale_by_id: dict[str, dict[str, Any]] = {}
            for sale_row in session.execute(
                select(SaleDocument)
                .where(SaleDocument.workspace_owner_id == wid)
                .order_by(SaleDocument.updated_at.desc())
            ).scalars():
                sale_data = _json_object(sale_row.data)
                sale_item = _sales_document_data(sale_row)
                sale_item["amount_value"] = _sales_decimal(sale_row.amount)
                sale_by_id[str(sale_row.id)] = sale_item
                keys = {
                    str(sale_data.get("counterparty_id") or sale_row.counterparty_id or "").strip(),
                    str(sale_data.get("client") or "").strip().lower(),
                }
                for key in keys:
                    if key and key not in sale_by_key:
                        sale_by_key[key] = sale_item
            subscribers = list(
                session.execute(
                    select(TelegramSubscriber)
                    .where(TelegramSubscriber.workspace_owner_id == wid)
                    .order_by(TelegramSubscriber.requested_at.desc())
                ).scalars()
            )

            def digits(value: Any) -> str:
                return "".join(ch for ch in str(value or "") if ch.isdigit())

            def chat_for_counterparty(counterparty: Counterparty | None) -> str:
                if counterparty is None:
                    return ""
                extra = _counterparty_extra(counterparty)
                telegram = str(extra.get("telegram") or "").lstrip("@").lower()
                phone_digits = digits(counterparty.phone)
                name_key = str(counterparty.name or "").strip().lower()
                for subscriber in subscribers:
                    if telegram and str(subscriber.username or "").lstrip("@").lower() == telegram:
                        return f"Telegram @{subscriber.username}" if subscriber.username else "Telegram"
                    sub_phone = digits(subscriber.phone)
                    if phone_digits and sub_phone and sub_phone.endswith(phone_digits[-9:]):
                        return f"Telegram {subscriber.display_name or subscriber.phone}".strip()
                    if name_key and str(subscriber.display_name or "").strip().lower() == name_key:
                        return f"Telegram {subscriber.display_name}".strip()
                return str(extra.get("telegram") or "")

            rows = list(
                session.execute(
                    select(CrmRecord)
                    .where(CrmRecord.workspace_owner_id == wid)
                    .order_by(CrmRecord.updated_at.desc())
                ).scalars()
            )
            all_deal_options = [
                {
                    "id": row.id,
                    "title": row.title,
                    "client": str(_json_object(row.data).get("client") or ""),
                }
                for row in rows
                if row.item_type == "deal"
            ]
            deal_options = [item for item in all_deal_options if next((row.status for row in rows if row.id == item["id"]), "") not in {"won", "lost", "done", "archived"}]
            deal_title_by_id = {item["id"]: item["title"] for item in all_deal_options}
            for row in rows:
                item = _crm_record_data(row)
                stage = _crm_stage_for_value(item["stage_id"] or item["stage"], crm_stages)
                item["stage_id"] = stage["id"]
                item["stage"] = stage["title"]
                item["related_deal_title"] = deal_title_by_id.get(item["related_deal_id"], "")
                if not item["lead_source"]:
                    item["lead_source"] = item["contact_type"] or "Ручной ввод"
                counterparty = counterparty_by_id.get(str(item.get("counterparty_id") or ""))
                if counterparty is None and item["client"]:
                    counterparty = counterparty_by_name.get(item["client"].strip().lower())
                if counterparty is not None:
                    item["client_id"] = counterparty.id
                    item["client_href"] = f"/clients?client={counterparty.id}#client-card"
                else:
                    item["client_id"] = ""
                    item["client_href"] = ""
                record_data = _json_object(row.data)
                linked_sale = (
                    sale_by_id.get(str(record_data.get("related_sale_id") or ""))
                    or sale_by_key.get(str(item.get("counterparty_id") or ""))
                    or sale_by_key.get(str(item["client"] or "").strip().lower())
                )
                item["order_number"] = str(linked_sale.get("number") or "") if linked_sale else ""
                item["order_amount"] = str(linked_sale.get("amount") or "") if linked_sale else ""
                item["order_currency"] = str(linked_sale.get("currency") or item["currency"] or "UZS") if linked_sale else item["currency"]
                linked_doc_type = str(linked_sale.get("doc_type") or "sale") if linked_sale else ""
                item["related_document_type"] = linked_doc_type
                item["related_document_label"] = {"order": "Заказ", "return": "Возврат"}.get(linked_doc_type, "Продажа") if linked_sale else ""
                item["kanban_amount"] = item["order_amount"] if linked_sale else item["amount"]
                item["kanban_amount_value"] = linked_sale.get("amount_value", Decimal("0")) if linked_sale else item["amount_value"]
                item["chat_label"] = item["chat_ref"] or chat_for_counterparty(counterparty)
                messenger_channels: list[dict[str, str]] = []
                seen_channels: set[str] = set()

                def add_messenger_channel(channel: str, contact: str = "", topic: str = "", status_label: str = "") -> None:
                    channel_clean = str(channel or "").strip()
                    if not channel_clean:
                        return
                    channel_key = channel_clean.casefold()
                    if channel_key in seen_channels:
                        return
                    seen_channels.add(channel_key)
                    channel_slug = {
                        "telegram": "telegram",
                        "instagram": "instagram",
                        "whatsapp": "whatsapp",
                        "facebook": "facebook",
                        "сайт": "site",
                        "site": "site",
                    }.get(channel_key, "all")
                    messenger_channels.append(
                        {
                            "channel": channel_clean,
                            "contact": str(contact or item["chat_label"] or "").strip(),
                            "topic": str(topic or "").strip(),
                            "status_label": str(status_label or "").strip(),
                            "href": f"/messengers?channel={quote(channel_slug)}#messages",
                        }
                    )

                for thread in messenger_threads_by_client.get(str(item["client"] or "").strip().casefold(), []):
                    add_messenger_channel(
                        str(thread.get("channel") or ""),
                        str(thread.get("contact") or thread.get("username") or thread.get("phone") or ""),
                        str(thread.get("topic") or ""),
                        str(thread.get("status_label") or ""),
                    )
                inferred_channel_text = " ".join([item["lead_source"], item["contact_type"], item["chat_label"]]).casefold()
                inferred_channels = (
                    ("Telegram", "telegram"),
                    ("Instagram", "instagram"),
                    ("WhatsApp", "whatsapp"),
                    ("Facebook", "facebook"),
                    ("Сайт", "сайт"),
                )
                for channel_label, channel_token in inferred_channels:
                    if channel_token in inferred_channel_text:
                        add_messenger_channel(channel_label, item["chat_label"])
                item["messenger_channels"] = messenger_channels
                item["action_state"] = _crm_due_state(item["due_date"], item["status"], today_iso)
                item["activity_state"] = _crm_activity_state(row, crm_activity_settings)
                item["edit_payload"] = _crm_record_edit_payload(item)
                hay = " ".join(
                    [
                        item["title"],
                        item["client"],
                        item["responsible"],
                        item["status_label"],
                        item["stage"],
                        item["lead_source"],
                        item["service_type"],
                        item["priority_label"],
                        item["next_step"],
                        item["note"],
                    ]
                ).lower()
                if q_clean and q_clean not in hay:
                    continue
                if filters["client"] and item["client"] != filters["client"]:
                    continue
                if filters["responsible"] and item["responsible"] != filters["responsible"]:
                    continue
                if filters["status"] != "all" and item["status"] != filters["status"]:
                    continue
                if filters["priority"] != "all" and item["priority"] != filters["priority"]:
                    continue
                if item["item_type"] == "history":
                    history_date = str(item.get("date") or item.get("due_date") or "")
                    if filters["date_from"] and history_date and history_date < filters["date_from"]:
                        continue
                    if filters["date_to"] and history_date and history_date > filters["date_to"]:
                        continue
                if item["status"] == "archived":
                    crm_archive_records.append(item)
                    continue
                crm_records.append(item)
                if item["item_type"] == "history":
                    crm_history_records.append(item)
                    continue
                if item["status"] not in {"done", "won", "lost"}:
                    crm_summary["active_count"] += 1
                    if item["item_type"] == "deal":
                        crm_summary["open_deals"] += 1
                    if item["action_state"]["id"] == "overdue":
                        crm_summary["overdue_count"] += 1
                    elif item["action_state"]["id"] == "today":
                        crm_summary["today_count"] += 1
                    if item["due_date"] or item["next_step"]:
                        crm_next_actions.append(item)
                    if item["item_type"] in {"deal", "task"} and (not item["due_date"] or not item["next_step"]):
                        crm_summary["missing_next_step_count"] += 1
                    if isinstance(item["kanban_amount_value"], Decimal):
                        summary_pipeline_total += item["kanban_amount_value"]
                if item["status"] == "won":
                    summary_won += 1
                elif item["status"] == "lost":
                    summary_lost += 1
                column = crm_stage_map.get(item["stage_id"]) or crm_stage_map[crm_stages[0]["id"]]
                column["records"].append(item)
                column["total_value"] += item["kanban_amount_value"] if isinstance(item["kanban_amount_value"], Decimal) else Decimal("0")
            crm_options = {
                "clients": sorted({row.name for row in counterparty_rows if row.name} | {item["client"] for item in crm_records if item["client"]}),
                "responsibles": sorted({item["responsible"] for item in crm_records if item["responsible"]} | ({str(request.session.get("user", {}).get("name") or "")} if request.session.get("user") else set())),
                "deals": deal_options,
            }
        crm_kanban_columns = []
        for stage in crm_stages:
            column = crm_stage_map[stage["id"]]
            column["total"] = _sales_money_label(column["total_value"])
            crm_kanban_columns.append(column)
        closed_count = summary_won + summary_lost
        crm_summary["pipeline_total"] = _sales_money_label(summary_pipeline_total)
        crm_summary["won_count"] = summary_won
        crm_summary["conversion_rate"] = f"{round(summary_won * 100 / closed_count)}%" if closed_count else "0%"
        action_rank = {"overdue": 0, "today": 1, "soon": 2, "planned": 3, "": 4}
        crm_next_actions = sorted(
            crm_next_actions,
            key=lambda item: (
                action_rank.get(str(item.get("action_state", {}).get("id") or ""), 5),
                str(item.get("due_date") or "9999-12-31"),
                str(item.get("title") or ""),
            ),
        )[:6]
        return tpl(
            request,
            "home_business_module.html",
            variant="user",
            active="crm",
            module=_business_module_context("crm"),
            crm_filters=filters,
            crm_options=crm_options,
            crm_records=crm_records,
            crm_history_records=crm_history_records,
            crm_archive_records=crm_archive_records,
            crm_summary=crm_summary,
            crm_next_actions=crm_next_actions,
            crm_pipeline_stages=crm_stages,
            crm_pipeline_stage_lines=_crm_stage_lines(crm_stages),
            crm_kanban_columns=crm_kanban_columns,
            crm_lead_sources=list(CRM_LEAD_SOURCES),
            crm_service_types=list(CRM_SERVICE_TYPES),
            crm_priorities=CRM_PRIORITY_LABELS,
            today=today_iso,
            flash_ok=request.query_params.get("msg"),
            flash_err=_module_flash_error(request),
        )

    TELEPHONY_CALL_STATUS = {
        "answered": "Отвечен",
        "missed": "Пропущен",
        "planned": "Запланирован",
        "blocked": "Заблокирован",
    }
    TELEPHONY_NUMBER_STATUS = {
        "active": "Активен",
        "paused": "Пауза",
        "offline": "Отключен",
    }
    TELEPHONY_PROVIDER_STATUS = {
        "online": "Онлайн",
        "offline": "Оффлайн",
        "testing": "Проверяется",
        "draft": "Черновик",
    }

    def _telephony_settings_payload(workspace_owner_id: str) -> dict[str, Any]:
        data = load_workspace_settings(workspace_owner_id)
        for key in TELEPHONY_SETTING_LIST_KEYS:
            if not isinstance(data.get(key), list):
                data[key] = []
        return data

    def _telephony_user_name(request: Request) -> str:
        user = request.session.get("user") or {}
        return str(user.get("name") or user.get("login") or "").strip()

    def _telephony_status_label(kind: str, value: str) -> str:
        clean = str(value or "").strip()
        if kind == "call":
            return TELEPHONY_CALL_STATUS.get(clean, clean or "-")
        if kind == "number":
            return TELEPHONY_NUMBER_STATUS.get(clean, clean or "-")
        return TELEPHONY_PROVIDER_STATUS.get(clean, clean or "-")

    def _telephony_direction_label(value: str) -> str:
        return {"incoming": "Входящий", "outgoing": "Исходящий"}.get(str(value or "").strip(), value or "-")

    def _telephony_recording_public_url(value: Any) -> str:
        raw = str(value or "").strip()
        if raw.startswith("/static/uploads/telephony/"):
            return raw
        parsed = urlparse(raw)
        if parsed.scheme in {"http", "https"} and parsed.netloc:
            return raw
        return ""

    def _telephony_rows_with_labels(rows: list[dict[str, Any]], kind: str) -> list[dict[str, Any]]:
        out: list[dict[str, Any]] = []
        for row in rows:
            item = dict(row)
            item["status_label"] = _telephony_status_label(kind, str(item.get("status") or ""))
            if kind == "call":
                item["direction_label"] = _telephony_direction_label(str(item.get("direction") or ""))
                item["recording_url"] = _telephony_recording_public_url(
                    item.get("recording_url") or item.get("recording_path")
                )
                item["has_recording"] = bool(item["recording_url"] or item.get("recording_path"))
            out.append(item)
        return out

    def _telephony_filter_rows(
        rows: list[dict[str, Any]],
        filters: dict[str, str],
        *,
        kind: str,
    ) -> list[dict[str, Any]]:
        q = filters.get("q", "").casefold()
        provider = filters.get("provider", "").casefold()
        responsible = filters.get("responsible", "").casefold()
        status = filters.get("status", "all")
        result: list[dict[str, Any]] = []
        for row in rows:
            haystack = " ".join(str(row.get(key) or "") for key in row.keys()).casefold()
            if q and q not in haystack:
                continue
            if provider and provider not in haystack:
                continue
            if responsible and responsible not in str(row.get("responsible") or "").casefold():
                continue
            if status and status != "all" and str(row.get("status") or "") != status:
                continue
            result.append(row)
        if kind == "call":
            result.sort(key=lambda item: str(item.get("started_at") or item.get("created_at") or ""), reverse=True)
        return result

    def _telephony_client_options(workspace_owner_id: str) -> list[str]:
        with session_scope() as session:
            rows = session.execute(
                select(Counterparty)
                .where(
                    Counterparty.workspace_owner_id == workspace_owner_id,
                    Counterparty.kind.in_(["client", "both"]),
                )
                .order_by(Counterparty.name.asc())
            ).scalars()
            return [row.name for row in rows if row.name]

    def _telephony_save_item(workspace_owner_id: str, key: str, item: dict[str, Any]) -> None:
        data = _telephony_settings_payload(workspace_owner_id)
        items = data.get(key)
        if not isinstance(items, list):
            items = []
        if not item.get("id"):
            item["id"] = str(uuid.uuid4())
        item["created_at"] = item.get("created_at") or datetime.now(timezone.utc).isoformat()
        items.insert(0, item)
        data[key] = items[:500]
        save_workspace_settings(workspace_owner_id, data)

    def _telephony_webhook_token(data: dict[str, Any]) -> tuple[str, bool]:
        token = str(data.get("telephony_webhook_token") or "").strip()
        if token:
            return token, False
        token = secrets.token_urlsafe(24)
        data["telephony_webhook_token"] = token
        return token, True

    def _telephony_public_base_url(request: Request) -> str:
        request_base = str(request.base_url).strip().rstrip("/")
        configured = str(get_settings().auth_url or "").strip().rstrip("/")
        if request_base and "up.railway.app" not in request_base and "localhost" not in request_base and "127.0.0.1" not in request_base:
            return request_base
        if configured and not configured.startswith(("http://", "https://")):
            configured = "https://" + configured
        if configured and "your-app.up.railway.app" not in configured and "up.railway.app" not in configured:
            return configured
        return request_base

    def _telephony_integration_context(request: Request, workspace_owner_id: str, data: dict[str, Any]) -> dict[str, Any]:
        token, changed = _telephony_webhook_token(data)
        if changed:
            save_workspace_settings(workspace_owner_id, data)
        base_url = _telephony_public_base_url(request)
        endpoint_url = f"{base_url}/api/telephony/calls/ingest"
        contacts_sync_url = f"{base_url}/api/telephony/contacts/sync"
        recording_upload_url = f"{base_url}/api/telephony/recordings/upload"
        webhook_url = f"{endpoint_url}?workspace_owner_id={quote(workspace_owner_id)}&token={quote(token)}"
        return {
            "workspace_owner_id": workspace_owner_id,
            "endpoint_url": endpoint_url,
            "contacts_sync_url": contacts_sync_url,
            "recording_upload_url": recording_upload_url,
            "webhook_url": webhook_url,
            "token": token,
            "token_header": "X-UPOS-SIP-Token",
        }

    def _telephony_sip_accounts_for_softphone(data: dict[str, Any]) -> list[dict[str, Any]]:
        providers = data.get("telephony_providers") if isinstance(data.get("telephony_providers"), list) else []
        accounts: list[dict[str, Any]] = []
        for provider in providers:
            if not isinstance(provider, dict):
                continue
            host, port, endpoint = _telephony_parse_sip_host(
                provider.get("host") or provider.get("endpoint"),
                provider.get("port") or "5060",
            )
            login = str(provider.get("login") or provider.get("account") or provider.get("extension") or "").strip()
            password = str(provider.get("password") or "").strip()
            if not host or not login:
                continue
            server = host if port == 5060 else endpoint
            name = str(provider.get("name") or provider.get("label") or "").strip()
            accounts.append(
                {
                    "id": str(provider.get("id") or login),
                    "provider_id": str(provider.get("id") or ""),
                    "server": server,
                    "host": host,
                    "port": str(port),
                    "extension": login,
                    "auth_id": str(provider.get("auth_id") or login).strip(),
                    "password": password,
                    "display_name": str(provider.get("display_name") or provider.get("employee") or name or login).strip(),
                    "transport": str(provider.get("transport") or "UDP").strip().upper() or "UDP",
                    "secure_media": str(provider.get("secure_media") or "Disabled").strip() or "Disabled",
                    "expire_time": int(str(provider.get("expire_time") or "300").strip() or "300"),
                    "label": name or f"Account {login}",
                    "codec": str(provider.get("codec") or "G711").strip(),
                    "status": str(provider.get("status") or "").strip(),
                }
            )
        return accounts

    def _telephony_first(payload: dict[str, Any], *keys: str) -> Any:
        for key in keys:
            value = payload.get(key)
            if value not in (None, ""):
                return value
        return ""

    def _telephony_iso_time(value: Any, fallback: str = "") -> str:
        raw = str(value or "").strip()
        if not raw:
            return fallback
        if raw.isdigit():
            try:
                return datetime.fromtimestamp(int(raw), tz=timezone.utc).isoformat()
            except (OverflowError, OSError, ValueError):
                return fallback or raw
        return raw

    def _telephony_normalize_direction(payload: dict[str, Any]) -> str:
        raw = str(_telephony_first(payload, "direction", "dir", "call_direction", "type", "inout")).strip().lower()
        if raw in {"in", "inbound", "incoming", "input", "входящий"}:
            return "incoming"
        if raw in {"out", "outbound", "outgoing", "output", "исходящий"}:
            return "outgoing"
        event = str(payload.get("event") or "").strip().lower()
        if "incoming" in event or "inbound" in event:
            return "incoming"
        if "outgoing" in event or "outbound" in event:
            return "outgoing"
        return "outgoing"

    def _telephony_normalize_status(payload: dict[str, Any]) -> str:
        raw = str(_telephony_first(payload, "status", "call_status", "disposition", "result", "event")).strip().lower()
        compact = raw.replace("-", "_").replace(" ", "_")
        if compact in {"answered", "answer", "connected", "complete", "completed", "ended", "hangup", "success"}:
            return "answered"
        if compact in {"missed", "no_answer", "noanswer", "busy", "failed", "fail", "cancelled", "canceled", "rejected", "abandoned"}:
            return "missed"
        if compact in {"blocked", "blacklist", "forbidden"}:
            return "blocked"
        duration = str(_telephony_first(payload, "duration", "duration_sec", "billsec", "seconds")).strip()
        if duration and duration not in {"0", "0.0"}:
            return "answered"
        if compact in {"ring", "ringing", "start", "started", "new", "created"}:
            return "planned"
        return "planned"

    def _telephony_request_token(payload: dict[str, Any], request: Request) -> str:
        return str(
            payload.get("token")
            or payload.get("auth_token")
            or payload.get("secret")
            or request.headers.get("X-UPOS-SIP-Token")
            or request.headers.get("X-Telephony-Token")
            or ""
        ).strip()

    def _telephony_token_allowed(workspace_owner_id: str, provided_token: str) -> bool:
        expected_global = str(os.getenv("UPOS_SIP_TOKEN") or "").strip()
        if expected_global and provided_token == expected_global:
            return True
        data = _telephony_settings_payload(workspace_owner_id)
        expected_workspace = str(data.get("telephony_webhook_token") or "").strip()
        if expected_workspace and provided_token == expected_workspace:
            return True
        return not expected_global and not expected_workspace and provided_token == "upos-local-sip"

    def _telephony_default_workspace_owner_id() -> str:
        with session_scope() as session:
            row = session.execute(
                select(User)
                .where(User.role != "admin", User.employer_user_id.is_(None))
                .order_by(User.created_at.asc())
            ).scalars().first()
            return str(row.id if row else "").strip()

    def _telephony_normalized_phone(value: Any) -> str:
        return re.sub(r"\D+", "", str(value or ""))

    def _telephony_phone_matches(left: Any, right: Any) -> bool:
        left_digits = _telephony_normalized_phone(left)
        right_digits = _telephony_normalized_phone(right)
        if not left_digits or not right_digits:
            return False
        if len(left_digits) >= 9 and len(right_digits) >= 9:
            return left_digits[-9:] == right_digits[-9:]
        return left_digits == right_digits

    def _telephony_upsert_web_client(
        workspace_owner_id: str,
        payload: dict[str, Any],
        *,
        phone: str = "",
        name: str = "",
    ) -> dict[str, Any] | None:
        phone_value: Any = phone or _telephony_first(
            payload,
            "phone",
            "mobile",
            "phone_number",
            "telephone",
            "remote_ext",
            "caller",
            "caller_id",
            "from",
        )
        if not phone_value:
            phone_list = payload.get("phones") or payload.get("phone_numbers") or payload.get("phoneNumbers") or payload.get("numbers")
            if isinstance(phone_list, list) and phone_list:
                first_phone = phone_list[0]
                if isinstance(first_phone, dict):
                    phone_value = _telephony_first(first_phone, "number", "phone", "value", "normalized_number")
                else:
                    phone_value = first_phone
        if isinstance(phone_value, dict):
            phone_value = _telephony_first(phone_value, "number", "phone", "value", "normalized_number")
        raw_phone = str(
            phone_value
            or ""
        ).strip()
        clean_phone = _telephony_normalized_phone(raw_phone)
        if not clean_phone:
            return None
        clean_name = str(
            name
            or _telephony_first(payload, "client", "client_name", "display_name", "name", "contact_name", "full_name")
            or ""
        ).strip()
        generated_name = f"Контакт +{clean_phone}"
        display_name = clean_name or generated_name
        external_id = f"phone:{clean_phone}"
        now_iso = datetime.now(timezone.utc).isoformat()
        created = False
        with session_scope() as session:
            row = session.execute(
                select(Counterparty).where(
                    Counterparty.workspace_owner_id == workspace_owner_id,
                    Counterparty.external_source == "upos-sip",
                    Counterparty.external_id == external_id,
                )
            ).scalars().first()
            if row is None:
                candidates = session.execute(
                    select(Counterparty)
                    .where(Counterparty.workspace_owner_id == workspace_owner_id)
                    .order_by(Counterparty.updated_at.desc())
                ).scalars().all()
                row = next(
                    (
                        candidate
                        for candidate in candidates
                        if _telephony_phone_matches(
                            candidate.phone or _counterparty_extra(candidate).get("phone"),
                            clean_phone,
                        )
                    ),
                    None,
                )
            if row is None:
                created = True
                row = Counterparty(
                    id=str(uuid.uuid4()),
                    workspace_owner_id=workspace_owner_id,
                    kind="client",
                    name=display_name,
                    phone=raw_phone[:64],
                    tax_id="",
                    external_source="upos-sip",
                    external_id=external_id,
                    data={},
                )
                session.add(row)
            else:
                has_client, has_supplier = _counterparty_role_flags(row.kind)
                row.kind = _counterparty_kind_from_flags(True, has_supplier)
                current_name = str(row.name or "").strip()
                if clean_name and (
                    not current_name
                    or current_name == str(row.phone or "").strip()
                    or current_name.startswith("Контакт +")
                ):
                    row.name = clean_name
                row.phone = raw_phone[:64] or row.phone
            extra = _counterparty_extra(row)
            extra.update(
                {
                    "phone": raw_phone,
                    "is_client": True,
                    "source": "telephony",
                    "telephony_synced_at": now_iso,
                    "email": str(_telephony_first(payload, "email", "mail") or extra.get("email") or "").strip(),
                    "telegram": str(_telephony_first(payload, "telegram", "telegram_username") or extra.get("telegram") or "").strip(),
                    "address": str(_telephony_first(payload, "address", "location") or extra.get("address") or "").strip(),
                    "note": str(_telephony_first(payload, "note", "comment", "description") or extra.get("note") or "").strip(),
                }
            )
            row.data = extra
            flag_modified(row, "data")
            session.flush()
            return {
                "id": row.id,
                "name": str(row.name or display_name).strip(),
                "phone": str(row.phone or raw_phone).strip(),
                "created": created,
            }

    def _telephony_contact_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
        raw_contacts: Any = payload.get("contacts") or payload.get("items") or payload.get("clients")
        if isinstance(raw_contacts, str):
            try:
                raw_contacts = json.loads(raw_contacts)
            except (TypeError, ValueError):
                raw_contacts = []
        if isinstance(raw_contacts, dict):
            mapped_contacts: list[dict[str, Any]] = []
            for key, value in raw_contacts.items():
                if isinstance(value, dict):
                    item = dict(value)
                    item.setdefault("phone", str(key))
                else:
                    item = {"phone": str(key), "name": str(value or "")}
                mapped_contacts.append(item)
            raw_contacts = mapped_contacts
        if isinstance(raw_contacts, list):
            return [dict(item) for item in raw_contacts if isinstance(item, dict)]
        return [payload]

    def _telephony_sync_contacts(workspace_owner_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        synced: list[dict[str, Any]] = []
        skipped = 0
        for contact in _telephony_contact_rows(payload):
            item = _telephony_upsert_web_client(workspace_owner_id, contact)
            if item is None:
                skipped += 1
            else:
                synced.append(item)
        return {
            "synced": len(synced),
            "created": sum(1 for item in synced if item.get("created")),
            "updated": sum(1 for item in synced if not item.get("created")),
            "skipped": skipped,
            "clients": synced,
        }

    def _telephony_client_name_by_phone(workspace_owner_id: str, phone: str) -> str:
        clean_phone = _telephony_normalized_phone(phone)
        if not clean_phone:
            return ""
        with session_scope() as session:
            clients = session.execute(
                select(Counterparty)
                .where(Counterparty.workspace_owner_id == workspace_owner_id)
                .order_by(Counterparty.updated_at.desc())
            ).scalars().all()
            for client in clients:
                extra = _counterparty_extra(client)
                client_phone = _telephony_normalized_phone(client.phone or extra.get("phone") or "")
                if _telephony_phone_matches(clean_phone, client_phone):
                    return str(client.name or "").strip()
        return ""

    def _telephony_upsert_device_entry(data: dict[str, Any], payload: dict[str, Any], now_iso: str) -> dict[str, Any] | None:
        raw_device_id = str(
            payload.get("device_id")
            or payload.get("app_instance_id")
            or payload.get("machine")
            or payload.get("machine_name")
            or ""
        ).strip()
        if not raw_device_id:
            return None
        devices = data.get("telephony_devices") if isinstance(data.get("telephony_devices"), list) else []
        device_id = f"upos-sip:{raw_device_id}"
        item = {
            "id": device_id,
            "device_id": raw_device_id,
            "name": str(payload.get("device_name") or payload.get("machine") or payload.get("machine_name") or "Upos Sip").strip(),
            "account": str(payload.get("account") or payload.get("sip_account") or "").strip(),
            "status": str(payload.get("device_status") or "online").strip() or "online",
            "last_seen_at": now_iso,
            "app_version": str(payload.get("app_version") or "").strip(),
            "platform": str(payload.get("platform") or "").strip(),
            "operator": str(payload.get("operator") or payload.get("responsible") or "").strip(),
        }
        existing_index = next((idx for idx, row in enumerate(devices) if str(row.get("id") or "") == device_id), None)
        if existing_index is None:
            item["created_at"] = now_iso
            devices.insert(0, item)
        else:
            existing = dict(devices[existing_index])
            existing.update({key: val for key, val in item.items() if val not in ("", None)})
            devices.pop(existing_index)
            devices.insert(0, existing)
            item = existing
        data["telephony_devices"] = devices[:200]
        return item

    def _telephony_upsert_recording_entry(
        data: dict[str, Any],
        call_item: dict[str, Any],
        payload: dict[str, Any],
        now_iso: str,
    ) -> dict[str, Any] | None:
        recording_path = str(_telephony_first(payload, "recording_path", "recording_url", "record_url", "recording") or call_item.get("recording_path") or "").strip()
        if not recording_path:
            return None
        recordings = data.get("telephony_recordings") if isinstance(data.get("telephony_recordings"), list) else []
        recording_id = str(_telephony_first(payload, "recording_id", "record_id") or call_item.get("external_id") or call_item.get("id") or recording_path).strip()
        item = {
            "id": f"upos-sip-recording:{recording_id}",
            "call_id": str(call_item.get("id") or call_item.get("external_id") or "").strip(),
            "client": str(call_item.get("client") or "").strip(),
            "phone": str(call_item.get("phone") or "").strip(),
            "path": recording_path,
            "duration": str(call_item.get("duration") or _telephony_first(payload, "duration_sec", "duration", "billsec") or "").strip(),
            "started_at": str(call_item.get("started_at") or _telephony_first(payload, "started_at", "start_time", "timestamp") or "").strip(),
            "created_at": str(_telephony_first(payload, "recording_created_at", "record_created_at") or now_iso).strip(),
            "device_id": str(_telephony_first(payload, "device_id", "app_instance_id", "machine", "machine_name")).strip(),
            "account": str(_telephony_first(payload, "account", "sip_account", "extension", "exten")).strip(),
            "updated_at": now_iso,
        }
        existing_index = next((idx for idx, row in enumerate(recordings) if str(row.get("id") or "") == item["id"]), None)
        if existing_index is None:
            recordings.insert(0, item)
        else:
            existing = dict(recordings[existing_index])
            existing.update({key: val for key, val in item.items() if val not in ("", None)})
            recordings.pop(existing_index)
            recordings.insert(0, existing)
            item = existing
        data["telephony_recordings"] = recordings[:500]
        return item

    def _telephony_upsert_call_from_payload(workspace_owner_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        data = _telephony_settings_payload(workspace_owner_id)
        calls = data.get("telephony_calls") if isinstance(data.get("telephony_calls"), list) else []
        call_id = str(_telephony_first(payload, "call_id", "id", "uuid", "call_uuid", "callid", "uniqueid", "linkedid")).strip()
        external_id = f"upos-sip:{call_id}" if call_id else str(uuid.uuid4())
        existing_index = next(
            (idx for idx, row in enumerate(calls) if str(row.get("external_id") or row.get("id") or "") == external_id),
            None,
        )
        existing_call = dict(calls[existing_index]) if existing_index is not None else {}
        direction = _telephony_normalize_direction(payload)
        event_name = str(payload.get("event") or "").strip().lower()
        has_direction_signal = any(
            payload.get(key) not in (None, "") for key in ("direction", "dir", "call_direction", "type", "inout")
        ) or any(marker in event_name for marker in ("incoming", "inbound", "outgoing", "outbound"))
        if existing_call and not has_direction_signal:
            direction = str(existing_call.get("direction") or direction)
        raw_phone = str(
            _telephony_first(
                payload,
                "phone",
                "remote_ext",
                "caller",
                "caller_id",
                "callerid",
                "from",
                "src",
                "source",
            )
        ).strip()
        if direction == "outgoing":
            raw_phone = str(_telephony_first(payload, "phone", "remote_ext", "to", "dst", "destination", "callee", "called", "from", "src")).strip()
        raw_phone = raw_phone or str(existing_call.get("phone") or "").strip()
        clean_name = str(_telephony_first(payload, "client", "client_name", "display_name", "name", "contact_name")).strip()
        synced_client = _telephony_upsert_web_client(workspace_owner_id, payload, phone=raw_phone, name=clean_name)
        if synced_client:
            clean_name = str(synced_client.get("name") or clean_name).strip()
        if not clean_name:
            clean_name = str(existing_call.get("client") or "").strip() or _telephony_client_name_by_phone(workspace_owner_id, raw_phone)

        now_iso = datetime.now(timezone.utc).isoformat()
        status = _telephony_normalize_status(payload)
        has_status_signal = any(
            payload.get(key) not in (None, "")
            for key in ("status", "call_status", "disposition", "result", "duration", "duration_sec", "billsec", "seconds")
        ) or event_name in {
            "answered",
            "connected",
            "completed",
            "ended",
            "hangup",
            "missed",
            "no_answer",
            "busy",
            "failed",
            "cancelled",
            "rejected",
        }
        if existing_call and not has_status_signal:
            status = str(existing_call.get("status") or status)
        device = _telephony_upsert_device_entry(data, payload, now_iso)
        item = {
            "id": external_id,
            "external_id": external_id,
            "client": clean_name or raw_phone or "Неизвестный клиент",
            "phone": raw_phone,
            "direction": direction,
            "started_at": _telephony_iso_time(
                _telephony_first(payload, "started_at", "start_time", "call_start", "timestamp", "time", "date", "created_at"),
                str(existing_call.get("started_at") or now_iso),
            ),
            "ended_at": _telephony_iso_time(
                _telephony_first(payload, "ended_at", "end_time", "call_end", "finished_at"),
                str(existing_call.get("ended_at") or ""),
            ),
            "duration": str(
                _telephony_first(payload, "duration", "duration_sec", "billsec", "seconds")
                or existing_call.get("duration")
                or ""
            ).strip(),
            "responsible": str(
                _telephony_first(payload, "responsible", "operator", "employee", "agent", "manager")
                or existing_call.get("responsible")
                or "Upos Sip"
            ).strip(),
            "status": status,
            "note": str(_telephony_first(payload, "note", "comment", "description")).strip(),
            "source": "upos-sip",
            "account": str(_telephony_first(payload, "account", "sip_account", "extension", "exten")).strip(),
            "device_id": str(_telephony_first(payload, "device_id", "app_instance_id", "machine", "machine_name")).strip(),
            "device_name": str((device or {}).get("name") or _telephony_first(payload, "device_name", "machine", "machine_name")).strip(),
            "recording_path": str(
                _telephony_first(payload, "recording_path", "recording_url", "record_url", "recording")
                or existing_call.get("recording_path")
                or ""
            ).strip(),
            "updated_at": now_iso,
        }
        if existing_index is None:
            item["created_at"] = now_iso
            calls.insert(0, item)
        else:
            existing = dict(calls[existing_index])
            existing.update({key: val for key, val in item.items() if val not in ("", None)})
            existing["updated_at"] = now_iso
            calls.pop(existing_index)
            calls.insert(0, existing)
            item = existing

        data["telephony_calls"] = calls[:500]
        _telephony_upsert_recording_entry(data, item, payload, now_iso)
        save_workspace_settings(workspace_owner_id, data)
        item["direction_label"] = _telephony_direction_label(item["direction"])
        item["status_label"] = _telephony_status_label("call", item["status"])
        item["recording_url"] = _telephony_recording_public_url(item.get("recording_path"))
        return item

    def _telephony_parse_sip_host(raw_host: Any, raw_port: Any = "5060") -> tuple[str, int, str]:
        value = str(raw_host or "").strip()
        port_text = str(raw_port or "5060").strip() or "5060"
        try:
            port = int(port_text)
        except ValueError:
            port = 5060

        if value:
            candidate = value
            if "://" not in candidate:
                candidate = "sip://" + candidate
            parsed = urlparse(candidate)
            host = (parsed.hostname or value).strip()
            if parsed.port:
                port = parsed.port
            host = host.strip("[]").strip()
        else:
            host = ""

        if host.casefold().startswith(("http://", "https://", "sip://", "sips://")):
            return _telephony_parse_sip_host(host, port)
        return host, port, f"{host}:{port}" if host else ""

    def _telephony_update_provider_test(workspace_owner_id: str, provider_id: str) -> tuple[str, str]:
        data = _telephony_settings_payload(workspace_owner_id)
        providers = data.get("telephony_providers") if isinstance(data.get("telephony_providers"), list) else []
        target = next((item for item in providers if str(item.get("id")) == str(provider_id)), None)
        if not target:
            return "offline", "Провайдер не найден"
        host, port, endpoint = _telephony_parse_sip_host(target.get("host") or target.get("endpoint"), target.get("port") or "5060")
        if host:
            target["host"] = host
            target["port"] = str(port)
            target["endpoint"] = endpoint
        transport = str(target.get("transport") or "UDP").strip().upper()
        status = "offline"
        message = "Не указан SIP-сервер"
        if host:
            try:
                if transport in {"TCP", "TLS"}:
                    with socket.create_connection((host, port), timeout=3):
                        pass
                    message = f"{transport} соединение с {host}:{port} установлено"
                else:
                    probe = (
                        f"REGISTER sip:{host} SIP/2.0\r\n"
                        f"Via: SIP/2.0/UDP upos.local;branch=z9hG4bK{uuid.uuid4().hex[:8]}\r\n"
                        f"From: <sip:{target.get('login') or 'upos'}@{host}>\r\n"
                        f"To: <sip:{target.get('login') or 'upos'}@{host}>\r\n"
                        f"Call-ID: {uuid.uuid4().hex}@upos\r\n"
                        "CSeq: 1 REGISTER\r\n"
                        "Max-Forwards: 70\r\n"
                        "Content-Length: 0\r\n\r\n"
                    )
                    with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
                        sock.settimeout(3)
                        sock.sendto(probe.encode("utf-8"), (host, port))
                    message = f"Тест REGISTER отправлен на {host}:{port}"
                status = "online"
            except OSError as exc:
                message = str(exc) or "SIP-сервер не отвечает"
        target["status"] = status
        target["last_test_message"] = message
        target["last_test_at"] = datetime.now(timezone.utc).isoformat()
        data["telephony_providers"] = providers
        save_workspace_settings(workspace_owner_id, data)
        return status, message

    def _messenger_settings_payload(workspace_owner_id: str) -> dict[str, Any]:
        data = load_workspace_settings(workspace_owner_id)
        for key in ("messenger_campaigns", "messenger_templates"):
            if not isinstance(data.get(key), list):
                data[key] = []
        if not isinstance(data.get("messenger_auto_reply"), dict):
            data["messenger_auto_reply"] = {
                "enabled": False,
                "greeting": "Здравствуйте, мы получили ваше сообщение и скоро ответим.",
            }
        return data

    def _messenger_status_label(value: str) -> str:
        return {
            "active": "Активный",
            "waiting": "Ожидает",
            "paused": "Пауза",
            "archived": "Архив",
            "draft": "Черновик",
            "scheduled": "Запланирована",
            "sent": "Отправлена",
        }.get(str(value or ""), str(value or "") or "-")

    def _messenger_client_options(workspace_owner_id: str) -> list[str]:
        with session_scope() as session:
            rows = session.execute(
                select(Counterparty)
                .where(Counterparty.workspace_owner_id == workspace_owner_id)
                .order_by(Counterparty.name.asc())
            ).scalars().all()
        names = []
        for row in rows:
            has_client, _ = _counterparty_role_flags(row.kind)
            if has_client and row.name:
                names.append(row.name)
        return names

    def _messenger_match_client(
        clients: list[Counterparty],
        *,
        phone: str = "",
        username: str = "",
        display_name: str = "",
    ) -> Counterparty | None:
        clean_phone = re.sub(r"\D+", "", str(phone or ""))
        clean_username = str(username or "").lstrip("@").lower()
        clean_name = str(display_name or "").strip().lower()
        for client in clients:
            extra = _counterparty_extra(client)
            client_phone = re.sub(r"\D+", "", str(client.phone or extra.get("phone") or ""))
            client_telegram = str(extra.get("telegram") or extra.get("telegram_username") or "").lstrip("@").lower()
            if clean_phone and client_phone and clean_phone.endswith(client_phone[-9:]):
                return client
            if clean_username and client_telegram and clean_username == client_telegram:
                return client
            if clean_name and client.name.strip().lower() == clean_name:
                return client
        return None

    def _messenger_payload_dict(value: Any) -> dict[str, Any]:
        return value if isinstance(value, dict) else {}

    def _messenger_avatar_url(payload: dict[str, Any]) -> str:
        for key in ("avatar_url", "photo_url", "telegram_photo_url", "profile_photo_url", "photo_path"):
            raw = str(payload.get(key) or "").strip()
            if not raw:
                continue
            if raw.startswith(("http://", "https://", "/")):
                return raw
            return "/static/" + raw.lstrip("/")
        return ""

    def _messenger_presence(updated_at: Any, status_value: str = "") -> tuple[str, str]:
        if str(status_value or "") == "waiting":
            return "waiting", "Ожидает"
        if isinstance(updated_at, datetime):
            seen_at = updated_at if updated_at.tzinfo else updated_at.replace(tzinfo=timezone.utc)
            age_seconds = (datetime.now(timezone.utc) - seen_at).total_seconds()
            if age_seconds <= 10 * 60:
                return "online", "В сети"
            if age_seconds <= 24 * 60 * 60:
                return "recent", "Недавно был(а)"
        return "offline", "Не в сети"

    def _messenger_threads_from_sources(workspace_owner_id: str) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        with session_scope() as session:
            clients = session.execute(
                select(Counterparty)
                .where(Counterparty.workspace_owner_id == workspace_owner_id)
                .order_by(Counterparty.name.asc())
            ).scalars().all()
            subscribers = session.execute(
                select(TelegramSubscriber)
                .where(TelegramSubscriber.workspace_owner_id == workspace_owner_id)
                .order_by(TelegramSubscriber.requested_at.desc())
            ).scalars().all()
            chats = session.execute(
                select(TelegramChat)
                .where(TelegramChat.workspace_owner_id == workspace_owner_id)
                .order_by(TelegramChat.last_seen_at.desc())
            ).scalars().all()

        seen_chat_ids: set[str] = set()
        for sub in subscribers:
            title = str(sub.display_name or sub.username or sub.phone or sub.telegram_user_id or "Telegram").strip()
            username = str(sub.username or "").strip()
            client = _messenger_match_client(
                clients,
                phone=str(sub.phone or ""),
                username=username,
                display_name=title,
            )
            status_value = "active" if str(sub.status or "") == "approved" else "waiting"
            payload = _messenger_payload_dict(sub.contact_payload)
            avatar_url = _messenger_avatar_url(payload)
            presence, presence_label = _messenger_presence(sub.decided_at or sub.requested_at, status_value)
            is_new_contact = client is None or status_value == "waiting"
            rows.append(
                {
                    "id": f"telegram-sub-{sub.id}",
                    "channel": "Telegram",
                    "contact": client.name if client else title,
                    "client": client.name if client else "",
                    "username": username,
                    "phone": str(sub.phone or ""),
                    "topic": ("@" + username) if username else str(sub.phone or "Заявка Telegram"),
                    "last_message": "Клиент ожидает ответа" if status_value == "waiting" else "Telegram-чат готов к работе",
                    "responsible": "",
                    "status": status_value,
                    "status_label": _messenger_status_label(status_value),
                    "presence": presence,
                    "presence_label": presence_label,
                    "avatar_url": avatar_url,
                    "avatar_ttl_days": 5,
                    "is_new": is_new_contact,
                    "updated_at": sub.decided_at or sub.requested_at,
                    "messages": [
                        {"author": title, "text": "Контакт Telegram синхронизирован с клиентской базой.", "kind": "in"},
                        {"author": "UPOS", "text": "Можно прикрепить диалог к клиенту CRM и отвечать из центра сообщений.", "kind": "system"},
                    ],
                }
            )
            seen_chat_ids.add(str(sub.chat_id))

        for chat in chats:
            if str(chat.chat_id) in seen_chat_ids:
                continue
            status_value = "active" if chat.is_enabled else "waiting"
            presence, presence_label = _messenger_presence(chat.last_seen_at, status_value)
            rows.append(
                {
                    "id": f"telegram-chat-{chat.id}",
                    "channel": "Telegram",
                    "contact": str(chat.title or chat.chat_id or "Telegram"),
                    "client": "",
                    "topic": str(chat.chat_type or "chat"),
                    "last_message": "Чат найден ботом. Включите его для ручных сообщений.",
                    "responsible": "",
                    "status": status_value,
                    "status_label": _messenger_status_label(status_value),
                    "presence": presence,
                    "presence_label": presence_label,
                    "avatar_url": "",
                    "avatar_ttl_days": 5,
                    "is_new": False,
                    "updated_at": chat.last_seen_at,
                    "messages": [
                        {"author": "Telegram", "text": "Чат доступен после обновления списка чатов.", "kind": "system"},
                    ],
                }
            )

        social_data = load_workspace_settings(workspace_owner_id).get("social_links")
        social_links = social_data if isinstance(social_data, dict) else {}

        def _social_value(*keys: str) -> str:
            for key in keys:
                value = str(social_links.get(key) or "").strip()
                if value:
                    return value
            return ""

        def _social_enabled(*keys: str) -> bool:
            return any(_social_value(key) for key in keys)

        social_thread_specs = [
            {
                "channel": "Instagram",
                "contact": _social_value("instagram_login", "instagram_business_id", "instagram_url") or "Instagram Direct",
                "topic": _social_value("instagram_url", "instagram_login") or "Direct и комментарии",
                "configured": _social_enabled(
                    "instagram_url",
                    "instagram_business_id",
                    "instagram_access_token",
                    "instagram_login",
                    "instagram_password",
                ),
            },
            {
                "channel": "WhatsApp",
                "contact": _social_value("whatsapp_phone", "whatsapp_business_id") or "WhatsApp Business",
                "topic": _social_value("whatsapp_phone") or "Чаты WhatsApp",
                "configured": _social_enabled("whatsapp_phone", "whatsapp_business_id", "whatsapp_access_token"),
                "phone": _social_value("whatsapp_phone"),
            },
            {
                "channel": "Facebook",
                "contact": _social_value("facebook_page_id", "facebook_url") or "Facebook Page",
                "topic": _social_value("facebook_url") or "Messenger страницы",
                "configured": _social_enabled("facebook_url", "facebook_page_id", "facebook_access_token"),
            },
            {
                "channel": "Сайт",
                "contact": _social_value("site_chat_widget_id", "website_url") or "Сайт чат",
                "topic": _social_value("website_url") or "Виджет сайта",
                "configured": _social_enabled("website_url", "site_chat_widget_id", "site_chat_greeting", "site_chat_enabled"),
            },
        ]

        for spec in social_thread_specs:
            channel = str(spec["channel"])
            channel_slug = {
                "Instagram": "instagram",
                "WhatsApp": "whatsapp",
                "Facebook": "facebook",
                "Сайт": "site",
            }.get(channel, _messenger_channel_key(channel).lower())
            configured = bool(spec.get("configured"))
            client = _messenger_match_client(
                clients,
                phone=str(spec.get("phone") or ""),
                username=str(spec.get("contact") or ""),
                display_name=str(spec.get("contact") or ""),
            )
            status_value = "active" if configured else "waiting"
            presence, presence_label = _messenger_presence(datetime.now(timezone.utc), status_value)
            messages = []
            if configured:
                messages = [
                    {
                        "author": str(spec.get("contact") or channel),
                        "text": f"Канал {channel} привязан к UPOS и готов принимать обращения.",
                        "kind": "in",
                    },
                    {
                        "author": "UPOS",
                        "text": "Диалог можно отправить в CRM, связать с клиентом и продолжить переписку из центра сообщений.",
                        "kind": "system",
                    },
                ]
            rows.append(
                {
                    "id": f"{channel_slug}-integration",
                    "channel": channel,
                    "contact": client.name if client else str(spec.get("contact") or channel),
                    "client": client.name if client else "",
                    "topic": str(spec.get("topic") or channel),
                    "last_message": (
                        f"{channel} подключен к центру сообщений"
                        if configured
                        else f"Подключите {channel} в Настройки -> Соцсети"
                    ),
                    "responsible": "",
                    "status": status_value,
                    "status_label": _messenger_status_label(status_value),
                    "presence": presence,
                    "presence_label": presence_label,
                    "avatar_url": "",
                    "avatar_ttl_days": 5,
                    "is_new": False,
                    "updated_at": datetime.now(timezone.utc),
                    "messages": messages,
                }
            )

        rows.sort(key=lambda item: item.get("updated_at") or datetime.min.replace(tzinfo=timezone.utc), reverse=True)
        return rows

    def _messenger_filter_rows(rows: list[dict[str, Any]], filters: dict[str, str]) -> list[dict[str, Any]]:
        q = str(filters.get("q") or "").strip().lower()
        channel = str(filters.get("channel") or "").strip().lower()
        responsible = str(filters.get("responsible") or "").strip().lower()
        status = str(filters.get("status") or "all").strip().lower() or "all"
        result = []
        for row in rows:
            haystack = " ".join(
                str(row.get(key) or "") for key in ("channel", "contact", "client", "topic", "last_message", "title", "name")
            ).lower()
            if q and q not in haystack:
                continue
            if channel and channel not in str(row.get("channel") or "").lower():
                continue
            if responsible and responsible not in str(row.get("responsible") or "").lower():
                continue
            if status != "all" and status != str(row.get("status") or "").lower():
                continue
            result.append(row)
        return result

    def _messenger_channel_key(value: str) -> str:
        channel = str(value or "").strip().lower()
        if "telegram" in channel:
            return "Telegram"
        if "instagram" in channel:
            return "Instagram"
        if "whatsapp" in channel or "whatsap" in channel:
            return "WhatsApp"
        if "facebook" in channel:
            return "Facebook"
        if "site" in channel or "сайт" in channel or "web" in channel:
            return "Сайт"
        return str(value or "").strip()

    def _messenger_save_item(workspace_owner_id: str, key: str, item: dict[str, Any]) -> None:
        data = _messenger_settings_payload(workspace_owner_id)
        rows = data.get(key) if isinstance(data.get(key), list) else []
        now = datetime.now(timezone.utc).isoformat()
        item = dict(item)
        item["id"] = item.get("id") or str(uuid.uuid4())
        item["created_at"] = item.get("created_at") or now
        item["updated_at"] = now
        item["status_label"] = _messenger_status_label(str(item.get("status") or ""))
        data[key] = [item] + [row for row in rows if str(row.get("id") or "") != item["id"]]
        save_workspace_settings(workspace_owner_id, data)

    def _messenger_delete_item(workspace_owner_id: str, key: str, item_id: str) -> None:
        data = _messenger_settings_payload(workspace_owner_id)
        rows = data.get(key) if isinstance(data.get(key), list) else []
        data[key] = [row for row in rows if str(row.get("id") or "") != str(item_id or "")]
        save_workspace_settings(workspace_owner_id, data)

    @app.get("/telephony", response_class=HTMLResponse, name="telephony_get")
    def telephony_get(
        request: Request,
        q: str = "",
        provider: str = "",
        responsible: str = "",
        status: str = "all",
    ):
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        filters = {
            "q": q.strip(),
            "provider": provider.strip(),
            "responsible": responsible.strip(),
            "status": status.strip() or "all",
        }
        user_name = _telephony_user_name(request)
        data = _telephony_settings_payload(wid)
        integration_context = _telephony_integration_context(request, wid, data)
        providers = _telephony_rows_with_labels(list(data.get("telephony_providers") or []), "provider")
        numbers = _telephony_rows_with_labels(list(data.get("telephony_numbers") or []), "number")
        calls = _telephony_rows_with_labels(list(data.get("telephony_calls") or []), "call")
        devices = list(data.get("telephony_devices") or [])
        recordings = []
        for raw_recording in list(data.get("telephony_recordings") or []):
            recording = dict(raw_recording)
            recording["recording_url"] = _telephony_recording_public_url(recording.get("path"))
            recordings.append(recording)
        provider_options = sorted(
            {
                "SIP",
                "АТС",
                "Telegram Call",
                "WhatsApp Call",
                *(str(item.get("name") or "").strip() for item in providers),
            }
            - {""}
        )
        responsible_options = sorted(
            {
                user_name,
                *(str(item.get("responsible") or "").strip() for item in calls + numbers),
                *(str(item.get("employee") or "").strip() for item in numbers),
            }
            - {""}
        )
        client_options = _telephony_client_options(wid)
        return tpl(
            request,
            "home_business_module.html",
            variant="user",
            active="telephony",
            module=_business_module_context("telephony"),
            telephony_filters=filters,
            telephony_options={
                "providers": provider_options,
                "responsibles": responsible_options,
                "clients": client_options,
            },
            telephony_calls=_telephony_filter_rows(calls, filters, kind="call"),
            telephony_numbers=_telephony_filter_rows(numbers, filters, kind="number"),
            telephony_providers=_telephony_filter_rows(providers, filters, kind="provider"),
            telephony_devices=devices,
            telephony_recordings=recordings,
            telephony_dashboard=telephony_account_dashboard(wid),
            telephony_integration=integration_context,
            flash_ok=request.query_params.get("msg"),
            flash_err=_module_flash_error(request),
        )

    @app.post("/telephony/calls/save", name="telephony_call_save")
    async def telephony_call_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/telephony?err=csrf#calls", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        client = str(form.get("client") or "").strip()
        phone = str(form.get("phone") or "").strip()
        if not client or not phone:
            return RedirectResponse(url="/telephony?error=" + quote("Клиент и номер телефона обязательны") + "#calls", status_code=302)
        _telephony_upsert_web_client(wid, {"name": client, "phone": phone}, phone=phone, name=client)
        _telephony_save_item(
            wid,
            "telephony_calls",
            {
                "client": client,
                "phone": phone,
                "direction": str(form.get("direction") or "outgoing").strip(),
                "started_at": str(form.get("started_at") or datetime.now(timezone.utc).isoformat()).strip(),
                "responsible": str(form.get("responsible") or _telephony_user_name(request)).strip(),
                "status": str(form.get("status") or "planned").strip(),
                "note": str(form.get("note") or "").strip(),
            },
        )
        return RedirectResponse(url="/telephony?msg=saved#calls", status_code=302)

    @app.post("/telephony/numbers/save", name="telephony_number_save")
    async def telephony_number_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/telephony?err=csrf#numbers", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        number = str(form.get("number") or "").strip()
        if not number:
            return RedirectResponse(url="/telephony?error=" + quote("Номер телефона обязателен") + "#numbers", status_code=302)
        _telephony_save_item(
            wid,
            "telephony_numbers",
            {
                "number": number,
                "provider": str(form.get("provider") or "SIP").strip(),
                "employee": str(form.get("employee") or "").strip(),
                "responsible": str(form.get("employee") or "").strip(),
                "mode": str(form.get("mode") or "Рабочие часы").strip(),
                "forwarding": str(form.get("forwarding") or "").strip(),
                "status": "active",
            },
        )
        return RedirectResponse(url="/telephony?msg=saved#numbers", status_code=302)

    @app.post("/telephony/providers/save", name="telephony_provider_save")
    async def telephony_provider_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/telephony?err=csrf#integrations", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        name = str(form.get("name") or "").strip()
        raw_host = str(form.get("host") or "").strip()
        host, port, endpoint = _telephony_parse_sip_host(raw_host, form.get("port") or "5060")
        login = str(form.get("login") or "").strip()
        if not name or not host or not login:
            return RedirectResponse(url="/telephony?error=" + quote("Название, SIP-сервер и логин обязательны") + "#integrations", status_code=302)
        _telephony_save_item(
            wid,
            "telephony_providers",
            {
                "name": name,
                "kind": "SIP",
                "host": host,
                "port": str(port),
                "endpoint": endpoint,
                "login": login,
                "account": login,
                "password": str(form.get("password") or "").strip(),
                "codec": str(form.get("codec") or "G711").strip(),
                "transport": str(form.get("transport") or "UDP").strip(),
                "employee": str(form.get("employee") or "").strip(),
                "status": "draft",
                "note": str(form.get("note") or "").strip(),
            },
        )
        return RedirectResponse(url="/telephony?msg=saved#integrations", status_code=302)

    @app.post("/telephony/providers/test", name="telephony_provider_test")
    async def telephony_provider_test(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/telephony?err=csrf#integrations", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        status_value, message = _telephony_update_provider_test(wid, str(form.get("provider_id") or ""))
        key = "saved" if status_value == "online" else "sip_test_failed"
        return RedirectResponse(url=f"/telephony?msg={quote(key)}&note={quote(message)}#integrations", status_code=302)

    @app.api_route("/api/telephony/calls/ingest", methods=["GET", "POST"], name="telephony_calls_ingest_api")
    async def telephony_calls_ingest_api(request: Request):
        payload: dict[str, Any] = {key: value for key, value in request.query_params.items()}
        if request.method == "POST":
            content_type = str(request.headers.get("content-type") or "").lower()
            try:
                if "application/json" in content_type:
                    body = await request.json()
                    if isinstance(body, dict):
                        payload.update(body)
                else:
                    form = await request.form()
                    payload.update({str(key): value for key, value in form.items()})
            except Exception:
                logger.exception("[telephony] failed to parse webhook payload")

        wid = str(_telephony_first(payload, "workspace_owner_id", "workspace_id", "owner_id", "account_id")).strip()
        if not valid_workspace_owner_id(wid):
            wid = _telephony_default_workspace_owner_id()
        if not valid_workspace_owner_id(wid):
            return JSONResponse({"ok": False, "error": "workspace_required"}, status_code=400)

        provided_token = _telephony_request_token(payload, request)
        if not _telephony_token_allowed(wid, provided_token):
            return JSONResponse({"ok": False, "error": "bad_token"}, status_code=401)

        event = str(payload.get("event") or "").strip().lower()
        if event == "test":
            return JSONResponse({"ok": True, "workspace_owner_id": wid, "message": "connected"})
        if event in {"contacts", "contacts_sync", "sync_contacts", "contact_created", "contact_updated"}:
            result = _telephony_sync_contacts(wid, payload)
            return JSONResponse({"ok": True, "workspace_owner_id": wid, **result})
        if event in {"heartbeat", "device_heartbeat", "ping"}:
            data = _telephony_settings_payload(wid)
            device = _telephony_upsert_device_entry(data, payload, datetime.now(timezone.utc).isoformat())
            save_workspace_settings(wid, data)
            contacts_result = _telephony_sync_contacts(wid, payload) if payload.get("contacts") else None
            return JSONResponse({"ok": True, "workspace_owner_id": wid, "device": device, "contacts": contacts_result})

        call_item = _telephony_upsert_call_from_payload(wid, payload)
        return JSONResponse({"ok": True, "workspace_owner_id": wid, "call": call_item})

    @app.post("/api/telephony/contacts/sync", name="telephony_contacts_sync_api")
    async def telephony_contacts_sync_api(request: Request):
        payload: dict[str, Any] = {key: value for key, value in request.query_params.items()}
        try:
            if "application/json" in str(request.headers.get("content-type") or "").lower():
                body = await request.json()
                if isinstance(body, list):
                    payload["contacts"] = body
                elif isinstance(body, dict):
                    payload.update(body)
            else:
                form = await request.form()
                payload.update({str(key): value for key, value in form.items()})
        except Exception:
            return JSONResponse({"ok": False, "error": "invalid_payload"}, status_code=400)
        wid = str(_telephony_first(payload, "workspace_owner_id", "workspace_id", "owner_id", "account_id")).strip()
        if not valid_workspace_owner_id(wid):
            wid = _telephony_default_workspace_owner_id()
        if not valid_workspace_owner_id(wid):
            return JSONResponse({"ok": False, "error": "workspace_required"}, status_code=400)
        if not _telephony_token_allowed(wid, _telephony_request_token(payload, request)):
            return JSONResponse({"ok": False, "error": "bad_token"}, status_code=401)
        result = _telephony_sync_contacts(wid, payload)
        return JSONResponse({"ok": True, "workspace_owner_id": wid, **result})

    @app.post("/api/telephony/recordings/upload", name="telephony_recording_upload_api")
    async def telephony_recording_upload_api(request: Request):
        try:
            form = await request.form()
        except Exception:
            return JSONResponse({"ok": False, "error": "invalid_multipart"}, status_code=400)
        payload = {str(key): value for key, value in form.items() if key not in {"file", "recording_file", "audio"}}
        wid = str(_telephony_first(payload, "workspace_owner_id", "workspace_id", "owner_id", "account_id")).strip()
        if not valid_workspace_owner_id(wid):
            wid = _telephony_default_workspace_owner_id()
        if not valid_workspace_owner_id(wid):
            return JSONResponse({"ok": False, "error": "workspace_required"}, status_code=400)
        if not _telephony_token_allowed(wid, _telephony_request_token(payload, request)):
            return JSONResponse({"ok": False, "error": "bad_token"}, status_code=401)
        if not _telephony_first(payload, "call_id", "id", "uuid", "call_uuid", "callid", "uniqueid", "linkedid") and not _telephony_first(
            payload, "phone", "remote_ext", "caller", "caller_id", "to", "from", "src", "dst"
        ):
            return JSONResponse({"ok": False, "error": "call_reference_required"}, status_code=400)
        upload = form.get("file") or form.get("recording_file") or form.get("audio")
        filename = str(getattr(upload, "filename", "") or "").strip()
        if not upload or not filename or not hasattr(upload, "read"):
            return JSONResponse({"ok": False, "error": "recording_required"}, status_code=400)
        suffix = Path(filename).suffix.lower()
        content_type = str(getattr(upload, "content_type", "") or "").lower()
        suffix_by_type = {
            "audio/mpeg": ".mp3",
            "audio/mp4": ".m4a",
            "audio/ogg": ".ogg",
            "audio/wav": ".wav",
            "audio/x-wav": ".wav",
            "audio/webm": ".webm",
        }
        if suffix not in {".mp3", ".m4a", ".ogg", ".wav", ".webm"}:
            suffix = suffix_by_type.get(content_type, "")
        if not suffix:
            return JSONResponse({"ok": False, "error": "unsupported_recording"}, status_code=415)
        content = await upload.read(50 * 1024 * 1024 + 1)
        if not content:
            return JSONResponse({"ok": False, "error": "empty_recording"}, status_code=400)
        if len(content) > 50 * 1024 * 1024:
            return JSONResponse({"ok": False, "error": "recording_too_large"}, status_code=413)
        workspace_dir = re.sub(r"[^a-zA-Z0-9_-]", "_", wid)
        target_dir = BASE_DIR / "static" / "uploads" / "telephony" / workspace_dir
        target_dir.mkdir(parents=True, exist_ok=True)
        target = target_dir / f"{uuid.uuid4().hex}{suffix}"
        target.write_bytes(content)
        recording_url = f"/static/uploads/telephony/{workspace_dir}/{target.name}"
        payload["recording_url"] = recording_url
        payload["recording_path"] = recording_url
        payload["event"] = str(payload.get("event") or "recording")
        call_item = _telephony_upsert_call_from_payload(wid, payload)
        return JSONResponse(
            {
                "ok": True,
                "workspace_owner_id": wid,
                "recording_url": recording_url,
                "call": call_item,
            }
        )

    @app.get("/api/telephony/config", name="telephony_softphone_config_api")
    def telephony_softphone_config_api(request: Request):
        wid = str(
            request.query_params.get("workspace_owner_id")
            or request.query_params.get("workspace_id")
            or request.headers.get("X-UPOS-Workspace")
            or ""
        ).strip()
        if not valid_workspace_owner_id(wid):
            wid = _telephony_default_workspace_owner_id()
        if not valid_workspace_owner_id(wid):
            return JSONResponse({"ok": False, "error": "workspace_required"}, status_code=400)
        provided_token = _telephony_request_token({key: value for key, value in request.query_params.items()}, request)
        if not _telephony_token_allowed(wid, provided_token):
            return JSONResponse({"ok": False, "error": "bad_token"}, status_code=401)
        data = _telephony_settings_payload(wid)
        integration = _telephony_integration_context(request, wid, data)
        return JSONResponse(
            {
                "ok": True,
                "workspace_owner_id": wid,
                "endpoint_url": integration["endpoint_url"],
                "contacts_sync_url": integration["contacts_sync_url"],
                "recording_upload_url": integration["recording_upload_url"],
                "token": integration["token"],
                "accounts": _telephony_sip_accounts_for_softphone(data),
            }
        )

    @app.post("/api/telephony/softphone/login", name="telephony_softphone_login_api")
    async def telephony_softphone_login_api(request: Request):
        try:
            payload = await request.json()
            if not isinstance(payload, dict):
                payload = {}
        except Exception:
            form = await request.form()
            payload = {str(key): value for key, value in form.items()}

        account_id = str(payload.get("account_id") or payload.get("organization") or payload.get("workspace") or "").strip()
        login_value = str(payload.get("login") or payload.get("username") or "").strip()
        password = str(payload.get("password") or "").strip()
        user = verify_login(login_value, password, account_id)
        if not user:
            return JSONResponse({"ok": False, "error": "bad_credentials"}, status_code=401)

        wid = str(user.get("workspace_owner_id") or user.get("account_owner_id") or user.get("user_id") or "").strip()
        if not valid_workspace_owner_id(wid):
            return JSONResponse({"ok": False, "error": "workspace_required"}, status_code=400)
        data = _telephony_settings_payload(wid)
        integration = _telephony_integration_context(request, wid, data)
        accounts = _telephony_sip_accounts_for_softphone(data)
        return JSONResponse(
            {
                "ok": True,
                "workspace_owner_id": wid,
                "endpoint_url": integration["endpoint_url"],
                "contacts_sync_url": integration["contacts_sync_url"],
                "recording_upload_url": integration["recording_upload_url"],
                "config_url": f"{_telephony_public_base_url(request)}/api/telephony/config",
                "token": integration["token"],
                "token_header": integration["token_header"],
                "user": {
                    "id": str(user.get("user_id") or ""),
                    "login": str(user.get("username") or login_value),
                    "name": str(user.get("name") or login_value),
                    "role": str(user.get("role") or ""),
                },
                "accounts": accounts,
            }
        )

    @app.get("/api/telephony/dashboard", name="telephony_dashboard_api")
    def telephony_dashboard_api(request: Request):
        wid, redir = _product_workspace_owner(request)
        if redir:
            return JSONResponse({"ok": False, "error": "login_required"}, status_code=401)
        assert wid is not None
        return JSONResponse({"ok": True, "dashboard": telephony_account_dashboard(wid)})

    @app.post("/api/telephony/incoming-call", name="telephony_incoming_call_api")
    async def telephony_incoming_call_api(request: Request):
        try:
            payload = await request.json()
        except Exception:
            payload = {}
        user = request.session.get("user") or {}
        wid = str(payload.get("workspace_owner_id") or user.get("workspace_owner_id") or user.get("user_id") or "").strip()
        if not valid_workspace_owner_id(wid):
            return JSONResponse({"ok": False, "error": "workspace_required"}, status_code=401)
        raw_phone = str(payload.get("phone") or payload.get("from") or payload.get("caller") or "").strip()
        clean_phone = re.sub(r"\D+", "", raw_phone)
        client_name = str(payload.get("client") or "").strip()
        synced_client = _telephony_upsert_web_client(wid, payload, phone=raw_phone, name=client_name)
        if synced_client:
            client_name = str(synced_client.get("name") or client_name).strip()
        with session_scope() as session:
            if not client_name and clean_phone:
                clients = session.execute(
                    select(Counterparty)
                    .where(Counterparty.workspace_owner_id == wid)
                    .order_by(Counterparty.updated_at.desc())
                ).scalars().all()
                for client in clients:
                    extra = _counterparty_extra(client)
                    client_phone = re.sub(r"\D+", "", str(client.phone or extra.get("phone") or ""))
                    if client_phone and clean_phone.endswith(client_phone[-9:]):
                        client_name = client.name
                        break
        call_item = {
            "client": client_name or "Неизвестный клиент",
            "phone": raw_phone,
            "direction": "incoming",
            "started_at": str(payload.get("started_at") or datetime.now(timezone.utc).isoformat()[:16]),
            "responsible": str(payload.get("responsible") or "").strip(),
            "status": str(payload.get("status") or "missed").strip() or "missed",
            "note": str(payload.get("note") or "Входящий webhook").strip(),
            "source": "webhook",
        }
        _telephony_save_item(wid, "telephony_calls", call_item)
        call_item["direction_label"] = _telephony_direction_label(call_item["direction"])
        call_item["status_label"] = _telephony_status_label("call", call_item["status"])
        return JSONResponse({"ok": True, "call": call_item, "client": client_name})

    @app.get("/messengers", response_class=HTMLResponse, name="messengers_get")
    def messengers_get(
        request: Request,
        q: str = "",
        channel: str = "",
        responsible: str = "",
        status: str = "all",
    ):
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        filters = {
            "q": q.strip(),
            "channel": channel.strip() or "Telegram",
            "responsible": responsible.strip(),
            "status": status.strip() or "all",
        }
        user = request.session.get("user") or {}
        channel_options = ["Telegram", "Instagram", "WhatsApp", "Facebook", "Сайт"]
        responsible_options = sorted({str(user.get("name") or "").strip()} - {""})
        data = _messenger_settings_payload(wid)
        all_threads = _messenger_threads_from_sources(wid)
        channel_message_counts: dict[str, int] = {}
        for item in all_threads:
            key = _messenger_channel_key(str(item.get("channel") or ""))
            message_count = len(item.get("messages") or [])
            if key:
                channel_message_counts[key] = channel_message_counts.get(key, 0) + message_count
        threads = _messenger_filter_rows(all_threads, filters)
        campaigns = _messenger_filter_rows(
            [dict(row, status_label=_messenger_status_label(str(row.get("status") or ""))) for row in data.get("messenger_campaigns") or []],
            filters,
        )
        templates_rows = _messenger_filter_rows(
            [dict(row, status_label=_messenger_status_label(str(row.get("status") or "active"))) for row in data.get("messenger_templates") or []],
            filters,
        )
        auto_reply = data.get("messenger_auto_reply") if isinstance(data.get("messenger_auto_reply"), dict) else {}
        threads_json = json.dumps(threads, ensure_ascii=False, default=str).replace("</", "<\\/")
        templates_json = json.dumps(templates_rows, ensure_ascii=False, default=str).replace("</", "<\\/")
        return tpl(
            request,
            "home_business_module.html",
            variant="user",
            active="messengers",
            module=_business_module_context("messengers"),
            messenger_filters=filters,
            messenger_options={
                "channels": channel_options,
                "responsibles": responsible_options,
                "clients": _messenger_client_options(wid),
            },
            messenger_channel_message_counts=channel_message_counts,
            messenger_threads=threads,
            messenger_threads_json=threads_json,
            messenger_campaigns=campaigns,
            messenger_templates=templates_rows,
            messenger_templates_json=templates_json,
            messenger_auto_reply=auto_reply,
            flash_ok=request.query_params.get("msg"),
            flash_err=_module_flash_error(request),
        )

    @app.post("/messengers/campaigns/save", name="messenger_campaign_save")
    async def messenger_campaign_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/messengers?err=csrf#campaigns", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        title = str(form.get("title") or "").strip()
        if not title:
            return RedirectResponse(url="/messengers?error=" + quote("Название кампании обязательно") + "#campaigns", status_code=302)
        status_value = str(form.get("action") or "draft").strip()
        status_value = "scheduled" if status_value == "launch" else "draft"
        _messenger_save_item(
            wid,
            "messenger_campaigns",
            {
                "title": title,
                "channel": str(form.get("channel") or "Telegram").strip(),
                "segment": str(form.get("segment") or "").strip(),
                "template": str(form.get("template") or "").strip(),
                "message": str(form.get("message") or "").strip(),
                "run_at": str(form.get("run_at") or "").strip(),
                "status": status_value,
                "note": str(form.get("note") or "").strip(),
            },
        )
        return RedirectResponse(url="/messengers?msg=saved#campaigns", status_code=302)

    @app.post("/messengers/templates/save", name="messenger_template_save")
    async def messenger_template_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/messengers?err=csrf#templates", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        title = str(form.get("title") or "").strip()
        text_value = str(form.get("text") or "").strip()
        if not title or not text_value:
            return RedirectResponse(url="/messengers?error=" + quote("Название и текст шаблона обязательны") + "#templates", status_code=302)
        _messenger_save_item(
            wid,
            "messenger_templates",
            {
                "title": title,
                "channel": str(form.get("channel") or "Telegram").strip(),
                "scenario": str(form.get("scenario") or "reply").strip(),
                "text": text_value,
                "preview": text_value[:120],
                "status": str(form.get("status") or "active").strip() or "active",
            },
        )
        return RedirectResponse(url="/messengers?msg=saved#templates", status_code=302)

    @app.post("/messengers/templates/{template_id}/delete", name="messenger_template_delete")
    async def messenger_template_delete(request: Request, template_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/messengers?err=csrf#templates", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        _messenger_delete_item(wid, "messenger_templates", template_id)
        return RedirectResponse(url="/messengers?msg=deleted#templates", status_code=302)

    @app.post("/messengers/auto-reply/save", name="messenger_auto_reply_save")
    async def messenger_auto_reply_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/messengers?err=csrf#telegram", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        data = _messenger_settings_payload(wid)
        data["messenger_auto_reply"] = {
            "enabled": str(form.get("auto_reply_enabled") or "") == "1",
            "greeting": str(form.get("greeting") or "").strip(),
        }
        save_workspace_settings(wid, data)
        return RedirectResponse(url="/messengers?msg=saved#telegram", status_code=302)

    @app.get("/products/catalog")
    def products_catalog_redirect():
        return RedirectResponse(url="/products#catalog", status_code=302)

    @app.get("/products/pricelists")
    def products_pricelists_redirect():
        return RedirectResponse(url="/products#price-types", status_code=302)

    @app.get("/warehouse/stocks")
    def warehouse_stocks_redirect():
        return RedirectResponse(url="/warehouse#stocks", status_code=302)

    @app.get("/warehouse/movements")
    def warehouse_movements_redirect():
        return RedirectResponse(url="/warehouse#adjustments", status_code=302)

    @app.get("/warehouse/transfers")
    def warehouse_transfers_redirect():
        return RedirectResponse(url="/warehouse#transfers", status_code=302)

    @app.get("/clients/routes")
    def clients_routes_redirect():
        return RedirectResponse(url="/clients#routes", status_code=302)

    @app.get("/crm/history")
    def crm_history_redirect():
        return RedirectResponse(url="/crm#history", status_code=302)

    @app.get("/telephony/integrations")
    def telephony_integrations_redirect():
        return RedirectResponse(url="/telephony#integrations", status_code=302)

    @app.get("/messengers/telegram")
    def messengers_telegram_redirect():
        return RedirectResponse(url="/messengers#telegram", status_code=302)

    @app.post("/crm/save", name="crm_save")
    async def crm_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/crm?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        title = str(form.get("title") or "").strip()
        if not title:
            return RedirectResponse(url="/crm?error=" + quote("Название записи обязательно") + "#tasks", status_code=302)
        data, amount, currency = _crm_record_payload(form)
        stages = _crm_workspace_stages(wid)
        _crm_apply_stage(data, stages)
        status = str(form.get("status") or "new").strip() or "new"
        validation_error = _crm_open_record_error(data, status, stages)
        if validation_error:
            target_hash = {"task": "tasks", "deal": "deals", "history": "history"}.get(str(data.get("item_type") or "task"), "tasks")
            return RedirectResponse(url=f"/crm?error={quote(validation_error)}#{target_hash}", status_code=302)
        _crm_append_activity(data, "Карточка создана", _crm_actor_name(request))
        with session_scope() as session:
            counterparty_id = None
            if data["client"]:
                counterparty = _resolve_counterparty(session, wid, name=data["client"], role="client")
                counterparty_id = counterparty.id if counterparty else None
            row = CrmRecord(
                id=str(uuid.uuid4()),
                workspace_owner_id=wid,
                item_type=str(data.get("item_type") or "task"),
                title=title,
                counterparty_id=counterparty_id,
                status=status,
                due_date=str(data.get("due_date") or ""),
                amount=amount,
                currency=currency,
                data=data,
            )
            session.add(row)
            if counterparty_id:
                _sync_counterparty_crm_status_from_stage(session, wid, row, stages)
        target_hash = {"task": "tasks", "deal": "deals", "history": "history"}.get(str(data.get("item_type") or "task"), "tasks")
        return RedirectResponse(url=f"/crm?msg=saved#{target_hash}", status_code=302)

    @app.post("/crm/{record_id}/update", name="crm_update")
    async def crm_update(request: Request, record_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/crm?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        title = str(form.get("title") or "").strip()
        if not title:
            return RedirectResponse(url="/crm?error=" + quote("Название записи обязательно") + "#tasks", status_code=302)
        data, amount, currency = _crm_record_payload(form)
        stages = _crm_workspace_stages(wid)
        _crm_apply_stage(data, stages)
        status = str(form.get("status") or "new").strip() or "new"
        validation_error = _crm_open_record_error(data, status, stages)
        if validation_error:
            target_hash = {"task": "tasks", "deal": "deals", "history": "history"}.get(str(data.get("item_type") or "task"), "tasks")
            return RedirectResponse(url=f"/crm?error={quote(validation_error)}#{target_hash}", status_code=302)
        with session_scope() as session:
            row = session.get(CrmRecord, record_id)
            if not row or row.workspace_owner_id != wid:
                return RedirectResponse(url="/crm?error=" + quote("CRM-запись не найдена") + "#tasks", status_code=302)
            counterparty_id = None
            if data["client"]:
                counterparty = _resolve_counterparty(session, wid, name=data["client"], role="client")
                counterparty_id = counterparty.id if counterparty else None
            old_data = _json_object(row.data)
            old_stage = str(old_data.get("stage") or "")
            old_log = old_data.get("activity_log") if isinstance(old_data.get("activity_log"), list) else []
            data["activity_log"] = list(old_log)
            change_bits = []
            if row.title != title:
                change_bits.append("название")
            if old_stage != str(data.get("stage") or ""):
                change_bits.append(f"этап: {old_stage or '-'} → {data.get('stage') or '-'}")
            _crm_append_activity(data, "Карточка изменена", _crm_actor_name(request), ", ".join(change_bits))
            row.title = title
            row.item_type = str(data.get("item_type") or "task")
            row.counterparty_id = counterparty_id
            row.status = status
            row.due_date = str(data.get("due_date") or "")
            row.amount = amount
            row.currency = currency
            row.data = data
            flag_modified(row, "data")
            if counterparty_id:
                _sync_counterparty_crm_status_from_stage(session, wid, row, stages)
        target_hash = {"task": "tasks", "deal": "deals", "history": "history"}.get(str(data.get("item_type") or "task"), "tasks")
        return RedirectResponse(url=f"/crm?msg=saved#{target_hash}", status_code=302)

    @app.post("/crm/{record_id}/complete", name="crm_complete")
    async def crm_complete(request: Request, record_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/crm?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        with session_scope() as session:
            row = session.get(CrmRecord, record_id)
            if row and row.workspace_owner_id == wid:
                data = _json_object(row.data).copy()
                stages = _crm_workspace_stages(wid)
                result = str(form.get("result") or "").strip()
                if row.item_type == "deal":
                    row.status = "lost" if result == "lost" else "won"
                    if row.status == "lost":
                        lost_reason = str(form.get("lost_reason") or "").strip()
                        if not lost_reason:
                            return RedirectResponse(url="/crm?error=" + quote("Укажите причину потери сделки") + "#deals", status_code=302)
                        data["lost_reason"] = lost_reason[:300]
                    target_stage = _crm_stage_for_client_status("lost" if row.status == "lost" else "our_client", stages)
                    _crm_apply_stage(data, stages, target_stage["id"])
                else:
                    row.status = "done"
                data["next_step"] = ""
                _crm_append_activity(
                    data,
                    "Сделка потеряна" if row.status == "lost" else ("Сделка выиграна" if row.status == "won" else "Задача завершена"),
                    _crm_actor_name(request),
                    str(data.get("lost_reason") or ""),
                )
                row.data = data
                flag_modified(row, "data")
                if row.item_type == "deal":
                    _sync_counterparty_crm_status_from_stage(session, wid, row, stages)
                target_hash = {"task": "tasks", "deal": "deals", "history": "history"}.get(row.item_type, "tasks")
                return RedirectResponse(url=f"/crm?msg=saved#{target_hash}", status_code=302)
        return RedirectResponse(url="/crm?error=" + quote("CRM-запись не найдена") + "#tasks", status_code=302)

    @app.post("/crm/stages/save", name="crm_stages_save")
    async def crm_stages_save(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/crm?err=csrf#deals", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        raw_lines = str(form.get("stages") or "").splitlines()
        stages = _crm_clean_stages(raw_lines)
        data = load_workspace_settings(wid)
        data["crm_pipeline_stages"] = stages
        save_workspace_settings(wid, data)
        return RedirectResponse(url="/crm?msg=saved#deals", status_code=302)

    @app.post("/crm/{record_id}/stage", name="crm_stage_update")
    async def crm_stage_update(request: Request, record_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return JSONResponse({"ok": False, "error": "csrf"}, status_code=403)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return JSONResponse({"ok": False, "error": "auth"}, status_code=401)
        assert wid is not None
        stage_id = str(form.get("stage_id") or "").strip()
        stages = _crm_workspace_stages(wid)
        stage = _crm_stage_for_value(stage_id, stages)
        lost_reason = str(form.get("lost_reason") or "").strip()
        if _client_crm_status_from_stage(stage) == "lost" and not lost_reason:
            return JSONResponse({"ok": False, "error": "lost_reason_required"}, status_code=400)
        with session_scope() as session:
            row = session.get(CrmRecord, record_id)
            if not row or row.workspace_owner_id != wid:
                return JSONResponse({"ok": False, "error": "not_found"}, status_code=404)
            data = _json_object(row.data).copy()
            previous_stage = str(data.get("stage") or "")
            _crm_apply_stage(data, stages, stage["id"])
            if lost_reason:
                data["lost_reason"] = lost_reason[:300]
            _crm_append_activity(data, "Этап изменён", _crm_actor_name(request), f"{previous_stage or '-'} → {stage['title']}")
            row.data = data
            flag_modified(row, "data")
            if stage["id"] == "won":
                row.status = "won"
            elif stage["id"] == "lost":
                row.status = "lost"
            elif row.status in {"new", "won", "lost"}:
                row.status = "in_progress"
            _sync_counterparty_crm_status_from_stage(session, wid, row, stages)
        return JSONResponse({"ok": True, "stage": stage})

    @app.post("/crm/{record_id}/archive", name="crm_archive")
    async def crm_archive(request: Request, record_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return JSONResponse({"ok": False, "error": "csrf"}, status_code=403)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return JSONResponse({"ok": False, "error": "auth"}, status_code=401)
        assert wid is not None
        with session_scope() as session:
            row = session.get(CrmRecord, record_id)
            if not row or row.workspace_owner_id != wid:
                return JSONResponse({"ok": False, "error": "not_found"}, status_code=404)
            data = _json_object(row.data).copy()
            data["archived_status"] = row.status
            data["archived_at"] = datetime.now(timezone.utc).isoformat()
            _crm_append_activity(data, "Перенесено в архив", _crm_actor_name(request))
            row.status = "archived"
            row.data = data
            flag_modified(row, "data")
        return JSONResponse({"ok": True})

    @app.post("/crm/{record_id}/restore", name="crm_restore")
    async def crm_restore(request: Request, record_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/crm?err=csrf#deals", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        with session_scope() as session:
            row = session.get(CrmRecord, record_id)
            if not row or row.workspace_owner_id != wid:
                return RedirectResponse(url="/crm?error=" + quote("CRM-запись не найдена") + "#deals", status_code=302)
            data = _json_object(row.data).copy()
            restored_status = str(data.pop("archived_status", "in_progress") or "in_progress")
            data.pop("archived_at", None)
            _crm_append_activity(data, "Восстановлено из архива", _crm_actor_name(request))
            row.status = restored_status if restored_status != "archived" else "in_progress"
            row.data = data
            flag_modified(row, "data")
        return RedirectResponse(url="/crm?msg=saved#deals", status_code=302)

    @app.post("/crm/{record_id}/tag", name="crm_tag_add")
    async def crm_tag_add(request: Request, record_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return JSONResponse({"ok": False, "error": "csrf"}, status_code=403)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return JSONResponse({"ok": False, "error": "auth"}, status_code=401)
        assert wid is not None
        tag = re.sub(r"\s+", " ", str(form.get("tag") or "")).strip(" #\t\r\n")[:24]
        if not tag:
            return JSONResponse({"ok": False, "error": "empty"}, status_code=400)
        with session_scope() as session:
            row = session.get(CrmRecord, record_id)
            if not row or row.workspace_owner_id != wid:
                return JSONResponse({"ok": False, "error": "not_found"}, status_code=404)
            data = _json_object(row.data).copy()
            tags = [str(item or "").strip() for item in data.get("tags", []) if str(item or "").strip()] if isinstance(data.get("tags"), list) else []
            if tag.casefold() not in {item.casefold() for item in tags}:
                tags.append(tag)
            data["tags"] = tags[:8]
            _crm_append_activity(data, "Добавлен тег", _crm_actor_name(request), tag)
            row.data = data
            flag_modified(row, "data")
        return JSONResponse({"ok": True, "tags": tags[:8]})

    @app.post("/crm/{record_id}/activity", name="crm_activity_add")
    async def crm_activity_add(request: Request, record_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return JSONResponse({"ok": False, "error": "csrf"}, status_code=403)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return JSONResponse({"ok": False, "error": "auth"}, status_code=401)
        assert wid is not None
        kind = str(form.get("kind") or "comment").strip().lower()
        if kind not in {"chat", "task", "comment"}:
            return JSONResponse({"ok": False, "error": "invalid_kind"}, status_code=400)
        text = re.sub(r"\s+", " ", str(form.get("text") or "")).strip()[:500]
        if not text:
            return JSONResponse({"ok": False, "error": "empty"}, status_code=400)
        due_date = str(form.get("due_date") or "").strip()[:10]
        assignee = re.sub(r"\s+", " ", str(form.get("assignee") or "")).strip()[:120]
        actor = _crm_actor_name(request)
        action = {"chat": "Сообщение", "task": "Задача", "comment": "Комментарий"}[kind]
        event = {
            "at": datetime.now(timezone.utc).isoformat(),
            "action": action,
            "actor": actor,
            "detail": text,
            "kind": kind,
            "due_date": due_date,
            "assignee": assignee,
        }
        with session_scope() as session:
            row = session.get(CrmRecord, record_id)
            if not row or row.workspace_owner_id != wid:
                return JSONResponse({"ok": False, "error": "not_found"}, status_code=404)
            data = _json_object(row.data).copy()
            raw = data.get("activity_log") if isinstance(data.get("activity_log"), list) else []
            events = [item for item in raw if isinstance(item, dict)][-99:]
            events.append(event)
            data["activity_log"] = events
            row.data = data
            flag_modified(row, "data")
        return JSONResponse({"ok": True, "event": event})

    @app.post("/crm/{record_id}/delete", name="crm_delete")
    async def crm_delete(request: Request, record_id: str):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/crm?err=csrf", status_code=302)
        wid, redir = _product_workspace_owner(request)
        if redir:
            return redir
        assert wid is not None
        with session_scope() as session:
            row = session.get(CrmRecord, record_id)
            if row and row.workspace_owner_id == wid:
                target_hash = {"task": "tasks", "deal": "deals", "history": "history"}.get(row.item_type, "tasks")
                session.delete(row)
                return RedirectResponse(url=f"/crm?msg=deleted#{target_hash}", status_code=302)
        return RedirectResponse(url="/crm?msg=deleted#tasks", status_code=302)

    @app.get("/reports", response_class=HTMLResponse)
    def home_reports(request: Request):
        u = request.session.get("user") or {}
        wid = str(u.get("workspace_owner_id") or u.get("user_id") or "").strip()
        loc = resolve_locale(request, u)
        ws = load_workspace_settings(wid) if valid_workspace_owner_id(wid) else {}
        courier_debt_limits = ws.get("delivery_debt_limits") if isinstance(ws.get("delivery_debt_limits"), dict) else {}
        tz_name = normalize_workspace_timezone(str(ws.get("timezone") or ""))
        report_period = _period_from_request(request, tz_name, default="month")
        visible_employee_id = _visible_employee_id(u)
        pnl = get_pnl_data(
            wid,
            report_period["start"],
            report_period["end"],
            visible_employee_id=visible_employee_id,
        )
        balance_sheet = load_treasury(wid, visible_employee_id=visible_employee_id)
        templates_by_id = {
            t["id"]: t for t in localize_treasury_templates(loc, list_templates_public())
        }

        def _amount(raw: object) -> float:
            try:
                return round(float(raw or 0), 2)
            except (TypeError, ValueError):
                return 0.0

        def _format_amount(value: float) -> str:
            value = _amount(value)
            text = f"{value:,.0f}"
            return text.replace(",", " ")

        def _tone(value: float) -> str:
            if value > 0:
                return "positive"
            if value < 0:
                return "negative"
            return "neutral"

        def _sum_by_currency(rows: list[dict]) -> dict[str, float]:
            totals: dict[str, float] = {}
            for row in rows:
                ccy = str(row.get("currency") or "").strip().upper() or "USD"
                totals[ccy] = round(totals.get(ccy, 0.0) + _amount(row.get("amount")), 2)
            return totals

        def _prepare_rows(rows: list[dict]) -> list[dict]:
            max_amount = max((_amount(row.get("amount")) for row in rows), default=0.0)
            prepared: list[dict] = []
            for row in rows:
                value = _amount(row.get("amount"))
                width = 0 if max_amount <= 0 else max(6, round(value / max_amount * 100))
                ccy = str(row.get("currency") or "").strip().upper() or "USD"
                prepared.append(
                    {
                        "name": str(row.get("name") or translate(loc, "reports.uncategorized")),
                        "amount": value,
                        "amount_label": _format_amount(value),
                        "currency": ccy,
                        "count": int(row.get("count") or 0),
                        "bar_width": width,
                    }
                )
            return prepared

        income_rows = _prepare_rows(pnl.get("income") or [])
        expense_rows = _prepare_rows(pnl.get("expense") or [])
        income_by_currency = _sum_by_currency(pnl.get("income") or [])
        expense_by_currency = _sum_by_currency(pnl.get("expense") or [])

        balance_totals: dict[str, float] = {}
        balance_accounts: list[dict] = []
        for pocket in balance_sheet.get("pockets") or []:
            if not isinstance(pocket, dict):
                continue
            entries: list[dict] = []
            for entry in pocket.get("entries") or []:
                if not isinstance(entry, dict):
                    continue
                ccy = str(entry.get("currency") or "").strip().upper()
                if not ccy:
                    continue
                amount = _amount(entry.get("amount"))
                balance_totals[ccy] = round(balance_totals.get(ccy, 0.0) + amount, 2)
                entries.append(
                    {
                        "currency": ccy,
                        "amount": amount,
                        "amount_label": _format_amount(amount),
                    }
                )

            template_id = str(pocket.get("template_id") or "custom")
            template_meta = templates_by_id.get(template_id) or {}
            icon_key = str(pocket.get("icon") or template_meta.get("icon") or "custom")
            balance_accounts.append(
                {
                    "label": str(
                        pocket.get("label")
                        or template_meta.get("title")
                        or translate(loc, "reports.account_default"),
                    ),
                    "kind": str(
                        template_meta.get("title") or translate(loc, "reports.account_kind_default"),
                    ),
                    "note": str(pocket.get("note") or ""),
                    "icon": icon_key,
                    "entries": entries,
                }
            )

        display_currency = str(balance_sheet.get("display_currency") or "USD").strip().upper() or "USD"
        pnl_currencies = set(income_by_currency) | set(expense_by_currency)
        if display_currency in pnl_currencies:
            primary_currency = display_currency
        elif pnl_currencies:
            primary_currency = sorted(pnl_currencies)[0]
        elif display_currency in balance_totals:
            primary_currency = display_currency
        else:
            primary_currency = sorted(balance_totals)[0] if balance_totals else display_currency

        all_pnl_currencies = sorted(pnl_currencies)
        if primary_currency in all_pnl_currencies:
            all_pnl_currencies.remove(primary_currency)
            all_pnl_currencies.insert(0, primary_currency)

        currency_summary = []
        for ccy in all_pnl_currencies:
            income_value = income_by_currency.get(ccy, 0.0)
            expense_value = expense_by_currency.get(ccy, 0.0)
            net_value = round(income_value - expense_value, 2)
            currency_summary.append(
                {
                    "currency": ccy,
                    "income": _format_amount(income_value),
                    "expense": _format_amount(expense_value),
                    "net": _format_amount(net_value),
                    "tone": _tone(net_value),
                }
            )

        balance_summary = [
            {
                "currency": ccy,
                "amount": _format_amount(amount),
            }
            for ccy, amount in sorted(
                balance_totals.items(),
                key=lambda item: (item[0] != primary_currency, item[0]),
            )
        ]

        income_total = income_by_currency.get(primary_currency, 0.0)
        expense_total = expense_by_currency.get(primary_currency, 0.0)
        net_total = round(income_total - expense_total, 2)
        report_data = {
            "period_label": _period_label(loc, report_period),
            "period_preset": report_period["preset"],
            "period_date_from": report_period["date_from"],
            "period_date_to": report_period["date_to"],
            "primary_currency": primary_currency,
            "income_total": _format_amount(income_total),
            "expense_total": _format_amount(expense_total),
            "net_total": _format_amount(net_total),
            "net_tone": _tone(net_total),
            "income_rows": income_rows,
            "expense_rows": expense_rows,
            "income_count": sum(item["count"] for item in income_rows),
            "expense_count": sum(item["count"] for item in expense_rows),
            "currency_summary": currency_summary,
            "balance_summary": balance_summary,
            "balance_accounts": balance_accounts,
            "account_count": len(balance_accounts),
            "balance_currency_count": len(balance_summary),
            "has_pnl": bool(income_rows or expense_rows),
        }
        delivery_debts: list[dict[str, Any]] = []
        delivery_shipments: list[dict[str, Any]] = []
        delivery_payments: list[dict[str, Any]] = []
        delivery_shipment_totals: dict[str, Any] = {"currencies": []}
        business_reports: dict[str, Any] = {
            "sales": {"summary": {}, "top_clients": [], "top_products": [], "days": []},
            "stock": {"summary": {}, "rows": []},
            "stock_analysis": {"summary": {}, "rows": []},
            "receivables": {"summary": {}, "rows": []},
            "calls": {"summary": {}, "rows": []},
        }
        reports_usd_rate = PRODUCT_USD_RATE

        def _report_date(data: dict[str, Any], created_at: Any) -> str:
            raw = str(data.get("date") or data.get("created_at") or "").strip()
            if raw:
                return raw[:10]
            try:
                return created_at.date().isoformat()
            except Exception:
                return ""

        def _report_in_period(date_value: str) -> bool:
            clean = str(date_value or "")[:10]
            date_from = str(report_period.get("date_from") or "")
            date_to = str(report_period.get("date_to") or "")
            if date_from and clean and clean < date_from:
                return False
            if date_to and clean and clean > date_to:
                return False
            return True

        def _report_money(raw: Any, currency: str = "UZS") -> str:
            return f"{_sales_money_label(raw)} {str(currency or 'UZS').upper()}"

        if valid_workspace_owner_id(wid):
            try:
                recompute_delivery_debts(wid)
                delivery_debts = list_courier_debts(wid, include_zero=True)
                delivery_shipments = list_delivery_shipments(wid)
                delivery_payments = [
                    tx
                    for tx in list_transactions(wid, limit=5000)
                    if str(tx.get("type") or "") == "income"
                    and str(tx.get("category") or "") == COURIER_PAYMENT_CATEGORY
                    and bool((tx.get("data") or {}).get("courier_payment") or tx.get("supplier") or tx.get("client"))
                ]
                delivery_shipment_totals = shipment_totals(wid)
            except Exception:
                logger.exception("[upos] delivery debts failed for reports; wid=%s", wid)
            try:
                with session_scope() as session:
                    sales_rows = session.execute(
                        select(SaleDocument)
                        .where(SaleDocument.workspace_owner_id == wid)
                        .order_by(SaleDocument.created_at.desc())
                    ).scalars().all()
                    purchase_rows = session.execute(
                        select(PurchaseDocument)
                        .where(PurchaseDocument.workspace_owner_id == wid)
                        .order_by(PurchaseDocument.created_at.desc())
                    ).scalars().all()
                    operations = session.execute(
                        select(WarehouseOperation)
                        .where(WarehouseOperation.workspace_owner_id == wid)
                        .order_by(WarehouseOperation.created_at.desc())
                    ).scalars().all()
                    products = session.execute(
                        select(Product)
                        .where(Product.workspace_owner_id == wid)
                        .order_by(Product.name.asc())
                    ).scalars().all()
                    crm_rows = session.execute(
                        select(CrmRecord)
                        .where(CrmRecord.workspace_owner_id == wid)
                        .order_by(CrmRecord.created_at.desc())
                    ).scalars().all()

                product_by_id = {str(product.id): product for product in products}
                product_by_name = {str(product.name or "").strip().casefold(): product for product in products if product.name}

                def report_product_unit_cost(line: dict[str, Any]) -> Decimal:
                    product = product_by_id.get(str(line.get("product_id") or ""))
                    if product is None:
                        product = product_by_name.get(str(line.get("product") or "").strip().casefold())
                    if product is None or not _product_uses_stock(product):
                        return Decimal("0")
                    product_data = _json_object(product.data)
                    stock_prices = [
                        _sales_decimal(stock.get("price"))
                        for stock in product_data.get("stocks", [])
                        if isinstance(stock, dict) and _sales_decimal(stock.get("price")) > 0
                    ]
                    if stock_prices:
                        return sum(stock_prices, Decimal("0")) / Decimal(len(stock_prices))
                    return _sales_decimal(
                        product_data.get("purchase_price")
                        or product_data.get("cost")
                        or product_data.get("last_purchase_price")
                    )

                report_rate = _workspace_usd_uzs_rate(wid)
                reports_usd_rate = report_rate

                def report_to_primary(value: Any, currency: str) -> Decimal:
                    return _convert_product_currency(
                        _sales_decimal(value),
                        str(currency or primary_currency).upper(),
                        primary_currency,
                        report_rate,
                    )

                sale_docs: list[SaleDocument] = []
                sales_report_rows: list[dict[str, Any]] = []
                top_clients: dict[str, Decimal] = {}
                top_products: dict[str, Decimal] = {}
                day_totals: dict[str, Decimal] = {}
                receivables: dict[str, dict[str, Any]] = {}
                gross_sales_total = Decimal("0")
                returns_total = Decimal("0")
                orders_total = Decimal("0")
                paid_total = Decimal("0")
                debt_total = Decimal("0")
                cost_total = Decimal("0")
                profit_total = Decimal("0")
                sale_count = 0
                return_count = 0
                order_count = 0
                for row in sales_rows:
                    data = _json_object(row.data)
                    doc_date = _report_date(data, row.created_at)
                    if not _report_in_period(doc_date):
                        continue
                    doc_type = str(data.get("doc_type") or "sale")
                    amount = _sales_decimal(row.amount)
                    currency = str(row.currency or data.get("currency") or "UZS").upper()
                    amount_primary = report_to_primary(amount, currency)
                    client = str(data.get("client") or "Без клиента").strip() or "Без клиента"
                    paid_amount = min(amount, _sales_decimal(data.get("paid_amount")))
                    debt_amount = max(Decimal("0"), amount - paid_amount)
                    lines = data.get("lines") if isinstance(data.get("lines"), list) else []
                    document_cost_uzs = sum(
                        (_sales_decimal(line.get("quantity")) * report_product_unit_cost(line) for line in lines if isinstance(line, dict)),
                        Decimal("0"),
                    )
                    document_cost = report_to_primary(document_cost_uzs, "UZS")
                    sign = Decimal("-1") if doc_type == "return" else Decimal("1")
                    if doc_type == "sale":
                        sale_count += 1
                        sale_docs.append(row)
                        gross_sales_total += amount_primary
                        paid_total += report_to_primary(paid_amount, currency)
                        debt_total += report_to_primary(debt_amount, currency)
                        cost_total += document_cost
                        profit_total += amount_primary - document_cost
                        top_clients[client] = top_clients.get(client, Decimal("0")) + amount_primary
                        day_totals[doc_date or "-"] = day_totals.get(doc_date or "-", Decimal("0")) + amount_primary
                    elif doc_type == "return":
                        return_count += 1
                        returns_total += amount_primary
                        cost_total -= document_cost
                        profit_total -= amount_primary - document_cost
                        top_clients[client] = top_clients.get(client, Decimal("0")) - amount_primary
                        day_totals[doc_date or "-"] = day_totals.get(doc_date or "-", Decimal("0")) - amount_primary
                    else:
                        order_count += 1
                        orders_total += amount_primary
                    for line in data.get("lines") if isinstance(data.get("lines"), list) else []:
                        product_name = str(line.get("product") or "Без товара").strip() or "Без товара"
                        if doc_type in {"sale", "return"}:
                            top_products[product_name] = top_products.get(product_name, Decimal("0")) + sign * _sales_decimal(line.get("quantity"))
                    if doc_type == "sale" and debt_amount > 0:
                        rec = receivables.setdefault(
                            client,
                            {
                                "client": client,
                                "debt": Decimal("0"),
                                "currency": currency,
                                "last_payment": str(data.get("paid_at") or data.get("payment_date") or ""),
                                "last_sale_date": doc_date,
                            },
                        )
                        rec["debt"] += debt_amount
                        rec["last_sale_date"] = max(str(rec.get("last_sale_date") or ""), doc_date)
                    view = _sales_document_data(row)
                    sales_report_rows.append(
                        {
                            "date": doc_date or "-",
                            "number": row.number,
                            "type": _sales_doc_type_label(doc_type),
                            "type_key": doc_type,
                            "client": client,
                            "warehouse": str(data.get("warehouse") or "Основной склад"),
                            "amount": _report_money(amount, currency),
                            "paid": _report_money(paid_amount, currency) if doc_type == "sale" else "-",
                            "debt": _report_money(debt_amount, currency) if doc_type == "sale" else "-",
                            "cost": _report_money(document_cost, primary_currency) if doc_type in {"sale", "return"} else "-",
                            "profit": _report_money(sign * (amount_primary - document_cost), primary_currency) if doc_type in {"sale", "return"} else "-",
                            "status": view.get("status_label") or "-",
                            "source_sale_number": str(data.get("source_sale_number") or ""),
                        }
                    )

                net_sales_total = gross_sales_total - returns_total
                max_day = max((float(value) for value in day_totals.values()), default=0.0)
                business_reports["sales"] = {
                    "summary": {
                        "period": report_data["period_label"],
                        "amount": _report_money(net_sales_total, primary_currency),
                        "gross_sales": _report_money(gross_sales_total, primary_currency),
                        "returns": _report_money(returns_total, primary_currency),
                        "orders": _report_money(orders_total, primary_currency),
                        "paid": _report_money(paid_total, primary_currency),
                        "debt": _report_money(debt_total, primary_currency),
                        "cost": _report_money(cost_total, primary_currency),
                        "profit": _report_money(profit_total, primary_currency),
                        "count": sale_count,
                        "return_count": return_count,
                        "order_count": order_count,
                    },
                    "top_clients": [
                        {"name": name, "amount": _report_money(amount, primary_currency)}
                        for name, amount in sorted(top_clients.items(), key=lambda item: item[1], reverse=True)[:10]
                        if amount > 0
                    ],
                    "top_products": [
                        {"name": name, "quantity": _sales_money_label(quantity)}
                        for name, quantity in sorted(top_products.items(), key=lambda item: item[1], reverse=True)[:10]
                        if quantity > 0
                    ],
                    "days": [
                        {
                            "date": date,
                            "amount": _report_money(amount, primary_currency),
                            "bar_width": 0 if max_day <= 0 else max(4, round(float(amount) / max_day * 100)),
                        }
                        for date, amount in sorted(day_totals.items())
                    ],
                    "rows": sales_report_rows[:200],
                }

                stock_movements: dict[str, dict[str, Any]] = {}
                for product in products:
                    data = _json_object(product.data)
                    current_qty = Decimal("0")
                    for item in data.get("stocks") if isinstance(data.get("stocks"), list) else []:
                        current_qty += _sales_decimal(item.get("quantity"))
                    cost = _sales_decimal(data.get("purchase_price") or data.get("cost") or data.get("last_purchase_price"))
                    stock_movements[product.name] = {
                        "product": product.name,
                        "incoming": Decimal("0"),
                        "outgoing": Decimal("0"),
                        "begin": current_qty,
                        "end": current_qty,
                        "cost": cost,
                        "currency": str(data.get("currency") or "UZS").upper(),
                    }
                for op in operations:
                    data = _json_object(op.data)
                    op_date = _report_date(data, op.created_at)
                    if not _report_in_period(op_date):
                        continue
                    product_name = str(data.get("product") or data.get("product_name") or "")
                    if not product_name and op.product_id:
                        matched = next((product.name for product in products if product.id == op.product_id), "")
                        product_name = matched
                    product_name = product_name or "Без товара"
                    row = stock_movements.setdefault(
                        product_name,
                        {
                            "product": product_name,
                            "incoming": Decimal("0"),
                            "outgoing": Decimal("0"),
                            "begin": Decimal("0"),
                            "end": Decimal("0"),
                            "cost": _sales_decimal(op.amount),
                            "currency": str(op.currency or "UZS").upper(),
                        },
                    )
                    quantity = _sales_decimal(op.quantity)
                    kind = str(op.operation_type or data.get("operation_type") or "").lower()
                    if kind in {"incoming", "in", "receipt", "purchase"}:
                        row["incoming"] += quantity
                        row["begin"] -= quantity
                    elif kind in {"outgoing", "out", "writeoff", "sale"}:
                        row["outgoing"] += quantity
                        row["begin"] += quantity
                    if op.amount:
                        row["cost"] = _sales_decimal(op.amount)
                stock_rows = []
                stock_total = Decimal("0")
                stock_analysis_rows = []
                stock_analysis_cost_total = Decimal("0")
                stock_analysis_revenue_total = Decimal("0")
                stock_analysis_qty_total = Decimal("0")
                product_id_by_name = {
                    str(product.name or "").strip().casefold(): str(product.id)
                    for product in products
                    if str(product.name or "").strip()
                }
                purchase_history_by_product: dict[str, list[dict[str, Any]]] = {}
                for purchase in purchase_rows:
                    purchase_data = _json_object(purchase.data)
                    purchase_date = _report_date(purchase_data, purchase.created_at)
                    purchase_currency = str(purchase.currency or purchase_data.get("currency") or "UZS").upper()
                    document_products: dict[str, dict[str, Any]] = {}
                    for raw_line in purchase_data.get("lines") or []:
                        if not isinstance(raw_line, dict):
                            continue
                        product_id = str(raw_line.get("product_id") or "").strip()
                        if not product_id:
                            product_id = product_id_by_name.get(
                                str(raw_line.get("product") or raw_line.get("product_name") or "").strip().casefold(),
                                "",
                            )
                        if not product_id:
                            continue
                        quantity = _sales_decimal(raw_line.get("quantity"))
                        price = _sales_decimal(raw_line.get("price"))
                        document_product = document_products.setdefault(
                            product_id,
                            {
                                "quantity": Decimal("0"),
                                "amount": Decimal("0"),
                                "unit": str(raw_line.get("unit") or "Штука"),
                                "warehouses": set(),
                            },
                        )
                        document_product["quantity"] += quantity
                        document_product["amount"] += quantity * price
                        document_product["warehouses"].add(
                            str(raw_line.get("warehouse") or purchase_data.get("warehouse") or "Основной склад")
                        )
                    for product_id, document_product in document_products.items():
                        quantity = _sales_decimal(document_product.get("quantity"))
                        amount = _sales_decimal(document_product.get("amount"))
                        average_price = amount / quantity if quantity else Decimal("0")
                        purchase_history_by_product.setdefault(product_id, []).append(
                            {
                                "purchase_id": str(purchase.id),
                                "number": str(purchase.number or "-"),
                                "date": purchase_date or "-",
                                "supplier": str(purchase_data.get("supplier") or "Без поставщика"),
                                "warehouse": ", ".join(sorted(document_product.get("warehouses") or {"Основной склад"})),
                                "quantity": _sales_money_label(quantity),
                                "unit": str(document_product.get("unit") or "Штука"),
                                "price": _report_money(average_price, purchase_currency),
                                "amount": _report_money(amount, purchase_currency),
                                "url": f"/warehouse?purchase_id={purchase.id}#purchases",
                            }
                        )
                for row in stock_movements.values():
                    total = _sales_decimal(row.get("end")) * _sales_decimal(row.get("cost"))
                    stock_total += total
                    stock_rows.append(
                        {
                            "product": row["product"],
                            "begin": _sales_money_label(row.get("begin")),
                            "incoming": _sales_money_label(row.get("incoming")),
                            "outgoing": _sales_money_label(row.get("outgoing")),
                            "end": _sales_money_label(row.get("end")),
                            "cost": _report_money(row.get("cost"), row.get("currency")),
                            "total": _report_money(total, row.get("currency")),
                        }
                    )
                for product in products:
                    data = _json_object(product.data)
                    item = _product_data(product)
                    qty = _sales_decimal(item.get("quantity"))
                    stocks = item.get("stocks") if isinstance(item.get("stocks"), list) else []
                    negative_stocks = [
                        stock
                        for stock in stocks
                        if isinstance(stock, dict) and _sales_decimal(stock.get("quantity")) < 0
                    ]
                    if qty == 0 and not negative_stocks:
                        continue
                    stock_warehouses = ", ".join(
                        sorted(
                            {
                                str(stock.get("warehouse") or "Основной склад")
                                for stock in stocks
                                if isinstance(stock, dict)
                            }
                        )
                    ) or "Основной склад"
                    if negative_stocks:
                        negative_details = ", ".join(
                            f"{str(stock.get('warehouse') or 'Основной склад')} ({_sales_money_label(stock.get('quantity'))})"
                            for stock in negative_stocks
                        )
                        stock_warehouses = f"{stock_warehouses}. Отрицательный остаток: {negative_details}"
                    cost = Decimal("0")
                    cost_count = Decimal("0")
                    for stock in stocks:
                        if not isinstance(stock, dict):
                            continue
                        price = _sales_decimal(stock.get("price"))
                        if price:
                            cost += price
                            cost_count += Decimal("1")
                    avg_cost = cost / cost_count if cost_count else _sales_decimal(data.get("purchase_price") or data.get("cost") or data.get("last_purchase_price"))
                    sale_price = _sales_decimal(item.get("sale_price"))
                    cost_sum = qty * avg_cost
                    revenue_sum = _convert_product_currency(
                        qty * sale_price,
                        item.get("sale_currency") or "UZS",
                        "UZS",
                        report_rate,
                    )
                    stock_analysis_qty_total += qty
                    stock_analysis_cost_total += cost_sum
                    stock_analysis_revenue_total += revenue_sum
                    stock_analysis_rows.append(
                        {
                            "product": item["name"],
                            "product_id": str(item["id"]),
                            "photo_url": item["photo_url"],
                            "quantity": f"{_sales_money_label(qty)} {item['unit']}",
                            "quantity_raw": _decimal_plain_text(qty),
                            "warehouse": stock_warehouses,
                            "avg_cost": _report_money(avg_cost, "UZS"),
                            "avg_cost_uzs": _decimal_plain_text(avg_cost),
                            "cost_sum": _report_money(cost_sum, "UZS"),
                            "cost_sum_uzs": _decimal_plain_text(cost_sum),
                            "expected_revenue": _report_money(revenue_sum, "UZS"),
                            "expected_revenue_uzs": _decimal_plain_text(revenue_sum),
                            "category": item["category"],
                            "group": item["group"],
                            "brand": item["brand"],
                            "is_negative": bool(negative_stocks),
                            "is_critical": bool(_sales_decimal(item.get("min_stock")) > 0 and qty <= _sales_decimal(item.get("min_stock"))),
                        }
                    )
                stock_analysis_rows.sort(key=lambda item: item["product"].lower())
                business_reports["stock"] = {
                    "summary": {"items": len(stock_rows), "total": _report_money(stock_total, primary_currency)},
                    "rows": stock_rows[:50],
                }
                business_reports["stock_analysis"] = {
                    "summary": {
                        "items": len(stock_analysis_rows),
                        "quantity": f"{_sales_money_label(stock_analysis_qty_total)} шт",
                        "cost_total": _report_money(stock_analysis_cost_total, "UZS"),
                        "cost_total_uzs": _decimal_plain_text(stock_analysis_cost_total),
                        "revenue_total": _report_money(stock_analysis_revenue_total, "UZS"),
                        "revenue_total_uzs": _decimal_plain_text(stock_analysis_revenue_total),
                        "negative_count": sum(1 for item in stock_analysis_rows if item.get("is_negative")),
                        "critical_count": sum(1 for item in stock_analysis_rows if item.get("is_critical")),
                    },
                    "rows": stock_analysis_rows[:200],
                    "history": purchase_history_by_product,
                }

                today = datetime.now(timezone.utc).date()
                receivable_rows = []
                receivable_total = Decimal("0")
                for row in receivables.values():
                    debt = _sales_decimal(row.get("debt"))
                    receivable_total += debt
                    last_date = str(row.get("last_sale_date") or "")
                    try:
                        overdue_days = max(0, (today - datetime.fromisoformat(last_date).date()).days)
                    except Exception:
                        overdue_days = 0
                    receivable_rows.append(
                        {
                            "client": row["client"],
                            "debt": _report_money(debt, row.get("currency")),
                            "last_payment": row.get("last_payment") or "-",
                            "overdue_days": overdue_days,
                        }
                    )
                receivable_rows.sort(key=lambda item: item["overdue_days"], reverse=True)
                business_reports["receivables"] = {
                    "summary": {"clients": len(receivable_rows), "total": _report_money(receivable_total, primary_currency)},
                    "rows": receivable_rows[:50],
                }

                settings_payload = _telephony_settings_payload(wid)
                calls = settings_payload.get("telephony_calls") if isinstance(settings_payload.get("telephony_calls"), list) else []
                period_calls = [call for call in calls if _report_in_period(str(call.get("started_at") or call.get("created_at") or "")[:10])]
                answered = [call for call in period_calls if str(call.get("status") or "") == "answered"]
                missed = [call for call in period_calls if str(call.get("status") or "") == "missed"]
                call_deals = [
                    row
                    for row in crm_rows
                    if "звон" in str(_json_object(row.data).get("lead_source") or _json_object(row.data).get("contact_type") or "").lower()
                ]
                conversion = 0 if not period_calls else round(len(call_deals) / len(period_calls) * 100)
                business_reports["calls"] = {
                    "summary": {
                        "total": len(period_calls),
                        "answered": len(answered),
                        "missed": len(missed),
                        "conversion": conversion,
                    },
                    "rows": [
                        {
                            "client": str(call.get("client") or "-"),
                            "phone": str(call.get("phone") or "-"),
                            "status": _telephony_status_label("call", str(call.get("status") or "")),
                            "responsible": str(call.get("responsible") or "-"),
                            "date": str(call.get("started_at") or call.get("created_at") or "")[:16],
                        }
                        for call in period_calls[:20]
                    ],
                }
            except Exception:
                logger.exception("[upos] business reports failed; wid=%s", wid)
        return tpl(
            request,
            "home_reports.html",
            variant="user",
            active="home_reports",
            pnl=pnl,
            balance_sheet=balance_sheet,
            reports=report_data,
            delivery_debts=delivery_debts,
            delivery_shipments=delivery_shipments,
            delivery_payments=delivery_payments,
            delivery_shipment_totals=delivery_shipment_totals,
            courier_debt_limits=courier_debt_limits,
            business_reports=business_reports,
            reports_usd_rate=_decimal_plain_text(reports_usd_rate),
        )

    @app.get("/adjustments", response_class=HTMLResponse, name="home_adjustments")
    def home_adjustments(request: Request):
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        raw = json.dumps(_adjustments_payload(oid), ensure_ascii=False).replace("</", "<\\/")
        return tpl(
            request,
            "home_adjustments.html",
            variant="user",
            active="home_adjustments",
            inside_organization=True,
            selected_organization_id=oid,
            adjustments_bootstrap_json=raw,
        )


    def _adjustments_path(workspace_owner_id: str) -> Path:
        workspace_dir = Path(ensure_client_workspace(workspace_owner_id))
        return workspace_dir / "adjustments_log.json"

    def _load_adjustments_log(workspace_owner_id: str) -> list[dict[str, Any]]:
        path = _adjustments_path(workspace_owner_id)
        if not path.exists():
            return []
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return []
        return data if isinstance(data, list) else []

    def _save_adjustments_log(workspace_owner_id: str, rows: list[dict[str, Any]]) -> None:
        path = _adjustments_path(workspace_owner_id)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(rows[-500:], ensure_ascii=False, indent=2), encoding="utf-8")

    def _adjustment_money(raw: Any) -> Decimal:
        text_value = str(raw or "0").replace(" ", "").replace("\u202f", "").replace(",", ".")
        try:
            return Decimal(text_value)
        except Exception:
            return Decimal("0")

    def _adjustment_amount_label(raw: Any) -> str:
        amount = _adjustment_money(raw).quantize(Decimal("0.01"))
        if amount == amount.to_integral_value():
            return f"{int(amount):,}".replace(",", " ")
        return f"{amount:,.2f}".replace(",", " ").replace(".", ",").rstrip("0").rstrip(",")

    def _append_adjustment_log(
        workspace_owner_id: str,
        *,
        section: str,
        target_id: str,
        target_name: str,
        currency: str,
        old_amount: Any,
        new_amount: Any,
        note: str,
        actor: dict[str, Any] | None,
    ) -> dict[str, Any]:
        old_dec = _adjustment_money(old_amount)
        new_dec = _adjustment_money(new_amount)
        row = {
            "id": str(uuid.uuid4()),
            "section": str(section or "").strip(),
            "target_id": str(target_id or "").strip(),
            "target_name": str(target_name or "").strip() or "—",
            "currency": str(currency or "UZS").strip().upper() or "UZS",
            "old_amount": float(old_dec),
            "new_amount": float(new_dec),
            "delta": float(new_dec - old_dec),
            "old_amount_label": _adjustment_amount_label(old_dec),
            "new_amount_label": _adjustment_amount_label(new_dec),
            "delta_label": _adjustment_amount_label(new_dec - old_dec),
            "note": str(note or "").strip(),
            "actor_name": str((actor or {}).get("name") or (actor or {}).get("username") or "").strip(),
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        rows = _load_adjustments_log(workspace_owner_id)
        rows.append(row)
        _save_adjustments_log(workspace_owner_id, rows)
        return row

    def _adjustments_payload(workspace_owner_id: str) -> dict[str, Any]:
        treasury = load_treasury(workspace_owner_id)
        accounts: list[dict[str, Any]] = []
        for pocket in treasury.get("pockets") or []:
            if not isinstance(pocket, dict):
                continue
            for entry in pocket.get("entries") or []:
                if not isinstance(entry, dict):
                    continue
                currency = str(entry.get("currency") or "UZS").strip().upper() or "UZS"
                amount = _adjustment_money(entry.get("amount"))
                accounts.append(
                    {
                        "id": str(pocket.get("id") or ""),
                        "name": str(pocket.get("label") or pocket.get("name") or "Счёт"),
                        "currency": currency,
                        "amount": float(amount),
                        "amount_label": _adjustment_amount_label(amount),
                    }
                )

        employees = _hr_employees_with_adjustments(workspace_owner_id, datetime.now().strftime("%Y-%m-%d"))
        employee_rows = [
            {
                "id": str(emp.get("id") or ""),
                "name": str(emp.get("full_name") or ""),
                "amount": float(_adjustment_money(emp.get("salary_due"))),
                "amount_label": _adjustment_amount_label(emp.get("salary_due")),
            }
            for emp in employees
            if str(emp.get("status") or "active") == "active"
        ]

        latest_manual: dict[tuple[str, str, str], Decimal] = {}
        for row in _load_adjustments_log(workspace_owner_id):
            section = str(row.get("section") or "")
            if section not in {"couriers", "suppliers"}:
                continue
            key = (
                section,
                str(row.get("target_id") or row.get("target_name") or ""),
                str(row.get("currency") or "UZS").upper(),
            )
            latest_manual[key] = _adjustment_money(row.get("new_amount"))

        try:
            recompute_delivery_debts(workspace_owner_id)
        except Exception:
            logger.exception("[upos] recompute_delivery_debts failed before adjustments payload")
        courier_rows = []
        for row in list_courier_debts(workspace_owner_id, include_zero=True):
            name = str(row.get("courier_name") or "Без имени")
            currency = str(row.get("currency") or "UZS").upper()
            amount = latest_manual.get(("couriers", name, currency), _adjustment_money(row.get("debt") or row.get("debt_amount") or 0))
            courier_rows.append(
                {
                    "id": name,
                    "name": name,
                    "currency": currency,
                    "amount": float(amount),
                    "amount_label": _adjustment_amount_label(amount),
                }
            )

        supplier_totals: dict[tuple[str, str], Decimal] = {}
        for tx in list_transactions(workspace_owner_id, limit=5000):
            supplier = str(tx.get("supplier") or "").strip()
            if not supplier:
                continue
            currency = str(tx.get("currency") or "UZS").strip().upper() or "UZS"
            amount = _adjustment_money(tx.get("amount"))
            sign = Decimal("1") if str(tx.get("type") or "") == "expense" else Decimal("-1")
            supplier_totals[(supplier, currency)] = supplier_totals.get((supplier, currency), Decimal("0")) + amount * sign
        supplier_rows = [
            {
                "id": f"{name}|{currency}",
                "name": name,
                "currency": currency,
                "amount": float(latest_manual.get(("suppliers", f"{name}|{currency}", currency), amount)),
                "amount_label": _adjustment_amount_label(latest_manual.get(("suppliers", f"{name}|{currency}", currency), amount)),
            }
            for (name, currency), amount in sorted(supplier_totals.items())
        ]

        history = list(reversed(_load_adjustments_log(workspace_owner_id)[-100:]))
        return {
            "ok": True,
            "accounts": accounts,
            "employees": employee_rows,
            "couriers": courier_rows,
            "suppliers": supplier_rows,
            "history": history,
        }

    @app.get("/api/adjustments")
    def api_adjustments_get(request: Request):
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        return _adjustments_payload(oid)

    @app.post("/api/adjustments/account")
    async def api_adjustments_account_save(request: Request):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        body = await request.json()
        account_id = str(body.get("account_id") or "").strip()
        currency = str(body.get("currency") or "UZS").strip().upper() or "UZS"
        new_amount = _adjustment_money(body.get("amount"))
        note = str(body.get("note") or "").strip()
        treasury = load_treasury(oid)
        target_name = ""
        old_amount = Decimal("0")
        found = False
        for pocket in treasury.get("pockets") or []:
            if str(pocket.get("id") or "") != account_id:
                continue
            target_name = str(pocket.get("label") or pocket.get("name") or "Счёт")
            entries = pocket.setdefault("entries", [])
            entry = next((item for item in entries if str(item.get("currency") or "").upper() == currency), None)
            if entry is None:
                entry = {"id": str(uuid.uuid4()), "currency": currency, "amount": 0}
                entries.append(entry)
            old_amount = _adjustment_money(entry.get("amount"))
            entry["amount"] = float(new_amount)
            found = True
            break
        if not found:
            return JSONResponse({"error": "account_not_found"}, status_code=404)
        save_treasury(oid, treasury)
        _append_adjustment_log(
            oid,
            section="accounts",
            target_id=account_id,
            target_name=target_name,
            currency=currency,
            old_amount=old_amount,
            new_amount=new_amount,
            note=note,
            actor=request.session.get("user") or {},
        )
        return _adjustments_payload(oid)

    @app.post("/api/adjustments/manual")
    async def api_adjustments_manual_save(request: Request):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        body = await request.json()
        section = str(body.get("section") or "").strip()
        if section not in {"couriers", "suppliers"}:
            return JSONResponse({"error": "invalid_section"}, status_code=400)
        _append_adjustment_log(
            oid,
            section=section,
            target_id=str(body.get("target_id") or body.get("target_name") or ""),
            target_name=str(body.get("target_name") or body.get("target_id") or ""),
            currency=str(body.get("currency") or "UZS").strip().upper() or "UZS",
            old_amount=body.get("old_amount") or 0,
            new_amount=body.get("amount") or 0,
            note=str(body.get("note") or "").strip(),
            actor=request.session.get("user") or {},
        )
        return _adjustments_payload(oid)

    @app.post("/api/adjustments/salary")
    async def api_adjustments_salary_save(request: Request):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        body = await request.json()
        employee_id = str(body.get("employee_id") or "").strip()
        amount = _adjustment_money(body.get("amount"))
        note = str(body.get("note") or "Корректировка зарплаты").strip()
        work_date = str(body.get("work_date") or datetime.now().strftime("%Y-%m-%d"))[:10]
        old_amount = Decimal("0")
        target_name = employee_id
        for emp in _hr_employees_with_adjustments(oid, work_date):
            if str(emp.get("id") or "") == employee_id:
                old_amount = _adjustment_money(emp.get("salary_due"))
                target_name = str(emp.get("full_name") or employee_id)
                break
        delta = amount - old_amount
        adj_type = "bonus" if delta >= 0 else "penalty"
        _save_hr_salary_adjustment(oid, employee_id, work_date, adj_type, str(abs(delta)), note)
        _append_adjustment_log(
            oid,
            section="employees",
            target_id=employee_id,
            target_name=target_name,
            currency="UZS",
            old_amount=old_amount,
            new_amount=amount,
            note=note,
            actor=request.session.get("user") or {},
        )
        return _adjustments_payload(oid)


    def _workspace_owner_or_redirect(request: Request):
        """Только владелец учётки бизнеса (не сотрудник и не админ)."""
        u = request.session.get("user") or {}
        if u.get("role") == "admin":
            return None, RedirectResponse(url="/admin", status_code=302)
        if u.get("org_scope") == "general":
            return None, RedirectResponse(url="/organizations", status_code=302)
        if u.get("is_employee") and not _can_manage_employees(u):
            return None, RedirectResponse(url="/schet", status_code=302)
        oid = u.get("account_owner_id") or u.get("user_id")
        if not oid:
            return None, RedirectResponse(url="/auth", status_code=302)
        return str(oid), None

    def _employees_owner_or_redirect(request: Request):
        u = request.session.get("user") or {}
        if u.get("role") == "admin":
            return None, RedirectResponse(url="/admin", status_code=302)
        if not _can_manage_employees(u):
            return None, RedirectResponse(url="/schet", status_code=302)
        oid = str(u.get("account_owner_id") or u.get("user_id") or "").strip()
        if not oid:
            return None, RedirectResponse(url="/auth", status_code=302)
        return oid, None

    def _employee_active_org_id(owner_id: str, user: dict | None) -> str:
        orgs = list_organizations(owner_id) if valid_workspace_owner_id(owner_id) else []
        active_id = str((user or {}).get("active_org_id") or (user or {}).get("workspace_owner_id") or "").strip()
        if active_id and any(str(org.get("id")) == active_id for org in orgs):
            return active_id
        default_org = next((org for org in orgs if org.get("is_default")), orgs[0] if orgs else None)
        return str(default_org.get("id") if default_org else owner_id)

    def _settings_owner_profile_sheet(request: Request) -> dict[str, str] | None:
        """Форма «Учётная запись»: только владелец рабочей области без привязки к работодателю."""
        u = request.session.get("user") or {}
        if str(u.get("role") or "") != "user":
            return None
        uid = str(u.get("user_id") or "").strip()
        if not uid:
            return None
        uname = str(u.get("username") or "").strip()
        rec = get_by_username(uname) if uname else None
        if not rec:
            return None
        return {
            "username": str(rec.get("username") or ""),
            "email": str(rec.get("email") or ""),
            "name": str(rec.get("name") or rec.get("username") or ""),
        }

    def _ensure_category_defaults_once(workspace_owner_id: str) -> None:
        if not valid_workspace_owner_id(workspace_owner_id):
            return
        data = load_workspace_settings(workspace_owner_id)
        if data.get("finance_categories_seeded_v2"):
            return
        seed_default_categories(workspace_owner_id)
        data["finance_categories_seeded_v2"] = True
        save_workspace_settings(workspace_owner_id, data)

    def _settings_storage_owner_id(user: dict | None) -> str:
        u = user or {}
        if u.get("org_scope") == "general":
            return str(u.get("account_owner_id") or u.get("user_id") or "").strip()
        return str(u.get("workspace_owner_id") or u.get("user_id") or "").strip()

    def _employee_management_context(owner_id: str, active_org_id: str = "") -> dict[str, object]:
        orgs_for_access = list_organizations(owner_id) if valid_workspace_owner_id(owner_id) else []
        default_org = next((org for org in orgs_for_access if org.get("is_default")), orgs_for_access[0] if orgs_for_access else None)
        role_workspace_id = (
            active_org_id.strip()
            if active_org_id.strip() and any(str(org.get("id")) == active_org_id.strip() for org in orgs_for_access)
            else str(default_org.get("id") if default_org else owner_id)
        )
        staff = list_employees_safe(owner_id, "") if valid_workspace_owner_id(owner_id) else []
        orgs_by_id = {str(org.get("id") or ""): org for org in orgs_for_access}
        access_accounts: list[dict[str, object]] = []
        accounts_by_id: dict[str, dict[str, object]] = {}
        for org in orgs_for_access:
            org_id = str(org.get("id") or "").strip()
            org_name = str(org.get("name") or "")
            if not org_id:
                continue
            pockets = []
            for pocket in load_treasury(org_id).get("pockets") or []:
                if not isinstance(pocket, dict):
                    continue
                row = {
                    "id": str(pocket.get("id") or ""),
                    "label": str(pocket.get("label") or ""),
                    "organization_id": org_id,
                    "organization_name": org_name,
                    "access_employees": pocket.get("access_employees") or [],
                    "access_employee_names": pocket.get("access_employee_names") or [],
                }
                if row["id"]:
                    accounts_by_id[str(row["id"])] = row
                    pockets.append(row)
            access_accounts.append({"organization": org, "accounts": pockets})
        for emp in staff:
            emp_org_ids = [str(x) for x in (emp.get("organization_ids") or []) if str(x or "").strip()]
            emp["organizations"] = [orgs_by_id[x] for x in emp_org_ids if x in orgs_by_id]
            emp["wallets"] = [
                accounts_by_id[x]
                for x in [str(a) for a in (emp.get("account_ids") or [])]
                if x in accounts_by_id
            ]
        employee_roles = list_roles_safe(role_workspace_id) if valid_workspace_owner_id(role_workspace_id) else []
        role_by_key = {str(role.get("key") or ""): str(role.get("id") or "") for role in employee_roles}
        employee_presets = [
            {
                "name": "Ген. директор",
                "login": "director",
                "email": "director@upos.local",
                "position": "Ген. директор",
                "role_id": role_by_key.get("general_director", ""),
                "hint": "видит всё и управляет всем",
            },
            {
                "name": "Администратор",
                "login": "admin",
                "email": "admin@upos.local",
                "position": "Администратор",
                "role_id": role_by_key.get("administrator", ""),
                "hint": "счета, сотрудники, категории",
            },
            {
                "name": "HR",
                "login": "hr",
                "email": "hr@upos.local",
                "position": "HR",
                "role_id": role_by_key.get("hr_manager", ""),
                "hint": "сотрудники и доступы",
            },
            {
                "name": "Бухгалтер",
                "login": "accountant",
                "email": "accountant@upos.local",
                "position": "Бухгалтер",
                "role_id": role_by_key.get("accountant", ""),
                "hint": "транзакции и отчёты",
            },
            {
                "name": "Кассир",
                "login": "cashier",
                "email": "cashier@upos.local",
                "position": "Кассир",
                "role_id": role_by_key.get("cashier", ""),
                "hint": "приход и расход в Кассе",
            },
        ]
        return {
            "employees": staff,
            "organizations_for_access": orgs_for_access,
            "access_accounts": access_accounts,
            "employee_roles": employee_roles,
            "employee_presets": employee_presets,
        }

    def _director_owner_or_redirect(request: Request):
        u = request.session.get("user") or {}
        if u.get("role") == "admin":
            return None, RedirectResponse(url="/admin", status_code=302)
        if not (_is_director(u) or _is_employee_general_director(u)):
            return None, RedirectResponse(url="/schet", status_code=302)
        owner_id = str(u.get("account_owner_id") or u.get("user_id") or "").strip()
        if not valid_workspace_owner_id(owner_id):
            return None, RedirectResponse(url="/auth", status_code=302)
        return owner_id, None

    def _refresh_org_session(
        request: Request,
        *,
        org_scope: str | None = None,
        active_org: dict | None = None,
    ) -> dict:
        u = dict(request.session.get("user") or {})
        owner_id = str(u.get("account_owner_id") or u.get("user_id") or "").strip()
        if u.get("is_employee") and not _is_employee_general_director(u):
            orgs = list_employee_organizations_safe(owner_id, str(u.get("user_id") or ""))
        else:
            orgs = list_organizations(owner_id) if valid_workspace_owner_id(owner_id) else []
        if active_org is None:
            active_id = str(u.get("active_org_id") or "").strip()
            active_org = next((org for org in orgs if str(org.get("id")) == active_id), None)
            active_org = active_org or next((org for org in orgs if org.get("is_default")), None)
            active_org = active_org or (orgs[0] if orgs else None)
        if active_org:
            u["workspace_owner_id"] = str(active_org["id"])
            u["active_org_id"] = str(active_org["id"])
            u["active_organization_name"] = str(active_org.get("name") or "")
        u["organizations"] = orgs
        u["can_switch_organizations"] = _is_director(u) or _is_employee_general_director(u) or len(orgs) > 1
        u["organization_switch_general_value"] = ORG_GENERAL_VALUE
        if org_scope:
            u["org_scope"] = org_scope
        request.session["user"] = u
        return u

    def _organization_integration_badges(organization_id: str) -> list[dict[str, object]]:
        try:
            settings = load_workspace_settings(organization_id)
        except Exception:
            logger.exception("[upos] load_workspace_settings failed for org card; oid=%s", organization_id)
            settings = {}
        return integration_badges(settings)

    def _director_amount_label(value: object) -> str:
        try:
            n = round(float(value or 0))
        except (TypeError, ValueError):
            n = 0
        return f"{n:,}".replace(",", " ")

    def _director_money_label(value: object, currency: str) -> str:
        return _director_amount_label(value) + f" {currency}"

    def _director_currency_chips(totals: dict[str, object] | None) -> list[dict[str, str]]:
        priority = {"UZS": 0, "USD": 1, "RUB": 2, "EUR": 3}
        rows: list[tuple[str, float]] = []
        for ccy_raw, value in (totals or {}).items():
            ccy = str(ccy_raw or "").strip().upper()
            if not ccy:
                continue
            try:
                amount = float(value or 0)
            except (TypeError, ValueError):
                amount = 0.0
            if abs(amount) < 0.5:
                continue
            rows.append((ccy, amount))
        rows.sort(key=lambda item: (priority.get(item[0], 99), item[0]))
        return [
            {
                "currency": ccy,
                "amount_label": _director_amount_label(amount),
                "label": _director_money_label(amount, ccy),
            }
            for ccy, amount in rows
        ]

    def _director_currency_options(totals: dict[str, object] | None, selected: str) -> list[dict[str, object]]:
        priority = {"UZS": 0, "USD": 1, "RUB": 2, "EUR": 3}
        selected_clean = str(selected or "USD").strip().upper() or "USD"
        codes = {"UZS", "USD", "RUB", "EUR", selected_clean}
        for ccy_raw in (totals or {}).keys():
            ccy = str(ccy_raw or "").strip().upper()
            if len(ccy) == 3:
                codes.add(ccy)
        ordered = sorted(codes, key=lambda code: (priority.get(code, 99), code))
        return [{"code": code, "selected": code == selected_clean} for code in ordered]

    def _director_sum_rows_by_currency(rows: list[dict[str, object]] | None) -> dict[str, float]:
        out: dict[str, float] = {}
        for row in rows or []:
            ccy = str(row.get("currency") or "").strip().upper() or "USD"
            if len(ccy) != 3:
                continue
            try:
                amount = float(row.get("amount") or 0)
            except (TypeError, ValueError):
                amount = 0.0
            out[ccy] = round(out.get(ccy, 0.0) + amount, 2)
        return out

    def _director_net_map(income: dict[str, float], expense: dict[str, float]) -> dict[str, float]:
        keys = set(income) | set(expense)
        return {ccy: round(income.get(ccy, 0.0) - expense.get(ccy, 0.0), 2) for ccy in keys}

    def _organization_dashboard_cards(
        owner_id: str,
        orgs: list[dict[str, object]],
        active_org_id: str,
        display_currency: str,
        loc: str,
    ) -> tuple[list[dict[str, object]], dict[str, object]]:
        cards: list[dict[str, object]] = []
        connected_integrations = 0
        active_name = translate(loc, "workspace.general")
        treasury_by_org: dict[str, dict[str, object]] = {}
        treasury_totals: dict[str, object] = {}
        treasury_display_total: object | None = None
        treasury_display_currency = str(display_currency or "USD").strip().upper() or "USD"
        pnl_by_org: dict[str, dict[str, object]] = {}
        month_income: dict[str, float] = {}
        month_expense: dict[str, float] = {}
        try:
            treasury = aggregate_director_treasury(owner_id, display_currency=treasury_display_currency)
            treasury_totals = treasury.get("consolidated_totals_by_currency") or {}
            treasury_display_total = treasury.get("approx_total_in_display")
            treasury_display_currency = str(treasury.get("display_currency") or treasury_display_currency).upper()
            treasury_by_org = {
                str(row.get("organization_id") or ""): row
                for row in treasury.get("organizations") or []
                if isinstance(row, dict)
            }
        except Exception:
            logger.exception("[upos] aggregate_director_treasury failed for organizations panel")
        try:
            settings = load_workspace_settings(owner_id)
            tz_name = normalize_workspace_timezone(str(settings.get("timezone") or ""))
            start_dt, end_dt, _now_local = current_month_local_bounds_utc(tz_name)
            pnl = get_director_consolidated_pnl(owner_id, start_dt, end_dt)
            month_income = _director_sum_rows_by_currency(pnl.get("income") or [])
            month_expense = _director_sum_rows_by_currency(pnl.get("expense") or [])
            pnl_by_org = {
                str(row.get("organization_id") or ""): row
                for row in pnl.get("by_organization") or []
                if isinstance(row, dict)
            }
        except Exception:
            logger.exception("[upos] get_director_consolidated_pnl failed for organizations panel")
        for org in orgs:
            card = dict(org)
            org_id = str(card.get("id") or "")
            integrations = _organization_integration_badges(org_id)
            treasury_row = treasury_by_org.get(org_id) or {}
            totals = treasury_row.get("totals_by_currency") if isinstance(treasury_row, dict) else {}
            pnl_row = pnl_by_org.get(org_id) or {}
            org_income = pnl_row.get("income_by_currency") if isinstance(pnl_row, dict) else {}
            org_expense = pnl_row.get("expense_by_currency") if isinstance(pnl_row, dict) else {}
            org_net = pnl_row.get("net_by_currency") if isinstance(pnl_row, dict) else {}
            if not isinstance(org_net, dict):
                org_net = _director_net_map(
                    org_income if isinstance(org_income, dict) else {},
                    org_expense if isinstance(org_expense, dict) else {},
                )
            card["integrations"] = integrations
            card["connected_integrations"] = sum(1 for item in integrations if item.get("active"))
            card["pocket_count"] = int(treasury_row.get("pocket_count") or 0) if isinstance(treasury_row, dict) else 0
            card["balance_chips"] = _director_currency_chips(totals if isinstance(totals, dict) else {})
            card["month_net_chips"] = _director_currency_chips(org_net)
            connected_integrations += sum(1 for item in integrations if item.get("active"))
            if org_id == active_org_id:
                active_name = str(card.get("name") or active_name)
            cards.append(card)
        month_net = _director_net_map(month_income, month_expense)
        return cards, {
            "organization_count": len(cards),
            "connected_integrations": connected_integrations,
            "available_integrations": len(cards) * len(INTEGRATION_PROVIDERS),
            "active_name": active_name,
            "total_balance_chips": _director_currency_chips(treasury_totals if isinstance(treasury_totals, dict) else {}),
            "total_display_label": (
                _director_money_label(treasury_display_total, treasury_display_currency)
                if treasury_display_total is not None
                else translate(loc, "general.no_rate")
            ),
            "total_display_currency": treasury_display_currency,
            "currency_options": _director_currency_options(
                treasury_totals if isinstance(treasury_totals, dict) else {},
                treasury_display_currency,
            ),
            "month_income_chips": _director_currency_chips(month_income),
            "month_expense_chips": _director_currency_chips(month_expense),
            "month_net_chips": _director_currency_chips(month_net),
            "total_pockets": sum(int(card.get("pocket_count") or 0) for card in cards),
        }

    def _localized_organization_store_error(loc: str, message: str) -> str:
        msg = (message or "").strip()
        if "Владелец" in msg:
            return translate(loc, "general.owner_not_found")
        if "таким названием" in msg:
            return translate(loc, "general.org_duplicate")
        if "не найд" in msg:
            return translate(loc, "general.org_not_found")
        return msg

    @app.get("/organizations", response_class=HTMLResponse, name="organizations_get")
    def organizations_get(request: Request):
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        u = _refresh_org_session(request, org_scope="general")
        orgs = list_organizations(owner_id)
        active_org_id = str(u.get("active_org_id") or "")
        display_currency = str(request.query_params.get("display_currency") or "USD").strip().upper()
        if len(display_currency) != 3:
            display_currency = "USD"
        loc = resolve_locale(request, request.session.get("user") or {})
        organization_cards, general_stats = _organization_dashboard_cards(
            owner_id,
            orgs,
            active_org_id,
            display_currency,
            loc,
        )
        return tpl(
            request,
            "home_organizations.html",
            variant="user",
            active="organizations",
            organizations=organization_cards,
            organization_active_id=active_org_id,
            general_stats=general_stats,
            flash_ok=request.query_params.get("msg"),
            flash_err=request.query_params.get("error"),
        )

    def _selected_org_for_general(owner_id: str, request: Request, user: dict | None) -> dict[str, Any] | None:
        orgs = list_organizations(owner_id)
        requested = str(request.query_params.get("organization_id") or "").strip()
        active_id = str((user or {}).get("active_org_id") or "").strip()
        selected_id = requested or active_id
        selected = next((org for org in orgs if str(org.get("id")) == selected_id), None)
        return selected or next((org for org in orgs if org.get("is_default")), None) or (orgs[0] if orgs else None)

    def _shipment_return_url(path: str, organization_id: str = "", **params: str) -> str:
        pairs: list[tuple[str, str]] = []
        if organization_id:
            pairs.append(("organization_id", organization_id))
        for key, value in params.items():
            clean = str(value or "").strip()
            if clean:
                pairs.append((key, clean))
        if not pairs:
            return path
        return path + "?" + "&".join(f"{quote(k)}={quote(v)}" for k, v in pairs)

    def _shipment_amount_label(raw: Any) -> str:
        try:
            value = float(raw or 0)
        except (TypeError, ValueError):
            value = 0.0
        return f"{value:,.0f}".replace(",", " ")

    def _shipment_daily_journal(shipments: list[dict[str, Any]]) -> list[dict[str, Any]]:
        grouped: dict[str, dict[str, Any]] = {}
        for row in shipments:
            day = str(row.get("shipment_date") or "").strip()
            if not day:
                continue
            bucket = grouped.setdefault(
                day,
                {
                    "date": day,
                    "count": 0,
                    "couriers": set(),
                    "currencies": {},
                    "numbers": [],
                    "statuses": set(),
                    "doc_statuses": set(),
                },
            )
            bucket["count"] += 1
            if row.get("number") is not None:
                bucket["numbers"].append(row.get("number"))
            courier = str(row.get("courier_name") or "").strip()
            if courier:
                bucket["couriers"].add(courier)
            ccy = str(row.get("currency") or "UZS").upper()
            totals = bucket["currencies"].setdefault(ccy, {"total": 0.0, "paid": 0.0, "debt": 0.0})
            totals["total"] += float(row.get("total_amount") or 0)
            totals["paid"] += float(row.get("paid_amount") or 0)
            totals["debt"] += float(row.get("debt_amount") or 0)
            status = str(row.get("status") or "").strip().lower()
            if status:
                bucket["statuses"].add(status)
            doc_status = str(row.get("doc_status") or "new").strip().lower()
            bucket["doc_statuses"].add("confirmed" if doc_status == "confirmed" else "new")
        journal: list[dict[str, Any]] = []
        for idx, day in enumerate(sorted(grouped.keys(), reverse=True), start=1):
            bucket = grouped[day]
            currencies = []
            total_amount = 0.0
            paid_amount = 0.0
            debt_amount = 0.0
            for ccy, totals in sorted(bucket["currencies"].items()):
                total_amount += float(totals["total"])
                paid_amount += float(totals["paid"])
                debt_amount += float(totals["debt"])
                currencies.append(
                    {
                        "currency": ccy,
                        "total": float(totals["total"]),
                        "paid": float(totals["paid"]),
                        "debt": float(totals["debt"]),
                        "total_label": _shipment_amount_label(totals["total"]),
                        "paid_label": _shipment_amount_label(totals["paid"]),
                        "debt_label": _shipment_amount_label(totals["debt"]),
                    }
                )
            numbers = sorted([int(n) for n in bucket["numbers"] if str(n).isdigit()])
            statuses = sorted(bucket["statuses"])
            doc_statuses = sorted(bucket["doc_statuses"])
            doc_status = "confirmed" if doc_statuses and set(doc_statuses) == {"confirmed"} else "new"
            journal.append(
                {
                    "day_index": idx,
                    "date": day,
                    "count": int(bucket["count"]),
                    "couriers": sorted(bucket["couriers"]),
                    "courier_count": len(bucket["couriers"]),
                    "numbers_label": f"{numbers[0]}-{numbers[-1]}" if len(numbers) > 1 else (str(numbers[0]) if numbers else ""),
                    "currencies": currencies,
                    "statuses": statuses,
                    "doc_status": doc_status,
                    "doc_status_label": "Подтверждён" if doc_status == "confirmed" else "Новый",
                    "total_amount": total_amount,
                    "paid_amount": paid_amount,
                    "debt_amount": debt_amount,
                }
            )
        return journal

    def _shipment_document_status(rows: list[dict[str, Any]]) -> dict[str, str]:
        confirmed = bool(rows) and all(str(row.get("doc_status") or "").lower() == "confirmed" for row in rows)
        doc_status = "confirmed" if confirmed else "new"
        return {
            "value": doc_status,
            "label": "Подтверждён" if confirmed else "Новый",
        }

    def _shipment_day_summary(shipments: list[dict[str, Any]], day: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, str]]:
        clean_day = str(day or "").strip()[:10]
        rows = [row for row in shipments if str(row.get("shipment_date") or "") == clean_day]
        grouped: dict[tuple[str, str], dict[str, Any]] = {}
        for row in rows:
            courier = str(row.get("courier_name") or "").strip() or "Без имени"
            ccy = str(row.get("currency") or "UZS").upper()
            bucket = grouped.setdefault(
                (courier, ccy),
                {"courier_name": courier, "currency": ccy, "count": 0, "total": 0.0, "paid": 0.0, "debt": 0.0},
            )
            bucket["count"] += 1
            bucket["total"] += float(row.get("total_amount") or 0)
            bucket["paid"] += float(row.get("paid_amount") or 0)
            bucket["debt"] += float(row.get("debt_amount") or 0)
        summary = []
        for bucket in sorted(grouped.values(), key=lambda item: (item["courier_name"], item["currency"])):
            summary.append(
                {
                    **bucket,
                    "total_label": _shipment_amount_label(bucket["total"]),
                    "paid_label": _shipment_amount_label(bucket["paid"]),
                    "debt_label": _shipment_amount_label(bucket["debt"]),
                }
            )
        return rows, summary, _shipment_document_status(rows)

    @app.get("/organizations/shipments", response_class=HTMLResponse, name="organizations_shipments_get")
    def organizations_shipments_get(request: Request):
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        u = _refresh_org_session(request, org_scope="general")
        orgs = list_organizations(owner_id)
        selected = _selected_org_for_general(owner_id, request, u)
        selected_org_id = str((selected or {}).get("id") or owner_id)
        try:
            recompute_delivery_debts(selected_org_id)
        except Exception:
            logger.exception("[upos] recompute_delivery_debts failed before shipments view")
        shipments_rows = list_delivery_shipments(selected_org_id)
        return tpl(
            request,
            "home_organizations_shipments.html",
            variant="user",
            active="organizations_shipments",
            organizations_for_filters=orgs,
            selected_organization=selected,
            selected_organization_id=selected_org_id,
            shipments=shipments_rows,
            shipment_daily_journal=_shipment_daily_journal(shipments_rows),
            shipment_totals=shipment_totals(selected_org_id),
            courier_debts=list_courier_debts(selected_org_id, include_zero=True),
            hr_employees=list_hr_employees(owner_id),
            today=datetime.now().strftime("%Y-%m-%d"),
            flash_ok=request.query_params.get("msg"),
            flash_err=request.query_params.get("error"),
        )

    @app.post("/organizations/shipments/create", name="organizations_shipments_create")
    async def organizations_shipments_create(request: Request):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations/shipments?err=csrf", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        org_id = str(form.get("organization_id") or "").strip()
        org = get_organization(owner_id, org_id)
        if not org:
            return RedirectResponse(url="/organizations/shipments?error=org", status_code=302)
        shipment_date = str(form.get("shipment_date") or "").strip()
        employee_ids = list(form.getlist("employee_id"))
        courier_names = list(form.getlist("courier_name"))
        product_names = list(form.getlist("product_name"))
        amounts = list(form.getlist("amount"))
        currencies = list(form.getlist("currency"))
        notes = list(form.getlist("note"))
        size = max(len(employee_ids), len(courier_names), len(product_names), len(amounts), len(currencies), len(notes))
        rows: list[dict[str, Any]] = []
        for idx in range(size):
            rows.append(
                {
                    "employee_id": str(employee_ids[idx] if idx < len(employee_ids) else "").strip(),
                    "courier_name": str(courier_names[idx] if idx < len(courier_names) else "").strip(),
                    "product_name": str(product_names[idx] if idx < len(product_names) else "").strip(),
                    "amount": str(amounts[idx] if idx < len(amounts) else "").strip(),
                    "currency": str(currencies[idx] if idx < len(currencies) else "UZS").strip(),
                    "note": str(notes[idx] if idx < len(notes) else "").strip(),
                }
            )
        try:
            create_delivery_shipments(
                str(org["id"]),
                rows,
                shipment_date,
                employee_workspace_owner_id=owner_id,
            )
        except ValueError as exc:
            return RedirectResponse(
                url=_shipment_return_url("/organizations/shipments", str(org["id"]), error=str(exc) or "shipments"),
                status_code=302,
            )
        return RedirectResponse(
            url=_shipment_return_url("/organizations/shipments", str(org["id"]), msg="shipments_created"),
            status_code=302,
        )

    @app.get("/api/shipments/courier-debts")
    def api_shipments_courier_debts(request: Request):
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        try:
            recompute_delivery_debts(oid)
        except Exception:
            logger.exception("[upos] recompute_delivery_debts failed before API debts")
        return {
            "category": COURIER_PAYMENT_CATEGORY,
            "courier_debts": list_courier_debts(oid, include_zero=True),
        }

    @app.get("/organizations/shipments/day/{day}", response_class=HTMLResponse, name="organizations_shipments_day")
    def organizations_shipments_day(request: Request, day: str):
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        u = _refresh_org_session(request, org_scope="general")
        orgs = list_organizations(owner_id)
        selected = _selected_org_for_general(owner_id, request, u)
        selected_org_id = str((selected or {}).get("id") or owner_id)
        try:
            recompute_delivery_debts(selected_org_id)
        except Exception:
            logger.exception("[upos] recompute_delivery_debts failed before shipments day view")
        shipments_rows, courier_summary, document_status = _shipment_day_summary(list_delivery_shipments(selected_org_id, limit=2000), day)
        return tpl(
            request,
            "home_shipments_day.html",
            variant="user",
            active="organizations_shipments",
            organizations_for_filters=orgs,
            selected_organization=selected,
            selected_organization_id=selected_org_id,
            day=str(day or "")[:10],
            shipments=shipments_rows,
            courier_summary=courier_summary,
            document_status=document_status,
            inside_organization=False,
        )

    @app.post("/organizations/shipments/day/{day}/update", name="organizations_shipments_day_update")
    async def organizations_shipments_day_update(request: Request, day: str):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations/shipments?err=csrf", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        org_id = str(form.get("organization_id") or "").strip()
        org = get_organization(owner_id, org_id)
        if not org:
            return RedirectResponse(url="/organizations/shipments?error=org", status_code=302)
        clean_day = str(day or "").strip()[:10]
        return_shipment_id = str(form.get("return_shipment_id") or "").strip()
        return_hash = f"#shipment-{quote(return_shipment_id)}" if return_shipment_id else f"#day-{quote(clean_day)}"
        try:
            ok = update_delivery_shipment_document(
                str(org["id"]),
                clean_day,
                _shipment_rows_from_form(form),
                employee_workspace_owner_id=owner_id,
            )
        except ValueError as exc:
            return RedirectResponse(
                url=f"{_shipment_return_url('/organizations/shipments', str(org['id']), error=str(exc) or 'shipments')}{return_hash}",
                status_code=302,
            )
        msg = "shipment_updated" if ok else "shipment_not_found"
        return RedirectResponse(
            url=f"{_shipment_return_url('/organizations/shipments', str(org['id']), msg=msg)}{return_hash}",
            status_code=302,
        )

    @app.post("/organizations/shipments/day/{day}/confirm", name="organizations_shipments_day_confirm")
    async def organizations_shipments_day_confirm(request: Request, day: str):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations/shipments?err=csrf", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        org_id = str(form.get("organization_id") or "").strip()
        org = get_organization(owner_id, org_id)
        if not org:
            return RedirectResponse(url="/organizations/shipments?error=org", status_code=302)
        clean_day = str(day or "").strip()[:10]
        ok = confirm_delivery_shipment_document(str(org["id"]), clean_day)
        msg = "shipment_confirmed" if ok else "shipment_not_found"
        return RedirectResponse(
            url=f"{_shipment_return_url('/organizations/shipments', str(org['id']), msg=msg)}#day-{quote(clean_day)}",
            status_code=302,
        )

    @app.post("/organizations/shipments/day/{day}/delete", name="organizations_shipments_day_delete")
    async def organizations_shipments_day_delete(request: Request, day: str):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations/shipments?err=csrf", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        org_id = str(form.get("organization_id") or "").strip()
        org = get_organization(owner_id, org_id)
        if not org:
            return RedirectResponse(url="/organizations/shipments?error=org", status_code=302)
        clean_day = str(day or "").strip()[:10]
        ok = delete_delivery_shipment_document(str(org["id"]), clean_day)
        msg = "shipment_deleted" if ok else "shipment_not_found"
        return RedirectResponse(
            url=_shipment_return_url("/organizations/shipments", str(org["id"]), msg=msg),
            status_code=302,
        )

    @app.post("/organizations/shipments/shipment/{shipment_id}/delete", name="organizations_shipment_delete")
    async def organizations_shipment_delete(request: Request, shipment_id: str):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations/shipments?err=csrf", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        org_id = str(form.get("organization_id") or "").strip()
        org = get_organization(owner_id, org_id)
        if not org:
            return RedirectResponse(url="/organizations/shipments?error=org", status_code=302)
        deleted_day = delete_delivery_shipment(str(org["id"]), shipment_id)
        msg = "shipment_deleted" if deleted_day else "shipment_not_found"
        url = _shipment_return_url("/organizations/shipments", str(org["id"]), msg=msg)
        if deleted_day:
            url = f"{url}#day-{quote(deleted_day)}"
        return RedirectResponse(url=url, status_code=302)

    @app.post("/organizations/shipments/shipment/{shipment_id}/confirm", name="organizations_shipment_confirm")
    async def organizations_shipment_confirm(request: Request, shipment_id: str):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations/shipments?err=csrf", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        org_id = str(form.get("organization_id") or "").strip()
        org = get_organization(owner_id, org_id)
        if not org:
            return RedirectResponse(url="/organizations/shipments?error=org", status_code=302)
        confirmed_day = confirm_delivery_shipment(str(org["id"]), shipment_id)
        msg = "shipment_confirmed" if confirmed_day else "shipment_not_found"
        url = _shipment_return_url("/organizations/shipments", str(org["id"]), msg=msg)
        return RedirectResponse(url=f"{url}#shipment-{quote(str(shipment_id or '').strip())}", status_code=302)

    @app.get("/organizations/hr", response_class=HTMLResponse, name="organizations_hr_get")
    def organizations_hr_get(request: Request):
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        u = _refresh_org_session(request, org_scope="general")
        active_org_id = _employee_active_org_id(owner_id, u)
        selected_date = str(request.query_params.get("date") or datetime.now().strftime("%Y-%m-%d"))[:10]
        return tpl(
            request,
            "home_organizations_hr.html",
            variant="user",
            active="organizations_hr",
            employees=_hr_employees_with_adjustments(active_org_id, selected_date),
            positions=list_positions(active_org_id),
            selected_organization_id=active_org_id,
            selected_date=selected_date,
            flash_ok=request.query_params.get("msg"),
            flash_err=request.query_params.get("error"),
        )

    @app.get("/organizations/hr/salary-act/{employee_id}", response_class=HTMLResponse, name="organizations_hr_salary_act_detail")
    def organizations_hr_salary_act_detail(request: Request, employee_id: str):
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        u = _refresh_org_session(request, org_scope="general")
        active_org_id = _employee_active_org_id(owner_id, u)
        selected_date = str(request.query_params.get("date") or datetime.now().strftime("%Y-%m-%d"))[:10]
        try:
            act = _salary_employee_act_payload(active_org_id, selected_date, employee_id)
        except ValueError as exc:
            raise StarletteHTTPException(status_code=404, detail=str(exc)) from exc
        return tpl(
            request,
            "home_hr_salary_act_detail.html",
            variant="user",
            active="organizations_hr",
            inside_organization=False,
            selected_organization_id=active_org_id,
            selected_date=selected_date,
            act=act,
        )

    async def _save_hr_photo(owner_id: str, photo: UploadFile | None) -> str:
        if photo is None or not str(photo.filename or "").strip():
            return ""
        content = await photo.read()
        if not content:
            return ""
        img = Image.open(io.BytesIO(content))
        w, h = img.size
        sz = min(w, h)
        img = img.crop(((w - sz) // 2, (h - sz) // 2, (w + sz) // 2, (h + sz) // 2))
        img = img.resize((500, 500), Image.Resampling.LANCZOS)
        rel_dir = "hr-photos"
        out_dir = BASE_DIR / "static" / rel_dir
        out_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{owner_id}_{uuid.uuid4().hex[:10]}.png"
        rel_path = f"{rel_dir}/{filename}"
        img.save(BASE_DIR / "static" / rel_path, "PNG")
        return rel_path

    @app.post("/organizations/hr/employees/create", name="organizations_hr_employee_create")
    async def organizations_hr_employee_create(
        request: Request,
        csrf_token: str = Form(default=""),
        first_name: str = Form(default=""),
        last_name: str = Form(default=""),
        position_id: str = Form(default=""),
        position: str = Form(default=""),
        passport_series: str = Form(default=""),
        passport_number: str = Form(default=""),
        monthly_salary: str = Form(default="0"),
        is_courier: str = Form(default=""),
        hired_at: str = Form(default=""),
        photo: UploadFile | None = File(default=None),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations/hr?err=csrf", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        active_org_id = _employee_active_org_id(owner_id, request.session.get("user") or {})
        try:
            photo_path = await _save_hr_photo(active_org_id, photo)
            create_hr_employee(
                active_org_id,
                {
                    "first_name": first_name,
                    "last_name": last_name,
                    "position_id": position_id,
                    "position": position,
                    "passport_series": passport_series,
                    "passport_number": passport_number,
                    "monthly_salary": monthly_salary,
                    "is_courier": is_courier,
                    "hired_at": hired_at,
                    "photo_path": photo_path,
                },
            )
        except ValueError as exc:
            return RedirectResponse(url=f"/organizations/hr?error={quote(str(exc) or 'hr')}", status_code=302)
        except Exception as exc:
            logger.exception("[upos] HR employee create failed")
            return RedirectResponse(url=f"/organizations/hr?error={quote(type(exc).__name__)}", status_code=302)
        return RedirectResponse(url="/organizations/hr?msg=employee_created", status_code=302)

    @app.post("/organizations/hr/employees/update", name="organizations_hr_employee_update")
    async def organizations_hr_employee_update(
        request: Request,
        csrf_token: str = Form(default=""),
        employee_id: str = Form(default=""),
        first_name: str = Form(default=""),
        last_name: str = Form(default=""),
        position_id: str = Form(default=""),
        position: str = Form(default=""),
        passport_series: str = Form(default=""),
        passport_number: str = Form(default=""),
        monthly_salary: str = Form(default="0"),
        is_courier: str = Form(default=""),
        hired_at: str = Form(default=""),
        photo: UploadFile | None = File(default=None),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations/hr?err=csrf", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        active_org_id = _employee_active_org_id(owner_id, request.session.get("user") or {})
        try:
            photo_path = await _save_hr_photo(active_org_id, photo)
            updated = update_hr_employee(
                active_org_id,
                employee_id,
                {
                    "first_name": first_name,
                    "last_name": last_name,
                    "position_id": position_id,
                    "position": position,
                    "passport_series": passport_series,
                    "passport_number": passport_number,
                    "monthly_salary": monthly_salary,
                    "is_courier": is_courier,
                    "hired_at": hired_at,
                    "photo_path": photo_path,
                },
            )
        except ValueError as exc:
            return RedirectResponse(url=f"/organizations/hr?error={quote(str(exc) or 'hr')}", status_code=302)
        except Exception as exc:
            logger.exception("[upos] HR employee update failed")
            return RedirectResponse(url=f"/organizations/hr?error={quote(type(exc).__name__)}", status_code=302)
        return RedirectResponse(
            url=f"/organizations/hr?msg={'employee_updated' if updated else 'not_found'}",
            status_code=302,
        )

    @app.post("/organizations/hr/employees/{employee_id}/dismiss", name="organizations_hr_employee_dismiss")
    def organizations_hr_employee_dismiss(
        request: Request,
        employee_id: str,
        csrf_token: str = Form(default=""),
        dismissed_at: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations/hr?err=csrf", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        active_org_id = _employee_active_org_id(owner_id, request.session.get("user") or {})
        ok = dismiss_hr_employee(active_org_id, employee_id, dismissed_at)
        return RedirectResponse(url=f"/organizations/hr?msg={'employee_dismissed' if ok else 'not_found'}", status_code=302)

    @app.post("/organizations/hr/employees/{employee_id}/restore", name="organizations_hr_employee_restore")
    def organizations_hr_employee_restore(
        request: Request,
        employee_id: str,
        csrf_token: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations/hr?err=csrf#staff", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        active_org_id = _employee_active_org_id(owner_id, request.session.get("user") or {})
        ok = restore_hr_employee(active_org_id, employee_id)
        return RedirectResponse(url=f"/organizations/hr?msg={'employee_restored' if ok else 'not_found'}#staff", status_code=302)

    @app.post("/organizations/hr/employees/{employee_id}/delete", name="organizations_hr_employee_delete")
    def organizations_hr_employee_delete(
        request: Request,
        employee_id: str,
        csrf_token: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations/hr?err=csrf#staff", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        active_org_id = _employee_active_org_id(owner_id, request.session.get("user") or {})
        ok = delete_hr_employee_permanently(active_org_id, employee_id)
        if ok:
            _purge_hr_salary_records(active_org_id, employee_id)
        return RedirectResponse(url=f"/organizations/hr?msg={'employee_deleted' if ok else 'not_found'}#staff", status_code=302)

    @app.post("/organizations/hr/attendance", name="organizations_hr_attendance")
    def organizations_hr_attendance(
        request: Request,
        csrf_token: str = Form(default=""),
        employee_id: str = Form(default=""),
        work_date: str = Form(default=""),
        status: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations/hr?err=csrf", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        active_org_id = _employee_active_org_id(owner_id, request.session.get("user") or {})
        try:
            ok = set_hr_attendance(active_org_id, employee_id, work_date, status)
        except ValueError as exc:
            return RedirectResponse(url=f"/organizations/hr?date={quote(work_date)}&error={quote(str(exc))}", status_code=302)
        return RedirectResponse(
            url=f"/organizations/hr?date={quote(work_date)}&msg={'attendance_saved' if ok else 'not_found'}",
            status_code=302,
        )

    def _hr_attendance_records_from_form(form) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        for employee_id in form.getlist("employee_id"):
            eid = str(employee_id or "").strip()
            if not eid:
                continue
            records.append(
                {
                    "employee_id": eid,
                    "status": str(form.get(f"status_{eid}") or "absent").strip(),
                    "note": str(form.get(f"note_{eid}") or "").strip(),
                }
            )
        return records

    @app.post("/organizations/hr/attendance/report", name="organizations_hr_attendance_report")
    async def organizations_hr_attendance_report(request: Request):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        work_date = str(form.get("work_date") or "").strip()
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations/hr?err=csrf#calendar", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        active_org_id = _employee_active_org_id(owner_id, request.session.get("user") or {})
        try:
            report = set_hr_attendance_day(active_org_id, work_date, _hr_attendance_records_from_form(form))
        except ValueError as exc:
            return RedirectResponse(url=f"/organizations/hr?date={quote(work_date)}&error={quote(str(exc))}#calendar", status_code=302)
        telegram_status = "attendance_report_saved"
        try:
            from upos.telegram_notifier import send_hr_attendance_report

            tg = send_hr_attendance_report(active_org_id, report)
            if tg.get("ok"):
                telegram_status = "attendance_report_sent"
            elif tg.get("error"):
                telegram_status = f"telegram_{tg.get('error')}"
        except Exception:
            logger.exception("[upos] HR attendance Telegram report failed")
            telegram_status = "telegram_failed"
        return RedirectResponse(url=f"/organizations/hr?date={quote(work_date)}&msg={quote(telegram_status)}#calendar", status_code=302)

    def _hr_money(raw: Any) -> Decimal:
        text_raw = str(raw or "0").strip()
        text_raw = text_raw.replace("\u00a0", "").replace("\u202f", "").replace(" ", "").replace("'", "")
        if "," in text_raw and "." in text_raw:
            text_raw = text_raw.replace(",", "")
        elif text_raw.count(",") > 1:
            text_raw = text_raw.replace(",", "")
        elif "," in text_raw:
            text_raw = text_raw.replace(",", ".")
        try:
            value = Decimal(text_raw or "0").quantize(Decimal("0.01"))
        except Exception as exc:
            raise ValueError("invalid_amount") from exc
        if value < 0:
            raise ValueError("invalid_amount")
        return value

    def _hr_amount_out(raw: Any) -> float:
        return float(Decimal(str(raw or "0")).quantize(Decimal("0.01")))

    def _hr_amount_label(raw: Any) -> str:
        value = _hr_amount_out(raw)
        return f"{value:,.0f}".replace(",", " ")

    def _hr_salary_adjustments_path(workspace_owner_id: str) -> Path:
        workspace_dir = CLIENT_WORKSPACES_DIR / workspace_owner_id
        workspace_dir.mkdir(parents=True, exist_ok=True)
        return workspace_dir / "hr_salary_adjustments.json"

    def _load_hr_salary_adjustments(workspace_owner_id: str) -> list[dict[str, Any]]:
        path = _hr_salary_adjustments_path(workspace_owner_id)
        if not path.exists():
            return []
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return []
        return raw if isinstance(raw, list) else []

    def _save_hr_salary_adjustments(workspace_owner_id: str, rows: list[dict[str, Any]]) -> None:
        path = _hr_salary_adjustments_path(workspace_owner_id)
        path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")

    def _purge_hr_salary_records(workspace_owner_id: str, employee_id: str) -> None:
        clean_employee_id = str(employee_id or "").strip()
        if not clean_employee_id:
            return
        rows = _load_hr_salary_adjustments(workspace_owner_id)
        filtered_rows = [row for row in rows if str(row.get("employee_id") or "").strip() != clean_employee_id]
        if len(filtered_rows) != len(rows):
            _save_hr_salary_adjustments(workspace_owner_id, filtered_rows)

        workspace_dir = CLIENT_WORKSPACES_DIR / workspace_owner_id
        path = workspace_dir / "hr_salary_acts.json"
        if not path.exists():
            return
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return
        if not isinstance(raw, list):
            return
        changed = False
        cleaned_acts: list[dict[str, Any]] = []
        for act in raw:
            if not isinstance(act, dict):
                cleaned_acts.append(act)
                continue
            employees = act.get("employees")
            if not isinstance(employees, list):
                cleaned_acts.append(act)
                continue
            kept = [
                emp
                for emp in employees
                if not isinstance(emp, dict) or str(emp.get("id") or "").strip() != clean_employee_id
            ]
            if len(kept) != len(employees):
                changed = True
            if kept:
                cleaned_acts.append({**act, "employees": kept})
            else:
                changed = True
        if changed:
            path.write_text(json.dumps(cleaned_acts, ensure_ascii=False, indent=2), encoding="utf-8")

    def _month_prefix_from_date(raw: str) -> str:
        value = str(raw or "")[:10]
        return value[:7] if len(value) >= 7 else datetime.now().strftime("%Y-%m")

    def _salary_month_label(raw: Any) -> str:
        value = str(raw or "").strip()[:10]
        try:
            dt = datetime.strptime(value[:7] + "-01", "%Y-%m-%d")
        except ValueError:
            dt = datetime.now()
        names = [
            "",
            "Январь",
            "Февраль",
            "Март",
            "Апрель",
            "Май",
            "Июнь",
            "Июль",
            "Август",
            "Сентябрь",
            "Октябрь",
            "Ноябрь",
            "Декабрь",
        ]
        return f"{names[dt.month]} {dt.year}"

    def _salary_period_month_ranges(date_from: str, date_to: str) -> list[dict[str, str]]:
        start = _salary_clean_date(date_from) or _salary_month_labels(date_from)[0]
        end = _salary_clean_date(date_to) or start
        if end < start:
            start, end = end, start
        try:
            cursor = datetime.strptime(start, "%Y-%m-%d").replace(day=1)
            last = datetime.strptime(end, "%Y-%m-%d").replace(day=1)
        except ValueError:
            return []
        rows: list[dict[str, str]] = []
        while cursor <= last:
            first = f"{cursor.year:04d}-{cursor.month:02d}-01"
            last_day = calendar.monthrange(cursor.year, cursor.month)[1]
            month_end = f"{cursor.year:04d}-{cursor.month:02d}-{last_day:02d}"
            rows.append(
                {
                    "month_key": f"{cursor.year:04d}-{cursor.month:02d}",
                    "month_label": _salary_month_label(first),
                    "date_from": max(start, first),
                    "date_to": min(end, month_end),
                    "month_date": first,
                }
            )
            next_month = cursor.month + 1
            next_year = cursor.year
            if next_month > 12:
                next_month = 1
                next_year += 1
            cursor = cursor.replace(year=next_year, month=next_month, day=1)
        return rows

    def _apply_hr_salary_adjustments(
        workspace_owner_id: str,
        employees: list[dict[str, Any]],
        selected_date: str,
        date_from: Any = None,
        date_to: Any = None,
    ) -> list[dict[str, Any]]:
        period_from, period_to = _salary_period_labels(selected_date, date_from, date_to)
        grouped: dict[str, list[dict[str, Any]]] = {str(emp.get("id") or ""): [] for emp in employees}
        for row in _load_hr_salary_adjustments(workspace_owner_id):
            eid = str(row.get("employee_id") or "").strip()
            if eid not in grouped:
                continue
            work_date = str(row.get("work_date") or "").strip()[:10]
            if not work_date or work_date < period_from or work_date > period_to:
                continue
            amount = _hr_money(row.get("amount"))
            if amount <= 0:
                continue
            grouped[eid].append(
                {
                    **row,
                    "amount": _hr_amount_out(amount),
                    "amount_label": _hr_amount_label(amount),
                    "month_key": work_date[:7],
                    "month_label": _salary_month_label(work_date),
                    "type_label": "Штраф" if row.get("type") == "penalty" else "Начисление",
                }
            )
        for emp in employees:
            rows = grouped.get(str(emp.get("id") or ""), [])
            base_due = _hr_money(emp.get("salary_due"))
            bonus = sum((_hr_money(row.get("amount")) for row in rows if row.get("type") == "bonus"), Decimal("0.00"))
            penalty = sum((_hr_money(row.get("amount")) for row in rows if row.get("type") == "penalty"), Decimal("0.00"))
            due = max(Decimal("0.00"), base_due + bonus - penalty)
            emp["salary_base_due"] = _hr_amount_out(base_due)
            emp["salary_base_due_label"] = _hr_amount_label(base_due)
            emp["salary_bonus"] = _hr_amount_out(bonus)
            emp["salary_bonus_label"] = _hr_amount_label(bonus)
            emp["salary_penalty"] = _hr_amount_out(penalty)
            emp["salary_penalty_label"] = _hr_amount_label(penalty)
            emp["salary_adjustments"] = rows
            emp["salary_adjustments_count"] = len(rows)
            emp["salary_due"] = _hr_amount_out(due)
            emp["salary_due_label"] = _hr_amount_label(due)
        return employees

    def _hr_employees_with_adjustments(
        workspace_owner_id: str,
        selected_date: str,
        date_from: Any = None,
        date_to: Any = None,
    ) -> list[dict[str, Any]]:
        period_from, period_to = _salary_period_labels(selected_date, date_from, date_to)
        employees = list_hr_employees(workspace_owner_id, period_from or selected_date)
        if date_from or date_to:
            employees = _recalculate_hr_salary_period(employees, period_from, period_to)
        return _apply_hr_salary_adjustments(workspace_owner_id, employees, selected_date, period_from, period_to)

    def _salary_month_labels(selected_date: str) -> tuple[str, str]:
        raw = str(selected_date or "").strip()[:10] or datetime.now().strftime("%Y-%m-%d")
        try:
            dt = datetime.strptime(raw, "%Y-%m-%d")
        except ValueError:
            dt = datetime.now()
        last_day = calendar.monthrange(dt.year, dt.month)[1]
        return f"{dt.year:04d}-{dt.month:02d}-01", f"{dt.year:04d}-{dt.month:02d}-{last_day:02d}"

    def _salary_clean_date(raw: Any) -> str | None:
        value = str(raw or "").strip()[:10]
        try:
            datetime.strptime(value, "%Y-%m-%d")
        except ValueError:
            return None
        return value

    def _salary_period_labels(selected_date: str, date_from: Any = None, date_to: Any = None) -> tuple[str, str]:
        if date_from or date_to:
            fallback_from, fallback_to = _salary_month_labels(selected_date)
            start = _salary_clean_date(date_from) or fallback_from
            end = _salary_clean_date(date_to) or start or fallback_to
            if end < start:
                start, end = end, start
            return start, end
        return _salary_month_labels(selected_date)

    def _recalculate_hr_salary_period(employees: list[dict[str, Any]], date_from: str, date_to: str) -> list[dict[str, Any]]:
        for emp in employees:
            attendance = {
                str(day): status
                for day, status in (emp.get("attendance") or {}).items()
                if date_from <= str(day) <= date_to
            }
            notes = {
                str(day): note
                for day, note in (emp.get("attendance_notes") or {}).items()
                if date_from <= str(day) <= date_to
            }
            present_days = sum(1 for status in attendance.values() if status == "present")
            absent_days = sum(1 for status in attendance.values() if status == "absent")
            salary = _hr_money(emp.get("monthly_salary"))
            base_due = Decimal("0.00")
            for day, status in attendance.items():
                if status != "present":
                    continue
                try:
                    dt = datetime.strptime(day, "%Y-%m-%d")
                except ValueError:
                    continue
                days_in_month = calendar.monthrange(dt.year, dt.month)[1]
                if days_in_month > 0:
                    base_due += salary / Decimal(days_in_month)
            base_due = base_due.quantize(Decimal("0.01"))
            emp["attendance"] = dict(sorted(attendance.items()))
            emp["attendance_notes"] = dict(sorted(notes.items()))
            emp["present_days"] = present_days
            emp["absent_days"] = absent_days
            emp["salary_due"] = _hr_amount_out(base_due)
            emp["salary_due_label"] = _hr_amount_label(base_due)
        return employees

    def _salary_employee_base_month_rows(
        workspace_owner_id: str,
        employee_id: str,
        date_from: str,
        date_to: str,
    ) -> list[dict[str, Any]]:
        clean_employee_id = str(employee_id or "").strip()
        rows: list[dict[str, Any]] = []
        for month in _salary_period_month_ranges(date_from, date_to):
            month_date = month.get("month_date") or month.get("date_from") or date_from
            employees = list_hr_employees(workspace_owner_id, month_date)
            employee = next((emp for emp in employees if str(emp.get("id") or "") == clean_employee_id), None)
            if employee is None:
                continue
            month_from = str(month.get("date_from") or date_from)
            month_to = str(month.get("date_to") or date_to)
            attendance = {
                str(day): status
                for day, status in (employee.get("attendance") or {}).items()
                if month_from <= str(day) <= month_to
            }
            present_days = sum(1 for status in attendance.values() if status == "present")
            absent_days = sum(1 for status in attendance.values() if status == "absent")
            salary = _hr_money(employee.get("monthly_salary"))
            base_due = Decimal("0.00")
            try:
                dt = datetime.strptime(month_date, "%Y-%m-%d")
                days_in_month = calendar.monthrange(dt.year, dt.month)[1]
            except ValueError:
                days_in_month = 0
            if days_in_month > 0 and present_days > 0:
                base_due = (salary / Decimal(days_in_month) * Decimal(present_days)).quantize(Decimal("0.01"))
            if not present_days and not absent_days and base_due <= 0:
                continue
            rows.append(
                {
                    "type": "base",
                    "type_label": "Зарплата по табелю",
                    "amount": _hr_amount_out(base_due),
                    "amount_label": _hr_amount_label(base_due),
                    "date_from": month_from,
                    "date_to": month_to,
                    "month_key": str(month.get("month_key") or month_from[:7]),
                    "month_label": str(month.get("month_label") or _salary_month_label(month_from)),
                    "present_days": present_days,
                    "absent_days": absent_days,
                    "monthly_salary_label": _hr_amount_label(salary),
                }
            )
        return rows

    def _salary_employee_payments(
        workspace_owner_id: str,
        selected_date: str,
        date_from: Any = None,
        date_to: Any = None,
    ) -> dict[str, dict[str, float]]:
        settings = load_workspace_settings(workspace_owner_id)
        tz_name = normalize_workspace_timezone(str(settings.get("timezone") or ""))
        date_from, date_to = _salary_period_labels(selected_date, date_from, date_to)
        start_utc, end_utc, *_ = period_local_bounds_utc(
            tz_name,
            preset="custom",
            date_from=date_from,
            date_to=date_to,
        )
        employee_expr = Transaction.data.op("->>")("hr_employee_id")
        with session_scope() as session:
            salary_category_ids = [
                str(x)
                for x in session.scalars(
                    select(FinanceCategory.id).where(
                        FinanceCategory.workspace_owner_id == workspace_owner_id,
                        FinanceCategory.type == "expense",
                        FinanceCategory.name == "Зарплата",
                    )
                ).all()
            ]
            category_filter = Transaction.category == "Зарплата"
            if salary_category_ids:
                category_filter = or_(category_filter, Transaction.category_id.in_(salary_category_ids))
            stmt = (
                select(employee_expr, Transaction.currency, func.sum(Transaction.amount))
                .where(
                    Transaction.workspace_owner_id == workspace_owner_id,
                    Transaction.type == "expense",
                    Transaction.status == "confirmed",
                    category_filter,
                    func.coalesce(employee_expr, "") != "",
                )
                .group_by(employee_expr, Transaction.currency)
            )
            if start_utc is not None:
                stmt = stmt.where(Transaction.created_at >= start_utc)
            if end_utc is not None:
                stmt = stmt.where(Transaction.created_at < end_utc)
            rows = session.execute(stmt).all()
        out: dict[str, dict[str, float]] = {}
        for employee_id, currency, amount in rows:
            eid = str(employee_id or "").strip()
            ccy = str(currency or "").strip().upper()
            if not eid or not ccy:
                continue
            out.setdefault(eid, {})[ccy] = float(Decimal(str(amount or "0")).quantize(Decimal("0.01")))
        return out

    def _hr_account_name_map(workspace_owner_id: str) -> dict[str, str]:
        names: dict[str, str] = {}
        treasury = load_treasury(workspace_owner_id)
        for pocket in treasury.get("pockets") or []:
            if not isinstance(pocket, dict):
                continue
            pid = str(pocket.get("id") or "").strip()
            if pid:
                names[pid] = str(pocket.get("label") or pocket.get("name") or pid)
        with session_scope() as session:
            rows = session.execute(
                select(FinanceAccount).where(FinanceAccount.workspace_owner_id == workspace_owner_id),
            ).scalars().all()
            for account in rows:
                names[str(account.id)] = str(account.name or account.id)
        return names

    def _salary_category_filter(workspace_owner_id: str):
        with session_scope() as session:
            salary_category_ids = [
                str(x)
                for x in session.scalars(
                    select(FinanceCategory.id).where(
                        FinanceCategory.workspace_owner_id == workspace_owner_id,
                        FinanceCategory.type == "expense",
                        FinanceCategory.name == "Зарплата",
                    )
                ).all()
            ]
        category_filter = Transaction.category == "Зарплата"
        if salary_category_ids:
            category_filter = or_(category_filter, Transaction.category_id.in_(salary_category_ids))
        return category_filter

    def _salary_employee_payment_rows(
        workspace_owner_id: str,
        selected_date: str,
        employee_id: str,
        date_from: Any = None,
        date_to: Any = None,
    ) -> list[dict[str, Any]]:
        settings = load_workspace_settings(workspace_owner_id)
        tz_name = normalize_workspace_timezone(str(settings.get("timezone") or ""))
        date_from, date_to = _salary_period_labels(selected_date, date_from, date_to)
        start_utc, end_utc, *_ = period_local_bounds_utc(
            tz_name,
            preset="custom",
            date_from=date_from,
            date_to=date_to,
        )
        employee_expr = Transaction.data.op("->>")("hr_employee_id")
        category_filter = _salary_category_filter(workspace_owner_id)
        with session_scope() as session:
            stmt = select(Transaction).where(
                Transaction.workspace_owner_id == workspace_owner_id,
                Transaction.type == "expense",
                Transaction.status == "confirmed",
                category_filter,
                employee_expr == str(employee_id or "").strip(),
            )
            if start_utc is not None:
                stmt = stmt.where(Transaction.created_at >= start_utc)
            if end_utc is not None:
                stmt = stmt.where(Transaction.created_at < end_utc)
            rows = session.execute(stmt.order_by(Transaction.created_at.desc())).scalars().all()
        account_names = _hr_account_name_map(workspace_owner_id)
        out: list[dict[str, Any]] = []
        for tx in rows:
            account_id = str(tx.from_account_id or tx.from_pocket_id or "").strip()
            created = tx.created_at
            if created and created.tzinfo is None:
                created = created.replace(tzinfo=timezone.utc)
            local_created = created.astimezone(ZoneInfo(tz_name)) if created else None
            created_label = local_created.strftime("%d.%m.%Y %H:%M") if local_created else ""
            created_date = local_created.strftime("%Y-%m-%d") if local_created else str(selected_date or "")[:10]
            out.append(
                {
                    "number": tx.number,
                    "created_label": created_label,
                    "month_key": created_date[:7],
                    "month_label": _salary_month_label(created_date),
                    "amount": _hr_amount_out(tx.amount),
                    "amount_label": _hr_amount_label(tx.amount),
                    "currency": str(tx.currency or "UZS").upper(),
                    "account": account_names.get(account_id, account_id[:8] if account_id else "—"),
                    "note": tx.note or "",
                }
            )
        return out

    def _salary_employee_act_payload(
        workspace_owner_id: str,
        selected_date: str,
        employee_id: str,
        date_from: Any = None,
        date_to: Any = None,
    ) -> dict[str, Any]:
        clean_employee_id = str(employee_id or "").strip()
        period_from, period_to = _salary_period_labels(selected_date, date_from, date_to)
        employees = _hr_employees_with_adjustments(workspace_owner_id, selected_date, period_from, period_to)
        employee = next((emp for emp in employees if str(emp.get("id") or "") == clean_employee_id), None)
        if employee is None:
            raise ValueError("employee_not_found")
        payments_by_currency = _salary_employee_payments(workspace_owner_id, selected_date, period_from, period_to).get(clean_employee_id, {})
        payment_rows = _salary_employee_payment_rows(workspace_owner_id, selected_date, clean_employee_id, period_from, period_to)
        salary_base_months = _salary_employee_base_month_rows(workspace_owner_id, clean_employee_id, period_from, period_to)
        salary_due = _hr_money(employee.get("salary_due"))
        paid_uzs = Decimal(str(payments_by_currency.get("UZS", 0))).quantize(Decimal("0.01"))
        balance_uzs = salary_due - paid_uzs
        return {
            "employee": employee,
            "date": str(selected_date or "")[:10],
            "date_from": period_from,
            "date_to": period_to,
            "months": _salary_period_month_ranges(period_from, period_to),
            "salary_base_months": salary_base_months,
            "payments": payment_rows,
            "payments_by_currency": [
                {"currency": ccy, "amount_label": _hr_amount_label(amount)}
                for ccy, amount in sorted(payments_by_currency.items())
            ],
            "paid_uzs_label": _hr_amount_label(paid_uzs),
            "balance_uzs": _hr_amount_out(balance_uzs),
            "balance_uzs_label": _hr_amount_label(abs(balance_uzs)),
            "balance_state": "overpaid" if balance_uzs < 0 else "due" if balance_uzs > 0 else "closed",
            "salary_due_label": _hr_amount_label(salary_due),
        }

    def _hr_api_workspace_owner(request: Request) -> tuple[str | None, JSONResponse | None]:
        u = request.session.get("user") or {}
        if u.get("org_scope") != "general":
            return _treasury_workspace_owner(request)
        owner_id = str(u.get("account_owner_id") or u.get("user_id") or "").strip()
        if not valid_workspace_owner_id(owner_id):
            return None, JSONResponse({"error": "workspace"}, status_code=400)
        requested_org_id = str(request.query_params.get("organization_id") or "").strip()
        if requested_org_id:
            orgs = list_organizations(owner_id)
            if any(str(org.get("id")) == requested_org_id for org in orgs):
                return requested_org_id, None
        return _employee_active_org_id(owner_id, u), None

    @app.get("/api/hr/salary-employees")
    def api_hr_salary_employees(request: Request):
        oid, err = _hr_api_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        selected_date = str(request.query_params.get("date") or datetime.now().strftime("%Y-%m-%d")).strip()[:10]
        employees = _hr_employees_with_adjustments(oid, selected_date)
        payments = _salary_employee_payments(oid, selected_date)
        payload: list[dict[str, Any]] = []
        for emp in employees:
            employee_id = str(emp.get("id") or "")
            paid_by_currency = payments.get(employee_id, {})
            salary_due = _hr_money(emp.get("salary_due"))
            paid_uzs = Decimal(str(paid_by_currency.get("UZS", 0))).quantize(Decimal("0.01"))
            balance_uzs = salary_due - paid_uzs
            payload.append(
                {
                    "id": employee_id,
                    "name": emp.get("full_name") or "",
                    "position": emp.get("position") or "",
                    "status": emp.get("status") or "active",
                    "salary_due": _hr_amount_out(salary_due),
                    "salary_due_label": _hr_amount_label(salary_due),
                    "paid_by_currency": paid_by_currency,
                    "paid_uzs": _hr_amount_out(paid_uzs),
                    "paid_uzs_label": _hr_amount_label(paid_uzs),
                    "balance_uzs": _hr_amount_out(balance_uzs),
                    "balance_uzs_label": _hr_amount_label(balance_uzs),
                }
            )
        return {"ok": True, "date": selected_date, "employees": payload}

    @app.get("/api/hr/salary-act/{employee_id}")
    def api_hr_salary_act(request: Request, employee_id: str):
        oid, err = _hr_api_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        selected_date = str(request.query_params.get("date") or datetime.now().strftime("%Y-%m-%d")).strip()[:10]
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        try:
            act = _salary_employee_act_payload(oid, selected_date, employee_id, date_from, date_to)
        except ValueError as exc:
            return JSONResponse({"ok": False, "error": str(exc)}, status_code=404)
        return {"ok": True, "date": selected_date, "act": act}

    def _save_hr_salary_adjustment(
        workspace_owner_id: str,
        work_date: str,
        employee_id: str,
        adjustment_type: str,
        amount: str,
        comment: str,
    ) -> None:
        clean_date = str(work_date or "").strip()[:10] or datetime.now().strftime("%Y-%m-%d")
        clean_employee_id = str(employee_id or "").strip()
        clean_type = str(adjustment_type or "").strip()
        if clean_type not in {"bonus", "penalty"}:
            raise ValueError("salary_adjustment_type_required")
        clean_amount = _hr_money(amount)
        if clean_amount <= 0:
            raise ValueError("salary_adjustment_amount_required")
        employees = list_hr_employees(workspace_owner_id, clean_date)
        employee = next((emp for emp in employees if str(emp.get("id") or "") == clean_employee_id), None)
        if employee is None:
            raise ValueError("salary_adjustment_employee_required")
        rows = _load_hr_salary_adjustments(workspace_owner_id)
        rows.append(
            {
                "id": uuid.uuid4().hex,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "work_date": clean_date,
                "employee_id": clean_employee_id,
                "employee_name": employee.get("full_name"),
                "type": clean_type,
                "amount": _hr_amount_out(clean_amount),
                "comment": str(comment or "").strip()[:500],
            }
        )
        _save_hr_salary_adjustments(workspace_owner_id, rows)

    def _save_hr_salary_act(workspace_owner_id: str, work_date: str, employee_ids: list[str]) -> int:
        clean_ids = [str(item or "").strip() for item in employee_ids if str(item or "").strip()]
        if not clean_ids:
            raise ValueError("salary_act_employees_required")
        employees = _hr_employees_with_adjustments(workspace_owner_id, work_date)
        selected = [emp for emp in employees if str(emp.get("id") or "") in clean_ids]
        if not selected:
            raise ValueError("salary_act_employees_required")
        workspace_dir = CLIENT_WORKSPACES_DIR / workspace_owner_id
        workspace_dir.mkdir(parents=True, exist_ok=True)
        path = workspace_dir / "hr_salary_acts.json"
        acts: list[dict[str, Any]] = []
        if path.exists():
            try:
                raw = json.loads(path.read_text(encoding="utf-8"))
                if isinstance(raw, list):
                    acts = raw
            except Exception:
                acts = []
        act = {
            "id": uuid.uuid4().hex,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "work_date": str(work_date or "")[:10],
            "employees": [
                {
                    "id": emp.get("id"),
                    "name": emp.get("full_name"),
                    "position": emp.get("position"),
                    "monthly_salary": emp.get("monthly_salary"),
                    "present_days": emp.get("present_days"),
                    "absent_days": emp.get("absent_days"),
                    "salary_base_due": emp.get("salary_base_due"),
                    "salary_bonus": emp.get("salary_bonus"),
                    "salary_penalty": emp.get("salary_penalty"),
                    "salary_adjustments": emp.get("salary_adjustments") or [],
                    "salary_due": emp.get("salary_due"),
                }
                for emp in selected
            ],
        }
        acts.append(act)
        path.write_text(json.dumps(acts, ensure_ascii=False, indent=2), encoding="utf-8")
        return len(selected)

    @app.post("/organizations/hr/salary-act/save", name="organizations_hr_salary_act_save")
    async def organizations_hr_salary_act_save(request: Request):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        work_date = str(form.get("work_date") or "").strip() or datetime.now().strftime("%Y-%m-%d")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations/hr?err=csrf#salary", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        active_org_id = _employee_active_org_id(owner_id, request.session.get("user") or {})
        try:
            _save_hr_salary_act(active_org_id, work_date, list(form.getlist("employee_id")))
        except ValueError as exc:
            return RedirectResponse(url=f"/organizations/hr?date={quote(work_date)}&error={quote(str(exc))}#salary", status_code=302)
        return RedirectResponse(url=f"/organizations/hr?date={quote(work_date)}&msg=salary_act_saved#salary", status_code=302)

    @app.post("/organizations/hr/salary-adjustment/save", name="organizations_hr_salary_adjustment_save")
    async def organizations_hr_salary_adjustment_save(request: Request):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        work_date = str(form.get("work_date") or "").strip() or datetime.now().strftime("%Y-%m-%d")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations/hr?err=csrf#staff", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        active_org_id = _employee_active_org_id(owner_id, request.session.get("user") or {})
        try:
            _save_hr_salary_adjustment(
                active_org_id,
                work_date,
                str(form.get("employee_id") or ""),
                str(form.get("adjustment_type") or ""),
                str(form.get("amount") or ""),
                str(form.get("comment") or ""),
            )
        except ValueError as exc:
            return RedirectResponse(url=f"/organizations/hr?date={quote(work_date)}&error={quote(str(exc))}#staff", status_code=302)
        return RedirectResponse(url=f"/organizations/hr?date={quote(work_date)}&msg=salary_adjustment_saved#staff", status_code=302)

    def _current_org_html_owner(request: Request):
        u = request.session.get("user") or {}
        if u.get("role") == "admin":
            return None, RedirectResponse(url="/admin", status_code=302)
        if u.get("org_scope") == "general":
            return None, RedirectResponse(url="/organizations", status_code=302)
        oid = str(u.get("workspace_owner_id") or u.get("user_id") or "").strip()
        if not valid_workspace_owner_id(oid):
            return None, RedirectResponse(url="/auth", status_code=302)
        return oid, None

    def _shipment_rows_from_form(form) -> list[dict[str, Any]]:
        shipment_ids = list(form.getlist("shipment_id"))
        employee_ids = list(form.getlist("employee_id"))
        courier_names = list(form.getlist("courier_name"))
        product_names = list(form.getlist("product_name"))
        amounts = list(form.getlist("amount"))
        currencies = list(form.getlist("currency"))
        notes = list(form.getlist("note"))
        size = max(len(shipment_ids), len(employee_ids), len(courier_names), len(product_names), len(amounts), len(currencies), len(notes))
        rows: list[dict[str, Any]] = []
        for idx in range(size):
            rows.append(
                {
                    "shipment_id": str(shipment_ids[idx] if idx < len(shipment_ids) else "").strip(),
                    "employee_id": str(employee_ids[idx] if idx < len(employee_ids) else "").strip(),
                    "courier_name": str(courier_names[idx] if idx < len(courier_names) else "").strip(),
                    "product_name": str(product_names[idx] if idx < len(product_names) else "").strip(),
                    "amount": str(amounts[idx] if idx < len(amounts) else "").strip(),
                    "currency": str(currencies[idx] if idx < len(currencies) else "UZS").strip(),
                    "note": str(notes[idx] if idx < len(notes) else "").strip(),
                }
            )
        return rows

    @app.get("/shipments", response_class=HTMLResponse, name="home_shipments")
    def home_shipments(request: Request):
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        u = request.session.get("user") or {}
        try:
            recompute_delivery_debts(oid)
        except Exception:
            logger.exception("[upos] recompute_delivery_debts failed before organization shipments view")
        selected = {
            "id": oid,
            "name": str(u.get("active_organization_name") or u.get("workspace_owner_name") or "Организация"),
        }
        shipments_rows = list_delivery_shipments(oid)
        return tpl(
            request,
            "home_organizations_shipments.html",
            variant="user",
            active="home_shipments",
            inside_organization=True,
            selected_organization=selected,
            selected_organization_id=oid,
            organizations_for_filters=[selected],
            shipments=shipments_rows,
            shipment_daily_journal=_shipment_daily_journal(shipments_rows),
            shipment_totals=shipment_totals(oid),
            courier_debts=list_courier_debts(oid, include_zero=True),
            hr_employees=list_hr_employees(oid),
            today=datetime.now().strftime("%Y-%m-%d"),
            flash_ok=request.query_params.get("msg"),
            flash_err=request.query_params.get("error"),
        )

    @app.post("/shipments/create", name="home_shipments_create")
    async def home_shipments_create(request: Request):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/shipments?err=csrf", status_code=302)
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        shipment_date = str(form.get("shipment_date") or "").strip()
        try:
            create_delivery_shipments(
                oid,
                _shipment_rows_from_form(form),
                shipment_date,
                employee_workspace_owner_id=oid,
            )
        except ValueError as exc:
            return RedirectResponse(url=f"/shipments?error={quote(str(exc) or 'shipments')}", status_code=302)
        return RedirectResponse(url="/shipments?msg=shipments_created", status_code=302)

    @app.get("/shipments/day/{day}", response_class=HTMLResponse, name="home_shipments_day")
    def home_shipments_day(request: Request, day: str):
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        u = request.session.get("user") or {}
        try:
            recompute_delivery_debts(oid)
        except Exception:
            logger.exception("[upos] recompute_delivery_debts failed before organization shipments day view")
        selected = {
            "id": oid,
            "name": str(u.get("active_organization_name") or u.get("workspace_owner_name") or "Организация"),
        }
        shipments_rows, courier_summary, document_status = _shipment_day_summary(list_delivery_shipments(oid, limit=2000), day)
        return tpl(
            request,
            "home_shipments_day.html",
            variant="user",
            active="home_shipments",
            inside_organization=True,
            selected_organization=selected,
            selected_organization_id=oid,
            organizations_for_filters=[selected],
            day=str(day or "")[:10],
            shipments=shipments_rows,
            courier_summary=courier_summary,
            document_status=document_status,
        )

    @app.post("/shipments/day/{day}/update", name="home_shipments_day_update")
    async def home_shipments_day_update(request: Request, day: str):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/shipments?err=csrf", status_code=302)
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        clean_day = str(day or "").strip()[:10]
        return_shipment_id = str(form.get("return_shipment_id") or "").strip()
        return_hash = f"#shipment-{quote(return_shipment_id)}" if return_shipment_id else f"#day-{quote(clean_day)}"
        try:
            ok = update_delivery_shipment_document(
                oid,
                clean_day,
                _shipment_rows_from_form(form),
                employee_workspace_owner_id=oid,
            )
        except ValueError as exc:
            return RedirectResponse(url=f"/shipments?error={quote(str(exc) or 'shipments')}{return_hash}", status_code=302)
        msg = "shipment_updated" if ok else "shipment_not_found"
        return RedirectResponse(url=f"/shipments?msg={quote(msg)}{return_hash}", status_code=302)

    @app.post("/shipments/day/{day}/confirm", name="home_shipments_day_confirm")
    async def home_shipments_day_confirm(request: Request, day: str):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/shipments?err=csrf", status_code=302)
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        clean_day = str(day or "").strip()[:10]
        ok = confirm_delivery_shipment_document(oid, clean_day)
        msg = "shipment_confirmed" if ok else "shipment_not_found"
        return RedirectResponse(url=f"/shipments?msg={quote(msg)}#day-{quote(clean_day)}", status_code=302)

    @app.post("/shipments/day/{day}/delete", name="home_shipments_day_delete")
    async def home_shipments_day_delete(request: Request, day: str):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/shipments?err=csrf", status_code=302)
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        clean_day = str(day or "").strip()[:10]
        ok = delete_delivery_shipment_document(oid, clean_day)
        msg = "shipment_deleted" if ok else "shipment_not_found"
        return RedirectResponse(url=f"/shipments?msg={quote(msg)}", status_code=302)

    @app.post("/shipments/shipment/{shipment_id}/delete", name="home_shipment_delete")
    async def home_shipment_delete(request: Request, shipment_id: str):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/shipments?err=csrf", status_code=302)
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        deleted_day = delete_delivery_shipment(oid, shipment_id)
        msg = "shipment_deleted" if deleted_day else "shipment_not_found"
        suffix = f"#day-{quote(deleted_day)}" if deleted_day else ""
        return RedirectResponse(url=f"/shipments?msg={quote(msg)}{suffix}", status_code=302)

    @app.post("/shipments/shipment/{shipment_id}/confirm", name="home_shipment_confirm")
    async def home_shipment_confirm(request: Request, shipment_id: str):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/shipments?err=csrf", status_code=302)
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        confirmed_day = confirm_delivery_shipment(oid, shipment_id)
        msg = "shipment_confirmed" if confirmed_day else "shipment_not_found"
        return RedirectResponse(url=f"/shipments?msg={quote(msg)}#shipment-{quote(str(shipment_id or '').strip())}", status_code=302)

    @app.get("/hr", response_class=HTMLResponse, name="home_hr")
    def home_hr(request: Request):
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        selected_date = str(request.query_params.get("date") or datetime.now().strftime("%Y-%m-%d"))[:10]
        return tpl(
            request,
            "home_organizations_hr.html",
            variant="user",
            active="home_hr",
            inside_organization=True,
            employees=_hr_employees_with_adjustments(oid, selected_date),
            positions=list_positions(oid),
            selected_organization_id=oid,
            selected_date=selected_date,
            flash_ok=request.query_params.get("msg"),
            flash_err=request.query_params.get("error"),
        )

    @app.get("/hr/salary-act/{employee_id}", response_class=HTMLResponse, name="home_hr_salary_act_detail")
    def home_hr_salary_act_detail(request: Request, employee_id: str):
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        selected_date = str(request.query_params.get("date") or datetime.now().strftime("%Y-%m-%d"))[:10]
        try:
            act = _salary_employee_act_payload(oid, selected_date, employee_id)
        except ValueError as exc:
            raise StarletteHTTPException(status_code=404, detail=str(exc)) from exc
        return tpl(
            request,
            "home_hr_salary_act_detail.html",
            variant="user",
            active="home_hr",
            inside_organization=True,
            selected_organization_id=oid,
            selected_date=selected_date,
            act=act,
        )

    @app.post("/hr/employees/create", name="home_hr_employee_create")
    async def home_hr_employee_create(
        request: Request,
        csrf_token: str = Form(default=""),
        first_name: str = Form(default=""),
        last_name: str = Form(default=""),
        position_id: str = Form(default=""),
        position: str = Form(default=""),
        passport_series: str = Form(default=""),
        passport_number: str = Form(default=""),
        monthly_salary: str = Form(default="0"),
        is_courier: str = Form(default=""),
        hired_at: str = Form(default=""),
        photo: UploadFile | None = File(default=None),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/hr?err=csrf", status_code=302)
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        try:
            photo_path = await _save_hr_photo(oid, photo)
            create_hr_employee(
                oid,
                {
                    "first_name": first_name,
                    "last_name": last_name,
                    "position_id": position_id,
                    "position": position,
                    "passport_series": passport_series,
                    "passport_number": passport_number,
                    "monthly_salary": monthly_salary,
                    "is_courier": is_courier,
                    "hired_at": hired_at,
                    "photo_path": photo_path,
                },
            )
        except ValueError as exc:
            return RedirectResponse(url=f"/hr?error={quote(str(exc) or 'hr')}", status_code=302)
        except Exception as exc:
            logger.exception("[upos] organization HR employee create failed")
            return RedirectResponse(url=f"/hr?error={quote(type(exc).__name__)}", status_code=302)
        return RedirectResponse(url="/hr?msg=employee_created", status_code=302)

    @app.post("/hr/employees/update", name="home_hr_employee_update")
    async def home_hr_employee_update(
        request: Request,
        csrf_token: str = Form(default=""),
        employee_id: str = Form(default=""),
        first_name: str = Form(default=""),
        last_name: str = Form(default=""),
        position_id: str = Form(default=""),
        position: str = Form(default=""),
        passport_series: str = Form(default=""),
        passport_number: str = Form(default=""),
        monthly_salary: str = Form(default="0"),
        is_courier: str = Form(default=""),
        hired_at: str = Form(default=""),
        photo: UploadFile | None = File(default=None),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/hr?err=csrf", status_code=302)
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        try:
            photo_path = await _save_hr_photo(oid, photo)
            updated = update_hr_employee(
                oid,
                employee_id,
                {
                    "first_name": first_name,
                    "last_name": last_name,
                    "position_id": position_id,
                    "position": position,
                    "passport_series": passport_series,
                    "passport_number": passport_number,
                    "monthly_salary": monthly_salary,
                    "is_courier": is_courier,
                    "hired_at": hired_at,
                    "photo_path": photo_path,
                },
            )
        except ValueError as exc:
            return RedirectResponse(url=f"/hr?error={quote(str(exc) or 'hr')}", status_code=302)
        except Exception as exc:
            logger.exception("[upos] organization HR employee update failed")
            return RedirectResponse(url=f"/hr?error={quote(type(exc).__name__)}", status_code=302)
        return RedirectResponse(url=f"/hr?msg={'employee_updated' if updated else 'not_found'}", status_code=302)

    @app.post("/hr/employees/{employee_id}/dismiss", name="home_hr_employee_dismiss")
    def home_hr_employee_dismiss(
        request: Request,
        employee_id: str,
        csrf_token: str = Form(default=""),
        dismissed_at: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/hr?err=csrf", status_code=302)
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        ok = dismiss_hr_employee(oid, employee_id, dismissed_at)
        return RedirectResponse(url=f"/hr?msg={'employee_dismissed' if ok else 'not_found'}", status_code=302)

    @app.post("/hr/employees/{employee_id}/restore", name="home_hr_employee_restore")
    def home_hr_employee_restore(
        request: Request,
        employee_id: str,
        csrf_token: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/hr?err=csrf#staff", status_code=302)
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        ok = restore_hr_employee(oid, employee_id)
        return RedirectResponse(url=f"/hr?msg={'employee_restored' if ok else 'not_found'}#staff", status_code=302)

    @app.post("/hr/employees/{employee_id}/delete", name="home_hr_employee_delete")
    def home_hr_employee_delete(
        request: Request,
        employee_id: str,
        csrf_token: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/hr?err=csrf#staff", status_code=302)
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        ok = delete_hr_employee_permanently(oid, employee_id)
        if ok:
            _purge_hr_salary_records(oid, employee_id)
        return RedirectResponse(url=f"/hr?msg={'employee_deleted' if ok else 'not_found'}#staff", status_code=302)

    @app.post("/hr/attendance", name="home_hr_attendance")
    def home_hr_attendance(
        request: Request,
        csrf_token: str = Form(default=""),
        employee_id: str = Form(default=""),
        work_date: str = Form(default=""),
        status: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/hr?err=csrf", status_code=302)
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        try:
            ok = set_hr_attendance(oid, employee_id, work_date, status)
        except ValueError as exc:
            return RedirectResponse(url=f"/hr?date={quote(work_date)}&error={quote(str(exc))}", status_code=302)
        return RedirectResponse(
            url=f"/hr?date={quote(work_date)}&msg={'attendance_saved' if ok else 'not_found'}",
            status_code=302,
        )

    @app.post("/hr/attendance/report", name="home_hr_attendance_report")
    async def home_hr_attendance_report(request: Request):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        work_date = str(form.get("work_date") or "").strip()
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/hr?err=csrf#calendar", status_code=302)
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        try:
            report = set_hr_attendance_day(oid, work_date, _hr_attendance_records_from_form(form))
        except ValueError as exc:
            return RedirectResponse(url=f"/hr?date={quote(work_date)}&error={quote(str(exc))}#calendar", status_code=302)
        telegram_status = "attendance_report_saved"
        try:
            from upos.telegram_notifier import send_hr_attendance_report

            tg = send_hr_attendance_report(oid, report)
            if tg.get("ok"):
                telegram_status = "attendance_report_sent"
            elif tg.get("error"):
                telegram_status = f"telegram_{tg.get('error')}"
        except Exception:
            logger.exception("[upos] HR attendance Telegram report failed")
            telegram_status = "telegram_failed"
        return RedirectResponse(url=f"/hr?date={quote(work_date)}&msg={quote(telegram_status)}#calendar", status_code=302)

    @app.post("/hr/salary-act/save", name="home_hr_salary_act_save")
    async def home_hr_salary_act_save(request: Request):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        work_date = str(form.get("work_date") or "").strip() or datetime.now().strftime("%Y-%m-%d")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/hr?err=csrf#salary", status_code=302)
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        try:
            _save_hr_salary_act(oid, work_date, list(form.getlist("employee_id")))
        except ValueError as exc:
            return RedirectResponse(url=f"/hr?date={quote(work_date)}&error={quote(str(exc))}#salary", status_code=302)
        return RedirectResponse(url=f"/hr?date={quote(work_date)}&msg=salary_act_saved#salary", status_code=302)

    @app.post("/hr/salary-adjustment/save", name="home_hr_salary_adjustment_save")
    async def home_hr_salary_adjustment_save(request: Request):
        form = await request.form()
        csrf_token = str(form.get("csrf_token") or "")
        work_date = str(form.get("work_date") or "").strip() or datetime.now().strftime("%Y-%m-%d")
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/hr?err=csrf#staff", status_code=302)
        oid, redir = _current_org_html_owner(request)
        if redir:
            return redir
        assert oid is not None
        try:
            _save_hr_salary_adjustment(
                oid,
                work_date,
                str(form.get("employee_id") or ""),
                str(form.get("adjustment_type") or ""),
                str(form.get("amount") or ""),
                str(form.get("comment") or ""),
            )
        except ValueError as exc:
            return RedirectResponse(url=f"/hr?date={quote(work_date)}&error={quote(str(exc))}#staff", status_code=302)
        return RedirectResponse(url=f"/hr?date={quote(work_date)}&msg=salary_adjustment_saved#staff", status_code=302)

    @app.get("/organizations/settings", response_class=HTMLResponse, name="organizations_settings_get")
    def organizations_settings_get(request: Request):
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        u = _refresh_org_session(request, org_scope="general")
        active_org_id = _employee_active_org_id(owner_id, u)
        loc = resolve_locale(request, request.session.get("user") or {})
        tab_raw = (request.query_params.get("tab") or "").strip().lower()
        if tab_raw:
            wid = _settings_storage_owner_id(u)
            if valid_workspace_owner_id(wid):
                data = load_workspace_settings(wid)
                _ensure_category_defaults_once(wid)
                categories = list_categories(wid)
            else:
                data = load_workspace_settings("")
                categories = []
            settings_can_manage_employees = _can_manage_employees(u)
            settings_can_manage_config = bool((not u.get("is_employee")) or _has_permission(u, "settings"))
            settings_can_manage_dictionary = bool(settings_can_manage_config or _has_permission(u, "dictionary"))
            settings_can_manage_roles = bool(
                settings_can_manage_config and (_is_director(u) or _is_employee_adminish(u)),
            )
            data["crm_activity"] = _crm_activity_settings(data)
            crm_settings_stage_lines = _crm_stage_lines(_crm_clean_stages(data.get("crm_pipeline_stages")))
            allowed_tabs = set()
            default_allowed_tab = "general"
            if settings_can_manage_config:
                allowed_tabs.add("general")
            if settings_can_manage_employees:
                allowed_tabs.add("employees")
                if default_allowed_tab not in allowed_tabs:
                    default_allowed_tab = "employees"
            if settings_can_manage_config:
                allowed_tabs.update({"telegram", "integrations", "crm"})
            if settings_can_manage_dictionary:
                allowed_tabs.add("dictionary")
                if default_allowed_tab not in allowed_tabs:
                    default_allowed_tab = "dictionary"
            if settings_can_manage_roles:
                allowed_tabs.add("roles")
            settings_default_tab = tab_raw if tab_raw in allowed_tabs else default_allowed_tab
            cur_tz = normalize_workspace_timezone(str(data.get("timezone") or ""))
            tz_extra = cur_tz if cur_tz not in curated_zone_ids() else None
            orgs_for_settings = list_organizations(owner_id)
            requested_telegram_org_id = str(request.query_params.get("organization_id") or "").strip()
            telegram_selected_org_id = (
                requested_telegram_org_id
                if any(str(org.get("id")) == requested_telegram_org_id for org in orgs_for_settings)
                else (active_org_id or str((orgs_for_settings[0] or {}).get("id") or "") if orgs_for_settings else "")
            )
            return tpl(
                request,
                "home_organizations_settings_panel.html",
                variant="user",
                active="organizations_settings",
                settings=data,
                categories=categories,
                settings_default_tab=settings_default_tab,
                timezone_groups=localized_timezone_groups(loc),
                settings_timezone_selected=cur_tz,
                settings_timezone_extra=tz_extra,
                settings_profile=_settings_owner_profile_sheet(request),
                settings_general_scope=True,
                integration_providers=INTEGRATION_PROVIDERS,
                settings_can_manage_settings=settings_can_manage_config,
                settings_can_manage_dictionary=settings_can_manage_dictionary,
                settings_show_advanced=settings_can_manage_config,
                settings_can_manage_employees=settings_can_manage_employees,
                settings_can_manage_roles=settings_can_manage_roles,
                crm_pipeline_stage_lines=crm_settings_stage_lines,
                staff_role_labels={k: translate(loc, f"staff.role.{k}") for k in STAFF_ROLE_LABELS},
                flash_ok=request.query_params.get("msg"),
                flash_err=request.query_params.get("error"),
                **_employee_management_context(owner_id, active_org_id),
                role_permission_labels=ROLE_PERMISSION_LABELS,
                role_button_permission_labels=ROLE_BUTTON_PERMISSION_LABELS,
                telegram_can_manage=_can_manage_telegram(u),
                telegram_can_approve=_is_director(u),
                telegram_organizations=orgs_for_settings,
                telegram_selected_organization_id=telegram_selected_org_id,
                settings_tabs_base_url=str(request.url_for("organizations_settings_get")),
                settings_tabs_home_url=str(request.url_for("organizations_settings_get")),
            )
        return tpl(
            request,
            "home_organizations_settings.html",
            variant="user",
            active="organizations_settings",
            settings_can_manage_employees=True,
            role_permission_labels=ROLE_PERMISSION_LABELS,
            role_button_permission_labels=ROLE_BUTTON_PERMISSION_LABELS,
            staff_role_labels={k: translate(loc, f"staff.role.{k}") for k in STAFF_ROLE_LABELS},
            flash_ok=request.query_params.get("msg"),
            flash_err=request.query_params.get("error"),
            **_employee_management_context(owner_id, active_org_id),
        )

    @app.get("/organizations/users", response_class=HTMLResponse, name="organizations_users_get")
    def organizations_users_get(request: Request):
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        u = _refresh_org_session(request, org_scope="general")
        active_org_id = _employee_active_org_id(owner_id, u)
        loc = resolve_locale(request, request.session.get("user") or {})
        categories: list[dict[str, Any]] = []
        if valid_workspace_owner_id(active_org_id):
            _ensure_category_defaults_once(active_org_id)
            categories = list_categories(active_org_id)
        return tpl(
            request,
            "home_organizations_users.html",
            variant="user",
            active="organizations_users",
            settings_can_manage_employees=True,
            role_permission_labels=ROLE_PERMISSION_LABELS,
            role_button_permission_labels=ROLE_BUTTON_PERMISSION_LABELS,
            staff_role_labels={k: translate(loc, f"staff.role.{k}") for k in STAFF_ROLE_LABELS},
            flash_ok=request.query_params.get("msg"),
            flash_err=request.query_params.get("error"),
            categories=categories,
            **_employee_management_context(owner_id, active_org_id),
        )

    @app.get("/organizations/kassa", response_class=HTMLResponse, name="organizations_kassa_get")
    def organizations_kassa(request: Request):
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        _refresh_org_session(request, org_scope="general")
        orgs = list_organizations(owner_id)
        default_org = next((o for o in orgs if o.get("is_default")), orgs[0] if orgs else None)
        tz = normalize_workspace_timezone("")
        if default_org:
            tz = normalize_workspace_timezone(
                str(load_workspace_settings(str(default_org["id"])).get("timezone") or ""),
            )
        return tpl(
            request,
            "home_organizations_kassa.html",
            variant="user",
            active="organizations_kassa",
            workspace_timezone=tz,
            organizations_for_filters=orgs,
            flash_err=request.query_params.get("error"),
        )

    @app.get("/organizations/open-kassa")
    def organizations_open_kassa(
        request: Request,
        organization_id: str,
        tx: str = "",
        tx_type: str = Query(default="", alias="type"),
    ):
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        oid = organization_id.strip()
        org = get_organization(owner_id, oid)
        if not org:
            loc = resolve_locale(request, request.session.get("user") or {})
            return RedirectResponse(
                url=f"/organizations/kassa?error={quote(translate(loc, 'general.org_not_found'), safe='')}",
                status_code=302,
            )
        tx_clean = tx.strip()
        if tx_clean:
            try:
                uuid.UUID(tx_clean)
            except ValueError:
                tx_clean = ""
        tx_type_clean = tx_type.strip().lower()
        if tx_type_clean not in {"income", "expense", "transfer"}:
            tx_type_clean = ""
        _refresh_org_session(request, org_scope="organization", active_org=org)
        params: dict[str, str] = {}
        if tx_clean:
            params["tx"] = tx_clean
        if tx_type_clean:
            params["type"] = tx_type_clean
        dest = f"/kassa?{urlencode(params)}" if params else "/kassa"
        return RedirectResponse(url=dest, status_code=302)

    @app.get("/organizations/schet", response_class=HTMLResponse, name="organizations_schet_get")
    def organizations_schet(request: Request):
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        _refresh_org_session(request, org_scope="general")
        orgs = list_organizations(owner_id)
        return tpl(
            request,
            "home_organizations_schet.html",
            variant="user",
            active="organizations_schet",
            organizations_for_filters=orgs,
        )

    @app.get("/organizations/reports", response_class=HTMLResponse, name="organizations_reports_get")
    def organizations_reports(request: Request):
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        _refresh_org_session(request, org_scope="general")
        orgs = list_organizations(owner_id)
        default_org = next((o for o in orgs if o.get("is_default")), orgs[0] if orgs else None)
        tz = normalize_workspace_timezone("")
        if default_org:
            tz = normalize_workspace_timezone(
                str(load_workspace_settings(str(default_org["id"])).get("timezone") or ""),
            )
        return tpl(
            request,
            "home_organizations_reports.html",
            variant="user",
            active="organizations_reports",
            workspace_timezone=tz,
            organizations_for_filters=orgs,
        )

    @app.get("/organizations/open-schet")
    def organizations_open_schet(request: Request, organization_id: str):
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        oid = organization_id.strip()
        org = get_organization(owner_id, oid)
        if not org:
            loc = resolve_locale(request, request.session.get("user") or {})
            return RedirectResponse(
                url=f"/organizations/schet?error={quote(translate(loc, 'general.org_not_found'), safe='')}",
                status_code=302,
            )
        _refresh_org_session(request, org_scope="organization", active_org=org)
        return RedirectResponse(url="/schet", status_code=302)

    @app.post("/organizations", name="organizations_create")
    def organizations_create(
        request: Request,
        csrf_token: str = Form(default=""),
        organization_name: str = Form(default=""),
        organization_note: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations?err=csrf", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        ok, msg_err, _org = create_organization(owner_id, organization_name, organization_note)
        _refresh_org_session(request, org_scope="general")
        if not ok:
            loc = resolve_locale(request, request.session.get("user") or {})
            return RedirectResponse(
                url=f"/organizations?error={quote(_localized_organization_store_error(loc, msg_err), safe='')}",
                status_code=302,
            )
        return RedirectResponse(url="/organizations?msg=created", status_code=302)

    @app.post("/organizations/{organization_id}/update", name="organizations_update")
    def organizations_update(
        request: Request,
        organization_id: str,
        csrf_token: str = Form(default=""),
        edit_organization_name: str = Form(default=""),
        edit_organization_note: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations?err=csrf", status_code=302)
        owner_id, redir = _director_owner_or_redirect(request)
        if redir:
            return redir
        assert owner_id is not None
        ok, msg_err, org = update_organization(
            owner_id,
            organization_id,
            name=edit_organization_name,
            note=edit_organization_note,
        )
        current = request.session.get("user") or {}
        active_org = org if org and str(current.get("active_org_id") or "") == organization_id else None
        _refresh_org_session(request, org_scope="general", active_org=active_org)
        if not ok:
            loc = resolve_locale(request, request.session.get("user") or {})
            return RedirectResponse(
                url=f"/organizations?error={quote(_localized_organization_store_error(loc, msg_err), safe='')}",
                status_code=302,
            )
        return RedirectResponse(url="/organizations?msg=updated", status_code=302)

    @app.post("/organizations/switch", name="organizations_switch")
    def organizations_switch(
        request: Request,
        csrf_token: str = Form(default=""),
        organization_id: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/organizations?err=csrf", status_code=302)
        u = request.session.get("user") or {}
        if u.get("role") == "admin":
            return RedirectResponse(url="/admin", status_code=302)
        owner_id = str(u.get("account_owner_id") or u.get("user_id") or "").strip()
        if not valid_workspace_owner_id(owner_id):
            return RedirectResponse(url="/auth", status_code=302)
        org_id = organization_id.strip()
        if u.get("is_employee") and not _is_employee_general_director(u):
            allowed_orgs = list_employee_organizations_safe(owner_id, str(u.get("user_id") or ""))
            org = next((item for item in allowed_orgs if str(item.get("id")) == org_id), None)
            if not org:
                _refresh_org_session(request, org_scope="organization")
                return RedirectResponse(url="/schet", status_code=302)
            _refresh_org_session(request, org_scope="organization", active_org=org)
            return RedirectResponse(url="/schet", status_code=302)
        if not (_is_director(u) or _is_employee_general_director(u)):
            return RedirectResponse(url="/schet", status_code=302)
        if org_id == ORG_GENERAL_VALUE:
            _refresh_org_session(request, org_scope="general")
            return RedirectResponse(url="/organizations", status_code=302)
        org = get_organization(owner_id, org_id)
        if not org:
            _refresh_org_session(request, org_scope="general")
            loc = resolve_locale(request, request.session.get("user") or {})
            return RedirectResponse(
                url=f"/organizations?error={quote(translate(loc, 'general.org_not_found'), safe='')}",
                status_code=302,
            )
        _refresh_org_session(request, org_scope="organization", active_org=org)
        return RedirectResponse(url="/schet", status_code=302)

    def _employees_return_base(request: Request, return_to_settings: str = "") -> str:
        marker = str(return_to_settings or "").strip().lower()
        user = request.session.get("user") or {}
        if marker in {"organizations", "general", "organizations_users"}:
            return "/organizations/users"
        if user.get("org_scope") == "general":
            return "/organizations/users"
        return "/settings?tab=employees"

    def _employees_return_url(request: Request, return_to_settings: str = "", **params: str) -> str:
        base = _employees_return_base(request, return_to_settings)
        pairs = [(key, str(value)) for key, value in params.items() if value]
        if not pairs:
            return base
        sep = "&" if "?" in base else "?"
        return base + sep + "&".join(f"{quote(key)}={quote(value)}" for key, value in pairs)

    def _settings_currency_codes(raw: object) -> list[str]:
        source = raw if isinstance(raw, list) else []
        out: list[str] = []
        seen: set[str] = set()
        for item in source:
            code = str(item or "").strip().upper()
            if len(code) != 3 or not code.isalpha() or code in seen:
                continue
            seen.add(code)
            out.append(code)
        return out

    def _apply_settings_currency_payload(
        data: dict[str, object],
        *,
        available_raw: object | None = None,
        enabled_raw: object | None = None,
    ) -> None:
        current_available = _settings_currency_codes(data.get("available_currencies"))
        current_enabled = _settings_currency_codes(data.get("enabled_currencies"))
        next_available = current_available
        if available_raw is not None:
            requested_available = _settings_currency_codes(available_raw)
            next_available = _settings_currency_codes(
                current_available + requested_available + current_enabled
            )
        if enabled_raw is not None:
            requested_enabled = _settings_currency_codes(enabled_raw)
            if not next_available:
                next_available = _settings_currency_codes(current_available + requested_enabled)
            available_set = set(next_available)
            next_enabled = [code for code in requested_enabled if code in available_set]
            if not next_enabled and next_available:
                next_enabled = [next_available[0]]
            data["enabled_currencies"] = next_enabled
        if next_available:
            data["available_currencies"] = next_available

    @app.get("/employees", response_class=HTMLResponse)
    def employees_get(request: Request):
        return RedirectResponse(
            url=_employees_return_url(
                request,
                msg=str(request.query_params.get("msg") or ""),
                error=str(request.query_params.get("error") or ""),
            ),
            status_code=302,
        )

    @app.post("/employees/create")
    async def employees_create(
        request: Request,
        csrf_token: str = Form(default=""),
        new_username: str = Form(...),
        new_password: str = Form(...),
        new_email: str = Form(default=""),
        new_name: str = Form(default=""),
        new_position: str = Form(default=""),
        new_staff_role: str = Form(default="viewer"),
        new_employee_role_id: str = Form(default=""),
        create_personal_wallet: str = Form(default=""),
        return_to_settings: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url=_employees_return_url(request, return_to_settings, error="csrf"), status_code=302)
        oid, redir = _employees_owner_or_redirect(request)
        if redir:
            return redir
        assert oid is not None
        sess = request.session.get("user") or {}
        active_org_id = _employee_active_org_id(oid, sess)
        form = await request.form()
        selected_org_ids = [str(x) for x in form.getlist("new_organization_ids") if str(x or "").strip()]
        selected_account_ids = [str(x) for x in form.getlist("new_account_ids") if str(x or "").strip()]
        ok, msg_err = add_employee(
            oid,
            new_username,
            new_password,
            new_email,
            new_name,
            new_position,
            new_staff_role,
            organization_id=active_org_id,
            employee_role_id=new_employee_role_id,
            organization_ids=selected_org_ids,
            account_ids=selected_account_ids,
        )
        if not ok:
            return RedirectResponse(
                url=_employees_return_url(request, return_to_settings, error=msg_err),
                status_code=302,
            )
        if create_personal_wallet:
            emp = next(
                (
                    row
                    for row in list_employees_safe(oid, "")
                    if str(row.get("username") or "").lower() == new_username.strip().lower()
                ),
                None,
            )
            if emp:
                wallet_org_id = selected_org_ids[0] if selected_org_ids else active_org_id
                treasury = load_treasury(wallet_org_id)
                pockets = list(treasury.get("pockets") or [])
                wallet_id = str(uuid.uuid4())
                label_name = str(emp.get("name") or emp.get("username") or new_username).strip()
                pockets.append(
                    {
                        "id": wallet_id,
                        "template_id": "cash",
                        "label": f"{label_name} - личный счёт",
                        "note": "",
                        "icon": "cash",
                        "owner_employee_id": str(emp.get("id") or ""),
                        "entries": [],
                    },
                )
                clean, _msg = validate_and_clean_treasury(
                    {
                        "version": 2,
                        "display_currency": treasury.get("display_currency") or "USD",
                        "pockets": pockets,
                    },
                )
                if clean is not None:
                    try:
                        save_treasury(wallet_org_id, clean)
                        update_employee(
                            oid,
                            new_username,
                            organization_id=wallet_org_id,
                            new_username=new_username,
                            new_password="",
                            new_email=new_email,
                            new_name=new_name,
                            position=new_position,
                            staff_role=new_staff_role,
                            employee_role_id=new_employee_role_id,
                            organization_ids=selected_org_ids or [wallet_org_id],
                            account_ids=[*selected_account_ids, wallet_id],
                        )
                    except ValueError:
                        logger.warning("[employees] personal wallet auto-create failed", exc_info=True)
        return RedirectResponse(url=_employees_return_url(request, return_to_settings, msg="created"), status_code=302)

    @app.post("/employees/update")
    async def employees_update(
        request: Request,
        csrf_token: str = Form(default=""),
        edit_old_username: str = Form(...),
        edit_username: str = Form(...),
        edit_password: str = Form(default=""),
        edit_email: str = Form(default=""),
        edit_name: str = Form(default=""),
        edit_position: str = Form(default=""),
        edit_staff_role: str = Form(default="viewer"),
        edit_employee_role_id: str = Form(default=""),
        return_to_settings: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url=_employees_return_url(request, return_to_settings, error="csrf"), status_code=302)
        oid, redir = _employees_owner_or_redirect(request)
        if redir:
            return redir
        assert oid is not None
        cur = request.session.get("user") or {}
        active_org_id = _employee_active_org_id(oid, cur)
        form = await request.form()
        selected_org_ids = [str(x) for x in form.getlist("edit_organization_ids") if str(x or "").strip()]
        selected_account_ids = [str(x) for x in form.getlist("edit_account_ids") if str(x or "").strip()]
        ok, msg_err, sess = update_employee(
            oid,
            edit_old_username,
            organization_id=active_org_id,
            new_username=edit_username,
            new_password=edit_password,
            new_email=edit_email,
            new_name=edit_name,
            position=edit_position,
            staff_role=edit_staff_role,
            employee_role_id=edit_employee_role_id,
            organization_ids=selected_org_ids,
            account_ids=selected_account_ids,
        )
        if not ok:
            return RedirectResponse(
                url=_employees_return_url(request, return_to_settings, error=msg_err),
                status_code=302,
            )
        if sess and str(cur.get("username", "")).lower() == edit_old_username.strip().lower():
            request.session["user"] = sess
        return RedirectResponse(url=_employees_return_url(request, return_to_settings, msg="updated"), status_code=302)

    @app.post("/employees/delete")
    def employees_delete(
        request: Request,
        csrf_token: str = Form(default=""),
        del_username: str = Form(...),
        return_to_settings: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url=_employees_return_url(request, return_to_settings, error="csrf"), status_code=302)
        oid, redir = _employees_owner_or_redirect(request)
        if redir:
            return redir
        assert oid is not None
        cur = request.session.get("user") or {}
        active_org_id = _employee_active_org_id(oid, cur)
        if str(cur.get("username", "")).lower() == del_username.strip().lower():
            loc_e = resolve_locale(request, cur)
            return RedirectResponse(
                url=_employees_return_url(request, return_to_settings, error=translate(loc_e, 'employees.cant_delete_self')),
                status_code=302,
            )
        ok, msg_err = delete_employee(oid, del_username, active_org_id)
        if not ok:
            return RedirectResponse(
                url=_employees_return_url(request, return_to_settings, error=msg_err),
                status_code=302,
            )
        return RedirectResponse(url=_employees_return_url(request, return_to_settings, msg="deleted"), status_code=302)

    @app.post("/employees/freeze")
    def employees_freeze(
        request: Request,
        csrf_token: str = Form(default=""),
        employee_username: str = Form(...),
        freeze_action: str = Form(default="freeze"),
        return_to_settings: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url=_employees_return_url(request, return_to_settings, error="csrf"), status_code=302)
        oid, redir = _employees_owner_or_redirect(request)
        if redir:
            return redir
        assert oid is not None
        cur = request.session.get("user") or {}
        if str(cur.get("username", "")).lower() == employee_username.strip().lower():
            loc_e = resolve_locale(request, cur)
            return RedirectResponse(
                url=_employees_return_url(request, return_to_settings, error=translate(loc_e, 'employees.cant_delete_self')),
                status_code=302,
            )
        frozen = freeze_action.strip().lower() != "unfreeze"
        ok, msg_err = set_employee_frozen(oid, employee_username, frozen)
        if not ok:
            return RedirectResponse(
                url=_employees_return_url(request, return_to_settings, error=msg_err),
                status_code=302,
            )
        return RedirectResponse(
            url=_employees_return_url(request, return_to_settings, msg=("frozen" if frozen else "unfrozen")),
            status_code=302,
        )

    @app.get("/settings", response_class=HTMLResponse, name="settings_get")
    def settings_get(request: Request):
        u = request.session.get("user") or {}
        settings_general_scope = bool(_is_director(u) and u.get("org_scope") == "general")
        if settings_general_scope:
            u = _refresh_org_session(request, org_scope="general")
        wid = _settings_storage_owner_id(u)
        if valid_workspace_owner_id(wid):
            data = load_workspace_settings(wid)
            if settings_general_scope:
                categories = []
            else:
                _ensure_category_defaults_once(wid)
                categories = list_categories(wid)
        else:
            data = load_workspace_settings("")
            categories = []
        social_links = data.get("social_links") if isinstance(data.get("social_links"), dict) else {}
        data["social_links"] = {
            "primary_channel": str(social_links.get("primary_channel") or "").strip(),
            "instagram_url": str(social_links.get("instagram_url") or "").strip(),
            "instagram_business_id": str(social_links.get("instagram_business_id") or "").strip(),
            "instagram_access_token": str(social_links.get("instagram_access_token") or "").strip(),
            "instagram_login": str(social_links.get("instagram_login") or "").strip(),
            "instagram_password": str(social_links.get("instagram_password") or "").strip(),
            "facebook_url": str(social_links.get("facebook_url") or "").strip(),
            "facebook_page_id": str(social_links.get("facebook_page_id") or "").strip(),
            "facebook_access_token": str(social_links.get("facebook_access_token") or "").strip(),
            "telegram_url": str(social_links.get("telegram_url") or "").strip(),
            "whatsapp_phone": str(social_links.get("whatsapp_phone") or "").strip(),
            "whatsapp_business_id": str(social_links.get("whatsapp_business_id") or "").strip(),
            "whatsapp_access_token": str(social_links.get("whatsapp_access_token") or "").strip(),
            "youtube_url": str(social_links.get("youtube_url") or "").strip(),
            "tiktok_url": str(social_links.get("tiktok_url") or "").strip(),
            "website_url": str(social_links.get("website_url") or "").strip(),
            "site_chat_enabled": str(social_links.get("site_chat_enabled") or "").strip(),
            "site_chat_widget_id": str(social_links.get("site_chat_widget_id") or "").strip(),
            "site_chat_greeting": str(social_links.get("site_chat_greeting") or "").strip(),
            "note": str(social_links.get("note") or "").strip(),
        }
        usd_rate = _sales_decimal(data.get("usd_uzs_rate"))
        data["usd_uzs_rate"] = _decimal_plain_text(usd_rate if usd_rate > 0 else PRODUCT_USD_RATE)
        data["crm_activity"] = _crm_activity_settings(data)
        crm_settings_stage_lines = _crm_stage_lines(_crm_clean_stages(data.get("crm_pipeline_stages")))
        tab_raw = (request.query_params.get("tab") or "").strip().lower()
        settings_can_manage_employees = _can_manage_employees(u)
        settings_can_manage_config = bool((not u.get("is_employee")) or _has_permission(u, "settings"))
        settings_can_manage_dictionary = bool(settings_can_manage_config or _has_permission(u, "dictionary"))
        settings_can_manage_roles = bool(
            settings_can_manage_config and (_is_director(u) or _is_employee_adminish(u)),
        )
        allowed_tabs = set()
        default_allowed_tab = "general"
        if settings_can_manage_config:
            allowed_tabs.add("general")
        if settings_can_manage_employees:
            allowed_tabs.add("employees")
            if default_allowed_tab not in allowed_tabs:
                default_allowed_tab = "employees"
        if settings_can_manage_config:
            allowed_tabs.update({"telegram", "integrations", "social", "crm"})
        if settings_can_manage_dictionary:
            allowed_tabs.add("dictionary")
            if default_allowed_tab not in allowed_tabs:
                default_allowed_tab = "dictionary"
        if settings_can_manage_roles:
            allowed_tabs.add("roles")
        settings_default_tab = tab_raw if tab_raw in allowed_tabs else ""
        cur_tz = normalize_workspace_timezone(str(data.get("timezone") or ""))
        tz_extra = cur_tz if cur_tz not in curated_zone_ids() else None
        loc = resolve_locale(request, u)
        owner_id = str(u.get("account_owner_id") or u.get("user_id") or "").strip()
        active_org_id = _employee_active_org_id(owner_id, u) if settings_can_manage_employees else ""
        orgs_for_settings = list_organizations(owner_id) if valid_workspace_owner_id(owner_id) else []
        requested_telegram_org_id = str(request.query_params.get("organization_id") or "").strip()
        telegram_selected_org_id = (
            requested_telegram_org_id
            if any(str(org.get("id")) == requested_telegram_org_id for org in orgs_for_settings)
            else (active_org_id or str((orgs_for_settings[0] or {}).get("id") or "") if orgs_for_settings else "")
        )
        employee_ctx = (
            _employee_management_context(owner_id, active_org_id)
            if settings_can_manage_employees and valid_workspace_owner_id(owner_id)
            else {
                "employees": [],
                "organizations_for_access": [],
                "access_accounts": [],
                "employee_roles": [],
                "employee_presets": [],
            }
        )
        return tpl(
            request,
            "settings.html",
            variant="user",
            active="settings",
            settings=data,
            categories=categories,
            settings_default_tab=settings_default_tab,
            timezone_groups=localized_timezone_groups(loc),
            settings_timezone_selected=cur_tz,
            settings_timezone_extra=tz_extra,
            settings_profile=_settings_owner_profile_sheet(request),
            settings_general_scope=settings_general_scope,
            integration_providers=INTEGRATION_PROVIDERS,
            settings_can_manage_settings=settings_can_manage_config,
            settings_can_manage_dictionary=settings_can_manage_dictionary,
            settings_show_advanced=settings_can_manage_config,
            settings_can_manage_employees=settings_can_manage_employees,
            settings_can_manage_roles=settings_can_manage_roles,
            crm_pipeline_stage_lines=crm_settings_stage_lines,
            staff_role_labels={k: translate(loc, f"staff.role.{k}") for k in STAFF_ROLE_LABELS},
            flash_ok=request.query_params.get("msg"),
            flash_err=request.query_params.get("error"),
            **employee_ctx,
            role_permission_labels=ROLE_PERMISSION_LABELS,
            role_button_permission_labels=ROLE_BUTTON_PERMISSION_LABELS,
            telegram_can_manage=_can_manage_telegram(u),
            telegram_can_approve=_is_director(u),
            telegram_organizations=orgs_for_settings,
            telegram_selected_organization_id=telegram_selected_org_id,
        )

    from upos.telegram_routes import register_telegram_routes

    register_telegram_routes(
        app,
        treasury_workspace_owner=_treasury_workspace_owner,
        is_director=_is_director,
        can_manage_telegram=_can_manage_telegram,
    )

    @app.post("/settings", name="settings_post")
    async def settings_post(
        request: Request,
        csrf_token: str = Form(default=""),
        upos_tg_bot_api_token: str = Form(default=""),
        theme: str = Form(default="light"),
        onec_base_url: str = Form(default=""),
        onec_username: str = Form(default=""),
        onec_password: str = Form(default=""),
        yespos_api_base_url: str = Form(default=""),
        yespos_api_key: str = Form(default=""),
        ibox_api_url: str = Form(default=""),
        ibox_api_key: str = Form(default=""),
        ibox_terminal_id: str = Form(default=""),
        clopos_api_base_url: str = Form(default=""),
        clopos_client_id: str = Form(default=""),
        clopos_client_secret: str = Form(default=""),
        clopos_brand: str = Form(default=""),
        clopos_integrator_id: str = Form(default=""),
        clopos_venue_id: str = Form(default=""),
        greenwhite_base_url: str = Form(default=""),
        greenwhite_username: str = Form(default=""),
        greenwhite_password: str = Form(default=""),
        greenwhite_project_code: str = Form(default="trade"),
        greenwhite_filial_id: str = Form(default=""),
        greenwhite_filial_code: str = Form(default=""),
        greenwhite_sync_days: str = Form(default="7"),
        greenwhite_sync_enabled: str = Form(default=""),
        active_settings_tab: str = Form(default="general"),
        workspace_timezone: str = Form(default=""),
        locale: str = Form(default="ru"),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/settings?err=csrf", status_code=302)
        u = request.session.get("user") or {}
        settings_general_scope = bool(_is_director(u) and u.get("org_scope") == "general")
        settings_can_manage_employees = _can_manage_employees(u)
        settings_can_manage_config = bool((not u.get("is_employee")) or _has_permission(u, "settings"))
        settings_can_manage_roles = bool(
            settings_can_manage_config and (_is_director(u) or _is_employee_adminish(u)),
        )
        if not settings_can_manage_config:
            return RedirectResponse(url="/settings?tab=employees", status_code=302)
        wid = _settings_storage_owner_id(u)
        try:
            data = load_workspace_settings(wid)
        except Exception:
            logger.exception("[upos] /settings POST load_workspace_settings wid=%s", wid)
            return RedirectResponse(url="/settings?err=settings_load", status_code=302)
        # Telegram bot tokens are verified and saved through /api/telegram/verify.
        # A normal settings save must not overwrite a connected bot with an empty form value.
        if theme in {"light", "dark", "emerald"}:
            data["theme"] = theme
        data["locale"] = normalize_locale(locale)
        tz_in = (workspace_timezone or "").strip()
        if tz_in:
            data["timezone"] = normalize_workspace_timezone(tz_in)
        if not settings_general_scope:
            data["integrations"]["onec"] = {
                "base_url": onec_base_url.strip(),
                "username": onec_username.strip(),
                "password": onec_password.strip(),
            }
            data["integrations"]["yespos"] = {
                "api_base_url": yespos_api_base_url.strip(),
                "api_key": yespos_api_key.strip(),
            }
            data["integrations"]["ibox"] = {
                "api_url": ibox_api_url.strip(),
                "api_key": ibox_api_key.strip(),
                "terminal_id": ibox_terminal_id.strip(),
            }
            prev_clopos = data["integrations"].get("clopos") or {}
            data["integrations"]["clopos"] = {
                "api_base_url": clopos_api_base_url.strip() or CLOPOS_DEFAULT_API_BASE_URL,
                "client_id": clopos_client_id.strip(),
                "client_secret": clopos_client_secret.strip(),
                "brand": clopos_brand.strip(),
                "integrator_id": clopos_integrator_id.strip(),
                "venue_id": clopos_venue_id.strip(),
                "token": str(prev_clopos.get("token") or ""),
                "expires_at": str(prev_clopos.get("expires_at") or ""),
            }
            prev_greenwhite = data["integrations"].get("greenwhite") or {}
            gw_filial_id = greenwhite_filial_id.strip()
            try:
                gw_sync_days = max(1, min(int(greenwhite_sync_days or "7"), 7))
            except ValueError:
                gw_sync_days = 7
            try:
                gw_export_timeout = max(5, min(int(prev_greenwhite.get("export_timeout") or 25), 60))
            except (TypeError, ValueError):
                gw_export_timeout = 25
            data["integrations"]["greenwhite"] = {
                "base_url": greenwhite_base_url.strip(),
                "username": greenwhite_username.strip(),
                "password": greenwhite_password.strip(),
                "project_code": greenwhite_project_code.strip() or "trade",
                "filial_id": gw_filial_id,
                "filial_code": greenwhite_filial_code.strip(),
                "sync_days": gw_sync_days,
                "export_timeout": gw_export_timeout,
                "token": "",
                "organization_id": gw_filial_id,
                "sync_enabled": greenwhite_sync_enabled == "1",
                "last_sync_at": str(prev_greenwhite.get("last_sync_at") or ""),
            }

        # Currencies
        form_data = await request.form()
        available_ccy = form_data.getlist("available_currencies")
        enabled_ccy = form_data.getlist("enabled_currencies")
        if available_ccy or enabled_ccy:
            _apply_settings_currency_payload(
                data,
                available_raw=available_ccy if available_ccy else None,
                enabled_raw=enabled_ccy if enabled_ccy else None,
            )

        save_workspace_settings(wid, data)
        if _is_director(u):
            owner_id = str(u.get("account_owner_id") or u.get("user_id") or "").strip()
            if valid_workspace_owner_id(owner_id):
                sync_common_settings(owner_id, data)
        allowed_tabs = {"general"}
        if settings_can_manage_employees:
            allowed_tabs.add("employees")
        if settings_can_manage_config:
            allowed_tabs.update({"telegram", "integrations", "social", "dictionary", "crm"})
        if settings_can_manage_roles:
            allowed_tabs.add("roles")
        tab = active_settings_tab.strip().lower() if active_settings_tab.strip().lower() in allowed_tabs else "general"
        resp = RedirectResponse(url=f"/settings?saved=1&tab={tab}", status_code=302)
        apply_locale_cookie(resp, str(data.get("locale") or "ru"))
        return resp

    @app.post("/settings/crm", name="settings_crm_post")
    async def settings_crm_post(request: Request):
        form = await request.form()
        if not csrf_matches_session(request, str(form.get("csrf_token") or "")):
            return RedirectResponse(url="/settings?err=csrf&tab=crm", status_code=302)
        u = request.session.get("user") or {}
        settings_can_manage_config = bool((not u.get("is_employee")) or _has_permission(u, "settings"))
        if not settings_can_manage_config:
            return RedirectResponse(url="/settings", status_code=302)
        wid = _settings_storage_owner_id(u)
        data = load_workspace_settings(wid)
        try:
            yellow_hours = max(1, min(int(str(form.get("crm_yellow_hours") or "24")), 720))
        except (TypeError, ValueError):
            yellow_hours = 24
        try:
            red_hours = max(yellow_hours + 1, min(int(str(form.get("crm_red_hours") or "48")), 1440))
        except (TypeError, ValueError):
            red_hours = 48
        data["crm_activity"] = {"yellow_hours": yellow_hours, "red_hours": red_hours}
        raw_lines = str(form.get("crm_pipeline_stages") or "").splitlines()
        data["crm_pipeline_stages"] = _crm_clean_stages(raw_lines)
        save_workspace_settings(wid, data)
        return_to = str(form.get("return_to") or "").strip()
        if not (return_to.startswith("/settings") or return_to.startswith("/organizations/settings")):
            return_to = "/settings"
        sep = "&" if "?" in return_to else "?"
        return RedirectResponse(url=f"{return_to}{sep}saved=1&tab=crm", status_code=302)

    @app.post("/settings/account", name="settings_account_post")
    def settings_account_post(
        request: Request,
        csrf_token: str = Form(default=""),
        profile_username: str = Form(default=""),
        profile_email: str = Form(default=""),
        profile_name: str = Form(default=""),
        profile_current_password: str = Form(default=""),
        profile_new_password: str = Form(default=""),
        profile_new_password_confirm: str = Form(default=""),
    ):
        if not csrf_matches_session(request, csrf_token):
            return RedirectResponse(url="/settings?err=csrf&tab=general", status_code=302)
        u = request.session.get("user") or {}
        loc = resolve_locale(request, u)
        if _settings_owner_profile_sheet(request) is None:
            return RedirectResponse(url="/settings?tab=general", status_code=302)
        sess_username = str(u.get("username") or "").strip()
        if not verify_login(sess_username, profile_current_password):
            return RedirectResponse(
                url=(
                    "/settings?tab=general&profile_error="
                    + quote(translate(loc, "settings.profile.err_bad_current"))
                ),
                status_code=302,
            )
        npw = profile_new_password.strip()
        npw_c = profile_new_password_confirm.strip()
        if npw or npw_c:
            if npw != npw_c:
                return RedirectResponse(
                    url=(
                        "/settings?tab=general&profile_error="
                        + quote(translate(loc, "settings.profile.err_pw_mismatch"))
                    ),
                    status_code=302,
                )
        else:
            npw = ""

        ok, err, new_sess = update_user(
            sess_username,
            new_username=profile_username.strip(),
            new_email=profile_email,
            new_name=profile_name.strip(),
            new_password=npw,
            new_role=None,
        )
        if not ok or not new_sess:
            msg = (err or "").strip() or translate(loc, "settings.profile.err_update")
            return RedirectResponse(
                url="/settings?tab=general&profile_error=" + quote(msg),
                status_code=302,
            )
        request.session["user"] = _attach_organization_context(new_sess, u)
        return RedirectResponse(url="/settings?tab=general&profile_saved=1", status_code=302)

    def _workspace_settings_owner_id(
        request: Request,
        *,
        allow_general: bool = False,
    ) -> tuple[str | None, JSONResponse | None]:
        u = request.session.get("user") or {}
        if u.get("role") == "admin":
            return None, JSONResponse({"error": "forbidden"}, status_code=403)
        if u.get("org_scope") == "general" and not allow_general:
            return None, JSONResponse({"error": "organization_required"}, status_code=409)
        wid = _settings_storage_owner_id(u)
        if not valid_workspace_owner_id(wid):
            return None, JSONResponse({"error": "workspace"}, status_code=400)
        return wid, None

    def _role_permissions_owner_id(request: Request) -> tuple[str | None, JSONResponse | None]:
        u = request.session.get("user") or {}
        if u.get("role") == "admin":
            return None, JSONResponse({"error": "forbidden"}, status_code=403)
        if not (_is_director(u) or (_is_employee_adminish(u) and _has_permission(u, "settings"))):
            return None, JSONResponse({"error": "forbidden"}, status_code=403)
        owner_id = str(u.get("account_owner_id") or u.get("user_id") or "").strip()
        if u.get("org_scope") == "general":
            wid = _employee_active_org_id(owner_id, u)
        else:
            wid = _settings_storage_owner_id(u)
        if not valid_workspace_owner_id(wid):
            return None, JSONResponse({"error": "workspace"}, status_code=400)
        return wid, None

    def _save_workspace_settings_from_user(request: Request, data: dict[str, Any]) -> None:
        u = request.session.get("user") or {}
        wid = _settings_storage_owner_id(u)
        save_workspace_settings(wid, data)
        if _is_director(u):
            owner_id = str(u.get("account_owner_id") or u.get("user_id") or "").strip()
            if valid_workspace_owner_id(owner_id):
                sync_common_settings(owner_id, data)

    @app.post("/api/settings/preferences")
    async def api_settings_preferences(request: Request):
        token = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, token):
            return JSONResponse({"error": "csrf"}, status_code=403)
        wid, err = _workspace_settings_owner_id(request, allow_general=True)
        if err:
            return err
        assert wid is not None
        try:
            body = await request.json()
        except Exception:
            body = {}
        if not isinstance(body, dict):
            body = {}
        data = load_workspace_settings(wid)
        changed_locale = False
        if "theme" in body:
            theme_val = str(body.get("theme") or "").strip()
            if theme_val in {"light", "dark", "emerald"}:
                data["theme"] = theme_val
        if "locale" in body:
            loc_val = normalize_locale(str(body.get("locale") or ""))
            if loc_val != normalize_locale(str(data.get("locale") or "")):
                changed_locale = True
            data["locale"] = loc_val
        if "timezone" in body:
            tz_in = str(body.get("timezone") or "").strip()
            if tz_in:
                data["timezone"] = normalize_workspace_timezone(tz_in)
        if "available_currencies" in body or "enabled_currencies" in body:
            raw_available = body.get("available_currencies") if "available_currencies" in body else None
            raw_enabled = body.get("enabled_currencies") if "enabled_currencies" in body else None
            _apply_settings_currency_payload(
                data,
                available_raw=raw_available if isinstance(raw_available, list) else None,
                enabled_raw=raw_enabled if isinstance(raw_enabled, list) else None,
            )
        if "usd_uzs_rate" in body:
            rate = _sales_decimal(body.get("usd_uzs_rate"))
            if rate <= 0:
                return JSONResponse({"error": "usd_uzs_rate"}, status_code=400)
            data["usd_uzs_rate"] = _decimal_plain_text(rate)
        _save_workspace_settings_from_user(request, data)
        out = {
            "ok": True,
            "theme": data.get("theme"),
            "locale": data.get("locale"),
            "timezone": data.get("timezone"),
            "available_currencies": data.get("available_currencies"),
            "enabled_currencies": data.get("enabled_currencies"),
            "usd_uzs_rate": data.get("usd_uzs_rate"),
            "reload": changed_locale,
        }
        resp = JSONResponse(out)
        if changed_locale:
            apply_locale_cookie(resp, str(data.get("locale") or "ru"))
        return resp

    def _integration_configured(key: str, block: dict[str, object]) -> bool:
        return integration_configured(key, block)

    def _apply_integration_connection(
        wid: str, key: str, block: dict[str, object], loc: str
    ) -> dict[str, object]:
        checked_at = datetime.now(timezone.utc).isoformat()
        ok = False
        message = translate(loc, "settings.integrations.not_configured")
        if key == "greenwhite":
            if not _integration_configured(key, block):
                block["connection_ok"] = False
                block["connection_message"] = message
                block["connection_checked_at"] = checked_at
                return {"key": key, "ok": False, "message": message}
            try:
                test_greenwhite_connection(wid)
                ok = True
                message = translate(loc, "settings.integrations.connected")
            except GreenWhiteError as exc:
                message = str(exc).strip() or translate(loc, "settings.integrations.not_connected")
        elif key == "clopos":
            if not _integration_configured(key, block):
                block["connection_ok"] = False
                block["connection_message"] = message
                block["connection_checked_at"] = checked_at
                return {"key": key, "ok": False, "message": message}
            try:
                payload = test_clopos_connection(block)
                ok = True
                message = translate(loc, "settings.integrations.connected")
                block["token"] = str(payload.get("token") or "")
                block["expires_at"] = str(payload.get("expires_at") or "")
            except CloposError as exc:
                block["token"] = ""
                block["expires_at"] = ""
                message = str(exc).strip() or translate(loc, "settings.integrations.not_connected")
        elif _integration_configured(key, block):
            ok = True
            message = translate(loc, "settings.integrations.connected")
        block["connection_ok"] = ok
        block["connection_message"] = message
        block["connection_checked_at"] = checked_at
        return {"key": key, "ok": ok, "message": message}

    @app.post("/api/settings/integrations")
    async def api_settings_integrations(request: Request):
        token = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, token):
            return JSONResponse({"error": "csrf"}, status_code=403)
        wid, err = _role_permissions_owner_id(request)
        if err:
            return err
        assert wid is not None
        u = request.session.get("user") or {}
        loc = resolve_locale(request, u)
        try:
            body = await request.json()
        except Exception:
            body = {}
        if not isinstance(body, dict):
            body = {}
        data = load_workspace_settings(wid)
        integrations = body.get("integrations") if isinstance(body.get("integrations"), dict) else {}
        prev_greenwhite = data.get("integrations", {}).get("greenwhite") or {}
        updated_keys: list[str] = []
        if isinstance(integrations.get("onec"), dict):
            onec = integrations["onec"]
            prev = data.get("integrations", {}).get("onec") or {}
            data["integrations"]["onec"] = {
                "base_url": str(onec.get("base_url") or "").strip(),
                "username": str(onec.get("username") or "").strip(),
                "password": str(onec.get("password") or "").strip(),
            }
            updated_keys.append("onec")
        if isinstance(integrations.get("yespos"), dict):
            yp = integrations["yespos"]
            data["integrations"]["yespos"] = {
                "api_base_url": str(yp.get("api_base_url") or "").strip(),
                "api_key": str(yp.get("api_key") or "").strip(),
            }
            updated_keys.append("yespos")
        if isinstance(integrations.get("ibox"), dict):
            ibox = integrations["ibox"]
            data["integrations"]["ibox"] = {
                "api_url": str(ibox.get("api_url") or "").strip(),
                "api_key": str(ibox.get("api_key") or "").strip(),
                "terminal_id": str(ibox.get("terminal_id") or "").strip(),
            }
            updated_keys.append("ibox")
        if isinstance(integrations.get("clopos"), dict):
            clopos = integrations["clopos"]
            prev = data.get("integrations", {}).get("clopos") or {}
            data["integrations"]["clopos"] = {
                "api_base_url": str(clopos.get("api_base_url") or "").strip() or CLOPOS_DEFAULT_API_BASE_URL,
                "client_id": str(clopos.get("client_id") or "").strip(),
                "client_secret": str(clopos.get("client_secret") or "").strip(),
                "brand": str(clopos.get("brand") or "").strip(),
                "integrator_id": str(clopos.get("integrator_id") or "").strip(),
                "venue_id": str(clopos.get("venue_id") or "").strip(),
                "token": str(prev.get("token") or ""),
                "expires_at": str(prev.get("expires_at") or ""),
            }
            updated_keys.append("clopos")
        if isinstance(integrations.get("greenwhite"), dict):
            gw = integrations["greenwhite"]
            gw_filial_id = str(gw.get("filial_id") or "").strip()
            try:
                gw_sync_days = max(1, min(int(gw.get("sync_days") or 7), 7))
            except (TypeError, ValueError):
                gw_sync_days = 7
            try:
                gw_export_timeout = max(5, min(int(prev_greenwhite.get("export_timeout") or 25), 60))
            except (TypeError, ValueError):
                gw_export_timeout = 25
            data["integrations"]["greenwhite"] = {
                "base_url": str(gw.get("base_url") or "").strip(),
                "username": str(gw.get("username") or "").strip(),
                "password": str(gw.get("password") or "").strip(),
                "project_code": str(gw.get("project_code") or "").strip() or "trade",
                "filial_id": gw_filial_id,
                "filial_code": str(gw.get("filial_code") or "").strip(),
                "sync_days": gw_sync_days,
                "export_timeout": gw_export_timeout,
                "token": "",
                "organization_id": gw_filial_id,
                "sync_enabled": bool(gw.get("sync_enabled")),
                "last_sync_at": str(prev_greenwhite.get("last_sync_at") or ""),
            }
            updated_keys.append("greenwhite")
        connection: dict[str, object] | None = None
        if updated_keys:
            key = updated_keys[-1]
            block = data.get("integrations", {}).get(key) or {}
            if isinstance(block, dict):
                connection = _apply_integration_connection(wid, key, block, loc)
        _save_workspace_settings_from_user(request, data)
        out: dict[str, object] = {"ok": True}
        if connection:
            out["connection"] = connection
        return out

    def _meta_graph_version() -> str:
        version = str(get_settings().meta_graph_version or "v20.0").strip().strip("/")
        return version if version.startswith("v") else f"v{version}"

    def _settings_social_redirect(message: str = "", error: str = "") -> RedirectResponse:
        params = {"tab": "social"}
        if message:
            params["msg"] = message
        if error:
            params["error"] = error
        return RedirectResponse(url="/settings?" + urlencode(params), status_code=302)

    def _meta_json_get(url: str, params: dict[str, str]) -> dict[str, Any]:
        full_url = url + ("&" if "?" in url else "?") + urlencode(params)
        req = UrlRequest(full_url, headers={"Accept": "application/json"})
        try:
            with urlopen(req, timeout=15) as response:
                payload = json.loads(response.read().decode("utf-8") or "{}")
        except HTTPError as exc:
            try:
                payload = json.loads(exc.read().decode("utf-8") or "{}")
                message = str((payload.get("error") or {}).get("message") or "")
            except Exception:
                message = str(exc)
            raise ValueError(message or "Meta вернул ошибку") from exc
        except (URLError, TimeoutError) as exc:
            raise ValueError("Meta сейчас недоступен. Попробуйте позже.") from exc
        if not isinstance(payload, dict):
            raise ValueError("Meta вернул неожиданный ответ")
        if isinstance(payload.get("error"), dict):
            raise ValueError(str(payload["error"].get("message") or "Meta вернул ошибку"))
        return payload

    def _meta_redirect_uri(request: Request) -> str:
        configured = str(get_settings().auth_url or "").strip().rstrip("/")
        if configured:
            return configured + "/settings/social/instagram/callback"
        return str(request.url_for("settings_instagram_callback"))

    def _meta_oauth_configured() -> tuple[str, str]:
        app_id = str(get_settings().meta_app_id or os.getenv("META_APP_ID") or os.getenv("FACEBOOK_APP_ID") or "").strip()
        app_secret = str(
            get_settings().meta_app_secret
            or os.getenv("META_APP_SECRET")
            or os.getenv("FACEBOOK_APP_SECRET")
            or ""
        ).strip()
        return app_id, app_secret

    @app.get("/settings/social/instagram/start", name="settings_instagram_start")
    def settings_instagram_start(request: Request):
        wid, err = _role_permissions_owner_id(request)
        if err:
            return _settings_social_redirect(error="Нет доступа к настройкам Instagram.")
        assert wid is not None
        app_id, app_secret = _meta_oauth_configured()
        if not app_id or not app_secret:
            return _settings_social_redirect(
                error="Для входа через Instagram укажите META_APP_ID и META_APP_SECRET в Railway Variables."
            )
        state = secrets.token_urlsafe(24)
        request.session["instagram_oauth_state"] = state
        request.session["instagram_oauth_workspace"] = wid
        scopes = [
            "pages_show_list",
            "pages_read_engagement",
            "pages_manage_metadata",
            "instagram_basic",
            "instagram_manage_messages",
        ]
        auth_url = f"https://www.facebook.com/{_meta_graph_version()}/dialog/oauth?" + urlencode(
            {
                "client_id": app_id,
                "redirect_uri": _meta_redirect_uri(request),
                "state": state,
                "response_type": "code",
                "scope": ",".join(scopes),
            }
        )
        return RedirectResponse(url=auth_url, status_code=302)

    @app.get("/settings/social/instagram/callback", name="settings_instagram_callback")
    def settings_instagram_callback(request: Request):
        wid, err = _role_permissions_owner_id(request)
        if err:
            return _settings_social_redirect(error="Нет доступа к настройкам Instagram.")
        assert wid is not None
        incoming_state = str(request.query_params.get("state") or "")
        expected_state = str(request.session.pop("instagram_oauth_state", "") or "")
        expected_wid = str(request.session.pop("instagram_oauth_workspace", "") or "")
        if not incoming_state or incoming_state != expected_state or expected_wid != wid:
            return _settings_social_redirect(error="Сессия входа Instagram устарела. Попробуйте еще раз.")
        if request.query_params.get("error"):
            return _settings_social_redirect(
                error=str(request.query_params.get("error_description") or request.query_params.get("error") or "")
            )
        code = str(request.query_params.get("code") or "").strip()
        if not code:
            return _settings_social_redirect(error="Meta не вернул код авторизации.")
        app_id, app_secret = _meta_oauth_configured()
        if not app_id or not app_secret:
            return _settings_social_redirect(error="META_APP_ID или META_APP_SECRET не настроены.")

        version = _meta_graph_version()
        graph_base = f"https://graph.facebook.com/{version}"
        try:
            token_payload = _meta_json_get(
                f"{graph_base}/oauth/access_token",
                {
                    "client_id": app_id,
                    "client_secret": app_secret,
                    "redirect_uri": _meta_redirect_uri(request),
                    "code": code,
                },
            )
            user_token = str(token_payload.get("access_token") or "").strip()
            if not user_token:
                return _settings_social_redirect(error="Meta не вернул access token.")
            pages_payload = _meta_json_get(
                f"{graph_base}/me/accounts",
                {
                    "access_token": user_token,
                    "fields": "id,name,access_token,instagram_business_account{id,username,profile_picture_url}",
                },
            )
        except ValueError as exc:
            return _settings_social_redirect(error=str(exc))

        selected_page = None
        selected_instagram = None
        for page in pages_payload.get("data") or []:
            if not isinstance(page, dict):
                continue
            ig_account = page.get("instagram_business_account")
            if isinstance(ig_account, dict) and ig_account.get("id"):
                selected_page = page
                selected_instagram = ig_account
                break
        data = load_workspace_settings(wid)
        social_links = data.get("social_links") if isinstance(data.get("social_links"), dict) else {}
        updated = dict(social_links)
        updated["primary_channel"] = updated.get("primary_channel") or "instagram"
        if selected_instagram:
            ig_id = str(selected_instagram.get("id") or "").strip()
            username = str(selected_instagram.get("username") or "").strip()
            updated["instagram_business_id"] = ig_id
            updated["instagram_login"] = ("@" + username.lstrip("@")) if username else updated.get("instagram_login", "")
            if username:
                updated["instagram_url"] = f"https://instagram.com/{username.lstrip('@')}"
        if selected_page:
            updated["instagram_access_token"] = str(selected_page.get("access_token") or user_token).strip()
            updated["facebook_page_id"] = str(selected_page.get("id") or "").strip()
            updated["facebook_access_token"] = str(selected_page.get("access_token") or "").strip()
        else:
            updated["instagram_access_token"] = user_token
        data["social_links"] = updated
        save_workspace_settings(wid, data)
        if _is_director(request.session.get("user") or {}):
            session_user = request.session.get("user") or {}
            owner_id = str(session_user.get("account_owner_id") or session_user.get("user_id") or "").strip()
            if valid_workspace_owner_id(owner_id) and owner_id != wid:
                sync_common_settings(owner_id, data)
        if selected_instagram:
            return _settings_social_redirect(message="instagram_connected")
        return _settings_social_redirect(
            message="instagram_token_saved",
            error="Токен сохранен, но Instagram Business аккаунт не найден. Проверьте связь Instagram с Facebook Page.",
        )

    @app.post("/api/settings/social-links")
    async def api_settings_social_links(request: Request):
        token = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, token):
            return JSONResponse({"error": "csrf"}, status_code=403)
        wid, err = _role_permissions_owner_id(request)
        if err:
            return err
        assert wid is not None
        try:
            body = await request.json()
        except Exception:
            body = {}
        if not isinstance(body, dict):
            body = {}
        raw_links = body.get("social_links") if isinstance(body.get("social_links"), dict) else {}
        data = load_workspace_settings(wid)
        data["social_links"] = {
            "primary_channel": str(raw_links.get("primary_channel") or "").strip(),
            "instagram_url": str(raw_links.get("instagram_url") or "").strip(),
            "instagram_business_id": str(raw_links.get("instagram_business_id") or "").strip(),
            "instagram_access_token": str(raw_links.get("instagram_access_token") or "").strip(),
            "instagram_login": str(raw_links.get("instagram_login") or "").strip(),
            "instagram_password": str(raw_links.get("instagram_password") or "").strip(),
            "facebook_url": str(raw_links.get("facebook_url") or "").strip(),
            "facebook_page_id": str(raw_links.get("facebook_page_id") or "").strip(),
            "facebook_access_token": str(raw_links.get("facebook_access_token") or "").strip(),
            "telegram_url": str(raw_links.get("telegram_url") or "").strip(),
            "whatsapp_phone": str(raw_links.get("whatsapp_phone") or "").strip(),
            "whatsapp_business_id": str(raw_links.get("whatsapp_business_id") or "").strip(),
            "whatsapp_access_token": str(raw_links.get("whatsapp_access_token") or "").strip(),
            "youtube_url": str(raw_links.get("youtube_url") or "").strip(),
            "tiktok_url": str(raw_links.get("tiktok_url") or "").strip(),
            "website_url": str(raw_links.get("website_url") or "").strip(),
            "site_chat_enabled": "1" if str(raw_links.get("site_chat_enabled") or "").strip() in {"1", "true", "on", "yes"} else "",
            "site_chat_widget_id": str(raw_links.get("site_chat_widget_id") or "").strip(),
            "site_chat_greeting": str(raw_links.get("site_chat_greeting") or "").strip(),
            "note": str(raw_links.get("note") or "").strip(),
        }
        _save_workspace_settings_from_user(request, data)
        return {"ok": True, "social_links": data["social_links"]}

    @app.post("/api/settings/account")
    async def api_settings_account(request: Request):
        from upos.users_store import email_valid

        token = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, token):
            return JSONResponse({"error": "csrf"}, status_code=403)
        u = request.session.get("user") or {}
        loc = resolve_locale(request, u)
        if _settings_owner_profile_sheet(request) is None:
            return JSONResponse({"error": "forbidden"}, status_code=403)
        try:
            body = await request.json()
        except Exception:
            body = {}
        if not isinstance(body, dict):
            body = {}
        sess_username = str(u.get("username") or "").strip()
        profile_username = str(body.get("username") or "").strip()
        profile_email = str(body.get("email") or "").strip()
        profile_name = str(body.get("name") or "").strip()
        profile_current_password = str(body.get("current_password") or "")
        profile_new_password = str(body.get("new_password") or "")
        profile_new_password_confirm = str(body.get("new_password_confirm") or "")

        if not profile_username:
            return JSONResponse(
                {"error": translate(loc, "settings.profile.err_login_required")},
                status_code=400,
            )
        if not email_valid(profile_email):
            return JSONResponse(
                {"error": translate(loc, "settings.profile.err_email_invalid")},
                status_code=400,
            )
        if not verify_login(sess_username, profile_current_password):
            return JSONResponse(
                {"error": translate(loc, "settings.profile.err_bad_current")},
                status_code=400,
            )
        npw = profile_new_password.strip()
        npw_c = profile_new_password_confirm.strip()
        if npw or npw_c:
            if len(npw) < 8:
                return JSONResponse(
                    {"error": translate(loc, "settings.profile.err_pw_short")},
                    status_code=400,
                )
            if npw != npw_c:
                return JSONResponse(
                    {"error": translate(loc, "settings.profile.err_pw_mismatch")},
                    status_code=400,
                )
        else:
            npw = ""

        ok, err, new_sess = update_user(
            sess_username,
            new_username=profile_username,
            new_email=profile_email,
            new_name=profile_name or profile_username,
            new_password=npw,
            new_role=None,
        )
        if not ok or not new_sess:
            msg = (err or "").strip() or translate(loc, "settings.profile.err_update")
            return JSONResponse({"error": msg}, status_code=400)
        request.session["user"] = _attach_organization_context(new_sess, u)
        return {
            "ok": True,
            "profile": {
                "username": profile_username,
                "email": profile_email,
                "name": profile_name or profile_username,
            },
        }

    @app.patch("/api/auth/me")
    async def api_auth_me_patch(request: Request):
        token = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, token):
            return JSONResponse({"error": "csrf"}, status_code=403)
        u = request.session.get("user") or {}
        uid = _session_user_id(u)
        if not uid:
            return JSONResponse({"error": "unauthorized"}, status_code=401)
        try:
            body = await request.json()
        except Exception:
            body = {}
        if not isinstance(body, dict):
            body = {}
        username = str(body.get("username") or "").strip()
        name = str(body.get("name") or "").strip()
        email = str(body.get("email") or "").strip() if "email" in body else None
        current_password = str(body.get("current_password") or "")
        new_password = str(body.get("new_password") or "")
        new_password_confirm = str(body.get("new_password_confirm") or "")
        ok, err, new_sess = update_self_account(
            uid,
            username=username,
            name=name or username,
            email=email,
            current_password=current_password,
            new_password=new_password,
            new_password_confirm=new_password_confirm,
        )
        if not ok or not new_sess:
            return JSONResponse({"error": err or "update_failed"}, status_code=400)
        request.session["user"] = _attach_organization_context(new_sess, u)
        return {
            "ok": True,
            "profile": {
                "username": str(new_sess.get("username") or username),
                "email": str(new_sess.get("email") or email or ""),
                "name": str(new_sess.get("name") or name or username),
            },
        }

    @app.post("/api/settings/roles")
    async def api_settings_roles(request: Request):
        token = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, token):
            return JSONResponse({"error": "csrf"}, status_code=403)
        u = request.session.get("user") or {}
        if not (_is_director(u) or (_is_employee_adminish(u) and _has_permission(u, "settings"))):
            return JSONResponse({"error": "forbidden"}, status_code=403)
        wid, err = _role_permissions_owner_id(request)
        if err:
            return err
        assert wid is not None
        try:
            body = await request.json()
        except Exception:
            body = {}
        roles_payload = body.get("roles") if isinstance(body, dict) and isinstance(body.get("roles"), dict) else {}
        roles = update_role_permissions(wid, roles_payload)
        return {"ok": True, "roles": roles}

    @app.get("/api/categories")
    def api_categories_get(request: Request):
        u = request.session.get("user") or {}
        wid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert wid is not None
        _ensure_category_defaults_once(wid)
        return {
            "ok": True,
            "categories": _filter_categories_for_user(u, list_categories(wid)),
            "restricted": _employee_category_access(u) is not None,
        }

    @app.get("/api/integrations/greenwhite/status")
    def api_greenwhite_status(request: Request):
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        return {"ok": True, "status": last_greenwhite_status(oid)}

    @app.post("/api/integrations/greenwhite/test")
    def api_greenwhite_test(request: Request):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        try:
            result = test_greenwhite_connection(oid)
        except GreenWhiteError as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        return {"ok": True, "result": result}

    @app.post("/api/integrations/greenwhite/sync")
    def api_greenwhite_sync(request: Request):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        oid, err = _treasury_workspace_owner(request)
        if err:
            return err
        assert oid is not None
        try:
            status = sync_greenwhite(oid)
        except GreenWhiteError as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        return {"ok": True, "status": status}

    @app.get("/admin", response_class=HTMLResponse)
    def admin_home(request: Request):
        return tpl(request, "admin_home.html", variant="admin", active="admin_home")

    @app.get("/admin/settings", response_class=HTMLResponse)
    def admin_settings(request: Request):
        return RedirectResponse(url="/admin", status_code=302)

    @app.post("/admin/settings")
    def admin_settings_post(
        request: Request,
        csrf_token: str = Form(default=""),
        locale: str = Form(default="ru"),
    ):
        return RedirectResponse(url="/admin", status_code=302)

    @app.get("/admin/users", response_class=HTMLResponse)
    def admin_users(request: Request):
        return RedirectResponse(url="/admin", status_code=302)

    @app.post("/admin/users/create")
    def admin_users_create(
        request: Request,
        csrf_token: str = Form(default=""),
        new_username: str = Form(...),
        new_password: str = Form(...),
        new_email: str = Form(...),
        new_name: str = Form(default=""),
        new_role: str = Form(default="user"),
    ):
        return RedirectResponse(url="/admin", status_code=302)

    @app.post("/admin/users/update")
    def admin_users_update(
        request: Request,
        csrf_token: str = Form(default=""),
        edit_old_username: str = Form(...),
        edit_username: str = Form(...),
        edit_password: str = Form(default=""),
        edit_email: str = Form(default=""),
        edit_name: str = Form(default=""),
        edit_role: str = Form(default="user"),
    ):
        return RedirectResponse(url="/admin", status_code=302)

    @app.post("/admin/users/delete")
    def admin_users_delete(
        request: Request,
        csrf_token: str = Form(default=""),
        del_username: str = Form(...),
    ):
        return RedirectResponse(url="/admin", status_code=302)

    @app.post("/api/categories/create")
    async def api_category_create(request: Request):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        u = request.session.get("user") or {}
        wid = str(u.get("workspace_owner_id") or u.get("user_id") or "")
        if not valid_workspace_owner_id(wid):
            return JSONResponse({"error": "unauthorized"}, status_code=401)
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "json"}, status_code=400)
        name = str(body.get("name", "")).strip()
        ctype = str(body.get("type", "expense")).strip()
        ctype = ctype if ctype in {"income", "expense", "transfer"} else "expense"
        if not name:
            return JSONResponse({"error": "name_required"}, status_code=400)
        cat_id = create_category(wid, name, ctype)
        locked = is_report_locked_category(name, ctype)
        return {
            "ok": True,
            "id": cat_id,
            "name": name,
            "type": ctype,
            "subcategories": [],
            "protected": locked,
            "protected_reason": "report" if locked else "",
        }

    @app.post("/api/categories/subcategories/add")
    async def api_category_subcategory_add(request: Request):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        u = request.session.get("user") or {}
        wid = str(u.get("workspace_owner_id") or u.get("user_id") or "")
        if not valid_workspace_owner_id(wid):
            return JSONResponse({"error": "unauthorized"}, status_code=401)
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "json"}, status_code=400)
        cat_id = str(body.get("id", "")).strip()
        name = str(body.get("name", "")).strip()
        try:
            category = add_category_subcategory(wid, cat_id, name)
        except ValueError as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        if category is None:
            return JSONResponse({"error": "not_found"}, status_code=404)
        return {"ok": True, "category": category}

    @app.post("/api/categories/subcategories/delete")
    async def api_category_subcategory_delete(request: Request):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        u = request.session.get("user") or {}
        wid = str(u.get("workspace_owner_id") or u.get("user_id") or "")
        if not valid_workspace_owner_id(wid):
            return JSONResponse({"error": "unauthorized"}, status_code=401)
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "json"}, status_code=400)
        cat_id = str(body.get("id", "")).strip()
        name = str(body.get("name", "")).strip()
        try:
            category = delete_category_subcategory(wid, cat_id, name)
        except ValueError as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        if category is None:
            return JSONResponse({"error": "not_found"}, status_code=404)
        return {"ok": True, "category": category}

    @app.post("/api/categories/update")
    async def api_category_update(request: Request):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        u = request.session.get("user") or {}
        wid = str(u.get("workspace_owner_id") or u.get("user_id") or "")
        if not valid_workspace_owner_id(wid):
            return JSONResponse({"error": "unauthorized"}, status_code=401)
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "json"}, status_code=400)
        cat_id = str(body.get("id", "")).strip()
        name = str(body.get("name", "")).strip()
        ctype = str(body.get("type", "expense")).strip()
        ctype = ctype if ctype in {"income", "expense", "transfer"} else "expense"
        if not cat_id:
            return JSONResponse({"error": "id_required"}, status_code=400)
        if not name:
            return JSONResponse({"error": "name_required"}, status_code=400)
        try:
            category = update_category(wid, cat_id, name, ctype)
        except ValueError as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        if category is None:
            return JSONResponse({"error": "not_found"}, status_code=404)
        return {"ok": True, "category": category}

    @app.post("/api/categories/delete")
    async def api_category_delete(request: Request):
        tok = request.headers.get("X-CSRF-Token") or request.headers.get("x-csrf-token") or ""
        if not csrf_matches_session(request, tok):
            return JSONResponse({"error": "csrf"}, status_code=403)
        u = request.session.get("user") or {}
        wid = str(u.get("workspace_owner_id") or u.get("user_id") or "")
        if not valid_workspace_owner_id(wid):
            return JSONResponse({"error": "unauthorized"}, status_code=401)
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "json"}, status_code=400)
        cat_id = str(body.get("id", "")).strip()
        if not cat_id:
            return JSONResponse({"error": "id_required"}, status_code=400)
        try:
            ok = delete_category(wid, cat_id)
        except ValueError as exc:
            return JSONResponse({"error": str(exc)}, status_code=400)
        return {"ok": ok}

    @app.post("/api/user/avatar")
    async def upload_avatar(request: Request, file: UploadFile = File(...)):
        u = request.session.get("user")
        if not u:
            return JSONResponse({"error": "unauthorized"}, status_code=401)
        uid = u.get("user_id")
        if not uid:
            return JSONResponse({"error": "invalid_session"}, status_code=401)
        try:
            content = await file.read()
            img = Image.open(io.BytesIO(content))
            # Crop to square
            w, h = img.size
            sz = min(w, h)
            left = (w - sz) // 2
            top = (h - sz) // 2
            right = left + sz
            bottom = top + sz
            img = img.crop((left, top, right, bottom))
            # Resize to 500x500
            img = img.resize((500, 500), Image.Resampling.LANCZOS)
            # Convert to PNG
            filename = f"{uid}_{uuid.uuid4().hex[:8]}.png"
            rel_path = f"avatars/{filename}"
            full_path = BASE_DIR / "static" / rel_path
            img.save(full_path, "PNG")
            # Update DB
            new_sess = save_user_avatar(uid, rel_path)
            if new_sess:
                request.session["user"] = _attach_organization_context(new_sess, u)
            return {"ok": True, "avatar_url": f"/static/{rel_path}"}
        except Exception as e:
            logger.exception("Avatar upload failed")
            return JSONResponse({"error": str(e)}, status_code=500)

    cookie_https_only = (
        os.environ.get("SESSION_HTTPS_ONLY", "").strip().lower()
        in ("1", "true", "yes")
        or os.environ.get("RAILWAY_ENVIRONMENT", "").strip().lower() == "production"
    )

    app.add_middleware(
        SessionMiddleware,
        secret_key=settings.auth_secret,
        session_cookie="upos_finance_session",
        max_age=60 * 60 * 24 * 14,
        same_site="lax",
        https_only=cookie_https_only,
    )

    return app


app = create_app()
